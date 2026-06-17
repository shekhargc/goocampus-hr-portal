"""
One-time import: Australia call notes from Excel into ops_call_notes.

Each Excel row goes into ops_call_notes with pathway='australia'.

Linking strategy:
  The Excel's "Candidate Name" cell contains both the candidate name AND
  the registration number, e.g.:
      " Dr. Yashika Bopaiah -GCAUSIP/2023/026"
      " Dr. Nikhil Paul Mathew -GCAUSIP/24-25/111"
  We parse out the GCAUSIP-prefixed reg number and store it in
  registration_number so the list page that JOINs against plab_clients
  surfaces the candidate name correctly.

  If the cell has no reg number, we fall back to a fuzzy name match
  against plab_clients WHERE pathway='australia' (lowercased trimmed
  prefix+first+last).

Idempotent: keyed on (registration_number, call_date, call_note). Re-running
with no marker change is a no-op. Bumping IMPORT_VERSION re-runs the upsert.
Bad rows (no reg, no name match) are skipped with a warning so one bad row
never blocks the whole import.

Performance: ~4,125 rows. We batch commits every 100 rows.

Hooked into app boot from app.py the same way as
import_australia_test_bookings.py.
"""

import logging
import os
import re
from datetime import datetime, date


EXCEL_FILENAME = 'All_Australia_Call_Notes.xlsx'
IMPORT_VERSION = 'au_call_notes_v1_initial_4124'

BATCH_SIZE = 100


H = {
    'cand_name': 'Candidate Name',
    'call_date': 'Call Date',
    'call_note': 'Call Notes',
    'added_by':  'Added User',
}


_REG_RE = re.compile(r'(GCAUSIP/\S+)', re.IGNORECASE)


def _extract_reg_number(candidate_cell):
    """Pull the GCAUSIP/... reg number out of the combined name+reg cell."""
    if not candidate_cell:
        return ''
    s = str(candidate_cell)
    m = _REG_RE.search(s)
    if not m:
        return ''
    raw = m.group(1).rstrip('.,;:-/ \t')
    return raw.upper().replace('GCAUSIP', 'GCAUSIP')


def _extract_name_only(candidate_cell):
    """Strip the reg-number tail and the leading 'Dr.'/'Mr.'/etc. prefix so
    we can fuzzy-match against plab_clients first_name + last_name.
    Returns a normalised lowercased string (or '')."""
    if not candidate_cell:
        return ''
    s = str(candidate_cell)
    # Remove the reg number portion (and trailing dash/space).
    s = _REG_RE.sub('', s)
    s = s.strip(' -\t')
    # Strip leading honorific.
    s = re.sub(r'^\s*(dr\.?|mr\.?|mrs\.?|ms\.?|miss)\s+', '', s, flags=re.IGNORECASE)
    # Collapse whitespace, lowercase.
    s = re.sub(r'\s+', ' ', s).strip().lower()
    return s


def _ss(v):
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _sd(v):
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def _build_name_index(conn):
    """Build {normalised_full_name: registration_number} for Australia clients
    so we can fuzzy-match the few rows whose cell lacks a GCAUSIP reg."""
    idx = {}
    for r in conn.execute(
        """SELECT registration_number, prefix, first_name, last_name
             FROM plab_clients WHERE pathway = 'australia'"""
    ).fetchall():
        first = (r['first_name'] or '').strip().lower()
        last = (r['last_name'] or '').strip().lower()
        if not first and not last:
            continue
        full = re.sub(r'\s+', ' ', f"{first} {last}").strip()
        if full and full not in idx:
            idx[full] = r['registration_number']
    return idx


def run_import_australia_call_notes_once(get_db_fn):
    """Import Australia call notes into ops_call_notes (pathway='australia')."""
    import openpyxl

    excel_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        EXCEL_FILENAME,
    )
    if not os.path.exists(excel_path):
        logging.info("Australia call-notes import: no Excel found, skipping")
        return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

    conn = get_db_fn()
    try:
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS _import_markers (key TEXT PRIMARY KEY, value TEXT)")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        marker = conn.execute(
            "SELECT value FROM _import_markers WHERE key = 'au_call_notes_import'"
        ).fetchone()
        if marker and marker['value'] == IMPORT_VERSION:
            logging.info(f"Australia call-notes import: already at {IMPORT_VERSION}, skipping")
            return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

        logging.info(f"Australia call-notes import: starting {IMPORT_VERSION}...")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h is not None}

        def cell(row, key):
            i = idx.get(H[key])
            return row[i] if i is not None and i < len(row) else None

        name_index = _build_name_index(conn)

        inserted = 0
        updated = 0
        skipped = 0
        errors = 0
        rows_since_commit = 0

        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row_cells):
                continue
            cand_cell = cell(row_cells, 'cand_name')
            reg_num = _extract_reg_number(cand_cell)
            if not reg_num:
                # Fuzzy name match fallback.
                fuzzy = _extract_name_only(cand_cell)
                if fuzzy and fuzzy in name_index:
                    reg_num = name_index[fuzzy]
            if not reg_num:
                skipped += 1
                logging.warning(
                    f"Australia call-notes: no reg/name match for {cand_cell!r}"
                )
                continue

            call_date = _sd(cell(row_cells, 'call_date'))
            call_note = _ss(cell(row_cells, 'call_note'))
            added_by = _ss(cell(row_cells, 'added_by'))

            if not call_note and not call_date:
                skipped += 1
                continue

            data = {
                'registration_number': reg_num,
                'call_date':           call_date,
                'call_note':           call_note,
                'added_by':            added_by,
                'pathway':             'australia',
            }

            try:
                existing = conn.execute(
                    "SELECT id FROM ops_call_notes "
                    "WHERE registration_number = ? AND call_date = ? "
                    "AND call_note = ? AND pathway = 'australia'",
                    (reg_num, call_date, call_note),
                ).fetchone()
                if existing:
                    sets = ', '.join(f"{k} = ?" for k in data.keys())
                    conn.execute(
                        f"UPDATE ops_call_notes SET {sets} WHERE id = ?",
                        list(data.values()) + [existing['id']],
                    )
                    updated += 1
                else:
                    cols = ', '.join(data.keys())
                    placeholders = ', '.join(['?'] * len(data))
                    conn.execute(
                        f"INSERT INTO ops_call_notes ({cols}) VALUES ({placeholders})",
                        list(data.values()),
                    )
                    inserted += 1
                rows_since_commit += 1
                if rows_since_commit >= BATCH_SIZE:
                    conn.commit()
                    rows_since_commit = 0
            except Exception as e:
                errors += 1
                logging.warning(f"Australia call-notes row error ({reg_num}): {e}")
                try:
                    conn.rollback()
                except Exception:
                    pass
                rows_since_commit = 0

        # Final commit for the tail.
        if rows_since_commit > 0:
            try:
                conn.commit()
            except Exception:
                pass

        # Marker
        try:
            conn.execute(
                "INSERT INTO _import_markers (key, value) VALUES ('au_call_notes_import', ?) "
                "ON CONFLICT (key) DO UPDATE SET value = excluded.value",
                (IMPORT_VERSION,),
            )
            conn.commit()
        except Exception:
            try:
                conn.execute("DELETE FROM _import_markers WHERE key = 'au_call_notes_import'")
                conn.execute(
                    "INSERT INTO _import_markers (key, value) VALUES ('au_call_notes_import', ?)",
                    (IMPORT_VERSION,),
                )
                conn.commit()
            except Exception as e:
                logging.warning(f"Australia call-notes: could not write marker: {e}")

        logging.info(
            f"Australia call-notes {IMPORT_VERSION} complete — "
            f"inserted={inserted} updated={updated} skipped={skipped} errors={errors}"
        )
        wb.close()
        return {'inserted': inserted, 'updated': updated, 'skipped': skipped, 'errors': errors}
    finally:
        try:
            conn.close()
        except Exception:
            pass
