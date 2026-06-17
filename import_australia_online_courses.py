"""
One-time import: Australia online courses & subscriptions from Excel
into ops_online_subscriptions (pathway='australia').

Each Excel row goes into ops_online_subscriptions with pathway='australia'.

Linking strategy:
  The Excel's "Candidate Name" cell contains both the candidate name
  AND the registration number, e.g.:
      " Dr. Ankitha Bhanu - GCAUSIP/25-26/08"
      " Dr. SALONI  MAHAJAN  - GCAUSIP/25-26/28"
  We parse out the GCAUSIP-prefixed reg number and store it in
  registration_number so existing list/edit pages that JOIN against
  plab_clients keep working unchanged.

Idempotent: keyed on (registration_number, online_subscription, issued_date).
Re-running with no marker change is a no-op. Bumping IMPORT_VERSION
re-runs the upsert. Bad rows (missing reg number) are skipped with a
warning so one bad row never blocks the whole import.

Hooked into app boot from app.py the same way as the other Australia
importers.
"""

import logging
import os
import re
from datetime import datetime, date


EXCEL_FILENAME = 'All_Australia_Online_Subscriptions.xlsx'
IMPORT_VERSION = 'au_online_courses_v1_initial_345'


H = {
    'cand_name':    'Candidate Name',
    'auto_number':  'Auto Number',
    'course_name':  'Course or Subscription Name',
    'booked_by':    'Booked By',
    'activation':   'Activation Type',
    'issued_date':  'Issued Date',
    'client_email': 'Client Email',
    'login_id':     'Login Id',
    'password':     'Password',
    'extra_notes':  'Notes',
}


_REG_RE = re.compile(r'(GCAUSIP/\S+)', re.IGNORECASE)


def _extract_reg_number(candidate_cell):
    """Pull the GCAUSIP/... reg number out of the combined name+reg cell.

    Returns '' if none found. The cell can have stray spaces and dashes
    around the reg number, so we use a generous regex and strip noise.
    """
    if not candidate_cell:
        return ''
    s = str(candidate_cell)
    m = _REG_RE.search(s)
    if not m:
        return ''
    raw = m.group(1).rstrip('.,;:-/ \t')
    return raw.upper().replace('GCAUSIP', 'GCAUSIP')


def _extract_candidate_name(candidate_cell):
    """Pull the human name out of the combined cell (strip Dr. + reg suffix)."""
    if not candidate_cell:
        return ''
    s = str(candidate_cell).strip()
    # Strip everything from the reg number onward
    m = _REG_RE.search(s)
    if m:
        s = s[:m.start()]
    # Drop trailing dash/whitespace
    s = s.rstrip(' -\t')
    # Drop leading "Dr." / "Dr "
    s = re.sub(r'^\s*Dr\.?\s+', '', s, flags=re.IGNORECASE)
    return ' '.join(s.split())  # collapse whitespace


def _fuzzy_match_reg(conn, candidate_cell):
    """Fallback: find a reg number in plab_clients by fuzzy name match.

    Returns the reg number or '' if no confident match. Uses
    LOWER(TRIM(prefix||first||last)) prefix-style comparison.
    """
    name = _extract_candidate_name(candidate_cell)
    if not name:
        return ''
    tokens = name.lower().split()
    if not tokens:
        return ''
    first = tokens[0]
    last = tokens[-1] if len(tokens) > 1 else ''
    try:
        rows = conn.execute(
            """SELECT registration_number, first_name, last_name
                 FROM plab_clients
                WHERE pathway = 'australia'
                  AND LOWER(TRIM(COALESCE(first_name,''))) = ?
                  AND LOWER(TRIM(COALESCE(last_name,''))) = ?""",
            (first, last),
        ).fetchall()
        if len(rows) == 1:
            return rows[0]['registration_number'] or ''
    except Exception:
        return ''
    return ''


def _ss(v):
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _sd(v):
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def run_import_australia_online_courses_once(get_db_fn):
    """Import Australia online courses into ops_online_subscriptions (pathway='australia')."""
    import openpyxl

    excel_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        EXCEL_FILENAME,
    )
    if not os.path.exists(excel_path):
        logging.info("Australia online-courses import: no Excel found, skipping")
        return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

    conn = get_db_fn()
    try:
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS _import_markers (key TEXT PRIMARY KEY, value TEXT)")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        marker = conn.execute(
            "SELECT value FROM _import_markers WHERE key = 'au_online_courses_import'"
        ).fetchone()
        if marker and marker['value'] == IMPORT_VERSION:
            logging.info(f"Australia online-courses import: already at {IMPORT_VERSION}, skipping")
            return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

        logging.info(f"Australia online-courses import: starting {IMPORT_VERSION}...")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h is not None}

        def cell(row, key):
            i = idx.get(H[key])
            return row[i] if i is not None and i < len(row) else None

        inserted = 0
        updated = 0
        skipped = 0
        errors = 0

        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row_cells):
                continue
            cand_cell = cell(row_cells, 'cand_name')
            reg_num = _extract_reg_number(cand_cell)
            if not reg_num:
                # Fallback to fuzzy name match against plab_clients
                reg_num = _fuzzy_match_reg(conn, cand_cell)
            if not reg_num:
                skipped += 1
                logging.warning(
                    f"Australia online-courses: no reg/name match for row "
                    f"candidate={cand_cell!r}"
                )
                continue

            course = _ss(cell(row_cells, 'course_name'))
            issued = _sd(cell(row_cells, 'issued_date'))
            if not course:
                skipped += 1
                continue

            # Compose notes: stash Login Id + Password + any extra notes
            login_id = _ss(cell(row_cells, 'login_id'))
            password = _ss(cell(row_cells, 'password'))
            extra_notes = _ss(cell(row_cells, 'extra_notes'))
            notes_bits = []
            if login_id:
                notes_bits.append(f"login: {login_id}")
            if password:
                notes_bits.append(f"password: {password}")
            if extra_notes:
                notes_bits.append(extra_notes)
            notes_val = ' | '.join(notes_bits)

            data = {
                'registration_number': reg_num,
                'online_subscription': course,
                'issued_date':         issued,
                'activation_type':     _ss(cell(row_cells, 'activation')),
                'notes':               notes_val,
                'client_email':        _ss(cell(row_cells, 'client_email')),
                'login_id':            login_id,
                'password':            password,
                'booked_by':           _ss(cell(row_cells, 'booked_by')),
                'pathway':             'australia',
            }

            try:
                existing = conn.execute(
                    "SELECT id FROM ops_online_subscriptions "
                    "WHERE registration_number = ? AND online_subscription = ? "
                    "AND COALESCE(issued_date,'') = ? AND pathway = 'australia'",
                    (reg_num, course, issued),
                ).fetchone()
                if existing:
                    sets = ', '.join(f"{k} = ?" for k in data.keys())
                    conn.execute(
                        f"UPDATE ops_online_subscriptions SET {sets} WHERE id = ?",
                        list(data.values()) + [existing['id']],
                    )
                    updated += 1
                else:
                    cols = ', '.join(data.keys())
                    placeholders = ', '.join(['?'] * len(data))
                    conn.execute(
                        f"INSERT INTO ops_online_subscriptions ({cols}) VALUES ({placeholders})",
                        list(data.values()),
                    )
                    inserted += 1
                conn.commit()
            except Exception as e:
                errors += 1
                logging.warning(
                    f"Australia online-courses row error ({reg_num} / {course}): {e}"
                )
                try:
                    conn.rollback()
                except Exception:
                    pass

        # Marker
        try:
            conn.execute(
                "INSERT INTO _import_markers (key, value) VALUES ('au_online_courses_import', ?) "
                "ON CONFLICT (key) DO UPDATE SET value = excluded.value",
                (IMPORT_VERSION,),
            )
            conn.commit()
        except Exception:
            try:
                conn.execute("DELETE FROM _import_markers WHERE key = 'au_online_courses_import'")
                conn.execute(
                    "INSERT INTO _import_markers (key, value) VALUES ('au_online_courses_import', ?)",
                    (IMPORT_VERSION,),
                )
                conn.commit()
            except Exception as e:
                logging.warning(f"Australia online-courses: could not write marker: {e}")

        logging.info(
            f"Australia online-courses {IMPORT_VERSION} complete — "
            f"inserted={inserted} updated={updated} skipped={skipped} errors={errors}"
        )
        wb.close()
        return {'inserted': inserted, 'updated': updated, 'skipped': skipped, 'errors': errors}
    finally:
        try:
            conn.close()
        except Exception:
            pass
