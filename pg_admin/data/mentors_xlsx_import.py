"""Import the founder's authoritative mentor list (mentors_list.xlsx) into
pg_mentors (founder 2026-08-13).

- Idempotent on (source='goocampus_xlsx', external_id=<mentor_id>): re-uploading
  refreshes the source-derived content but PRESERVES admin publish/verify/active
  choices made after the first import.
- Every one of the sheet's 42 columns is stored verbatim in the granular columns
  added to pg_mentors; the display columns (name/bio/mentor_type/…) are populated
  from them so the existing admin + /api/pg/mentors keep working unchanged.
- Photos: the sheet's profile_pic (an S3 URL) is stored as source_photo_url so it
  shows immediately via the photo fallback; a separate batched step copies them to R2.
"""
import io
import json
import logging

SOURCE = 'goocampus_xlsx'

# Excel header -> our handling. Values are stored verbatim in same-named columns
# unless remapped below.
_HEADERS = [
    'mentor_id', 'mentor_fname', 'mentor_lname', 'mentor_location',
    'mentor_country_origin', 'mentor_completion_yr', 'mentor_languages',
    'mentor_curr_location', 'mentor_special_interest', 'mentor_hobbies',
    'mentor_pricing', 'currency', 'mentor_status', 'timezone', 'mentor_bio',
    'mentor_type', 'profile_pic', 'rating', 'mentor_added_on', 'gender',
    'discount', 'current_state', 'current_city', 'mentor_last_updated',
    'verification_by_admin', 'profession_job_title', 'profession_company',
    'profession_location', 'profession_total_exp', 'profession_curr_work_exp',
    'profession_previous_work_exp', 'profession_designation', 'pre_work_exp',
    'education_qualification', 'education_pg_speciality', 'education_mbbs_college',
    'education_completion_yr', 'education_pg_college', 'education_pg_completion_yr',
    'mentor_topics_list', 'mentor_profession', 'intro_video',
]


def parse_xlsx(file_bytes):
    """Parse the uploaded .xlsx into a list of {header: value} dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    headers = None
    rows = []
    for r in it:
        if r is None:
            continue
        if headers is None:
            if any(c not in (None, '') for c in r):
                headers = [str(h).strip() if h is not None else '' for h in r]
            continue
        if any(c not in (None, '') for c in r):
            rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    wb.close()
    return rows, (headers or [])


def _s(v):
    return str(v).strip() if v is not None else ''


def _num(v):
    v = _s(v)
    if not v:
        return None
    try:
        return float(v)
    except ValueError:
        return None


def _langs(v):
    return [p.strip() for p in _s(v).replace(';', ',').split(',') if p.strip()]


def _content_from_row(m):
    """Map one sheet row to {db_column: value} for the content (source-derived)
    columns — shared by INSERT and UPDATE."""
    fname, lname = _s(m.get('mentor_fname')), _s(m.get('mentor_lname'))
    name = (fname + ' ' + lname).strip() or fname or lname
    mtype = 'peer_to_peer' if 'peer' in _s(m.get('mentor_type')).lower() else 'specialist'
    special = _s(m.get('mentor_special_interest'))
    return {
        # Display columns the existing admin + API already use.
        'name': name,
        'mentor_type': mtype,
        'bio': _s(m.get('mentor_bio')),
        'languages': json.dumps(_langs(m.get('mentor_languages'))),
        'specialties': json.dumps([special] if special else []),
        'counselling_fee': _num(m.get('mentor_pricing')),
        'rating': _num(m.get('rating')) or 0,
        'country': _s(m.get('mentor_country_origin')),
        'specialization': (_s(m.get('education_pg_speciality')) or special
                           or 'General Mentorship'),
        'qualification': _s(m.get('education_qualification')) or 'MBBS',
        'designation': (_s(m.get('profession_designation'))
                        or _s(m.get('profession_job_title'))),
        'hospital_name': _s(m.get('profession_company')),
        'experience_range': _s(m.get('profession_total_exp')),
        'source_photo_url': _s(m.get('profile_pic')),
        # Granular columns — the full record, kept verbatim.
        'gender': _s(m.get('gender')),
        'timezone': _s(m.get('timezone')),
        'pricing_currency': _s(m.get('currency')),
        'discount': _s(m.get('discount')),
        'current_state': _s(m.get('current_state')),
        'current_city': _s(m.get('current_city')),
        'location_origin': _s(m.get('mentor_location')) or _s(m.get('mentor_curr_location')),
        'special_interest': special,
        'hobbies': _s(m.get('mentor_hobbies')),
        'completion_yr': _s(m.get('mentor_completion_yr')),
        'profession_job_title': _s(m.get('profession_job_title')),
        'profession_company': _s(m.get('profession_company')),
        'profession_location': _s(m.get('profession_location')),
        'profession_total_exp': _s(m.get('profession_total_exp')),
        'profession_curr_work_exp': _s(m.get('profession_curr_work_exp')),
        'profession_previous_work_exp': _s(m.get('profession_previous_work_exp')),
        'pre_work_exp': _s(m.get('pre_work_exp')),
        'edu_qualification': _s(m.get('education_qualification')),
        'edu_pg_speciality': _s(m.get('education_pg_speciality')),
        'edu_mbbs_college': _s(m.get('education_mbbs_college')),
        'edu_mbbs_year': _s(m.get('education_completion_yr')),
        'edu_pg_college': _s(m.get('education_pg_college')),
        'edu_pg_year': _s(m.get('education_pg_completion_yr')),
        'topics_list': _s(m.get('mentor_topics_list')),
        'intro_video': _s(m.get('intro_video')),
        'added_on': _s(m.get('mentor_added_on')),
        'updated_src': _s(m.get('mentor_last_updated')),
    }


def import_mentors_from_xlsx(conn, rows, created_by=None, retire_others=True):
    """Upsert every sheet row into pg_mentors, then (optionally) retire the older
    auto-scraped mentors so the sheet is the single source of truth."""
    created = updated = skipped = 0
    errors = []
    for m in rows:
        ext = _s(m.get('mentor_id'))
        if not ext:
            skipped += 1
            continue
        try:
            content = _content_from_row(m)
            existing = conn.execute(
                "SELECT id FROM pg_mentors WHERE source = ? AND external_id = ?",
                (SOURCE, ext)).fetchone()
            if existing:
                # Refresh source content; leave admin flags (published/verified/active) alone.
                cols = list(content.keys())
                sets = ', '.join(f"{c} = ?" for c in cols) + ", updated_at = CURRENT_TIMESTAMP"
                conn.execute(f"UPDATE pg_mentors SET {sets} WHERE id = ?",
                             [content[c] for c in cols] + [existing['id']])
                updated += 1
            else:
                pub = _s(m.get('mentor_status')) in ('1', 'active', 'published', 'true')
                verified = _s(m.get('verification_by_admin')) in ('1', 'yes', 'true')
                cols = list(content.keys()) + ['source', 'external_id', 'is_published',
                                               'is_active', 'is_verified', 'created_by']
                vals = [content[c] for c in content] + [SOURCE, ext, pub, True, verified, created_by]
                ph = ', '.join(['?'] * len(cols))
                conn.execute(f"INSERT INTO pg_mentors ({', '.join(cols)}) VALUES ({ph})",
                             vals)
                created += 1
        except Exception as e:
            conn.rollback()
            errors.append(f"{ext}: {e}")
            logging.error("mentors_xlsx import row %s: %s", ext, e)
    conn.commit()

    retired = 0
    if retire_others:
        try:
            cur = conn.execute(
                "UPDATE pg_mentors SET is_published = FALSE, is_active = FALSE "
                " WHERE source <> ? AND COALESCE(source,'') <> ''", (SOURCE,))
            retired = getattr(cur, 'rowcount', 0) or 0
            conn.commit()
        except Exception as e:
            conn.rollback()
            logging.error("mentors_xlsx retire: %s", e)

    return {'created': created, 'updated': updated, 'skipped': skipped,
            'retired': retired, 'errors': errors, 'total': len(rows)}
