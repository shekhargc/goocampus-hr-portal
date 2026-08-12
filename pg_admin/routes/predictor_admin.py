"""Predictor Data admin — upload the NEET-PG closing-rank dataset that powers the
goocampus.in College Predictor.

Why this exists: the site shipped the 10 MB cut-off JSON inside its own repo, but
that file is .gitignored, so the deployed site never had it and the predictor fell
back to a tiny sample ("Preview mode"). goocampus.org is the backend for
goocampus.in, so the data belongs here — uploaded once a year by an admin instead
of committed into a front-end repo. (founder 2026-07-24)

Upload format: the same JSON the site used —
    {"year": 2025, "count": N, "rows": [{i,s,q,c,d,n,st,f,r1,r2,r3,r4,stray,sp,by,pen}, ...]}
Short keys are kept so the founder can upload the existing file unchanged.
"""
import io
import csv
import json
import logging
import re

from flask import request, redirect, url_for, flash, render_template
from core.auth import login_required
from core.users import get_user
from db import get_db

# Import in batches — 37k+ rows in one statement would blow the parameter limit.
_BATCH = 500

# Upload JSON key -> our column. Keeps the site's existing short-key format.
_FIELD_MAP = [
    ('i', 'institute'), ('s', 'authority'), ('q', 'quota'), ('c', 'category'),
    ('d', 'degree'), ('n', 'course'), ('st', 'state'),
    ('f', 'fee'), ('sp', 'stipend'), ('by', 'bond_years'), ('pen', 'penalty'),
    ('r1', 'r1'), ('r2', 'r2'), ('r3', 'r3'), ('r4', 'r4'), ('stray', 'stray'),
    # College-database columns (founder 2026-08-11) — stored verbatim on pg_cutoffs.
    ('it', 'institute_type'), ('seat', 'seat_type'), ('ct', 'course_type'), ('beds', 'beds'),
]
# 'scope' (Counselling Scope) is captured for DERIVATION only (authority_type /
# is_reference) — it is not a stored pg_cutoffs column, so it stays out of _FIELD_MAP.
_EXTRA_COL_TO_KEY = {'scope': 'scope'}


# ── Excel / CSV support ──────────────────────────────────────────────────────
# The founder's source data is a spreadsheet, not JSON. Rather than force a
# conversion every year, accept .xlsx/.csv directly and match the columns by
# HEADER NAME, tolerating the usual wording differences. (founder 2026-07-24)
_HEADER_ALIASES = {
    'institute':  ['institute', 'institution', 'college', 'college name', 'institute name',
                   'hospital', 'institute/college', 'name of college', 'name of institute'],
    'authority':  ['authority', 'counselling authority', 'counseling authority', 'counselling',
                   'source', 'conducted by', 'exam authority', 'type'],
    'quota':      ['quota', 'seat quota', 'quota type'],
    'category':   ['category', 'cat', 'seat category', 'reservation', 'caste category'],
    'degree':     ['degree', 'qualification', 'level', 'degree type'],
    'course':     ['course', 'course name', 'speciality', 'specialty', 'subject', 'branch',
                   'discipline', 'pg course'],
    'state':      ['state', 'state name', 'region', 'state/ut'],
    'fee':        ['fee', 'fees', 'course fee', 'annual fee', 'tuition', 'tuition fee'],
    'stipend':    ['stipend', 'monthly stipend', 'stipend per month',
                   'stipend yr 1', 'stipend year 1', 'stipend yr 1 mo'],
    'bond_years': ['bond', 'bond years', 'bond period', 'bond (years)', 'bond duration'],
    'penalty':    ['penalty', 'bond penalty', 'bond amount', 'penalty amount',
                   'bond course mid way', 'bond / course mid way'],
    # College-database columns from the founder's Master file (2026-08-11).
    'institute_type': ['institute type', 'institution type', 'college type',
                       'type of institute', 'type of institution'],
    'seat_type':      ['seat type', 'seat category type'],
    'course_type':    ['course type', 'clinical type'],
    'scope':          ['counselling scope', 'counseling scope', 'scope'],
    'beds':           ['beds', 'no of beds', 'number of beds', 'bed count', 'bed strength'],
    'r1':         ['r1', 'round 1', 'round1', 'round-1', 'r 1', 'closing rank r1', 'cr1'],
    'r2':         ['r2', 'round 2', 'round2', 'round-2', 'r 2', 'closing rank r2', 'cr2'],
    'r3':         ['r3', 'round 3', 'round3', 'round-3', 'r 3', 'closing rank r3', 'cr3'],
    'r4':         ['r4', 'round 4', 'round4', 'round-4', 'r 4', 'closing rank r4', 'cr4'],
    'stray':      ['stray', 'stray round', 'stray vacancy', 'stray vacancy round', 'svr',
                   'mop up', 'mop-up', 'mopup'],
}
# Reverse map for a fast lookup, normalised.
_ALIAS_TO_COL = {}
for _col, _aliases in _HEADER_ALIASES.items():
    for _a in _aliases:
        _ALIAS_TO_COL[_a] = _col


def _norm_header(h):
    """Lower-case, collapse punctuation/space so 'Round-1 ' == 'round 1'."""
    s = re.sub(r'[^a-z0-9 ]+', ' ', str(h or '').strip().lower())
    return re.sub(r'\s+', ' ', s).strip()


# Second-pass hints: if an exact alias didn't match, a header CONTAINING one of
# these fragments still maps. Ordered most-specific first so 'bond penalty' lands
# on penalty, not bond_years. Lets real-world headers like 'Allotted Quota',
# 'Quota Type', 'Seat Type (Quota)' work without me guessing every wording.
_CONTAINS_HINTS = [
    ('penalty', 'penalty'),
    ('stray', 'stray'),
    ('mop', 'stray'),
    ('stipend', 'stipend'),
    ('bond', 'bond_years'),
    ('quota', 'quota'),
    ('categ', 'category'),
    ('institut', 'institute'),
    ('college', 'institute'),
    ('hospital', 'institute'),
    ('special', 'course'),
    ('course', 'course'),
    ('subject', 'course'),
    ('branch', 'course'),
    ('authorit', 'authority'),
    ('counsel', 'authority'),
    ('state', 'state'),
    ('fee', 'fee'),
    ('degree', 'degree'),
]
# 'Round 3', 'R-3', 'Closing Rank Round 3' -> r3
_ROUND_RE = re.compile(r'(?:^|\b)(?:r|round)\s*[-_ ]?\s*([1-4])(?:\b|$)')


def _map_headers(headers):
    """Header row -> {column_index: our_column}.

    Pass 1 exact alias, pass 2 fuzzy (round-number regex, then keyword
    containment) so a spreadsheet doesn't have to use our exact wording.
    Unmatched headers are ignored and reported back to the admin."""
    mapping = {}
    taken = set()

    def _claim(idx, col):
        if col and col not in taken:
            mapping[idx] = col
            taken.add(col)
            return True
        return False

    for idx, h in enumerate(headers):
        _claim(idx, _ALIAS_TO_COL.get(_norm_header(h)))
    for idx, h in enumerate(headers):
        if idx in mapping:
            continue
        n = _norm_header(h)
        if not n:
            continue
        m = _ROUND_RE.search(n)
        if m and _claim(idx, f"r{m.group(1)}"):
            continue
        for frag, col in _CONTAINS_HINTS:
            if frag in n and _claim(idx, col):
                break
    return mapping


def _rows_from_tabular(headers, data_rows):
    """Turn a header row + data rows into the same short-key dicts the JSON uses,
    so both upload formats share one import path. Returns (rows, mapping, headers)."""
    mapping = _map_headers(headers)
    # Our short JSON keys, so downstream code is identical for both formats.
    col_to_key = {c: k for k, c in _FIELD_MAP}
    col_to_key.update(_EXTRA_COL_TO_KEY)   # 'scope' captured for derivation only
    rows = []
    for raw in data_rows:
        row = {}
        for idx, col in mapping.items():
            if idx < len(raw):
                row[col_to_key[col]] = raw[idx]
        if any(v not in (None, '') for v in row.values()):
            rows.append(row)
    return rows, mapping


def _read_upload(f):
    """Parse an uploaded .json / .xlsx / .csv into (rows, year, mapping, headers).
    Raises ValueError with a human message the admin page can show."""
    name = (f.filename or '').lower()
    if name.endswith('.json'):
        payload = json.load(f.stream)
        rows = payload.get('rows') if isinstance(payload, dict) else payload
        year = (payload.get('year') if isinstance(payload, dict) else None)
        if not isinstance(rows, list):
            raise ValueError('No "rows" list found in that JSON.')
        return rows, year, None, None

    if name.endswith('.csv'):
        text = f.stream.read().decode('utf-8-sig', errors='replace')
        reader = csv.reader(io.StringIO(text))
        all_rows = [r for r in reader if any((c or '').strip() for c in r)]
        if not all_rows:
            raise ValueError('That CSV is empty.')
        headers, data = all_rows[0], all_rows[1:]
    elif name.endswith(('.xlsx', '.xlsm')):
        try:
            import openpyxl
        except ImportError:
            raise ValueError('Excel support is unavailable on the server.')
        # Read into memory first — openpyxl needs a seekable/tellable stream and
        # the upload wrapper isn't guaranteed to be one.
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        headers = None
        data = []
        for r in it:
            if r is None:
                continue
            if headers is None:
                if any(c not in (None, '') for c in r):
                    headers = list(r)
                continue
            if any(c not in (None, '') for c in r):
                data.append(list(r))
        wb.close()
        if headers is None:
            raise ValueError('That spreadsheet has no header row.')
    elif name.endswith('.xls'):
        raise ValueError('Old .xls format — please "Save As" .xlsx and upload again.')
    else:
        raise ValueError('Unsupported file. Upload a .xlsx, .csv or .json file.')

    rows, mapping = _rows_from_tabular(headers, data)
    if not mapping:
        raise ValueError(
            'Could not recognise any columns. Found these headers: '
            + ', '.join(str(h) for h in headers[:15] if h)
            + '. Expected names like: College / Course / Category / State / Round 1 …')
    matched = set(mapping.values())
    if 'institute' not in matched and 'course' not in matched:
        raise ValueError('Neither a College nor a Course column was found — '
                         'check the header row.')
    if not ({'r1', 'r2', 'r3', 'r4', 'stray'} & matched):
        raise ValueError('No closing-rank columns found (Round 1 / R1 / Stray …).')
    return rows, None, mapping, headers


def _admin_only():
    user = get_user()
    return bool(user and user['is_admin']), user


def _num(v):
    """None-safe number: '' / None / junk -> None (never 0, which would be a lie)."""
    if v in (None, '', 'NA', 'N/A', '-'):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v):
    n = _num(v)
    return int(n) if n is not None else None


def _closing_rank(row):
    """The LAST round a seat was still available = the most generous closing rank,
    which is what 'is this within reach' should be judged against."""
    ranks = [_int(row.get(k)) for k in ('stray', 'r4', 'r3', 'r2', 'r1')]
    ranks = [r for r in ranks if r and r > 0]
    return max(ranks) if ranks else None


def _derive_tags(row):
    """From one Master row, derive (authority_type, degree_group, is_reference).

    - authority_type: 'allindia' (MCC / All-India Quota) | 'state' | '' (unknown).
      Read from Counselling Scope first ('AIQ MCC' / 'State' / 'Reference (SBP only)'),
      falling back to the authority string.
    - degree_group: 'mdms' (MD/MS) | 'dnb' (DNB/DNB-Diploma) | 'other'
      (PG-Diploma / MCh / MPH — excluded from both predictors). If the degree is
      blank (stipend-only reference rows), infer from Seat Type.
    - is_reference: 1 for 'Reference (SBP only)' rows — stipend/bond detail with no
      cut-off; kept for the college master but excluded from the predictor.
    DNB with no known scope defaults to 'allindia' (DNB counselling is centralised),
    so DNB colleges are still browsable under the all-India DNB tab.
    """
    scope = _norm_header(row.get('scope'))          # 'aiq mcc' / 'state' / 'reference sbp only'
    authority = str(row.get('s') or '').strip().lower()
    degree = str(row.get('d') or '').strip().upper()
    seat = _norm_header(row.get('seat'))            # 'md ms seats' / 'dnb seats' / …

    is_ref = 1 if 'reference' in scope else 0

    if 'mcc' in scope or 'aiq' in scope:
        at = 'allindia'
    elif scope == 'state':
        at = 'state'
    elif is_ref:
        at = ''
    elif 'mcc' in authority or 'all india' in authority:
        at = 'allindia'
    elif authority:
        at = 'state'
    else:
        at = ''

    if degree in ('MD', 'MS', 'MD/MS', 'MD-MS', 'MDMS'):
        dg = 'mdms'
    elif degree.startswith('DNB'):
        dg = 'dnb'
    elif not degree:
        if 'dnb' in seat:
            dg = 'dnb'
        elif 'md' in seat or 'ms' in seat:
            dg = 'mdms'
        else:
            dg = 'other'
    else:
        dg = 'other'

    if not at and dg == 'dnb':
        at = 'allindia'
    return at, dg, is_ref


def _rebuild_pg_colleges(conn, year):
    """Rebuild the pg_colleges master from the cut-off rows just loaded, then
    backfill college_id onto pg_cutoffs. Stable ids: UPSERT on
    (name_key, authority_type, state) so a college keeps its id across re-uploads.

    Done in TWO set-based SQL statements (aggregate-upsert + join-update) instead
    of a per-college Python loop — thousands of round-trips over 67k rows would
    blow the request timeout (the 2026-08-12 upload 502'd on exactly that)."""
    # 1) Aggregate distinct colleges and upsert in a single statement.
    #    degree_groups is built as a JSON text array (e.g. ["dnb","mdms"]) so the
    #    /api/pg/colleges LIKE '%"mdms"%' filter matches.
    conn.execute(
        "INSERT INTO pg_colleges "
        "  (name, name_key, institution_type, authority_type, state, degree_groups, beds, year) "
        "SELECT MIN(institute) AS name, "
        "       LOWER(TRIM(institute)) AS name_key, "
        "       COALESCE(MAX(NULLIF(TRIM(COALESCE(institute_type,'')),'')), '') AS institution_type, "
        "       COALESCE(authority_type,'') AS authority_type, "
        "       COALESCE(state,'') AS state, "
        "       to_json(COALESCE("
        "           array_agg(DISTINCT degree_group) FILTER (WHERE degree_group IN ('mdms','dnb')), "
        "           ARRAY[]::text[]))::text AS degree_groups, "
        "       MAX(beds) AS beds, "
        "       ? AS year "
        "  FROM pg_cutoffs "
        " WHERE year = ? AND COALESCE(institute,'') <> '' "
        " GROUP BY LOWER(TRIM(institute)), COALESCE(authority_type,''), COALESCE(state,'') "
        "ON CONFLICT (name_key, authority_type, state) DO UPDATE SET "
        "  name = EXCLUDED.name, institution_type = EXCLUDED.institution_type, "
        "  degree_groups = EXCLUDED.degree_groups, beds = EXCLUDED.beds, "
        "  year = EXCLUDED.year, updated_at = CURRENT_TIMESTAMP",
        (year, year))
    conn.commit()
    # 2) Backfill college_id onto the cut-off rows in one join-update.
    conn.execute(
        "UPDATE pg_cutoffs c SET college_id = col.college_id "
        "  FROM pg_colleges col "
        " WHERE c.year = ? "
        "   AND col.name_key = LOWER(TRIM(c.institute)) "
        "   AND col.authority_type = COALESCE(c.authority_type,'') "
        "   AND col.state = COALESCE(c.state,'')",
        (year,))
    conn.commit()
    n = conn.execute(
        "SELECT COUNT(*) AS c FROM pg_colleges WHERE year = ?", (year,)).fetchone()
    return (n['c'] if n else 0)


@login_required
def predictor_admin():
    """Show what's loaded + the upload form."""
    ok, _user = _admin_only()
    if not ok:
        flash('Admin access required', 'error')
        return redirect(url_for('dashboard'))
    conn = get_db()
    years, total, sample = [], 0, []
    try:
        years = [dict(r) for r in conn.execute(
            "SELECT year, COUNT(*) AS rows, MAX(created_at) AS loaded_at "
            "  FROM pg_cutoffs GROUP BY year ORDER BY year DESC").fetchall()]
        total = sum(y['rows'] for y in years)
        if total:
            sample = [dict(r) for r in conn.execute(
                "SELECT institute, course, quota, category, authority, state, closing_rank "
                "  FROM pg_cutoffs ORDER BY id DESC LIMIT 5").fetchall()]
    except Exception as e:
        logging.error("predictor_admin: %s", e)
        flash(f'Could not read the dataset: {e}', 'error')
    finally:
        conn.close()
    return render_template('pg_admin/predictor.html', years=years, total=total,
                           sample=sample, active_section='goocampus_in')


@login_required
def predictor_upload():
    """Import a cut-off JSON. Replaces that YEAR only (never touches other years)."""
    ok, user = _admin_only()
    if not ok:
        flash('Admin access required', 'error')
        return redirect(url_for('dashboard'))
    f = request.files.get('data_file')
    if not f or not f.filename:
        flash('Choose a file to upload (.xlsx, .csv or .json).', 'error')
        return redirect(url_for('pg_predictor_admin'))
    try:
        rows, file_year, mapping, headers = _read_upload(f)
    except ValueError as e:
        flash(str(e), 'error')
        return redirect(url_for('pg_predictor_admin'))
    except Exception as e:
        logging.error("predictor_upload parse: %s", e)
        flash(f'Could not read that file: {e}', 'error')
        return redirect(url_for('pg_predictor_admin'))

    if not rows:
        flash('No data rows found in the file.', 'error')
        return redirect(url_for('pg_predictor_admin'))
    try:
        year = int(file_year or request.form.get('year') or 0)
    except (TypeError, ValueError):
        year = 0
    if not year:
        flash('No year found in the file — type the year in the Year box and upload again.', 'error')
        return redirect(url_for('pg_predictor_admin'))

    # Storable columns + the three derived tags + closing_rank.
    cols = ['year'] + [c for _k, c in _FIELD_MAP] + \
           ['authority_type', 'degree_group', 'is_reference', 'closing_rank']
    placeholders = ','.join(['?'] * len(cols))
    sql = f"INSERT INTO pg_cutoffs ({','.join(cols)}) VALUES ({placeholders})"

    conn = get_db()
    inserted = skipped = colleges = 0
    try:
        # Replace this year only — re-uploading a corrected file is safe, and other
        # years (and everything else in the DB) are untouched.
        conn.execute("DELETE FROM pg_cutoffs WHERE year = ?", (year,))
        allvals = []
        for r in rows:
            if not isinstance(r, dict):
                skipped += 1
                continue
            vals = [year]
            for key, col in _FIELD_MAP:
                v = r.get(key)
                if col in ('fee', 'stipend', 'bond_years', 'penalty'):
                    vals.append(_num(v))
                elif col in ('r1', 'r2', 'r3', 'r4', 'stray', 'beds'):
                    vals.append(_int(v))
                else:
                    vals.append((str(v).strip() if v is not None else ''))
            at, dg, is_ref = _derive_tags(r)
            vals += [at, dg, is_ref, _closing_rank(r)]
            allvals.append(vals)
        # execute_values (multi-row INSERT) — psycopg2's default executemany sends
        # one statement PER ROW, which timed out the 67k-row upload (2026-08-12).
        if allvals:
            conn.execute_batch(sql, allvals, page_size=1000)
            inserted = len(allvals)
        conn.commit()
        # Build the college master + backfill college_id (founder 2026-08-11).
        try:
            colleges = _rebuild_pg_colleges(conn, year)
        except Exception as ce:
            conn.rollback()
            logging.error("pg_colleges rebuild: %s", ce)
        msg = (f'Loaded {inserted:,} rows for {year}'
               + (f' · {colleges:,} colleges' if colleges else '')
               + (f' ({skipped} skipped)' if skipped else '') + '.')
        if mapping:
            # Spreadsheet upload: show which of THEIR columns we used, so a
            # mis-read header is obvious immediately rather than after go-live.
            used = ', '.join(sorted(set(mapping.values())))
            msg += f' Columns matched from your file: {used}.'
            missed = [str(headers[i]) for i in range(len(headers or []))
                      if i not in mapping and headers[i] not in (None, '')]
            if missed:
                msg += f' Ignored: {", ".join(missed[:8])}.'
        flash(msg, 'success')
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        logging.error("predictor_upload: %s", e)
        flash(f'Import failed — nothing was changed: {e}', 'error')
    finally:
        conn.close()
    return redirect(url_for('pg_predictor_admin'))
