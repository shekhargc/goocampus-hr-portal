"""
One-time import: Australia payments from Excel into ops_payments.

Each Excel row goes into ops_payments with pathway='australia'.

The Australia Payments Excel has a dedicated "Registration Number"
column (no parsing needed), so linkage to plab_clients is direct.
If the dedicated column is empty for a row, we fall back to parsing
GCAUSIP/... out of the candidate-name cell, then to a fuzzy name
match against plab_clients WHERE pathway='australia'.

Idempotent: keyed on (registration_number, payment_date, instalment,
total_amount_paid). Re-running with no marker change is a no-op.
Bumping IMPORT_VERSION re-runs the upsert. Bad rows (no reg, no name
match) are skipped with a warning so one bad row never blocks the
whole import.

Hooked into app boot from app.py the same way as
import_australia_test_bookings.py.
"""

import logging
import os
import re
from datetime import datetime, date


EXCEL_FILENAME = 'All_Australia_Payments.xlsx'
IMPORT_VERSION = 'au_payments_v1_initial_380'


H = {
    'cand_name':         'Enter Candidate Name',
    'reg_number':        'Registration Number',
    'payment_date':      'Payment Date',
    'total_package':     'Total Package',
    'instalment':        'Instalment',
    'total_amt_paid':    'Total Amount Paid',
    'amount_paid':       'Amount Paid',
    'gst_paid':          'GST Paid',
    'payment_method':    'Payment Method',
    'notes':             'Notes',
}


_REG_RE = re.compile(r'(GCAUSIP/\S+)', re.IGNORECASE)


def _extract_reg_number(candidate_cell):
    """Pull the GCAUSIP/... reg number out of any text cell."""
    if not candidate_cell:
        return ''
    s = str(candidate_cell)
    m = _REG_RE.search(s)
    if not m:
        return ''
    raw = m.group(1).rstrip('.,;:-/ \t')
    return raw.upper()


def _ss(v):
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _sd(v):
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def _num(v):
    """Parse to float; tolerate strings like '1,50,000.00', '₹1,000', blanks."""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    # strip currency symbols, commas, spaces
    s = s.replace('₹', '').replace('Rs.', '').replace('Rs', '').replace(',', '').replace(' ', '')
    try:
        return float(s)
    except Exception:
        return None


def _fuzzy_match_reg(conn, candidate_cell):
    """Last-resort: try to map a candidate-name cell to a reg number via
    LOWER(TRIM(name)) prefix match against Australia plab_clients.
    Returns reg_number or ''.
    """
    if not candidate_cell:
        return ''
    s = str(candidate_cell).strip()
    # Strip leading 'Dr.' / 'Dr ' and split off any '-' or '/' suffix
    s_clean = re.sub(r'^\s*Dr\.?\s+', '', s, flags=re.IGNORECASE)
    s_clean = re.split(r'[-/]', s_clean, maxsplit=1)[0].strip().lower()
    if not s_clean:
        return ''
    parts = s_clean.split()
    if len(parts) < 1:
        return ''
    first = parts[0]
    last = parts[-1] if len(parts) > 1 else ''
    try:
        rows = conn.execute(
            """SELECT registration_number, first_name, last_name
                 FROM plab_clients
                WHERE pathway = 'australia'
                  AND LOWER(TRIM(first_name)) = ?
                  AND LOWER(TRIM(last_name)) = ?""",
            (first, last),
        ).fetchall()
        if len(rows) == 1:
            return rows[0]['registration_number'] or ''
    except Exception:
        return ''
    return ''


def run_import_australia_payments_once(get_db_fn):
    """Import Australia payments into ops_payments (pathway='australia')."""
    import openpyxl

    excel_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        EXCEL_FILENAME,
    )
    if not os.path.exists(excel_path):
        logging.info("Australia payments import: no Excel found, skipping")
        return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

    conn = get_db_fn()
    try:
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS _import_markers (key TEXT PRIMARY KEY, value TEXT)")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        marker = conn.execute(
            "SELECT value FROM _import_markers WHERE key = 'au_payments_import'"
        ).fetchone()
        if marker and marker['value'] == IMPORT_VERSION:
            logging.info(f"Australia payments import: already at {IMPORT_VERSION}, skipping")
            return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

        logging.info(f"Australia payments import: starting {IMPORT_VERSION}...")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h is not None}

        def cell(row, key):
            i = idx.get(H[key])
            return row[i] if i is not None and i < len(row) else None

        inserted = 0
        updated = 0
        skipped = 0
        errors = 0

        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row_cells):
                continue

            # 1. Prefer dedicated reg-number column.
            reg_num = _ss(cell(row_cells, 'reg_number')).upper()
            # 2. Fallback: parse out of candidate name cell.
            if not reg_num:
                reg_num = _extract_reg_number(cell(row_cells, 'cand_name'))
            # 3. Last resort: fuzzy name match against plab_clients.
            if not reg_num:
                reg_num = _fuzzy_match_reg(conn, cell(row_cells, 'cand_name'))
            if not reg_num:
                skipped += 1
                logging.warning(
                    f"Australia payments: no reg number for row, candidate cell = "
                    f"{_ss(cell(row_cells, 'cand_name'))!r} — skipped"
                )
                continue

            payment_date = _sd(cell(row_cells, 'payment_date'))
            instalment = _ss(cell(row_cells, 'instalment'))

            data = {
                'registration_number': reg_num,
                'payment_date':        payment_date,
                'amount_paid':         _num(cell(row_cells, 'amount_paid')),
                'gst_paid':            _num(cell(row_cells, 'gst_paid')),
                'total_amount_paid':   _num(cell(row_cells, 'total_amt_paid')),
                'instalment':          instalment,
                'payment_method':      _ss(cell(row_cells, 'payment_method')),
                'total_package':       _num(cell(row_cells, 'total_package')),
                'notes':               _ss(cell(row_cells, 'notes')),
                'pathway':             'australia',
            }

            try:
                # Idempotency key: registration + payment_date + instalment + total_amount_paid
                existing = conn.execute(
                    """SELECT id FROM ops_payments
                        WHERE registration_number = ?
                          AND COALESCE(payment_date, '') = ?
                          AND COALESCE(instalment, '') = ?
                          AND COALESCE(total_amount_paid, 0) = COALESCE(?, 0)
                          AND pathway = 'australia'""",
                    (reg_num, payment_date or '', instalment or '', data['total_amount_paid']),
                ).fetchone()
                if existing:
                    sets = ', '.join(f"{k} = ?" for k in data.keys())
                    conn.execute(
                        f"UPDATE ops_payments SET {sets} WHERE id = ?",
                        list(data.values()) + [existing['id']],
                    )
                    updated += 1
                else:
                    cols = ', '.join(data.keys())
                    placeholders = ', '.join(['?'] * len(data))
                    conn.execute(
                        f"INSERT INTO ops_payments ({cols}) VALUES ({placeholders})",
                        list(data.values()),
                    )
                    inserted += 1
                conn.commit()
            except Exception as e:
                errors += 1
                logging.warning(f"Australia payments row error ({reg_num} / {payment_date}): {e}")
                try:
                    conn.rollback()
                except Exception:
                    pass

        # Marker
        try:
            conn.execute(
                "INSERT INTO _import_markers (key, value) VALUES ('au_payments_import', ?) "
                "ON CONFLICT (key) DO UPDATE SET value = excluded.value",
                (IMPORT_VERSION,),
            )
            conn.commit()
        except Exception:
            try:
                conn.execute("DELETE FROM _import_markers WHERE key = 'au_payments_import'")
                conn.execute(
                    "INSERT INTO _import_markers (key, value) VALUES ('au_payments_import', ?)",
                    (IMPORT_VERSION,),
                )
                conn.commit()
            except Exception as e:
                logging.warning(f"Australia payments: could not write marker: {e}")

        logging.info(
            f"Australia payments {IMPORT_VERSION} complete — "
            f"inserted={inserted} updated={updated} skipped={skipped} errors={errors}"
        )
        wb.close()
        return {'inserted': inserted, 'updated': updated, 'skipped': skipped, 'errors': errors}
    finally:
        try:
            conn.close()
        except Exception:
            pass
