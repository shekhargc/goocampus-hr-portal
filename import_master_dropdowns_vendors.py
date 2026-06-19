"""
import_master_dropdowns_vendors.py
==================================

ONE importer that reads the new TWO-SHEET Zoho section files (a data sheet
[sheet 0] + a "Drop Downs"/"Drop Down" sheet [sheet 1]) and rebuilds, per
pathway:

  1) Field Manager dropdowns  -> lookup_options   (clean-replace keep-in-use)
  2) Centralised vendors      -> vendors_providers (clean-replace keep-in-use)
  3) Vendor->deliverable map  -> vendor_service_map (pairs from the DATA sheet)

DRY-RUN by default. Nothing is written unless DRY=0 (or --live) is passed.
The DB is shared with production, so DRY mode only PRINTS the plan.

Run:
  # dry-run (no DB writes) — default
  python3 import_master_dropdowns_vendors.py
  # against the shared DB read-only checks need DATABASE_URL set:
  DATABASE_URL=...  python3 import_master_dropdowns_vendors.py
  # live (writes) — run by the user after reviewing the plan
  DATABASE_URL=...  DRY=0  python3 import_master_dropdowns_vendors.py

Env:
  DRY=1 (default) | DRY=0 / --live   -> write mode
  AMC_DIR   default /Users/Santosh/Desktop/Zoho Data/AMC PGCP
  UK_DIR    default /Users/Santosh/Desktop/Zoho Data/UK GC
  PATHWAYS  e.g. "plab" or "australia" or "plab,australia" (default both)

WHY THE CATEGORY NAMES ARE WHAT THEY ARE
----------------------------------------
The forms read dropdowns via get_lookup_options(<category>, pathway=...).
The category names below were taken verbatim from:
  - app.py  PLAB render_template get_lookup_options('<cat>') calls
  - routes/operations/_form_lookups.py  section_*_lookups()  (AMC/Consulting)
so the values we seed are exactly the ones the section forms render.

A handful of "fields" are NOT backed by lookup_options at all — they are
driven by the centralised vendor DB through VENDOR_LOOKUP_MAP in app.py
(course_provider, subscription_type, course_name, research_provider,
ngo_vendor, cab_vendor, certification_body). Those Excel columns are routed
to VENDORS instead of lookup_options, and flagged in the report so the
behaviour is explicit rather than silently wrong.

CLEAN REPLACE KEEP-IN-USE (lookup_options & vendors)
----------------------------------------------------
For each (category, pathway):
  target = master values from the sheet.
  - master value missing in DB      -> ADD (is_active=true)
  - existing value not in master:
        used by any record for pathway -> KEEP active  (keep-in-use)
        not used                       -> DEACTIVATE (is_active=false)
  Never DELETE. Deactivation only.
"""

import os
import sys
import re

# ── mode ────────────────────────────────────────────────────────────
DRY = os.environ.get('DRY', '1') not in ('0', 'false', 'False')
if '--live' in sys.argv:
    DRY = False

AMC_DIR = os.environ.get('AMC_DIR', '/Users/Santosh/Desktop/Zoho Data/AMC PGCP')
UK_DIR = os.environ.get('UK_DIR', '/Users/Santosh/Desktop/Zoho Data/UK GC')

# Country labels for vendors_providers (per task spec).
COUNTRY = {'plab': 'UK Pathway', 'australia': 'AMC Pathway'}

# Tag-row labels that identify a column's role on a "Drop Downs" sheet.
VENDOR_TAGS = ('vendor/provider', 'vendor', 'provider')
DELIVERABLE_TAGS = ('deliverable',)            # 'Deliverable' (NOT 'Deliverable Type')
DELIVERABLE_TYPE_TAGS = ('deliverable type',)

# Header keywords that also mark a column as a vendor column (used as a
# secondary hint when the tag row is absent/inconsistent).
VENDOR_HDR_KEYWORDS = ('vendor', 'provider', 'centre', 'center', 'hospital')

# Booking-arrangement values that must NEVER become vendors. Mirrors the
# _NON_VENDOR guard already in import_amc_sections.py.
_NON_VENDOR = {
    'booked & paid by goocampus', 'booked and paid by goocampus',
    'booked by gc / paid by client', 'booking included in package',
    'booking included in the package', 'booked and paid by client',
    'paid by gc', 'paid by client', 'free entry', 'free', 'paid', 'not paid',
}

# Categories that are NOT lookup_options — they are vendor-backed via
# app.py VENDOR_LOOKUP_MAP. If an Excel column maps to one of these we route
# it to vendors_providers instead of lookup_options.
VENDOR_BACKED_CATEGORIES = {
    'course_provider', 'subscription_type', 'course_name',
    'research_provider', 'ngo_vendor', 'cab_vendor', 'certification_body',
}

# "Categories" that LOOK like lookups but are actually hardcoded Python lists
# in the app (the form does NOT read lookup_options for them), so seeding them
# would be a silent no-op. Report these instead of seeding.
#   job_confirmed_by -> app.py JOB_CONFIRMED_BY_OPTIONS (hardcoded list)
# (was {'job_confirmed_by'} — user asked to make Job Confirmed By an editable
#  Field-Manager dropdown, so it is now seeded like any other lookup and the
#  app reads get_lookup_options('job_confirmed_by') with a hardcoded fallback.)
HARDCODED_NONLOOKUP = set()


# ════════════════════════════════════════════════════════════════════
#  PER-FILE CONFIG
#  Each entry: a "section" with everything needed to process its file.
#    file       : Excel filename (under the pathway dir)
#    section    : human label + vendors_providers.category
#    vp_category: vendors_providers.category for this section
#    cols       : { excel_header : category }  field -> lookup_options category
#                 (only headers that are real dropdown fields)
#    data_reg   : data-sheet column holding reg/name (for vendor->deliverable
#                 pair extraction we read the DATA sheet directly)
#    vendor_col : data-sheet header that holds the vendor value (for pairs)
#    deliverable_col : data-sheet header that holds the deliverable value
#  Columns NOT in `cols` and not vendor/deliverable are reported as unmapped
#  only if they actually carry master values on the drop-downs sheet.
# ════════════════════════════════════════════════════════════════════

# ── AMC (australia) ──────────────────────────────────────────────────
AMC_SECTIONS = [
 {
  'file': 'AMC - All Academic Details (1).xlsx',
  'section': 'Academic Details', 'vp_category': None,
  'cols': {
    'IMG / FMG': 'img_fmg', 'Country': 'exam_country',
    'MBBS Status': 'mbbs_status', 'Internship Status': 'internship_status',
    'Internship Gap': 'internship_gap', 'Working Status': 'working_status',
    'Speciality Interest 1': 'medical_speciality',
    'Speciality Interest 2': 'medical_speciality',
  },
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'AMC - All Amc Registrations (1).xlsx',
  'section': 'AMC Registrations', 'vp_category': None,
  'cols': {'AMC Setup': 'amc_setup_status'},   # AMC-specific; flagged unmapped if no form cat
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'AMC - All Epic Registrations (1).xlsx',
  'section': 'EPIC Registrations', 'vp_category': None,
  'cols': {
    'EPIC Registration Status': 'epic_reg_status', 'EPIC Status': 'epic_status',
    'Notary Cam': 'notary_camp_status', 'Documents Stage': 'doc_stage',
    'Documents Stage Status': 'doc_stage_status',
  },
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'AMC - All Online Courses & Subscriptions (1).xlsx',
  'section': 'Online Subscriptions', 'vp_category': 'Online Subscriptions',
  'cols': {
    'Course or Subscription Name': 'subscription_type',   # VENDOR-BACKED -> routed to vendors
    'Booked By': 'subscription_booked_by',
    'Activation Type': 'activation_type',
  },
  'vendor_col': 'Course or Subscription Name',
  'deliverable_col': 'Course or Subscription Name',
 },
 {
  'file': 'AMC - All Payments (1).xlsx',
  'section': 'Payments', 'vp_category': None,
  'cols': {'Payment Method': 'payment_method', 'Instalment': 'instalment'},
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'AMC - All Test Bookings (1).xlsx',
  'section': 'Test Bookings', 'vp_category': None,
  'cols': {
    'Booked By': 'exam_booked_by', 'Exam': 'exam_name',
    'Exam Method': 'exam_method', 'Country': 'exam_country',
    'Exam Status': 'exam_status', 'Exam Result': 'exam_result',
    'Revaluation Applied': 'revaluation_option', 'Reval Result': 'reval_result',
  },
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'AMC - All Trainings (1).xlsx',
  'section': 'Training Programs', 'vp_category': 'Training Programs',
  'cols': {
    'Training Program': 'training_program', 'Course Type': 'coaching_course_type',
    'Training Method': 'coaching_method', 'Booked By': 'coaching_booked_by',
    'Batch Year': 'batch_year', 'Batch (Month)': 'batch_month',
    'Training Status': 'coaching_status',
    'Training Vendor Name': '__VENDOR__',   # vendor column, not a lookup
  },
  'vendor_col': 'Training Vendor Name', 'deliverable_col': 'Training Program',
 },
 {
  'file': 'AMC - Research and Publication Report (1).xlsx',
  'section': 'Research & Publications', 'vp_category': 'Research & Publications',
  'cols': {
    'Research Status': 'research_status', 'Author Position': 'author_position',
    'Service': 'research_service',
    'Research Provider': '__VENDOR__',   # vendor column
  },
  'vendor_col': 'Research Provider', 'deliverable_col': 'Service',
 },
 {
  'file': 'AMC - Webinar and Conferences Report (1).xlsx',
  'section': 'Webinars', 'vp_category': 'Webinars',
  'cols': {
    'Event Type': 'event_type', 'Participation Type': 'participation_type',
    'Event Value': 'event_value',
    'Provider Name': '__VENDOR__',   # vendor column
  },
  'vendor_col': 'Provider Name', 'deliverable_col': 'Event Type',
 },
 {
  'file': 'AMC - GC AUS Registration Report (1).xlsx',
  'section': 'Clients (Registration)', 'vp_category': None,
  'cols': {
    'Plan Type': 'plan_type', 'Account Status': 'account_status',
    'Counsellor Name': 'counsellor', 'Stage (Current Status)': 'current_stage',
    'Switched Program': 'switched_program', 'Joined Service': 'joined_stage',
    'Lead Source': 'lead_source',
    # CITY / STATE / Country present but not form dropdowns -> reported unmapped
  },
  'vendor_col': None, 'deliverable_col': None,
 },
]

# ── PLAB (plab) ──────────────────────────────────────────────────────
PLAB_SECTIONS = [
 {
  'file': 'PLAB - Academic Details (R).xlsx',
  'section': 'Academic Details', 'vp_category': None,
  'cols': {
    'IMG / FMG': 'img_fmg', 'Country': 'exam_country',
    'MBBS Status': 'mbbs_status', 'Internship Status': 'internship_status',
    'Internship Gap': 'internship_gap',
    'Speciality Interest 1': 'medical_speciality',
    'Speciality Interest 2': 'medical_speciality',
  },
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'PLAB - All English Login Details.xlsx',
  'section': 'English Logins', 'vp_category': None,
  'cols': {},  # no dropdown fields
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'PLAB - All Job Stages.xlsx',
  'section': 'Job Stages', 'vp_category': None,
  'cols': {'Job Confirmed By': 'job_confirmed_by'},  # now an editable lookup dropdown
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'PLAB - All Ngo Activities.xlsx',
  'section': 'NGO Activities', 'vp_category': 'NGO Activities',
  'cols': {
    'NGO Vendor Name': '__VENDOR__',
    'Activity Type': 'ngo_activity_type',
  },
  'vendor_col': 'NGO Vendor Name', 'deliverable_col': 'Service',
 },
 {
  'file': 'PLAB - All Uk Cab Bookings.xlsx',
  'section': 'UK Cab Bookings', 'vp_category': 'Cab Bookings',
  'cols': {'Vendor': '__VENDOR__'},
  'vendor_col': 'Vendor', 'deliverable_col': 'Services',
 },
 {
  'file': 'PLAB - All_Payments Report.xlsx',
  'section': 'Payments', 'vp_category': None,
  'cols': {'Payment Method': 'payment_method'},
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'PLAB - Epic Registration Status (R).xlsx',
  'section': 'EPIC Registrations', 'vp_category': None,
  'cols': {
    'EPIC Registration': 'epic_reg_status', 'EPIC Status': 'epic_status',
    'Notary Camp': 'notary_camp_status', 'Documents Stage': 'doc_stage',
    'Document Stage Status': 'doc_stage_status',
  },
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'PLAB - ICP Training Report List.xlsx',
  'section': 'Training Programs', 'vp_category': 'Training Programs',
  'cols': {
    'Training Program': 'training_program', 'Course Type': 'coaching_course_type',
    'Coaching Method': 'coaching_method', 'Coaching Status': 'coaching_status',
    'Training Attendance': 'coaching_attendance',
    'Training Vendor': '__VENDOR__',
  },
  'vendor_col': 'Training Vendor', 'deliverable_col': 'Training Program',
 },
 {
  'file': 'PLAB - Mentorship Report.xlsx',
  'section': 'Mentorship', 'vp_category': 'Mentorship',
  'cols': {
    'Program Provider': '__VENDOR__',
    'Candidate Attendance': 'mentorship_attendance',
    'Mentor Attendance': 'mentorship_attendance',
    'Session Confirmation': 'mentorship_confirmation',
    'Services Type': 'mentorship_payment_status',
  },
  'vendor_col': 'Program Provider', 'deliverable_col': 'Services',
 },
 {
  'file': 'PLAB - Online Courses (R).xlsx',
  'section': 'Online Courses', 'vp_category': 'Online Courses',
  'cols': {
    'Courses Name': 'course_name',       # VENDOR-BACKED (deliverables) -> vendors/services
    'Course Type': 'course_type',
    'Course Provider': '__VENDOR__',
    'Course Status': 'course_status',
    'Booked By': 'course_booked_by',
  },
  'vendor_col': 'Course Provider', 'deliverable_col': 'Courses Name',
 },
 {
  'file': 'PLAB - Online Subscriptions.xlsx',
  'section': 'Online Subscriptions', 'vp_category': 'Online Subscriptions',
  'cols': {
    'Online Subscription': 'subscription_type',  # VENDOR-BACKED -> vendors
    'Subscription Providers': '__VENDOR__',
    'Activation type': 'activation_type',
    'Booked By': 'subscription_booked_by',
  },
  'vendor_col': 'Subscription Providers', 'deliverable_col': 'Online Subscription',
 },
 {
  'file': 'PLAB - Research and Publication (R).xlsx',
  'section': 'Research & Publications', 'vp_category': 'Research & Publications',
  'cols': {
    'Research Status': 'research_status', 'Author Position': 'author_position',
    'Service': 'research_service',
    'Research Provider': '__VENDOR__',
  },
  'vendor_col': 'Research Provider', 'deliverable_col': 'Service',
 },
 {
  'file': 'PLAB - Test Bookings (R).xlsx',
  'section': 'Test Bookings', 'vp_category': None,
  'cols': {
    'Exam': 'exam_name', 'Exam Booked By': 'exam_booked_by',
    'Exam Method': 'exam_method', 'Country': 'exam_country',
    'Exam Status': 'exam_status', 'Exam Result': 'exam_result',
    'Are we applying for revaluation': 'revaluation_option',
    'Reval Result': 'reval_result',
    # 'City / State' present but no form category -> reported unmapped
  },
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'PLAB - UK Observership Report.xlsx',
  'section': 'UK Observerships', 'vp_category': None,
  'cols': {'Speciality': 'medical_speciality', 'Payment': 'observership_payment'},
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'PLAB - Visa and Travel (R).xlsx',
  'section': 'Visa & Travel', 'vp_category': None,
  'cols': {
    'Visa Application Status': 'visa_app_status',
    'Documents Collected': 'visa_doc_collected',
    'Documents Status': 'visa_doc_status', 'Visa Status': 'visa_status',
    'Visa Type': 'visa_type', 'Visa Period': 'visa_period',
    'Quarantine': 'quarantine', 'Lodging Type': 'lodging_type',
  },
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'PLAB - Webinar and Conferences.xlsx',
  'section': 'Webinars', 'vp_category': 'Webinars',
  'cols': {
    'Event Type': 'event_type', 'Event Value': 'event_value',
    'Participation Type': 'participation_type',
    'Provider': '__VENDOR__',
    # 'Webinar and Event Name' present but no form category -> reported unmapped
  },
  'vendor_col': 'Provider', 'deliverable_col': 'Event Type',
 },
 {
  'file': 'PLAB -All Gmc Registrations.xlsx',
  'section': 'GMC Registrations', 'vp_category': None,
  'cols': {
    'GMC Setup': 'gmc_setup_status', 'English Exam': 'gmc_english_exam',
    'License': 'gmc_license_status',
  },
  'vendor_col': None, 'deliverable_col': None,
 },
 {
  'file': 'PLAB - Registration list.xlsx',
  'section': 'Clients (Registration)', 'vp_category': None,
  'cols': {
    'Plan Type': 'plan_type', 'Joined Stage': 'joined_stage',
    'PLAB Stage (Current Status)': 'plab_stage', 'Account Status': 'account_status',
    'Switched Program': 'switched_program', 'Counsellor': 'counsellor',
    'Lead Source': 'lead_source', 'Operations Team Referral': 'operations_referral',
    # 'Upgraded To' -> new editable dropdown. (CITY / STATE are intentionally
    #  NOT seeded here: the client form already renders them as a State dropdown
    #  + City autocomplete backed by the dedicated `states`/`cities` tables —
    #  42,045 cities — so a lookup_options copy would just be orphan values.)
    'Upgraded To': 'upgraded_to',
  },
  'vendor_col': None, 'deliverable_col': None,
 },
]

PATHWAY_DIRS = {'australia': AMC_DIR, 'plab': UK_DIR}
PATHWAY_SECTIONS = {'australia': AMC_SECTIONS, 'plab': PLAB_SECTIONS}


# ════════════════════════════════════════════════════════════════════
#  category -> (ops_table, column) for the "in use?" check.
#  Derived from import_plab_sections.py / import_amc_sections.py mappings.
#  Used to decide keep-in-use vs deactivate for lookup_options values.
#  (Client-level categories map to plab_clients.)
# ════════════════════════════════════════════════════════════════════
CATEGORY_INUSE = {
    # Test bookings
    'exam_name': ('ops_test_bookings', 'exam'),
    'exam_method': ('ops_test_bookings', 'exam_type'),
    'exam_status': ('ops_test_bookings', 'exam_status'),
    'exam_result': ('ops_test_bookings', 'exam_result'),
    'exam_country': ('ops_test_bookings', 'country'),
    'exam_booked_by': ('ops_test_bookings', 'booked_by'),
    'revaluation_option': ('ops_test_bookings', 'revaluation'),
    'reval_result': ('ops_test_bookings', 'reval_result'),
    # Coaching / training
    'training_program': ('ops_coaching', 'english_training'),
    'coaching_course_type': ('ops_coaching', 'course_type'),
    'coaching_method': ('ops_coaching', 'coaching_method'),
    'coaching_status': ('ops_coaching', 'coaching_status'),
    'coaching_booked_by': ('ops_coaching', 'booked_by'),
    'coaching_attendance': ('ops_coaching', 'attendance'),
    'batch_year': ('ops_coaching', 'batch_year'),
    'batch_month': ('ops_coaching', 'batch_month'),
    # Online subscriptions
    'subscription_type': ('ops_online_subscriptions', 'online_subscription'),
    'activation_type': ('ops_online_subscriptions', 'activation_type'),
    'subscription_booked_by': ('ops_online_subscriptions', 'booked_by'),
    # Online courses
    'course_name': ('ops_online_courses', 'courses_name'),
    'course_type': ('ops_online_courses', 'course_type'),
    'course_status': ('ops_online_courses', 'course_status'),
    'course_booked_by': ('ops_online_courses', 'booked_by'),
    'course_provider': ('ops_online_courses', 'course_provider'),
    'certification_body': ('ops_online_courses', 'certification_body'),
    # Payments
    'payment_method': ('ops_payments', 'payment_method'),
    'instalment': ('ops_payments', 'instalment'),
    # EPIC
    'epic_reg_status': ('ops_epic_registration', 'epic_registration'),
    'epic_status': ('ops_epic_registration', 'epic_status'),
    'notary_camp_status': ('ops_epic_registration', 'notary_camp'),
    'doc_stage': ('ops_epic_registration', 'documents_stage'),
    'doc_stage_status': ('ops_epic_registration', 'document_stage_status'),
    # AMC registrations
    'amc_setup_status': ('ops_amc_registration', 'amc_setup'),
    # GMC
    'gmc_setup_status': ('ops_gmc_registration', 'gmc_setup'),
    'gmc_english_exam': ('ops_gmc_registration', 'english_exam'),
    'gmc_license_status': ('ops_gmc_registration', 'license'),
    # Webinars
    'event_type': ('ops_webinars_conferences', 'event_type'),
    'event_value': ('ops_webinars_conferences', 'event_value'),
    'participation_type': ('ops_webinars_conferences', 'participation_type'),
    # Research
    'research_status': ('ops_research_publication', 'research_status'),
    'research_provider': ('ops_research_publication', 'research_provider'),
    'author_position': ('ops_research_publication', 'author_position'),
    'research_service': ('ops_research_publication', 'service'),
    # Academic
    'img_fmg': ('ops_academic_details', 'img_fmg'),
    'mbbs_status': ('ops_academic_details', 'mbbs_status'),
    'internship_status': ('ops_academic_details', 'internship_status'),
    'internship_gap': ('ops_academic_details', 'internship_gap'),
    'working_status': ('ops_academic_details', 'working_status'),
    # Visa & travel
    'visa_app_status': ('ops_uk_visa_travel', 'visa_application_status'),
    'visa_doc_collected': ('ops_uk_visa_travel', 'documents_collected'),
    'visa_doc_status': ('ops_uk_visa_travel', 'documents_status'),
    'visa_status': ('ops_uk_visa_travel', 'visa_status'),
    'visa_type': ('ops_uk_visa_travel', 'visa_type'),
    'visa_period': ('ops_uk_visa_travel', 'visa_period'),
    'quarantine': ('ops_uk_visa_travel', 'quarantine'),
    'lodging_type': ('ops_uk_visa_travel', 'lodging_type'),
    # Observerships
    'observership_payment': ('ops_uk_observerships', 'payment'),
    'medical_speciality': ('ops_uk_observerships', 'speciality'),
    # Mentorship
    'mentorship_attendance': ('ops_mentorship', 'candidate_attendance'),
    'mentorship_confirmation': ('ops_mentorship', 'session_confirmation'),
    'mentorship_payment_status': ('ops_mentorship', 'amount_paid'),
    # NGO
    'ngo_activity_type': ('ops_ngo_activities', 'activity_type'),
    # Client-level
    'plan_type': ('plab_clients', 'plan_type'),
    'account_status': ('plab_clients', 'account_status'),
    'counsellor': ('plab_clients', 'counsellor'),
    'current_stage': ('plab_clients', 'current_stage'),
    'plab_stage': ('plab_clients', 'current_stage'),
    'switched_program': ('plab_clients', 'switched_program'),
    'joined_stage': ('plab_clients', 'joined_stage'),
    'lead_source': ('plab_clients', 'lead_source'),
    'operations_referral': ('plab_clients', 'operations_referral'),
    # newly-added client + job-stage dropdowns
    # (city/state handled by the dedicated states/cities tables, not lookups)
    'upgraded_to': ('plab_clients', 'upgraded_to'),
    'job_confirmed_by': ('ops_job_stage', 'job_confirmed_by'),
}

# Vendor category -> (ops_table, column) for the vendor "in use?" check.
VENDOR_INUSE = {
    'Training Programs':       ('ops_coaching', 'vendor_provider'),
    'Research & Publications': ('ops_research_publication', 'research_provider'),
    'Online Subscriptions':   ('ops_online_subscriptions', 'online_subscription'),
    'Online Courses':         ('ops_online_courses', 'course_provider'),
    'Webinars':               ('ops_webinars_conferences', 'event_name'),
    'NGO Activities':         ('ops_ngo_activities', 'ngo_vendor_name'),
    'Cab Bookings':           ('ops_uk_cab_bookings', 'vendor'),
    'Mentorship':             ('ops_mentorship', 'program_provider'),
}


# ════════════════════════════════════════════════════════════════════
#  Excel parsing
# ════════════════════════════════════════════════════════════════════
def _norm(v):
    return str(v).strip() if v is not None else ''


def read_dropdown_sheet(path):
    """Return (variant, headers, per_header_values, per_header_tag).

    Auto-detects layout:
      Variant A: row0 = headers, rows 1+ = values (no tag row).
      Variant B: row0 = tag row, row1 = headers, rows 2+ = values.
    Decision: whichever of row0/row1 best matches the DATA-sheet headers is
    the header row.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sn = wb.sheetnames
    data_ws = wb[sn[0]]
    data_hdr = set(_norm(c.value) for c in next(data_ws.iter_rows(min_row=1, max_row=1)) if c.value)
    dd = next((s for s in sn if 'drop' in s.lower()), None)
    if dd is None:
        wb.close()
        return None, [], {}, {}
    ws = wb[dd]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return None, [], {}, {}
    r0 = [_norm(c) for c in rows[0]]
    r1 = [_norm(c) for c in rows[1]] if len(rows) > 1 else []
    r0_match = sum(1 for c in r0 if c in data_hdr)
    r1_match = sum(1 for c in r1 if c in data_hdr)
    variant_b = r1_match > r0_match
    hdr_row = 1 if variant_b else 0
    val_start = 2 if variant_b else 1
    headers = [_norm(c) for c in rows[hdr_row]]
    per_vals, per_tag = {}, {}
    for ci, h in enumerate(headers):
        if not h:
            continue
        vals, seen = [], set()
        for r in rows[val_start:]:
            v = _norm(r[ci]) if ci < len(r) else ''
            if v and v.lower() not in seen:
                seen.add(v.lower())
                vals.append(v)
        per_vals.setdefault(h, vals)
        tag = _norm(rows[0][ci]) if (variant_b and ci < len(rows[0])) else ''
        per_tag.setdefault(h, tag)
    return ('B' if variant_b else 'A'), headers, per_vals, per_tag


def read_data_pairs(path, vendor_col, deliverable_col):
    """Distinct (vendor, deliverable) pairs from the DATA sheet (sheet 0)."""
    import openpyxl
    if not vendor_col or not deliverable_col:
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    try:
        headers = [_norm(h) for h in next(it)]
    except StopIteration:
        wb.close()
        return []
    hi = {h: i for i, h in enumerate(headers)}
    vi, di = hi.get(vendor_col), hi.get(deliverable_col)
    if vi is None or di is None:
        wb.close()
        return []
    pairs, seen = [], set()
    for r in it:
        v = _norm(r[vi]) if vi < len(r) else ''
        d = _norm(r[di]) if di < len(r) else ''
        if not v or v.lower() in _NON_VENDOR:
            continue
        if not d:
            continue
        key = (v.lower(), d.lower())
        if key not in seen:
            seen.add(key)
            pairs.append((v, d))
    wb.close()
    return pairs


# ════════════════════════════════════════════════════════════════════
#  DB helpers (read-only in DRY mode)
# ════════════════════════════════════════════════════════════════════
def get_conn():
    """Return a DatabaseConnection via db.get_db(), or None if unavailable."""
    try:
        from db import get_db
        return get_db()
    except Exception as e:
        print(f"  [warn] no DB connection ({e}); in-use checks skipped, "
              f"deactivation counts will be best-effort.")
        return None


def _table_has_column(conn, table, column):
    if conn is None:
        return False
    try:
        rows = conn.execute(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_name = ? AND column_name = ?", (table, column)).fetchall()
        if rows:
            return True
    except Exception:
        try: conn.rollback()
        except Exception: pass
    try:
        rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
        return any((r[1] if not isinstance(r, dict) else r.get('name')) == column for r in rows)
    except Exception:
        try: conn.rollback()
        except Exception: pass
        return False


def existing_lookup_values(conn, category, pathway):
    """{value_lower: is_active} for (category, pathway)."""
    if conn is None:
        return {}
    try:
        rows = conn.execute(
            "SELECT value, is_active FROM lookup_options "
            "WHERE category = ? AND COALESCE(pathway,'plab') = ?",
            (category, pathway)).fetchall()
        return {(_norm(r['value']).lower()): bool(r['is_active']) for r in rows}
    except Exception:
        try: conn.rollback()
        except Exception: pass
        return {}


def inuse_values(conn, table, column, pathway):
    """Set of lower-cased values currently used in ops table for the pathway."""
    if conn is None or not _table_has_column(conn, table, column):
        return set()
    try:
        rows = conn.execute(
            f"SELECT DISTINCT TRIM({column}) AS v FROM {table} "
            f"WHERE COALESCE(pathway,'plab') = ? "
            f"  AND {column} IS NOT NULL AND TRIM({column}) <> ''",
            (pathway,)).fetchall()
        return set(_norm(r['v']).lower() for r in rows if r['v'])
    except Exception:
        try: conn.rollback()
        except Exception: pass
        return set()


def existing_vendor_names(conn, country, category):
    if conn is None:
        return {}
    try:
        rows = conn.execute(
            "SELECT name, is_active FROM vendors_providers "
            "WHERE country = ? AND category = ?", (country, category)).fetchall()
        return {(_norm(r['name']).lower()): bool(r['is_active']) for r in rows}
    except Exception:
        try: conn.rollback()
        except Exception: pass
        return {}


# ════════════════════════════════════════════════════════════════════
#  Per-section processing (plan only; writes gated on DRY)
# ════════════════════════════════════════════════════════════════════
def is_vendor_column(header, tag):
    t = (tag or '').lower()
    h = (header or '').lower()
    if any(vt in t for vt in VENDOR_TAGS):
        return True
    if any(k in h for k in VENDOR_HDR_KEYWORDS):
        return True
    return False


def plan_section(conn, pathway, sec):
    folder = PATHWAY_DIRS[pathway]
    path = os.path.join(folder, sec['file'])
    out = {
        'pathway': pathway, 'section': sec['section'], 'file': sec['file'],
        'exists': os.path.exists(path),
        'variant': None, 'field_map': {}, 'lookup_plan': {},
        'vendor_cols': [], 'vendor_plan': {}, 'pairs': 0, 'pair_samples': [],
        'unmapped': [], 'vendor_backed_routed': [],
    }
    if not out['exists']:
        return out

    variant, headers, per_vals, per_tag = read_dropdown_sheet(path)
    out['variant'] = variant
    country = COUNTRY[pathway]

    cols = sec['cols']
    # 1) Field dropdowns -> lookup_options
    for h in headers:
        vals = per_vals.get(h, [])
        tag = per_tag.get(h, '')
        cat = cols.get(h)

        # Vendor columns (explicit __VENDOR__ in config, or tag/header hint).
        if cat == '__VENDOR__' or (cat is None and is_vendor_column(h, tag) and vals):
            if h not in [c for c, _ in out['vendor_cols']]:
                out['vendor_cols'].append((h, len(vals)))
            continue

        if cat in HARDCODED_NONLOOKUP:
            # Form uses a hardcoded Python list, not lookup_options — seeding
            # would be a no-op. Report it so the user can decide.
            if vals:
                out['unmapped'].append(
                    (h, len(vals), tag or f'(hardcoded list {cat}, NOT lookup-backed)',
                     per_vals.get(h, [])[:4]))
            continue

        if cat is None:
            # Unmapped column that actually carries master values -> flag it.
            if vals:
                out['unmapped'].append((h, len(vals), tag, per_vals.get(h, [])[:4]))
            continue

        # Vendor-backed "categories" are NOT lookup_options — route to vendors.
        if cat in VENDOR_BACKED_CATEGORIES:
            out['vendor_backed_routed'].append((h, cat))
            if h not in [c for c, _ in out['vendor_cols']]:
                out['vendor_cols'].append((h, len(vals)))
            continue

        out['field_map'][h] = cat
        if not vals:
            continue

        master = {v.lower(): v for v in vals}
        existing = existing_lookup_values(conn, cat, pathway)
        tbl_col = CATEGORY_INUSE.get(cat)
        used = inuse_values(conn, tbl_col[0], tbl_col[1], pathway) if tbl_col else set()

        adds = [v for k, v in master.items() if k not in existing]
        stale = [k for k in existing if k not in master]
        keep_in_use = [k for k in stale if k in used]
        deactivate = [k for k in stale if k not in used and existing.get(k)]
        out['lookup_plan'][cat] = {
            'master': len(master), 'add': len(adds),
            'deactivate': len(deactivate), 'keep_in_use': len(keep_in_use),
            'inuse_check': bool(tbl_col), 'inuse_table': (tbl_col or ('?', '?'))[0],
        }

    # 2) Vendors -> vendors_providers (clean replace keep-in-use)
    vp_cat = sec['vp_category']
    if vp_cat and out['vendor_cols']:
        # union of values across all vendor columns on the dropdown sheet
        vmaster = {}
        for h, _n in out['vendor_cols']:
            for v in per_vals.get(h, []):
                if v and v.lower() not in _NON_VENDOR:
                    vmaster[v.lower()] = v
        existing_v = existing_vendor_names(conn, country, vp_cat)
        vt = VENDOR_INUSE.get(vp_cat)
        used_v = inuse_values(conn, vt[0], vt[1], pathway) if vt else set()
        adds = [v for k, v in vmaster.items() if k not in existing_v]
        stale = [k for k in existing_v if k not in vmaster]
        keep = [k for k in stale if k in used_v]
        deact = [k for k in stale if k not in used_v and existing_v.get(k)]
        out['vendor_plan'] = {
            'category': vp_cat, 'country': country, 'master': len(vmaster),
            'add': len(adds), 'deactivate': len(deact), 'keep_in_use': len(keep),
            'inuse_check': bool(vt), 'sample': list(vmaster.values())[:6],
        }

    # 3) Vendor -> deliverable pairs from DATA sheet
    pairs = read_data_pairs(path, sec.get('vendor_col'), sec.get('deliverable_col'))
    out['pairs'] = len(pairs)
    out['pair_samples'] = pairs[:5]
    return out


# ════════════════════════════════════════════════════════════════════
#  Reporting
# ════════════════════════════════════════════════════════════════════
def print_section(o):
    print(f"\n{'─'*72}")
    print(f"FILE: {o['file']}")
    print(f"  pathway={o['pathway']}  section='{o['section']}'  "
          f"sheet_variant={o['variant']}  exists={o['exists']}")
    if not o['exists']:
        print("  !! FILE NOT FOUND — skipped")
        return
    if o['field_map']:
        print("  FIELD -> CATEGORY map (lookup_options):")
        for h, cat in o['field_map'].items():
            p = o['lookup_plan'].get(cat, {})
            if p:
                print(f"    - {h:38} -> {cat:24} "
                      f"master={p['master']:3} add={p['add']:3} "
                      f"deact={p['deactivate']:3} keep_in_use={p['keep_in_use']:3}"
                      f"{'' if p['inuse_check'] else '  [no in-use table!]'}")
            else:
                print(f"    - {h:38} -> {cat:24} (no master values on sheet)")
    if o['vendor_backed_routed']:
        print("  VENDOR-BACKED columns (routed to vendors, NOT lookup_options):")
        for h, cat in o['vendor_backed_routed']:
            print(f"    - {h:38} (would be category '{cat}' but it is vendor-driven)")
    if o['vendor_cols']:
        print("  VENDOR columns found:", ", ".join(f"{h}({n})" for h, n in o['vendor_cols']))
    if o['vendor_plan']:
        vp = o['vendor_plan']
        print(f"  VENDORS -> vendors_providers  category='{vp['category']}' "
              f"country='{vp['country']}'")
        print(f"    master={vp['master']} add={vp['add']} "
              f"deactivate={vp['deactivate']} keep_in_use={vp['keep_in_use']}"
              f"{'' if vp['inuse_check'] else '  [no in-use table!]'}")
        print(f"    sample: {vp['sample']}")
    elif o['vendor_cols'] and not o['vendor_plan']:
        print("  (vendor columns present but section has no vp_category — "
              "vendors NOT seeded for this section)")
    print(f"  VENDOR->DELIVERABLE pairs (from data sheet): {o['pairs']}")
    if o['pair_samples']:
        for v, d in o['pair_samples']:
            print(f"      {v}  ->  {d}")
    if o['unmapped']:
        print("  UNMAPPED columns with master values (NEED YOUR DECISION):")
        for h, n, tag, sample in o['unmapped']:
            print(f"    - {h:38} n={n:3} tag={tag!r:18} {sample}")


# ════════════════════════════════════════════════════════════════════
#  WRITE PATH  (only runs when DRY=0).  Idempotent: re-running adds nothing.
#  Strategy = clean-replace KEEP-IN-USE. Never DELETE; only INSERT / set
#  is_active. Works on SQLite (local) and Postgres (go-live) via db.py.
# ════════════════════════════════════════════════════════════════════
def _lookup_rows(conn, category, pathway):
    """{value_lower: {'value':..,'is_active':bool,'sort_order':int}}"""
    rows = conn.execute(
        "SELECT value, is_active, COALESCE(sort_order,0) AS so FROM lookup_options "
        "WHERE category = ? AND COALESCE(pathway,'plab') = ?",
        (category, pathway)).fetchall()
    return {_norm(r['value']).lower(): {
        'value': _norm(r['value']), 'is_active': bool(r['is_active']), 'so': r['so']}
        for r in rows}


def apply_lookups(conn, rconn, category, pathway, master_ordered):
    """master_ordered: list of values in sheet order. Returns counts dict.

    `conn` = WRITE connection (also reads uncommitted same-txn state).
    `rconn` = separate READ-ONLY connection for the in-use check, whose
    failure+rollback can never disturb the write transaction.
    """
    existing = _lookup_rows(conn, category, pathway)
    tc = CATEGORY_INUSE.get(category)
    used = inuse_values(rconn, tc[0], tc[1], pathway) if tc else set()
    max_so = max([r['so'] for r in existing.values()], default=0)
    added = react = deact = kept = 0
    for v in master_ordered:
        k = v.lower()
        if k in existing:
            if not existing[k]['is_active']:
                conn.execute(
                    "UPDATE lookup_options SET is_active = ? "
                    "WHERE category = ? AND COALESCE(pathway,'plab') = ? AND LOWER(TRIM(value)) = ?",
                    (True, category, pathway, k))
                react += 1
        else:
            max_so += 1
            conn.execute(
                "INSERT INTO lookup_options (category, label, value, pathway, sort_order, is_active) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (category, v, v, pathway, max_so, True))
            added += 1
    master_keys = {v.lower() for v in master_ordered}
    for k, r in existing.items():
        if k in master_keys:
            continue
        if k in used:
            kept += 1
        elif r['is_active']:
            conn.execute(
                "UPDATE lookup_options SET is_active = ? "
                "WHERE category = ? AND COALESCE(pathway,'plab') = ? AND LOWER(TRIM(value)) = ?",
                (False, category, pathway, k))
            deact += 1
    return {'add': added, 'reactivate': react, 'deactivate': deact, 'keep_in_use': kept}


def apply_vendors(conn, rconn, country, category, master_ordered, pathway):
    rows = conn.execute(
        "SELECT name, is_active, COALESCE(sort_order,0) AS so FROM vendors_providers "
        "WHERE country = ? AND category = ?", (country, category)).fetchall()
    existing = {_norm(r['name']).lower(): {'is_active': bool(r['is_active']), 'so': r['so']} for r in rows}
    vt = VENDOR_INUSE.get(category)
    used = inuse_values(rconn, vt[0], vt[1], pathway) if vt else set()
    max_so = max([r['so'] for r in existing.values()], default=0)
    added = react = deact = kept = 0
    for v in master_ordered:
        k = v.lower()
        if k in existing:
            if not existing[k]['is_active']:
                conn.execute("UPDATE vendors_providers SET is_active = ? "
                             "WHERE country = ? AND category = ? AND LOWER(TRIM(name)) = ?",
                             (True, country, category, k))
                react += 1
        else:
            max_so += 1
            conn.execute(
                "INSERT INTO vendors_providers (name, country, category, is_active, sort_order) "
                "VALUES (?, ?, ?, ?, ?)", (v, country, category, True, max_so))
            added += 1
    master_keys = {v.lower() for v in master_ordered}
    for k, r in existing.items():
        if k in master_keys:
            continue
        if k in used:
            kept += 1
        elif r['is_active']:
            conn.execute("UPDATE vendors_providers SET is_active = ? "
                         "WHERE country = ? AND category = ? AND LOWER(TRIM(name)) = ?",
                         (False, country, category, k))
            deact += 1
    return {'add': added, 'reactivate': react, 'deactivate': deact, 'keep_in_use': kept}


def apply_pairs(conn, country, category, pairs):
    """Upsert vendor_service_map (vendor_id, service_name) from (vendor, deliverable)
    pairs. Vendor resolved by name within (country, category). Pairs whose vendor
    isn't a known vendor are skipped + counted (curated-list integrity)."""
    # vendor name -> id
    vrows = conn.execute("SELECT id, name FROM vendors_providers WHERE country = ? AND category = ?",
                         (country, category)).fetchall()
    vid = {_norm(r['name']).lower(): r['id'] for r in vrows}
    added = skipped = existing_n = 0
    # cache existing service names per vendor_id
    have = {}
    for vendor, deliverable in pairs:
        k = vendor.lower()
        if k not in vid:
            skipped += 1
            continue
        v_id = vid[k]
        if not v_id:          # defensive: vendor row without a usable id
            skipped += 1
            continue
        if v_id not in have:
            er = conn.execute("SELECT service_name FROM vendor_service_map WHERE vendor_id = ?",
                              (v_id,)).fetchall()
            have[v_id] = {_norm(r['service_name']).lower() for r in er}
        if deliverable.lower() in have[v_id]:
            existing_n += 1
            continue
        conn.execute("INSERT INTO vendor_service_map (vendor_id, service_name) VALUES (?, ?)",
                     (v_id, deliverable))
        have[v_id].add(deliverable.lower())
        added += 1
    return {'add': added, 'skipped_no_vendor': skipped, 'already': existing_n}


def collect_pathway(pathway):
    """Aggregate, across ALL sections of a pathway, the UNION master per
    (category) / (vp_category) plus all vendor->deliverable pairs.

    Aggregating first (instead of applying per-section) is what makes the run
    idempotent: a category fed by two sections — e.g. medical_speciality from
    Academic (96) + Observerships (36) — is applied ONCE against the union, so
    the two sections can't fight over which values are active.
    Returns (lookups, vendors, pairs) where:
      lookups  : {category: [ordered unique values]}
      vendors  : {vp_category: [ordered unique vendor names]}
      pairs    : {vp_category: [(vendor, deliverable), ...]}
    """
    folder = PATHWAY_DIRS[pathway]
    lookups, vendors, pairs = {}, {}, {}
    for sec in PATHWAY_SECTIONS[pathway]:
        path = os.path.join(folder, sec['file'])
        if not os.path.exists(path):
            continue
        _, headers, per_vals, per_tag = read_dropdown_sheet(path)
        cols = sec['cols']
        vp_cat = sec['vp_category']
        for h in headers:
            vals = per_vals.get(h, [])
            tag = per_tag.get(h, '')
            cat = cols.get(h)
            is_vendor = (cat == '__VENDOR__' or cat in VENDOR_BACKED_CATEGORIES
                         or (cat is None and is_vendor_column(h, tag) and vals))
            if is_vendor:
                if vp_cat:
                    bucket = vendors.setdefault(vp_cat, {})
                    for v in vals:
                        if v and v.lower() not in _NON_VENDOR:
                            bucket.setdefault(v.lower(), v)
                continue
            if cat is None or cat in HARDCODED_NONLOOKUP or not vals:
                continue
            bucket = lookups.setdefault(cat, {})
            for v in vals:
                bucket.setdefault(v.lower(), v)
        if vp_cat:
            pr = read_data_pairs(path, sec.get('vendor_col'), sec.get('deliverable_col'))
            if pr:
                pairs.setdefault(vp_cat, []).extend(pr)
    # collapse the dedupe-dicts back to ordered lists
    lookups = {c: list(d.values()) for c, d in lookups.items()}
    vendors = {c: list(d.values()) for c, d in vendors.items()}
    return lookups, vendors, pairs


def run_live(pathways):
    conn = get_conn()
    rconn = get_conn()   # separate read-only conn for in-use checks (failure-isolated)
    if conn is None:
        print("!! No DB connection — cannot write. Set DATABASE_URL for prod.")
        return
    grand = {k: 0 for k in ('lk_add', 'lk_react', 'lk_deact', 'lk_keep',
                            'vn_add', 'vn_react', 'vn_deact', 'vn_keep',
                            'pairs_add', 'pairs_skip')}
    for pw in pathways:
        if pw not in PATHWAY_SECTIONS:
            print(f"[skip] unknown pathway '{pw}'"); continue
        print(f"\n{'#'*60}\n#  WRITING PATHWAY: {pw}\n{'#'*60}")
        lookups, vendors, pairs = collect_pathway(pw)
        country = COUNTRY[pw]
        # 1) dropdowns — one apply per category against the union
        for cat, vals in lookups.items():
            try:
                r = apply_lookups(conn, rconn, cat, pw, vals); conn.commit()
            except Exception as e:
                conn.rollback(); print(f"  lookup '{cat}' FAILED: {e}"); continue
            grand['lk_add'] += r['add']; grand['lk_react'] += r['reactivate']
            grand['lk_deact'] += r['deactivate']; grand['lk_keep'] += r['keep_in_use']
        # 2) vendors — one apply per vp_category
        for vp_cat, vals in vendors.items():
            try:
                r = apply_vendors(conn, rconn, country, vp_cat, vals, pw); conn.commit()
            except Exception as e:
                conn.rollback(); print(f"  vendors '{vp_cat}' FAILED: {e}"); continue
            grand['vn_add'] += r['add']; grand['vn_react'] += r['reactivate']
            grand['vn_deact'] += r['deactivate']; grand['vn_keep'] += r['keep_in_use']
        # 3) pairs — after vendors exist
        for vp_cat, pr in pairs.items():
            try:
                p = apply_pairs(conn, country, vp_cat, pr); conn.commit()
            except Exception as e:
                conn.rollback(); print(f"  pairs '{vp_cat}' FAILED: {e}"); continue
            grand['pairs_add'] += p['add']; grand['pairs_skip'] += p['skipped_no_vendor']
        print(f"  [{pw}] {len(lookups)} dropdown categories, {len(vendors)} vendor groups, "
              f"{sum(len(v) for v in pairs.values())} pairs processed")
    print(f"\n{'='*60}\n  LIVE WRITE COMPLETE\n{'='*60}")
    print(f"  lookup_options : add={grand['lk_add']} reactivate={grand['lk_react']} "
          f"deactivate={grand['lk_deact']} keep_in_use={grand['lk_keep']}")
    print(f"  vendors        : add={grand['vn_add']} reactivate={grand['vn_react']} "
          f"deactivate={grand['vn_deact']} keep_in_use={grand['vn_keep']}")
    print(f"  vendor_service_map pairs: add={grand['pairs_add']} skipped={grand['pairs_skip']}")
    conn.close()
    if rconn is not None:
        try: rconn.close()
        except Exception: pass


def main():
    pathways = os.environ.get('PATHWAYS', 'plab,australia').split(',')
    pathways = [p.strip() for p in pathways if p.strip()]
    print("=" * 72)
    print(f"  import_master_dropdowns_vendors.py   MODE={'DRY-RUN (no writes)' if DRY else 'LIVE (writes!)'}")
    print(f"  pathways={pathways}")
    print("=" * 72)
    if not DRY:
        print("\n!!! LIVE MODE — writing to the database now.\n")
        run_live([p for p in pathways if p in PATHWAY_SECTIONS])
        return

    conn = get_conn()

    # Cross-check: every targeted lookup category should be a real form category.
    FORM_CATEGORIES = set(CATEGORY_INUSE.keys()) | VENDOR_BACKED_CATEGORIES
    targeted = set()
    grand = {'lk_add': 0, 'lk_deact': 0, 'lk_keep': 0,
             'vn_add': 0, 'vn_deact': 0, 'vn_keep': 0, 'pairs': 0, 'unmapped': 0}
    unknown_cats = set()

    for pw in pathways:
        if pw not in PATHWAY_SECTIONS:
            print(f"\n[skip] unknown pathway '{pw}'")
            continue
        print(f"\n{'#'*72}\n#  PATHWAY: {pw}  (dir={PATHWAY_DIRS[pw]})\n{'#'*72}")
        for sec in PATHWAY_SECTIONS[pw]:
            o = plan_section(conn, pw, sec)
            print_section(o)
            for cat, p in o['lookup_plan'].items():
                targeted.add(cat)
                if cat not in FORM_CATEGORIES:
                    unknown_cats.add(cat)
                grand['lk_add'] += p['add']; grand['lk_deact'] += p['deactivate']
                grand['lk_keep'] += p['keep_in_use']
            if o['vendor_plan']:
                grand['vn_add'] += o['vendor_plan']['add']
                grand['vn_deact'] += o['vendor_plan']['deactivate']
                grand['vn_keep'] += o['vendor_plan']['keep_in_use']
            grand['pairs'] += o['pairs']
            grand['unmapped'] += len(o['unmapped'])

    print(f"\n{'='*72}\n  GRAND TOTALS (DRY — nothing written)\n{'='*72}")
    print(f"  lookup_options : add={grand['lk_add']}  deactivate={grand['lk_deact']}  "
          f"keep_in_use={grand['lk_keep']}")
    print(f"  vendors        : add={grand['vn_add']}  deactivate={grand['vn_deact']}  "
          f"keep_in_use={grand['vn_keep']}")
    print(f"  vendor->deliverable pairs (would upsert): {grand['pairs']}")
    print(f"  unmapped columns flagged: {grand['unmapped']}")

    # Category sanity check
    print(f"\n  CATEGORY SANITY CHECK")
    bad = sorted(c for c in targeted if c not in FORM_CATEGORIES)
    if bad:
        print(f"    !! categories targeted but NOT found in any form/_form_lookups: {bad}")
    else:
        print("    OK — every targeted lookup category maps to a known form category.")

    if conn is not None:
        try: conn.close()
        except Exception: pass


if __name__ == '__main__':
    main()
