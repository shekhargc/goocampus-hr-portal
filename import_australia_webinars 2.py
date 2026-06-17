"""
One-time import: Australia webinars & conferences from Excel into
ops_webinars_conferences.

Each Excel row goes into ops_webinars_conferences with pathway='australia'.

Linking strategy:
  The Excel's "Candidate Name" cell contains both the candidate name AND
  the registration number, e.g.:
      " Dr. Anvitha Rose George - GCAUSIP/24-25/107"
      " Dr.  Chinmay Sakharekar - GCAUSIP/25-26/36"
  We parse out the GCAUSIP-prefixed reg number and store it in
  registration_number so existing list/edit pages that JOIN against
  plab_clients keep working unchanged. If we cannot find the reg number
  inline, we fall back to a fuzzy name match against plab_clients with
  pathway='australia'.

Idempotent: keyed on (registration_number, event_name, start_date,
provider). Re-running with no marker change is a no-op. Bumping
IMPORT_VERSION re-runs the upsert. Bad rows (missing reg number, no name
match) are skipped with a warning so one bad row never blocks the whole
import.

Hooked into app boot from app.py the same way as
import_australia_test_bookings.py.
"""

import logging
import os
import re
from datetime import datetime, date


EXCEL_FILENAME = 'All_Australia_Webinars.xlsx'
IMPORT_VERSION = 'au_webinars_v1_initial_106'


H = {
    'cand_name':    'Candidate Name',
    'event_type':   'Event Type',
    'part_type':    'Participation Type',
    'event_value':  'Event Value',
    'start_date':   'Start Date',
    'end_date':     'End Date',
    'duration':     'Durantion (Number of Days)',
    'provider':     'Provider Name',
    'event_name':   'Webinar & Conference Name',
    'points':       'Points',
    'notes':        'Notes',
}


_REG_RE = re.compile(r'(GCAUSIP/\S+)', re.IGNORECASE)


def _extract_reg_number(candidate_cell):
    """Pull the GCAUSIP/... reg number out of the combined name+reg cell.

    Returns '' if none found. The cell can have stray spaces and dashes
    around the reg number, so we use a generous regex and strip noise.
    """
    if not candidate_cell:
        return ''
    s = str(candidate_cell)
    m = _REG_RE.search(s)
    if not m:
        return ''
    raw = m.group(1).rstrip('.,;:-/ \t')
    return raw.upper().replace('GCAUSIP', 'GCAUSIP')


def _extract_name_parts(candidate_cell):
    """Best-effort split of the candidate cell into (prefix, first, last).

    The cell looks like " Dr. Anvitha Rose George - GCAUSIP/24-25/107".
    We strip the trailing " - GCAUSIP/..." chunk, then split the leading
    "Dr."-style prefix off, and finally split the remaining words into
    first/last for fuzzy matching.
    """
    if not candidate_cell:
        return ('', '', '')
    s = str(candidate_cell).strip()
    # Cut off the reg-number suffix if present.
    s = re.sub(r'\s*-\s*GCAUSIP/.*$', '', s, flags=re.IGNORECASE).strip()
    # Pull a leading prefix (Dr., Mr., Mrs., Ms., etc.)
    prefix = ''
    m = re.match(r'^(Dr\.?|Mr\.?|Mrs\.?|Ms\.?|Miss)\s+', s, flags=re.IGNORECASE)
    if m:
        prefix = m.group(1)
        s = s[m.end():].strip()
    parts = [p for p in re.split(r'\s+', s) if p]
    if not parts:
        return (prefix, '', '')
    if len(parts) == 1:
        return (prefix, parts[0], '')
    return (prefix, parts[0], parts[-1])


def _ss(v):
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _sd(v):
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def _fuzzy_match_reg(conn, first, last):
    """Try to look up a registration_number from plab_clients by name.

    Australia pathway only. Match is case-insensitive on TRIM of first
    and last name. Returns '' if zero or more than one candidate matches
    (we don't guess between ambiguous matches).
    """
    if not first and not last:
        return ''
    try:
        if first and last:
            rows = conn.execute(
                """SELECT registration_number FROM plab_clients
                    WHERE pathway = 'australia'
                      AND LOWER(TRIM(first_name)) = LOWER(?)
                      AND LOWER(TRIM(last_name)) = LOWER(?)""",
                (first.strip(), last.strip()),
            ).fetchall()
        elif first:
            rows = conn.execute(
                """SELECT registration_number FROM plab_clients
                    WHERE pathway = 'australia'
                      AND LOWER(TRIM(first_name)) = LOWER(?)""",
                (first.strip(),),
            ).fetchall()
        else:
            rows = conn.execute(
                """SELECT registration_number FROM plab_clients
                    WHERE pathway = 'australia'
                      AND LOWER(TRIM(last_name)) = LOWER(?)""",
                (last.strip(),),
            ).fetchall()
        if len(rows) == 1:
            return rows[0]['registration_number'] or ''
        return ''
    except Exception:
        return ''


def run_import_australia_webinars_once(get_db_fn):
    """Import Australia webinars & conferences into ops_webinars_conferences."""
    import openpyxl

    excel_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        EXCEL_FILENAME,
    )
    if not os.path.exists(excel_path):
        logging.info("Australia webinars import: no Excel found, skipping")
        return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

    conn = get_db_fn()
    try:
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS _import_markers (key TEXT PRIMARY KEY, value TEXT)")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        marker = conn.execute(
            "SELECT value FROM _import_markers WHERE key = 'au_webinars_import'"
        ).fetchone()
        if marker and marker['value'] == IMPORT_VERSION:
            logging.info(f"Australia webinars import: already at {IMPORT_VERSION}, skipping")
            return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

        logging.info(f"Australia webinars import: starting {IMPORT_VERSION}...")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h is not None}

        def cell(row, key):
            i = idx.get(H[key])
            return row[i] if i is not None and i < len(row) else None

        inserted = 0
        updated = 0
        skipped = 0
        errors = 0

        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row_cells):
                continue
            cand_cell = cell(row_cells, 'cand_name')
            reg_num = _extract_reg_number(cand_cell)
            if not reg_num:
                # Fallback: fuzzy match by name against plab_clients.
                _prefix, first, last = _extract_name_parts(cand_cell)
                reg_num = _fuzzy_match_reg(conn, first, last)
            if not reg_num:
                skipped += 1
                logging.warning(
                    f"Australia webinars: could not resolve reg for cell {cand_cell!r}"
                )
                continue

            event_type = _ss(cell(row_cells, 'event_type'))
            start_date = _sd(cell(row_cells, 'start_date'))
            event_name = _ss(cell(row_cells, 'event_name'))
            provider = _ss(cell(row_cells, 'provider'))

            data = {
                'registration_number': reg_num,
                'event_type':          event_type,
                'start_date':          start_date,
                'end_date':            _sd(cell(row_cells, 'end_date')),
                'duration_days':       _ss(cell(row_cells, 'duration')),
                'event_value':         _ss(cell(row_cells, 'event_value')),
                'cpd_points':          _ss(cell(row_cells, 'points')),
                'event_name':          event_name,
                'participation_type':  _ss(cell(row_cells, 'part_type')),
                'notes':               _ss(cell(row_cells, 'notes')),
                'pathway':             'australia',
            }
            # Stash provider name in notes (no dedicated column on the table).
            if provider:
                existing_notes = data.get('notes', '')
                provider_note = f"Provider: {provider}"
                data['notes'] = (
                    f"{provider_note} | {existing_notes}" if existing_notes else provider_note
                )

            try:
                existing = conn.execute(
                    """SELECT id FROM ops_webinars_conferences
                        WHERE registration_number = ?
                          AND COALESCE(event_name, '') = ?
                          AND COALESCE(start_date, '') = ?
                          AND COALESCE(event_type, '') = ?
                          AND pathway = 'australia'""",
                    (reg_num, event_name, start_date, event_type),
                ).fetchone()
                if existing:
                    sets = ', '.join(f"{k} = ?" for k in data.keys())
                    conn.execute(
                        f"UPDATE ops_webinars_conferences SET {sets} WHERE id = ?",
                        list(data.values()) + [existing['id']],
                    )
                    updated += 1
                else:
                    cols = ', '.join(data.keys())
                    placeholders = ', '.join(['?'] * len(data))
                    conn.execute(
                        f"INSERT INTO ops_webinars_conferences ({cols}) VALUES ({placeholders})",
                        list(data.values()),
                    )
                    inserted += 1
                conn.commit()
            except Exception as e:
                errors += 1
                logging.warning(
                    f"Australia webinars row error ({reg_num} / {event_name}): {e}"
                )
                try:
                    conn.rollback()
                except Exception:
                    pass

        # Marker
        try:
            conn.execute(
                "INSERT INTO _import_markers (key, value) VALUES ('au_webinars_import', ?) "
                "ON CONFLICT (key) DO UPDATE SET value = excluded.value",
                (IMPORT_VERSION,),
            )
            conn.commit()
        except Exception:
            try:
                conn.execute("DELETE FROM _import_markers WHERE key = 'au_webinars_import'")
                conn.execute(
                    "INSERT INTO _import_markers (key, value) VALUES ('au_webinars_import', ?)",
                    (IMPORT_VERSION,),
                )
                conn.commit()
            except Exception as e:
                logging.warning(f"Australia webinars: could not write marker: {e}")

        logging.info(
            f"Australia webinars {IMPORT_VERSION} complete — "
            f"inserted={inserted} updated={updated} skipped={skipped} errors={errors}"
        )
        wb.close()
        return {'inserted': inserted, 'updated': updated, 'skipped': skipped, 'errors': errors}
    finally:
        try:
            conn.close()
        except Exception:
            pass
