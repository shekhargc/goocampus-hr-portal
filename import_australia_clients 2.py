"""
One-time import: Load Australia (AMC) clients from Excel into plab_clients.

Same pattern as import_excel_clients.py (PLAB importer). The user locked
2026-05-31 that Australia clients live in plab_clients with
pathway='australia' (legacy PLAB schema extended with a pathway column).

Excel source: GC_AUS_Registration_Report.xlsx (committed alongside the
PLAB Reg-List). 228 rows × 50 columns at the time of writing.

Idempotent: upserts by registration_number, skips rows with no reg number,
sets pathway='australia' on every row it touches.

Invoked from app.py on boot via run_import_australia_clients_once() with
a version marker so it only does work when the marker changes.
"""

import logging
import os
import re
from datetime import datetime, date


EXCEL_FILENAME = 'GC_AUS_Registration_Report.xlsx'
IMPORT_VERSION = 'au_v1_initial_228'


# Headers used in the user's Australia Excel (note the typo "Registation").
H = {
    'customer_id':       'Customer ID',
    'reg_number':        'Registation Number',          # NB: Excel typo (missing r)
    'candidate_name':    'Candidate Name',
    'mobile':            'Mobile Number',
    'email':             'Candidate Email',
    'reg_date':          'Registration Date (Payment Date)',
    'plan_type':         'Plan Type',
    'package':           'Package (Mention Actual Package)',
    'final_package':     'Final Package',
    'discount':          'Discount Allowed (Discount Offered)',
    'account_status':    'Account Status',
    'counsellor':        'Counsellor Name',
    'current_stage':     'Stage (Current Status)',
    'switched_program':  'Switched Program',
    'dob':               'D.O.B',
    'city':              'CITY',
    'state':             'STATE',
    'joined_service':    'Joined Service',
    'lead_source':       'Lead Source',
    'whatsapp':          'Whats App Number',
    'instagram':         'Instgram Account Name',         # NB: Excel typo (missing a)
    'facebook':          'Facebook Account Name',
    'linkedin':          'LinkedIn Account Name',
    'father_name':       'Fathers Name',
    'father_phone':      'Fathers Mobile Number',
    'mother_name':       'Mothers Name',
    'mother_phone':      'Mothers Mobile Number',
    'parents_email':     'Parents Email ID',
    'counsellor_number': 'Counsellor Number',
    'counsellor_email':  'Counsellor Email',
    'discount_notes':    'Discount Notes',
    'inst1_amt':         '1st Installment',
    'inst1_date':        '1st Installment Date',
    'inst1_note':        '1st Installment Note',
    'inst2_amt':         '2nd Installment',
    'inst2_date':        '2nd Installment Date',
    'inst2_note':        '2nd Installment Note',
    'inst3_amt':         '3rd Installment',
    'inst3_date':        '3rd Installment Date',
    'inst3_note':        '3rd Installment Note',
    'inst4_amt':         '4th Installment',
    'inst4_date':        '4th Installment Date',
    'inst4_note':        '4th Installment Note',
    'notes':             'Notes',
}


def _ss(v):
    """Safe string: strip, drop None, collapse stray int-floats to int."""
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _sf(v):
    """Safe float: 0 if parse fails or None."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '').replace('₹', '').replace('$', '')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _sd(v):
    """Safe date: ISO YYYY-MM-DD or empty."""
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


_PREFIX_TOKENS = {'Dr', 'Dr.', 'Mr', 'Mr.', 'Mrs', 'Mrs.', 'Ms', 'Ms.', 'Prof', 'Prof.'}


def _split_name(cname):
    """Split 'Dr. First Last' into (prefix, first, last). Defaults prefix='Dr.'."""
    cname = (cname or '').strip()
    if not cname:
        return 'Dr.', '', ''
    parts = cname.split()
    prefix = 'Dr.'
    if parts and parts[0].rstrip('.') in {'Dr', 'Mr', 'Mrs', 'Ms', 'Prof'}:
        prefix = parts[0] if parts[0].endswith('.') else parts[0] + '.'
        parts = parts[1:]
    if not parts:
        return prefix, '', ''
    if len(parts) == 1:
        return prefix, parts[0], ''
    return prefix, parts[0], ' '.join(parts[1:])


def run_import_australia_clients_once(get_db_fn):
    """Import Australia clients from Excel into plab_clients (pathway='australia').

    Args:
        get_db_fn: callable returning a fresh DB connection (matches the
            project's get_db() pattern so this module doesn't import db
            directly and stays decoupled).

    Returns:
        dict with inserted / updated / skipped / errors counts. Returns
        zeros (and logs a single info line) when the marker says we're
        already up to date.
    """
    import openpyxl

    excel_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        EXCEL_FILENAME,
    )
    if not os.path.exists(excel_path):
        logging.info("Australia import: no Excel found, skipping")
        return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

    conn = get_db_fn()
    try:
        # Marker table reused from the PLAB importer.
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS _import_markers (key TEXT PRIMARY KEY, value TEXT)")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        marker = conn.execute(
            "SELECT value FROM _import_markers WHERE key = 'au_excel_import'"
        ).fetchone()
        if marker and marker['value'] == IMPORT_VERSION:
            logging.info(f"Australia import: already at {IMPORT_VERSION}, skipping")
            return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

        logging.info(f"Australia import: starting {IMPORT_VERSION}...")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        # Build a header→col-index map so cell lookup by header name is O(1).
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h is not None}

        def cell(row, key):
            i = idx.get(H[key])
            return row[i] if i is not None and i < len(row) else None

        inserted = 0
        updated = 0
        skipped = 0
        errors = 0
        seen_regs = set()

        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row_cells):
                continue
            reg_num = _ss(cell(row_cells, 'reg_number'))
            if not reg_num:
                skipped += 1
                continue
            if reg_num in seen_regs:
                # Same reg number twice in the same file — keep the first row,
                # log the dupe but don't break.
                skipped += 1
                continue
            seen_regs.add(reg_num)

            prefix, first_name, last_name = _split_name(_ss(cell(row_cells, 'candidate_name')))

            data = {
                'customer_id':       _ss(cell(row_cells, 'customer_id')),
                'registration_number': reg_num,
                'registration_date': _sd(cell(row_cells, 'reg_date')),
                'prefix':            prefix,
                'first_name':        first_name or '(missing)',
                'last_name':         last_name,
                'mobile':            _ss(cell(row_cells, 'mobile')),
                'whatsapp1':         _ss(cell(row_cells, 'whatsapp')),
                'email':             _ss(cell(row_cells, 'email')),
                'dob':               _sd(cell(row_cells, 'dob')),
                'city':              _ss(cell(row_cells, 'city')),
                'state':             _ss(cell(row_cells, 'state')),
                'instagram':         _ss(cell(row_cells, 'instagram')),
                'facebook':          _ss(cell(row_cells, 'facebook')),
                'linkedin':          _ss(cell(row_cells, 'linkedin')),
                'father_name':       _ss(cell(row_cells, 'father_name')),
                'father_phone':      _ss(cell(row_cells, 'father_phone')),
                'mother_name':       _ss(cell(row_cells, 'mother_name')),
                'mother_phone':      _ss(cell(row_cells, 'mother_phone')),
                'parents_email':     _ss(cell(row_cells, 'parents_email')),
                'plan_type':         _ss(cell(row_cells, 'plan_type')),
                'account_status':    _ss(cell(row_cells, 'account_status')),
                'current_stage':     _ss(cell(row_cells, 'current_stage')),
                'switched_program':  _ss(cell(row_cells, 'switched_program')),
                'counsellor':        _ss(cell(row_cells, 'counsellor')),
                'counsellor_email':  _ss(cell(row_cells, 'counsellor_email')),
                'counsellor_number': _ss(cell(row_cells, 'counsellor_number')),
                'lead_source':       _ss(cell(row_cells, 'lead_source')),
                'package_amount':    _sf(cell(row_cells, 'package')),
                'discount_allowed':  _sf(cell(row_cells, 'discount')),
                'final_package':     _sf(cell(row_cells, 'final_package')),
                'additional_package_notes': _ss(cell(row_cells, 'discount_notes')),
                'inst1_amount':      _sf(cell(row_cells, 'inst1_amt')),
                'inst1_date':        _sd(cell(row_cells, 'inst1_date')),
                'inst1_note':        _ss(cell(row_cells, 'inst1_note')),
                'inst2_amount':      _sf(cell(row_cells, 'inst2_amt')),
                'inst2_date':        _sd(cell(row_cells, 'inst2_date')),
                'inst2_note':        _ss(cell(row_cells, 'inst2_note')),
                'inst3_amount':      _sf(cell(row_cells, 'inst3_amt')),
                'inst3_date':        _sd(cell(row_cells, 'inst3_date')),
                'inst3_note':        _ss(cell(row_cells, 'inst3_note')),
                'inst4_amount':      _sf(cell(row_cells, 'inst4_amt')),
                'inst4_date':        _sd(cell(row_cells, 'inst4_date')),
                'inst4_note':        _ss(cell(row_cells, 'inst4_note')),
                'pathway':           'australia',
            }

            try:
                existing = conn.execute(
                    "SELECT id FROM plab_clients WHERE registration_number = ?",
                    (reg_num,),
                ).fetchone()
                if existing:
                    sets = ', '.join(f"{k} = ?" for k in data.keys())
                    conn.execute(
                        f"UPDATE plab_clients SET {sets} WHERE id = ?",
                        list(data.values()) + [existing['id']],
                    )
                    updated += 1
                else:
                    cols = ', '.join(data.keys())
                    placeholders = ', '.join(['?'] * len(data))
                    conn.execute(
                        f"INSERT INTO plab_clients ({cols}) VALUES ({placeholders})",
                        list(data.values()),
                    )
                    inserted += 1
                conn.commit()
            except Exception as e:
                errors += 1
                logging.warning(f"Australia import row error ({reg_num}): {e}")
                try:
                    conn.rollback()
                except Exception:
                    pass

        # Persist the marker so reboots skip the work.
        try:
            conn.execute(
                "INSERT INTO _import_markers (key, value) VALUES ('au_excel_import', ?) "
                "ON CONFLICT (key) DO UPDATE SET value = excluded.value",
                (IMPORT_VERSION,),
            )
            conn.commit()
        except Exception:
            # SQLite doesn't support ON CONFLICT in older versions — fall back.
            try:
                conn.execute("DELETE FROM _import_markers WHERE key = 'au_excel_import'")
                conn.execute(
                    "INSERT INTO _import_markers (key, value) VALUES ('au_excel_import', ?)",
                    (IMPORT_VERSION,),
                )
                conn.commit()
            except Exception as e:
                logging.warning(f"Australia import: could not write marker: {e}")

        logging.info(
            f"Australia import {IMPORT_VERSION} complete — "
            f"inserted={inserted} updated={updated} skipped={skipped} errors={errors}"
        )
        wb.close()
        return {'inserted': inserted, 'updated': updated, 'skipped': skipped, 'errors': errors}
    finally:
        try:
            conn.close()
        except Exception:
            pass
