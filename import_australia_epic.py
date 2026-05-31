"""
One-time import: Australia EPIC registrations from Excel into
ops_epic_registration.

Each Excel row goes into ops_epic_registration with pathway='australia'.

Linking strategy:
  The Excel's "Enter Candidate Name" cell contains ONLY the candidate
  name (no embedded reg number, unlike the test-bookings sheet). So we
  resolve the candidate by case-insensitive trimmed name match against
  plab_clients WHERE pathway='australia'. We strip the leading "Dr." /
  "Dr" prefix and any stray whitespace, then try:
    1) exact match on LOWER(TRIM(prefix||' '||first_name||' '||last_name))
    2) exact match on LOWER(TRIM(first_name||' '||last_name))
  Rows that don't resolve are skipped with a warning so one bad row
  never blocks the whole import.

Idempotent: keyed on (registration_number, pathway). Re-running with no
marker change is a no-op. Bumping IMPORT_VERSION re-runs the upsert.

Hooked into app boot from app.py the same way as
import_australia_test_bookings.py.
"""

import logging
import os
import re
from datetime import datetime, date


EXCEL_FILENAME = 'All_Australia_Epic_Registrations.xlsx'
IMPORT_VERSION = 'au_epic_v1_initial_158'


H = {
    'cand_name':            'Enter Candidate Name',
    'epic_registration':    'EPIC Registration Status',
    'epic_status':          'EPIC Status',
    'notary_cam':           'Notary Cam',
    'reg_date':             'Registration Date',
    'docs_stage':           'Documents Stage',
    'docs_stage_status':    'Documents Stage Status',
    'notary_cam_login':     'Notary Cam Login',
    'added_user':           'Added User',
    'modified_user':        'Modified User',
    'added_time':           'Added Time',
    'login_id':             'Login ID',
    'sq1':                  'Secret Question 1',
    'sa1':                  'Secret Answer 1',
    'sq2':                  'Secret Question 2',
    'sa2':                  'Secret Answer 2',
    'sq3':                  'Secret Question 3',
    'sa3':                  'Secret Answer 3',
    'sq4':                  'Secret Question 4',
    'sa4':                  'Secret Answer 4',
    'login_pwd':            'Login Password',
    'notary_cam_password':  'Notary Cam Password',
    'epic_id_number':       'EPIC ID Number',
}


_REG_RE = re.compile(r'(GCAUSIP/\S+)', re.IGNORECASE)
_PREFIX_RE = re.compile(r'^\s*(dr\.?|mr\.?|mrs\.?|ms\.?|miss\.?)\s+', re.IGNORECASE)


def _extract_reg_number(candidate_cell):
    """If the cell happens to contain a GCAUSIP/... reg number, pull it.

    Returns '' if none found. Most EPIC cells are name-only — this is a
    belt-and-braces fallback in case some rows were entered with the
    test-bookings convention.
    """
    if not candidate_cell:
        return ''
    s = str(candidate_cell)
    m = _REG_RE.search(s)
    if not m:
        return ''
    raw = m.group(1).rstrip('.,;:-/ \t')
    return raw.upper()


def _normalize_name(s):
    """Lowercase, strip whitespace, collapse internal spaces, drop title."""
    if not s:
        return ''
    s = str(s)
    # Remove anything after the GCAUSIP marker if present
    s = _REG_RE.sub('', s)
    # Drop trailing dash before reg number (e.g. " - GCAUSIP/...")
    s = s.replace('-', ' ')
    s = _PREFIX_RE.sub('', s.strip())
    s = re.sub(r'\s+', ' ', s).strip().lower()
    return s


def _ss(v):
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _sd(v):
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def _build_name_index(conn):
    """Map normalized full-name -> registration_number for Australia clients.

    We build two lookups:
      - full: prefix + first + last (normalized)
      - short: first + last (normalized)
    `full` wins on collision; `short` is fallback.
    """
    full_idx = {}
    short_idx = {}
    rows = conn.execute(
        """SELECT registration_number,
                  COALESCE(prefix,'') AS prefix,
                  COALESCE(first_name,'') AS first_name,
                  COALESCE(last_name,'') AS last_name
             FROM plab_clients
            WHERE pathway = 'australia'
              AND registration_number IS NOT NULL
              AND registration_number != ''"""
    ).fetchall()
    for r in rows:
        full = _normalize_name(f"{r['prefix']} {r['first_name']} {r['last_name']}")
        short = _normalize_name(f"{r['first_name']} {r['last_name']}")
        if full and full not in full_idx:
            full_idx[full] = r['registration_number']
        if short and short not in short_idx:
            short_idx[short] = r['registration_number']
    return full_idx, short_idx


def run_import_australia_epic_once(get_db_fn):
    """Import Australia EPIC registrations into ops_epic_registration
    (pathway='australia').
    """
    import openpyxl

    excel_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        EXCEL_FILENAME,
    )
    if not os.path.exists(excel_path):
        logging.info("Australia EPIC import: no Excel found, skipping")
        return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

    conn = get_db_fn()
    try:
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS _import_markers (key TEXT PRIMARY KEY, value TEXT)")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        marker = conn.execute(
            "SELECT value FROM _import_markers WHERE key = 'au_epic_import'"
        ).fetchone()
        if marker and marker['value'] == IMPORT_VERSION:
            logging.info(f"Australia EPIC import: already at {IMPORT_VERSION}, skipping")
            return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

        logging.info(f"Australia EPIC import: starting {IMPORT_VERSION}...")
        full_idx, short_idx = _build_name_index(conn)

        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h is not None}

        def cell(row, key):
            i = idx.get(H[key])
            return row[i] if i is not None and i < len(row) else None

        inserted = 0
        updated = 0
        skipped = 0
        errors = 0

        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row_cells):
                continue

            cand_raw = cell(row_cells, 'cand_name')
            # Try reg-in-cell first (rare for EPIC, common for test bookings)
            reg_num = _extract_reg_number(cand_raw)
            if not reg_num:
                # Fuzzy name match against plab_clients
                norm = _normalize_name(cand_raw)
                if not norm:
                    skipped += 1
                    logging.warning(f"Australia EPIC: blank candidate name, skipping row")
                    continue
                reg_num = full_idx.get(norm) or short_idx.get(norm) or ''
                if not reg_num:
                    skipped += 1
                    logging.warning(
                        f"Australia EPIC: no plab_clients match for '{cand_raw}' "
                        f"(normalized='{norm}'), skipping"
                    )
                    continue

            data = {
                'registration_number':    reg_num,
                'epic_registration':      _ss(cell(row_cells, 'epic_registration')),
                'epic_status':            _ss(cell(row_cells, 'epic_status')),
                'notary_camp':            _ss(cell(row_cells, 'notary_cam')),
                'registration_date':      _sd(cell(row_cells, 'reg_date')),
                'documents_stage':        _ss(cell(row_cells, 'docs_stage')),
                'document_stage_status':  _ss(cell(row_cells, 'docs_stage_status')),
                'login_id':               _ss(cell(row_cells, 'login_id')),
                'login_pwd':              _ss(cell(row_cells, 'login_pwd')),
                'secret_question_1':      _ss(cell(row_cells, 'sq1')),
                'secret_answer_1':        _ss(cell(row_cells, 'sa1')),
                'secret_question_2':      _ss(cell(row_cells, 'sq2')),
                'secret_answer_2':        _ss(cell(row_cells, 'sa2')),
                'secret_question_3':      _ss(cell(row_cells, 'sq3')),
                'secret_answer_3':        _ss(cell(row_cells, 'sa3')),
                'secret_question_4':      _ss(cell(row_cells, 'sq4')),
                'secret_answer_4':        _ss(cell(row_cells, 'sa4')),
                'epic_id_number':         _ss(cell(row_cells, 'epic_id_number')),
                'notary_camp_login':      _ss(cell(row_cells, 'notary_cam_login')),
                'notary_camp_password':   _ss(cell(row_cells, 'notary_cam_password')),
                'pathway':                'australia',
            }

            try:
                existing = conn.execute(
                    "SELECT id FROM ops_epic_registration "
                    "WHERE registration_number = ? AND pathway = 'australia'",
                    (reg_num,),
                ).fetchone()
                if existing:
                    sets = ', '.join(f"{k} = ?" for k in data.keys())
                    conn.execute(
                        f"UPDATE ops_epic_registration SET {sets} WHERE id = ?",
                        list(data.values()) + [existing['id']],
                    )
                    updated += 1
                else:
                    cols = ', '.join(data.keys())
                    placeholders = ', '.join(['?'] * len(data))
                    conn.execute(
                        f"INSERT INTO ops_epic_registration ({cols}) VALUES ({placeholders})",
                        list(data.values()),
                    )
                    inserted += 1
                conn.commit()
            except Exception as e:
                errors += 1
                logging.warning(f"Australia EPIC row error ({reg_num}): {e}")
                try:
                    conn.rollback()
                except Exception:
                    pass

        # Marker
        try:
            conn.execute(
                "INSERT INTO _import_markers (key, value) VALUES ('au_epic_import', ?) "
                "ON CONFLICT (key) DO UPDATE SET value = excluded.value",
                (IMPORT_VERSION,),
            )
            conn.commit()
        except Exception:
            try:
                conn.execute("DELETE FROM _import_markers WHERE key = 'au_epic_import'")
                conn.execute(
                    "INSERT INTO _import_markers (key, value) VALUES ('au_epic_import', ?)",
                    (IMPORT_VERSION,),
                )
                conn.commit()
            except Exception as e:
                logging.warning(f"Australia EPIC: could not write marker: {e}")

        logging.info(
            f"Australia EPIC {IMPORT_VERSION} complete — "
            f"inserted={inserted} updated={updated} skipped={skipped} errors={errors}"
        )
        wb.close()
        return {'inserted': inserted, 'updated': updated, 'skipped': skipped, 'errors': errors}
    finally:
        try:
            conn.close()
        except Exception:
            pass
