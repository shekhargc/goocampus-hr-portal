"""
import_amc_sections.py — refresh AMC (australia) pathway data from Zoho exports.

Mirrors import_plab_sections.py but for pathway='australia' (the AMC pathway).
Steps (control with STEP=clients|sections|onboarding|lookups|all, default all):

  clients   -> upsert plab_clients (pathway='australia') from
               "GC AUS Registration Report (1).xlsx" (FK target, run FIRST).
  sections  -> for each section file: resolve client reg (embedded GCAUSIP
               reg, with padding/year variants -> canonical DB reg; else
               name/email/mobile match), wipe pathway='australia' rows in
               the target table, insert fresh rows.
  onboarding-> best-effort upsert into client_onboarding (welcome call/kit).
  lookups   -> add any NEW dropdown values + vendors/providers found in the
               imported data into lookup_options / vendors_providers
               (pathway='australia'), never deleting existing options.

Reg numbers are GCAUSIP/... . Section regs may be padded differently than
the client's stored reg (e.g. /08 vs /008, /2023 vs /23-24); we resolve to
the exact stored reg so the FK to plab_clients holds.

Run:
  DATABASE_URL=...  DOWNLOADS=/Users/Santosh/Downloads  python3 import_amc_sections.py
  DRY_RUN=1 -> parse + match only, no writes.
  SECTIONS=academic,test_bookings -> only some section keys.
"""
import os, re
from datetime import datetime, date

DL = os.environ.get('DOWNLOADS', '/Users/Santosh/Downloads')
PATHWAY = 'australia'
DRY = os.environ.get('DRY_RUN') in ('1', 'true', 'True')

# ── value converters ───────────────────────────────────────────────
def s(v):
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): return str(int(v))
    return str(v).strip() or None

def d(v):
    if v in (None, ''): return None
    if isinstance(v, (datetime, date)): return v.strftime('%Y-%m-%d')
    return str(v).strip() or None

def num(v):
    if v in (None, ''): return None
    try: return float(str(v).replace(',', '').replace('₹', '').strip())
    except (ValueError, TypeError): return None

def norm(x): return ' '.join((x or '').strip().lower().split())
def digits(x): return re.sub(r'\D', '', str(x or ''))

CONV = {'s': s, 'd': d, 'num': num}

# ── reg parsing + variants ─────────────────────────────────────────
_REG_RE = re.compile(r'(GC[A-Z]*/[0-9A-Za-z-]+/\d+)')
_REG_SPLIT = re.compile(r'^(GC[A-Z]*)/([0-9A-Za-z-]+)/(\d{1,4})$')

def extract_reg(text):
    if not text: return None
    m = _REG_RE.search(str(text))
    return m.group(1) if m else None

def strip_reg(text):
    if not text: return text
    return _REG_RE.sub('', str(text)).rstrip(' -').strip()

def reg_variants(reg):
    seen = set()
    def emit(v):
        if v not in seen: seen.add(v); return True
        return False
    if reg and emit(reg): yield reg
    m = _REG_SPLIT.match(reg or '')
    if not m: return
    prefix, middle, tail = m.group(1), m.group(2), m.group(3)
    try: n = int(tail)
    except ValueError: return
    middles = [middle]
    if '-' not in middle and middle.isdigit() and len(middle) == 4:
        y = int(middle)
        middles.append(f"{y%100:02d}-{(y+1)%100:02d}")
        middles.append(f"{(y-1)%100:02d}-{y%100:02d}")
    elif '-' in middle:
        a = middle.split('-', 1)[0]
        if a.isdigit() and len(a) == 2: middles.append(f"20{int(a):02d}")
    for md in middles:
        for pad in (3, 4, 2, 1):
            v = f"{prefix}/{md}/{n:0{pad}d}"
            if emit(v): yield v

# ── name splitting (for client upsert) ─────────────────────────────
_PREFIX_RE = re.compile(r'^(dr|dr\.|mr|mr\.|ms|ms\.|mrs|mrs\.|miss|prof|prof\.)\s+', re.I)
def split_name(full):
    full = strip_reg(full) or ''
    full = ' '.join(full.split())
    prefix = ''
    m = _PREFIX_RE.match(full)
    if m:
        prefix = full[:m.end()].strip()
        full = full[m.end():].strip()
    parts = full.split()
    if not parts: return prefix or 'Dr.', '', ''
    if len(parts) == 1: return prefix or 'Dr.', parts[0], ''
    return prefix or 'Dr.', parts[0], ' '.join(parts[1:])

def _pg():
    import psycopg2
    return psycopg2.connect(os.environ['DATABASE_URL'], connect_timeout=20,
        keepalives=1, keepalives_idle=30, keepalives_interval=10, keepalives_count=5)


# ── STEP: clients ──────────────────────────────────────────────────
CLIENT_FILE = "GC AUS Registration Report (1).xlsx"
# plab_clients col -> (excel header, kind)
CLIENT_MAP = {
    'customer_id': ('Customer ID', 's'),
    'registration_date': ('Registration Date (Payment Date)', 'd'),
    'mobile': ('Mobile Number', 's'), 'email': ('Candidate Email', 's'),
    'plan_type': ('Plan Type', 's'),
    'package_amount': ('Package (Mention Actual Package)', 'num'),
    'final_package': ('Final Package', 'num'),
    'discount_allowed': ('Discount Allowed (Discount Offer + Stage Discount)', 'num'),
    'account_status': ('Account Status', 's'), 'counsellor': ('Counsellor Name', 's'),
    'current_stage': ('Stage (Current Status)', 's'), 'switched_program': ('Switched Program', 's'),
    'dob': ('D.O.B', 'd'), 'city': ('CITY', 's'), 'joined_stage': ('Joined Service', 's'),
    'lead_source': ('Lead Source', 's'), 'state': ('STATE', 's'),
    'whatsapp1': ('Whats App Number', 's'),
    'dropped_date': ('Dropped Out / Switched Program Date', 'd'),
    'instagram': ('Instgram Account Name', 's'), 'facebook': ('Facebook Account Name', 's'),
    'linkedin': ('LinkedIn Account Name', 's'),
    'father_name': ('Fathers Name', 's'), 'father_phone': ('Fathers Mobile Number', 's'),
    'mother_name': ('Mothers Name', 's'), 'mother_phone': ('Mothers Mobile Number', 's'),
    'parents_email': ('Parents Email ID', 's'),
    'counsellor_number': ('Counsellor Number', 's'), 'counsellor_email': ('Counsellor Email', 's'),
    'portfolio_referral': ('GC Portfolio Client Name', 's'),
    'australia_referral': ('GC AUS Registration', 's'),
    'additional_package_notes': ('Discount Notes', 's'),
    'inst1_amount': ('1st Installment', 'num'), 'inst1_date': ('1st Installment Date', 'd'), 'inst1_note': ('1st Installment Note', 's'),
    'inst2_amount': ('2nd Installment', 'num'), 'inst2_date': ('2nd Installment Date', 'd'), 'inst2_note': ('2nd Installment Note', 's'),
    'inst3_amount': ('3rd Installment', 'num'), 'inst3_date': ('3rd Installment Date', 'd'), 'inst3_note': ('3rd Installment Note', 's'),
    'inst4_amount': ('4th Installment', 'num'), 'inst4_date': ('4th Installment Date', 'd'), 'inst4_note': ('4th Installment Note', 's'),
    'additional_notes': ('Notes', 's'),
}

def step_clients():
    import openpyxl
    path = os.path.join(DL, CLIENT_FILE)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else '' for h in next(it)]
    hi = {h: i for i, h in enumerate(headers)}
    def g(row, h):
        i = hi.get(h); return row[i] if (i is not None and i < len(row)) else None
    rows = [r for r in it if any(v not in (None, '') for v in r)]
    wb.close()

    conn = _pg(); cur = conn.cursor()
    cur.execute("SELECT registration_number FROM plab_clients WHERE COALESCE(pathway,'plab')='australia'")
    existing = set((r[0] or '').strip() for r in cur.fetchall())
    ins = upd = skip = 0
    for r in rows:
        reg = extract_reg(g(r, 'Registation Number')) or s(g(r, 'Registation Number'))
        if not reg: skip += 1; continue
        prefix, first, last = split_name(g(r, 'Candidate Name'))
        data = {'prefix': prefix, 'first_name': first, 'last_name': last}
        for col, (h, k) in CLIENT_MAP.items():
            data[col] = CONV[k](g(r, h))
        if DRY:
            (None)
            if reg in existing: upd += 1
            else: ins += 1
            continue
        if reg in existing:
            sets = ', '.join(f"{c}=%s" for c in data)
            cur.execute(f"UPDATE plab_clients SET {sets}, updated_at=NOW() WHERE registration_number=%s AND COALESCE(pathway,'plab')='australia'",
                        list(data.values()) + [reg]); upd += 1
        else:
            cols = ['registration_number', 'pathway'] + list(data.keys())
            vals = [reg, 'australia'] + list(data.values())
            cur.execute(f"INSERT INTO plab_clients ({','.join(cols)}) VALUES ({','.join(['%s']*len(cols))})", vals)
            existing.add(reg); ins += 1
    if not DRY: conn.commit()
    cur.close(); conn.close()
    print(f"[clients] inserted={ins} updated={upd} skipped(no reg)={skip}  (DRY={DRY})")


# ── index for section reg resolution ───────────────────────────────
def build_index():
    conn = _pg(); cur = conn.cursor()
    cur.execute("SELECT registration_number, prefix, first_name, last_name, email, mobile, id "
                "FROM plab_clients WHERE COALESCE(pathway,'plab')='australia'")
    by_reg = {}   # variant(norm) -> canonical reg
    by_name = {}; by_email = {}; by_mobile = {}; by_reg_cid = {}
    for reg, prefix, fn, ln, email, mob, cid in cur.fetchall():
        if not reg: continue
        canon = reg.strip()
        by_reg_cid[norm(canon)] = cid
        for v in reg_variants(canon):
            by_reg.setdefault(norm(v), canon)
        full = norm(f"{fn or ''} {ln or ''}")
        if full: by_name.setdefault(full, canon)
        full2 = norm(f"{ln or ''} {fn or ''}")
        if full2: by_name.setdefault(full2, canon)
        if email: by_email.setdefault(norm(email), canon)
        if mob:
            dm = digits(mob)[-10:]
            if dm: by_mobile.setdefault(dm, canon)
    cur.close(); conn.close()
    return by_reg, by_name, by_email, by_mobile, by_reg_cid

def _strip_prefix(text):
    """Drop a leading Dr./Mr./Ms. title so names match the index keys
    (which are built from first_name+last_name, prefix-less)."""
    if not text: return text
    return _PREFIX_RE.sub('', str(text).strip()).strip()

def resolve_reg(idx, reg_val, name_val, email_val, mobile_val):
    by_reg, by_name, by_email, by_mobile, _ = idx
    emb = extract_reg(name_val) or extract_reg(reg_val) or reg_val
    if emb:
        for v in reg_variants(emb):
            c = by_reg.get(norm(v))
            if c: return c
    nm = norm(_strip_prefix(strip_reg(name_val)))
    if nm and nm in by_name: return by_name[nm]
    if email_val and norm(email_val) in by_email: return by_email[norm(email_val)]
    if mobile_val:
        dm = digits(mobile_val)[-10:]
        if dm and dm in by_mobile: return by_mobile[dm]
    return None


# ── SECTIONS config: (file, table, reg_col, name_col, email_col, mobile_col, mapping) ──
SECTIONS = {
 'academic': ("All Academic Details (1).xlsx", "ops_academic_details", None, "Enter Name", None, None, {
    'img_fmg':('IMG / FMG','s'),'img_medical_college':('IMG Medical College Name','s'),
    'fmg_medical_college':('FMG Medical College Name','s'),'country':('Country','s'),
    'mbbs_status':('MBBS Status','s'),'mbbs_start_date':('MBBS Start Date','d'),'mbbs_end_date':('MBBS End Date','d'),
    'speciality_interest_1':('Speciality Interest 1','s'),'speciality_interest_2':('Speciality Interest 2','s'),
    'internship_status':('Internship Status','s'),'internship_hospital':('Internship Hospital','s'),
    'internship_location':('Internship Location (State/Country)','s'),
    'internship_hospital_2':('Internship Hospital 2','s'),
    'internship_location_2':('Internship Location 2 (State / Country)','s'),
    'internship_start_date':('Internship Start Date','d'),'internship_end_date':('Internship End Date','d'),
    'internship_gap':('Internship Gap','s'),'gap_in_months':('Internship Gap in Months','s'),
    'gap_reason':('Intership Gap Reason','s'),
    'working_status':('Working Status','s'),'working_hospital_name':('Working Hospital Name','s'),
    'additional_info':('Additional Info','s'),
 }),
 'amc_reg': ("All Amc Registrations (1).xlsx", "ops_amc_registration", None, "Enter Candidate Name", None, None, {
    'amc_reference_number':('AMC Reference Number','s'),'login_pwd':('Login Password','s'),
    'amc_setup':('AMC Setup','s'),'registration_date':('Registration Date','d'),
 }),
 'epic': ("All Epic Registrations (1).xlsx", "ops_epic_registration", None, "Enter Candidate Name", None, None, {
    'epic_registration':('EPIC Registration Status','s'),'epic_status':('EPIC Status','s'),
    'notary_camp':('Notary Cam','s'),'registration_date':('Registration Date','d'),
    'documents_stage':('Documents Stage','s'),'document_stage_status':('Documents Stage Status','s'),
    'notary_camp_login':('Notary Cam Login','s'),'login_id':('Login ID','s'),
    'secret_question_1':('Secret Question 1','s'),'secret_answer_1':('Secret Answer 1','s'),
    'secret_question_2':('Secret Question 2','s'),'secret_answer_2':('Secret Answer 2','s'),
    'secret_question_3':('Secret Question 3','s'),'secret_answer_3':('Secret Answer 3','s'),
    'secret_question_4':('Secret Question 4','s'),'secret_answer_4':('Secret Answer 4','s'),
    'login_pwd':('Login Password','s'),'notary_camp_password':('Notary Cam Password','s'),
    'epic_id_number':('EPIC ID Number','s'),
 }),
 'subscriptions': ("All Online Courses & Subscriptions (1).xlsx", "ops_online_subscriptions", None, "Candidate Name", "Client Email", None, {
    'online_subscription':('Course or Subscription Name','s'),'booked_by':('Booked By','s'),
    'activation_type':('Activation Type','s'),'issued_date':('Issued Date','d'),
    'client_email':('Client Email','s'),'login_id':('Login Id','s'),'password':('Password','s'),
    'notes':('Notes','s'),
 }),
 'payments': ("All Payments (1).xlsx", "ops_payments", "Registration Number", "Enter Candidate Name", None, None, {
    'payment_date':('Payment Date','d'),'total_package':('Total Package','num'),
    'instalment':('Instalment','s'),'total_amount_paid':('Total Amount Paid','num'),
    'amount_paid':('Amount Paid','num'),'gst_paid':('GST Paid','num'),
    'payment_method':('Payment Method','s'),'notes':('Notes','s'),
 }),
 'test_bookings': ("All Test Bookings (1).xlsx", "ops_test_bookings", None, "Enter Candidate Name", None, None, {
    'booked_by':('Booked By','s'),'exam':('Exam','s'),'booking_date':('Booking Date','d'),
    'exam_type':('Exam Method','s'),'exam_date':('Exam Date','d'),'country':('Country','s'),
    'city_state':('City / State','s'),'exam_status':('Exam Status','s'),'exam_result':('Exam Result','s'),
    'exam_result_date':('Exam Result Date','d'),'score':('Test Score','s'),
    'revaluation':('Revaluation Applied','s'),'reval_applied_date':('Reval Applied Date','d'),
    'reval_result':('Reval Result','s'),'reval_score':('Reval Score','s'),
 }),
 'training': ("All Trainings (1).xlsx", "ops_coaching", None, "Enter Candidate Name", "Candidate Email", None, {
    'english_training':('Training Program','s'),'course_type':('Course Type','s'),
    'coaching_method':('Training Method','s'),'coaching_status':('Training Status','s'),
    'batch_year':('Batch Year','s'),'batch_month':('Batch (Month)','s'),
    'start_date':('Start Date','d'),'end_date':('End Date','d'),
    'other_vendor':('Training Vendor Name','s'),
 }),
 'call_notes': ("Client Call Notes Report (2).xlsx", "ops_call_notes", None, "Candidate Name", None, None, {
    'call_date':('Call Date','d'),'call_note':('Call Notes','s'),'added_by':('Added User','s'),
 }),
 'research': ("Research and Publication Report (1).xlsx", "ops_research_publication", None, "Enter Candidate Name", None, None, {
    'research_status':('Research Status','s'),'research_start_date':('Research Start Date','d'),
    'research_topic':('Research Topic','s'),'research_batch':('Research Batch','s'),
    'research_end_date':('Research End Date','d'),'research_provider':('Research Provider','s'),
    'published_journal_name':('Published Journal Name','s'),'author_position':('Author Position','s'),
    'published_copy':('Upload Published Copy','s'),
 }),
 'webinars': ("Webinar and Conferences Report (1).xlsx", "ops_webinars_conferences", None, "Candidate Name", None, None, {
    'event_type':('Event Type','s'),'participation_type':('Participation Type','s'),
    'event_value':('Event Value','s'),'start_date':('Start Date','d'),'end_date':('End Date','d'),
    'duration_days':('Durantion (Number of Days)','s'),
    'event_name':('Webinar & Conference Name','s'),'cpd_points':('Points','s'),'notes':('Notes','s'),
 }),
}

def step_sections(idx):
    import openpyxl
    from psycopg2.extras import execute_values
    only = os.environ.get('SECTIONS')
    only = set(only.split(',')) if only else None
    summary = []
    for key, cfg in SECTIONS.items():
        if only and key not in only: continue
        fname, table, reg_col, name_col, email_col, mobile_col, mapping = cfg
        path = os.path.join(DL, fname)
        if not os.path.exists(path):
            print(f"[{key}] MISSING {fname}"); continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        hidx = {h: i for i, h in enumerate(headers)}
        cols = ['registration_number'] + list(mapping.keys()) + ['pathway']
        batch = []; unmatched = blank = 0; unm_samples = []
        for rcells in ws.iter_rows(min_row=2):
            vals = [c.value for c in rcells]
            if not any(v not in (None, '') for v in vals): continue
            def g(h):
                i = hidx.get(h); return vals[i] if (i is not None and i < len(vals)) else None
            reg_val = s(g(reg_col)) if reg_col else None
            name_val = s(g(name_col)) if name_col else None
            email_val = s(g(email_col)) if email_col else None
            mobile_val = s(g(mobile_col)) if mobile_col else None
            if not (reg_val or name_val or email_val or mobile_val): blank += 1; continue
            reg = resolve_reg(idx, reg_val, name_val, email_val, mobile_val)
            if not reg:
                unmatched += 1
                if len(unm_samples) < 8: unm_samples.append(name_val or reg_val)
                continue
            row_vals = [reg] + [CONV[k](g(h)) for (h, k) in mapping.values()] + ['australia']
            batch.append(tuple(row_vals))
        wb.close()
        if DRY:
            print(f"[{key:13s}] {table:24s} would_insert={len(batch):5d} unmatched={unmatched:4d} blank={blank}  {unm_samples[:3]}")
            summary.append((key, len(batch), unmatched)); continue
        conn = _pg(); conn.autocommit = False; cur = conn.cursor()
        try:
            cur.execute(f"DELETE FROM {table} WHERE COALESCE(pathway,'plab')='australia'")
            if batch:
                tmpl = "(" + ",".join(["%s"]*len(cols)) + ")"
                execute_values(cur, f"INSERT INTO {table} ({','.join(cols)}) VALUES %s", batch, template=tmpl, page_size=500)
            conn.commit()
            print(f"[{key:13s}] {table:24s} inserted={len(batch):5d} unmatched={unmatched:4d} blank={blank}")
            summary.append((key, len(batch), unmatched))
        except Exception as e:
            conn.rollback(); print(f"[{key}] FAILED: {e}"); summary.append((key, -1, unmatched))
        finally:
            cur.close(); conn.close()
    print("\n=== SECTION SUMMARY ===")
    for k, n, u in summary: print(f"  {k:14s} {n:6d} inserted, {u} unmatched")
    return summary


# ── STEP: onboarding (best-effort into client_onboarding) ──────────
def step_onboarding(idx):
    import openpyxl
    by_reg_cid = idx[4]
    path = os.path.join(DL, "OnBoarding Status (R) (1).xlsx")
    if not os.path.exists(path): print("[onboarding] MISSING file"); return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else '' for h in next(it)]
    hi = {h: i for i, h in enumerate(headers)}
    def g(row, h):
        i = hi.get(h); return row[i] if (i is not None and i < len(row)) else None
    upd = nomatch = 0
    conn = _pg(); cur = conn.cursor()
    for r in it:
        if not any(v not in (None, '') for v in r): continue
        reg = resolve_reg(idx, None, s(g(r, 'Enter Candidate Name')), s(g(r, 'Candidate Email')), s(g(r, 'Candidate Phone')))
        cid = by_reg_cid.get(norm(reg)) if reg else None
        if not cid: nomatch += 1; continue
        wc_by = s(g(r, 'Welcome Call by')); wc_date = d(g(r, 'Welcome Call Date'))
        kit_date = d(g(r, 'Welcome Kit Sent Date'))
        notes_bits = []
        for h in ['Welcome Kit', 'AMC Hand Book', 'AMC Hand Book sent Date', 'Brochure & Policy',
                  'SLA Copy', 'AMC Clinical Hand Book', 'AMC Clinical Hand Book Sent Date', 'Additional Goodies']:
            v = s(g(r, h))
            if v: notes_bits.append(f"{h}: {v}")
        wc_notes = ' | '.join(notes_bits) or None
        if DRY: upd += 1; continue
        cur.execute("SELECT id FROM client_onboarding WHERE client_id=%s", (cid,))
        ex = cur.fetchone()
        if ex:
            cur.execute("""UPDATE client_onboarding SET welcome_call_by=COALESCE(%s,welcome_call_by),
                welcome_call_date=COALESCE(%s,welcome_call_date), kit_delivered_date=COALESCE(%s,kit_delivered_date),
                welcome_call_notes=COALESCE(%s,welcome_call_notes), updated_at=NOW() WHERE client_id=%s""",
                (wc_by, wc_date, kit_date, wc_notes, cid))
        else:
            cur.execute("""INSERT INTO client_onboarding (client_id, registration_number, welcome_call_by,
                welcome_call_date, kit_delivered_date, welcome_call_notes) VALUES (%s,%s,%s,%s,%s,%s)""",
                (cid, reg, wc_by, wc_date, kit_date, wc_notes))
        upd += 1
    if not DRY: conn.commit()
    cur.close(); conn.close(); wb.close()
    print(f"[onboarding] upserted={upd} no_client_match={nomatch}  (DRY={DRY})")


# ── STEP: lookups + vendors (add new values only) ──────────────────
# (table, column) -> lookup category
LOOKUP_SRC = [
    ('ops_test_bookings','exam','exam_name'), ('ops_test_bookings','exam_type','exam_method'),
    ('ops_test_bookings','exam_status','exam_status'), ('ops_test_bookings','exam_result','exam_result'),
    ('ops_test_bookings','country','exam_country'), ('ops_test_bookings','booked_by','exam_booked_by'),
    ('ops_coaching','coaching_status','coaching_status'), ('ops_coaching','coaching_method','coaching_method'),
    ('ops_coaching','english_training','training_program'), ('ops_coaching','course_type','coaching_course_type'),
    ('ops_online_subscriptions','online_subscription','subscription_type'),
    ('ops_online_subscriptions','activation_type','activation_type'),
    ('ops_online_subscriptions','booked_by','subscription_booked_by'),
    ('ops_research_publication','research_status','research_status'),
    ('ops_research_publication','research_provider','research_provider'),
    ('ops_research_publication','author_position','author_position'),
    ('ops_webinars_conferences','event_type','event_type'), ('ops_webinars_conferences','event_value','event_value'),
    ('ops_webinars_conferences','participation_type','participation_type'),
    ('ops_epic_registration','epic_status','epic_status'),
    ('ops_academic_details','img_fmg','img_fmg'), ('ops_academic_details','mbbs_status','mbbs_status'),
    ('ops_academic_details','internship_status','internship_status'),
    ('ops_academic_details','working_status','working_status'),
    ('plab_clients','account_status','account_status'), ('plab_clients','current_stage','current_stage'),
    ('plab_clients','plan_type','plan_type'), ('plab_clients','counsellor','counsellor'),
    ('plab_clients','lead_source','lead_source'),
    ('ops_payments','payment_method','payment_method'), ('ops_payments','instalment','instalment'),
]
VENDOR_SRC = [
    ('ops_coaching','other_vendor','Training'), ('ops_research_publication','research_provider','Research'),
]

def step_lookups():
    conn = _pg(); cur = conn.cursor()
    added_lk = 0
    for table, col, cat in LOOKUP_SRC:
        try:
            cur.execute(f"SELECT DISTINCT {col} FROM {table} WHERE COALESCE(pathway,'plab')='australia' AND {col} IS NOT NULL AND TRIM({col})<>''")
        except Exception:
            conn.rollback(); continue
        vals = [r[0].strip() for r in cur.fetchall() if r[0] and r[0].strip()]
        cur.execute("SELECT LOWER(TRIM(value)) FROM lookup_options WHERE category=%s AND COALESCE(pathway,'plab')='australia'", (cat,))
        have = set(r[0] for r in cur.fetchall())
        for v in vals:
            if len(v) > 120: continue
            if v.lower() in have: continue
            if DRY: added_lk += 1; have.add(v.lower()); continue
            cur.execute("INSERT INTO lookup_options (category, label, value, pathway, is_active) VALUES (%s,%s,%s,'australia',true)", (cat, v, v))
            have.add(v.lower()); added_lk += 1
    # vendors
    added_vn = 0
    cur.execute("SELECT LOWER(TRIM(name)) FROM vendors_providers"); have_v = set(r[0] for r in cur.fetchall())
    for table, col, category in VENDOR_SRC:
        try:
            cur.execute(f"SELECT DISTINCT {col} FROM {table} WHERE COALESCE(pathway,'plab')='australia' AND {col} IS NOT NULL AND TRIM({col})<>''")
        except Exception:
            conn.rollback(); continue
        for r in cur.fetchall():
            nm = (r[0] or '').strip()
            if not nm or len(nm) > 120 or nm.lower() in have_v: continue
            if DRY: added_vn += 1; have_v.add(nm.lower()); continue
            cur.execute("INSERT INTO vendors_providers (name, category, is_active) VALUES (%s,%s,true)", (nm, category))
            have_v.add(nm.lower()); added_vn += 1
    if not DRY: conn.commit()
    cur.close(); conn.close()
    print(f"[lookups] new dropdown values added={added_lk}  new vendors added={added_vn}  (DRY={DRY})")


def main():
    step = os.environ.get('STEP', 'all')
    print(f"=== AMC (australia) refresh — STEP={step} DRY_RUN={DRY} ===")
    if step in ('all', 'clients'):
        step_clients()
    idx = build_index()
    print(f"index: {len(idx[0])} reg-variants, {len(idx[1])} names, {len(idx[4])} client_ids")
    if step in ('all', 'sections'):
        step_sections(idx)
    if step in ('all', 'onboarding'):
        step_onboarding(idx)
    if step in ('all', 'lookups'):
        step_lookups()


if __name__ == '__main__':
    main()
