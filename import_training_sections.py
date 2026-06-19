"""
import_training_sections.py — import Training pathway data (pathway='training').

Lightweight pathway: clients + payments only.
  1. step_clients()  — upsert plab_clients (pathway='training') from the
     Registration list (GCTRN reg numbers from the file).
  2. step_sections() — replace-per-section the ops_payments rows for
     pathway='training', reg-matched to those clients.

Run:
  DATABASE_URL=...  DRY_RUN=1  python3 import_training_sections.py   # plan only
  DATABASE_URL=...               python3 import_training_sections.py   # live
"""
import os
from import_amc_sections import (
    s, d, num, norm, digits, extract_reg, strip_reg, reg_variants,
    split_name, _strip_prefix, CONV, _pg,
)

DL = os.environ.get('TR_DIR', '/Users/Santosh/Desktop/Zoho Data/Training')
PATHWAY = 'training'
DRY = os.environ.get('DRY_RUN') in ('1', 'true', 'True')


def _resolve(fname):
    for cand in (fname, f"Training - {fname}", f"Training- {fname}"):
        p = os.path.join(DL, cand)
        if os.path.exists(p):
            return p
    return os.path.join(DL, fname)


CLIENT_FILE = "Training - Pathway Registration List.xlsx"
CLIENT_MAP = {
    'customer_id': ('Customer ID', 's'),
    'registration_date': ('Registration Date (Payment Date)', 'd'),
    'email': ('Candidate Email', 's'), 'mobile': ('Mobile Number', 's'),
    'whatsapp1': ('Whats App Number', 's'),
    'plan_type': ('Plan Type', 's'), 'counsellor': ('Counsellor Name', 's'),
    'dob': ('D.O.B', 'd'), 'city': ('CITY', 's'), 'state': ('STATE', 's'),
    'lead_source': ('Lead Source', 's'),
    'counsellor_email': ('Counsellor Email', 's'), 'counsellor_number': ('Counsellor Number', 's'),
    'additional_notes': ('Notes', 's'),
}


def step_clients():
    import openpyxl
    path = _resolve(CLIENT_FILE)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else '' for h in next(it)]
    hi = {h: i for i, h in enumerate(headers)}
    def g(row, h):
        i = hi.get(h)
        return row[i] if (i is not None and i < len(row)) else None
    rows = [r for r in it if any(v not in (None, '') for v in r)]
    wb.close()

    conn = _pg(); cur = conn.cursor()
    cur.execute("SELECT registration_number FROM plab_clients WHERE COALESCE(pathway,'plab')='training'")
    existing = set((r[0] or '').strip() for r in cur.fetchall())
    ins = upd = skip = 0
    for r in rows:
        reg = extract_reg(g(r, 'Registration Number')) or s(g(r, 'Registration Number'))
        if not reg:
            skip += 1; continue
        prefix, first, last = split_name(g(r, 'Candidate Name'))
        data = {'prefix': prefix, 'first_name': first, 'last_name': last}
        for col, (h, k) in CLIENT_MAP.items():
            data[col] = CONV[k](g(r, h))
        if DRY:
            upd += 1 if reg in existing else 0
            ins += 0 if reg in existing else 1
            continue
        if reg in existing:
            sets = ', '.join(f"{c}=%s" for c in data)
            cur.execute(f"UPDATE plab_clients SET {sets}, updated_at=NOW() "
                        f"WHERE registration_number=%s AND COALESCE(pathway,'plab')='training'",
                        list(data.values()) + [reg]); upd += 1
        else:
            cols = ['registration_number', 'pathway'] + list(data.keys())
            vals = [reg, 'training'] + list(data.values())
            cur.execute(f"INSERT INTO plab_clients ({','.join(cols)}) VALUES ({','.join(['%s']*len(cols))})", vals)
            existing.add(reg); ins += 1
    if not DRY:
        conn.commit()
    cur.close(); conn.close()
    print(f"[clients] inserted={ins} updated={upd} skipped(no reg)={skip}  (DRY={DRY})")


def build_index():
    conn = _pg(); cur = conn.cursor()
    cur.execute("SELECT registration_number, first_name, last_name, email, mobile "
                "FROM plab_clients WHERE COALESCE(pathway,'plab')='training'")
    by_reg, by_name, by_email, by_mobile = {}, {}, {}, {}
    for reg, fn, ln, email, mob in cur.fetchall():
        if not reg:
            continue
        canon = reg.strip()
        for v in reg_variants(canon):
            by_reg.setdefault(norm(v), canon)
        full = norm(f"{fn or ''} {ln or ''}")
        if full:
            by_name.setdefault(full, canon)
        if email:
            by_email.setdefault(norm(email), canon)
        if mob:
            dm = digits(mob)[-10:]
            if dm:
                by_mobile.setdefault(dm, canon)
    cur.close(); conn.close()
    return by_reg, by_name, by_email, by_mobile


def resolve_reg(idx, reg_val, name_val):
    by_reg, by_name, by_email, by_mobile = idx
    emb = extract_reg(name_val) or extract_reg(reg_val) or reg_val
    if emb:
        for v in reg_variants(emb):
            cc = by_reg.get(norm(v))
            if cc:
                return cc
    nm = norm(_strip_prefix(strip_reg(name_val)))
    if nm and nm in by_name:
        return by_name[nm]
    return None


SECTIONS = {
 'payments': ("Training- All Payments Report.xlsx", "ops_payments", "Registration Number", "Enter Candidate Name", {
    'total_package': ('Final Package', 'num'), 'instalment': ('Installment', 's'),
    'total_amount_paid': ('Total Amount Paid', 'num'), 'payment_date': ('Payment Date', 'd'),
    'amount_paid': ('Amount Paid', 'num'), 'payment_method': ('Payment Method', 's'),
    'gst_paid': ('GST Paid', 'num'), 'notes': ('Notes', 's'),
 }),
}


def step_sections():
    import openpyxl
    idx = build_index()
    print(f"[index] {len(idx[1])} names, {len(idx[0])} reg-variants")
    for key, (fname, table, reg_col, name_col, mapping) in SECTIONS.items():
        path = _resolve(fname)
        if not os.path.exists(path):
            print(f"[{key}] MISSING {fname}"); continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else '' for h in next(it)]
        hi = {h: i for i, h in enumerate(headers)}
        def g(row, h):
            i = hi.get(h)
            return row[i] if (i is not None and i < len(row)) else None
        batch = []; unmatched = 0
        for r in it:
            if not any(v not in (None, '') for v in r):
                continue
            reg_val = s(g(r, reg_col)) if reg_col else None
            name_val = s(g(r, name_col)) if name_col else None
            if not (reg_val or name_val):
                continue
            reg = resolve_reg(idx, reg_val, name_val)
            if not reg:
                unmatched += 1; continue
            row_vals = [reg] + [CONV[k](g(r, h)) for (h, k) in mapping.values()] + ['training']
            batch.append(tuple(row_vals))
        wb.close()
        cols = ['registration_number'] + list(mapping.keys()) + ['pathway']
        if DRY:
            print(f"[{key:9}] {table} would insert={len(batch)} unmatched={unmatched}")
            continue
        conn = _pg(); conn.autocommit = False; cur = conn.cursor()
        try:
            cur.execute(f"DELETE FROM {table} WHERE COALESCE(pathway,'plab')='training'")
            if batch:
                from psycopg2.extras import execute_values
                tmpl = "(" + ",".join(["%s"] * len(cols)) + ")"
                execute_values(cur, f"INSERT INTO {table} ({','.join(cols)}) VALUES %s",
                               batch, template=tmpl, page_size=500)
            conn.commit()
            print(f"[{key:9}] {table} inserted={len(batch)} unmatched={unmatched}")
        except Exception as e:
            conn.rollback(); print(f"[{key}] FAILED: {e}")
        finally:
            cur.close(); conn.close()


def run():
    print(f"=== Training import  DRY={DRY} ===")
    step_clients()
    step_sections()


if __name__ == '__main__':
    run()
