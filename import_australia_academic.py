"""
One-time import: Australia academic details from Excel into
ops_academic_details.

Each Excel row goes into ops_academic_details with pathway='australia'.

Linking strategy:
  The Excel's "Enter Name" cell contains both the candidate name AND the
  registration number, e.g.:
      " Dr. Yuvashree Krishnamurthy - GCAUSIP/24-25/079"
      " Dr. Akshaya Akhilesh - GCAUSIP/24-25/017"
  Note the SPACE-DASH-SPACE variant (vs. the test-bookings file which
  uses no leading space before the dash). We parse out the
  GCAUSIP-prefixed reg number and store it in registration_number so
  existing list/edit pages that JOIN against plab_clients keep working
  unchanged.

  If no GCAUSIP reg is found in the cell, we fall back to fuzzy name
  match against plab_clients WHERE pathway='australia' (LOWER(TRIM(name))
  match on prefix+first+last).

Idempotent: keyed on (registration_number, pathway). Re-running with no
marker change is a no-op. Bumping IMPORT_VERSION re-runs the upsert. Bad
rows (missing reg number, no match in plab_clients) are skipped with a
warning so one bad row never blocks the whole import.

Hooked into app boot from app.py the same way as
import_australia_test_bookings.py.
"""

import logging
import os
import re
from datetime import datetime, date


EXCEL_FILENAME = 'All_Australia_Academic_Details.xlsx'
IMPORT_VERSION = 'au_academic_v1_initial_72'


H = {
    'reg_date':         'Registration Date (Payment Date)',
    'name_reg':         'Enter Name',
    'img_fmg':          'IMG / FMG',
    'img_college':      'IMG Medical College Name',
    'fmg_college':      'FMG Medical College Name',
    'country':          'Country',
    'mbbs_status':      'MBBS Status',
    'mbbs_start':       'MBBS Start Date',
    'mbbs_end':         'MBBS End Date',
    'spec1':            'Speciality Interest 1',
    'spec2':            'Speciality Interest 2',
    'int_status':       'Internship Status',
    'int_hospital':     'Internship Hospital',
    'int_location':     'Internship Location (State/Country)',
    'int_hospital_2':   'Internship Hospital 2',
    'int_location_2':   'Internship Location 2 (State / Country)',
    'int_start':        'Internship Start Date',
    'int_end':          'Internship End Date',
    'int_gap':          'Internship Gap',
    'gap_months':       'Internship Gap in Months',
    'gap_reason':       'Intership Gap Reason',
    'working_status':   'Working Status',
    'working_hospital': 'Working Hospital Name',
    'additional_info':  'Additional Info',
}


_REG_RE = re.compile(r'(GCAUSIP/\S+)', re.IGNORECASE)


def _extract_reg_number(name_cell):
    """Pull the GCAUSIP/... reg number out of the combined name+reg cell.

    Returns '' if none found. The cell can have stray spaces and dashes
    around the reg number, so we use a generous regex and strip noise.
    """
    if not name_cell:
        return ''
    s = str(name_cell)
    m = _REG_RE.search(s)
    if not m:
        return ''
    raw = m.group(1).rstrip('.,;:-/ \t')
    return raw.upper().replace('GCAUSIP', 'GCAUSIP')


def _extract_name(name_cell):
    """Pull just the candidate name (before the dash) from the combined cell."""
    if not name_cell:
        return ''
    s = str(name_cell).strip()
    # Split on the first " - " or "-GCAUSIP"
    s = re.sub(r'\s*-\s*GCAUSIP.*$', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s*-GCAUSIP.*$', '', s, flags=re.IGNORECASE)
    # Drop leading prefix like "Dr."
    s = re.sub(r'^\s*Dr\.?\s*', '', s, flags=re.IGNORECASE)
    return s.strip()


def _fuzzy_match_reg(conn, name_cell):
    """Try to find a plab_clients reg number by name match (fallback)."""
    name = _extract_name(name_cell)
    if not name:
        return ''
    parts = name.split()
    if not parts:
        return ''
    first = parts[0].lower()
    last = parts[-1].lower() if len(parts) > 1 else ''
    try:
        row = conn.execute(
            """SELECT registration_number FROM plab_clients
                WHERE pathway = 'australia'
                  AND LOWER(TRIM(first_name)) = ?
                  AND LOWER(TRIM(last_name)) = ?
                LIMIT 1""",
            (first, last),
        ).fetchone()
        if row:
            return row['registration_number']
        # Try first name only
        row = conn.execute(
            """SELECT registration_number FROM plab_clients
                WHERE pathway = 'australia'
                  AND LOWER(TRIM(first_name)) = ?
                LIMIT 1""",
            (first,),
        ).fetchone()
        if row:
            return row['registration_number']
    except Exception:
        pass
    return ''


def _ss(v):
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _sd(v):
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def run_import_australia_academic_once(get_db_fn):
    """Import Australia academic details into ops_academic_details (pathway='australia')."""
    import openpyxl

    excel_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        EXCEL_FILENAME,
    )
    if not os.path.exists(excel_path):
        logging.info("Australia academic-details import: no Excel found, skipping")
        return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

    conn = get_db_fn()
    try:
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS _import_markers (key TEXT PRIMARY KEY, value TEXT)")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        marker = conn.execute(
            "SELECT value FROM _import_markers WHERE key = 'au_academic_import'"
        ).fetchone()
        if marker and marker['value'] == IMPORT_VERSION:
            logging.info(f"Australia academic-details import: already at {IMPORT_VERSION}, skipping")
            return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

        logging.info(f"Australia academic-details import: starting {IMPORT_VERSION}...")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h is not None}

        def cell(row, key):
            i = idx.get(H[key])
            return row[i] if i is not None and i < len(row) else None

        inserted = 0
        updated = 0
        skipped = 0
        errors = 0

        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row_cells):
                continue
            name_cell = cell(row_cells, 'name_reg')
            reg_num = _extract_reg_number(name_cell)
            if not reg_num:
                # Fall back to fuzzy name match
                reg_num = _fuzzy_match_reg(conn, name_cell)
            if not reg_num:
                skipped += 1
                logging.warning(f"Australia academic-details: no reg/name match for cell: {name_cell!r}")
                continue

            data = {
                'registration_number':  reg_num,
                'img_fmg':              _ss(cell(row_cells, 'img_fmg')),
                'img_medical_college':  _ss(cell(row_cells, 'img_college')),
                'fmg_medical_college':  _ss(cell(row_cells, 'fmg_college')),
                'country':              _ss(cell(row_cells, 'country')),
                'mbbs_status':          _ss(cell(row_cells, 'mbbs_status')),
                'mbbs_start_date':      _sd(cell(row_cells, 'mbbs_start')),
                'mbbs_end_date':        _sd(cell(row_cells, 'mbbs_end')),
                'speciality_interest_1': _ss(cell(row_cells, 'spec1')),
                'speciality_interest_2': _ss(cell(row_cells, 'spec2')),
                'internship_status':    _ss(cell(row_cells, 'int_status')),
                'internship_hospital':  _ss(cell(row_cells, 'int_hospital')),
                'internship_location':  _ss(cell(row_cells, 'int_location')),
                'internship_hospital_2': _ss(cell(row_cells, 'int_hospital_2')),
                'internship_location_2': _ss(cell(row_cells, 'int_location_2')),
                'internship_start_date': _sd(cell(row_cells, 'int_start')),
                'internship_end_date':   _sd(cell(row_cells, 'int_end')),
                'internship_gap':       _ss(cell(row_cells, 'int_gap')),
                'gap_in_months':        _ss(cell(row_cells, 'gap_months')),
                'gap_reason':           _ss(cell(row_cells, 'gap_reason')),
                'working_status':       _ss(cell(row_cells, 'working_status')),
                'working_hospital_name': _ss(cell(row_cells, 'working_hospital')),
                'additional_info':      _ss(cell(row_cells, 'additional_info')),
                'pathway':              'australia',
            }

            try:
                existing = conn.execute(
                    "SELECT id FROM ops_academic_details "
                    "WHERE registration_number = ? AND pathway = 'australia'",
                    (reg_num,),
                ).fetchone()
                if existing:
                    sets = ', '.join(f"{k} = ?" for k in data.keys())
                    conn.execute(
                        f"UPDATE ops_academic_details SET {sets} WHERE id = ?",
                        list(data.values()) + [existing['id']],
                    )
                    updated += 1
                else:
                    cols = ', '.join(data.keys())
                    placeholders = ', '.join(['?'] * len(data))
                    conn.execute(
                        f"INSERT INTO ops_academic_details ({cols}) VALUES ({placeholders})",
                        list(data.values()),
                    )
                    inserted += 1
                conn.commit()
            except Exception as e:
                errors += 1
                logging.warning(f"Australia academic-details row error ({reg_num}): {e}")
                try:
                    conn.rollback()
                except Exception:
                    pass

        # Marker
        try:
            conn.execute(
                "INSERT INTO _import_markers (key, value) VALUES ('au_academic_import', ?) "
                "ON CONFLICT (key) DO UPDATE SET value = excluded.value",
                (IMPORT_VERSION,),
            )
            conn.commit()
        except Exception:
            try:
                conn.execute("DELETE FROM _import_markers WHERE key = 'au_academic_import'")
                conn.execute(
                    "INSERT INTO _import_markers (key, value) VALUES ('au_academic_import', ?)",
                    (IMPORT_VERSION,),
                )
                conn.commit()
            except Exception as e:
                logging.warning(f"Australia academic-details: could not write marker: {e}")

        logging.info(
            f"Australia academic-details {IMPORT_VERSION} complete — "
            f"inserted={inserted} updated={updated} skipped={skipped} errors={errors}"
        )
        wb.close()
        return {'inserted': inserted, 'updated': updated, 'skipped': skipped, 'errors': errors}
    finally:
        try:
            conn.close()
        except Exception:
            pass
