"""
analyze_dropdowns.py — READ-ONLY extractor for the per-section "Drop Down"
sheets in the Zoho exports.

For each file it:
  1. Finds the data-sheet header (field labels).
  2. Locates the "Drop Down" sheet and, inside it, the labels row (the row
     that matches the data header) and the marker row directly above it
     (the row tagging Vendor/Provider, Deliverable, Deliverable Type).
  3. Walks each column collecting distinct non-empty values below the
     labels row.

It WRITES NOTHING. Output is a JSON map + a readable summary so we can
review exactly what dropdowns / vendors / deliverables would be seeded.

Run:
  DD_DIR="/Users/Santosh/Desktop/Zoho Data/UK GC" python3 analyze_dropdowns.py
"""
import os, glob, json, re

DD_DIR = os.environ.get('DD_DIR', '/Users/Santosh/Desktop/Zoho Data/UK GC')
PREFIX = os.environ.get('FILE_GLOB', 'PLAB *.xlsx')


def norm(x):
    return ' '.join(str(x or '').strip().lower().split())


def is_marker(cell):
    c = norm(cell)
    return ('vendor' in c) or ('deliverable' in c) or ('provider' in c)


def classify(marker):
    m = norm(marker)
    if 'vendor' in m or 'provider' in m:
        return 'vendor'
    if 'deliverable type' in m:
        return 'deliverable_type'
    if 'deliverable' in m:
        return 'deliverable'
    return None


def header_overlap(row, data_hdr):
    rs = {norm(c) for c in row if c not in (None, '')}
    return len(rs & set(data_hdr))


def analyze_file(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    data_ws = wb.worksheets[0]
    data_hdr = [norm(c) for c in next(data_ws.iter_rows(values_only=True)) if c not in (None, '')]
    dd = next((w for w in wb.worksheets if 'drop' in w.title.lower()), None)
    out = {'file': os.path.basename(path), 'data_sheet': data_ws.title,
           'dropdown_sheet': dd.title if dd else None, 'fields': {}, 'markers': {}}
    if not dd:
        wb.close()
        return out
    rows = list(dd.iter_rows(values_only=True))
    # labels row = best overlap with data header, scanning first 4 rows
    best_i, best_score = 0, -1
    for i in range(min(4, len(rows))):
        sc = header_overlap(rows[i], data_hdr)
        if sc > best_score:
            best_i, best_score = i, sc
    labels = rows[best_i]
    marker_row = rows[best_i - 1] if best_i >= 1 else [None] * len(labels)
    value_rows = rows[best_i + 1:]
    ncols = max(len(labels), len(marker_row))
    for ci in range(ncols):
        label = labels[ci] if ci < len(labels) else None
        if label in (None, ''):
            continue
        label = str(label).strip()
        # distinct values below, in order
        seen, vals = set(), []
        for r in value_rows:
            v = r[ci] if ci < len(r) else None
            if v in (None, ''):
                continue
            v = str(v).strip()
            if v and norm(v) not in seen:
                seen.add(norm(v)); vals.append(v)
        if not vals:
            continue
        out['fields'][label] = vals
        mk = classify(marker_row[ci]) if ci < len(marker_row) else None
        if mk:
            out['markers'].setdefault(mk, []).append(label)
    wb.close()
    return out


def main():
    files = sorted(glob.glob(os.path.join(DD_DIR, PREFIX)))
    # de-dup (the listing tool double-printed; real fs won't, but be safe)
    files = sorted(set(files))
    result = []
    all_vendors = {}        # vendor -> set(sections)
    all_deliverables = {}   # deliverable -> set(sections)
    for f in files:
        try:
            r = analyze_file(f)
        except Exception as e:
            print(f"ERR {os.path.basename(f)}: {e}")
            continue
        result.append(r)
        sec = r['file']
        for fld in r['markers'].get('vendor', []):
            for v in r['fields'].get(fld, []):
                all_vendors.setdefault(v, set()).add(sec)
        for fld in r['markers'].get('deliverable', []):
            for v in r['fields'].get(fld, []):
                all_deliverables.setdefault(v, set()).add(sec)

    # readable summary
    for r in result:
        print('\n' + '=' * 78)
        print(r['file'], ' | dropdown sheet:', r['dropdown_sheet'])
        if r['markers']:
            for kind, flds in r['markers'].items():
                print(f"   <{kind.upper()}> -> {flds}")
        for fld, vals in r['fields'].items():
            tag = ''
            for kind, flds in r['markers'].items():
                if fld in flds:
                    tag = f" [{kind}]"
            print(f"   - {fld}{tag}: {len(vals)} -> {vals[:10]}{' ...' if len(vals)>10 else ''}")

    print('\n' + '#' * 78)
    print(f"CENTRALISED VENDORS/PROVIDERS  ({len(all_vendors)} distinct):")
    for v in sorted(all_vendors):
        print(f"   - {v}   <- {sorted(all_vendors[v])}")
    print(f"\nDELIVERABLES (marked)  ({len(all_deliverables)} distinct)")

    with open('dropdowns_extract.json', 'w') as fh:
        json.dump({'sections': result,
                   'vendors': {k: sorted(v) for k, v in all_vendors.items()},
                   'deliverables': {k: sorted(v) for k, v in all_deliverables.items()}},
                  fh, indent=2)
    print("\nWrote dropdowns_extract.json")


if __name__ == '__main__':
    main()
