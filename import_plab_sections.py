"""
import_plab_sections.py — refresh PLAB pathway section data from Zoho exports.

For each of the 16 PLAB section files:
  1. Resolve each row to a client (registration_number) by:
     reg-number column -> candidate name -> email -> mobile.
  2. Wipe existing pathway='plab' rows in the target table.
  3. Insert fresh rows mapped from the Excel, pathway='plab'.

Clients themselves are upserted separately by import_excel_clients.py
(run that FIRST) so client_id / R2 documents / counsellor links stay
intact. This script only touches the section tables, which have no
downstream FK links.

Run:
  DATABASE_URL=... DOWNLOADS=/Users/Santosh/Downloads python3 import_plab_sections.py
Optional: SECTIONS=academic,gmc  to run only some sections.
"""
import os, re
from datetime import datetime, date
from db import get_db

DL = os.environ.get('UK_DIR') or os.environ.get('DOWNLOADS', '/Users/Santosh/Desktop/Zoho Data/UK GC')


def s(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def d(v):
    if v is None or v == '':
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip() or None


def num(v):
    if v in (None, ''):
        return None
    try:
        return float(str(v).replace(',', '').strip())
    except (ValueError, TypeError):
        return None


def norm(x):
    return ' '.join((x or '').strip().lower().split())


def digits(x):
    return re.sub(r'\D', '', str(x or ''))


def build_client_index(conn):
    rows = conn.execute(
        "SELECT registration_number, prefix, first_name, last_name, email, mobile "
        "FROM plab_clients WHERE COALESCE(pathway,'plab')='plab'"
    ).fetchall()
    by_name = {}
    by_email = {}
    by_mobile = {}
    by_reg = set()
    for r in rows:
        reg = r['registration_number']
        if not reg:
            continue
        by_reg.add(norm(reg))
        full = norm(f"{r['first_name'] or ''} {r['last_name'] or ''}")
        if full:
            by_name.setdefault(full, reg)
        if r['email']:
            by_email.setdefault(norm(r['email']), reg)
        if r['mobile']:
            dm = digits(r['mobile'])[-10:]
            if dm:
                by_mobile.setdefault(dm, reg)
    return by_reg, by_name, by_email, by_mobile


# Zoho lookup fields embed the reg number in the name string, e.g.
# "Dr.Hrudaya Charan K - GCUKIP/2023/057". Extract it directly.
_REG_RE = re.compile(r'(GC[A-Z]*/[0-9A-Za-z-]+/\d+)')


def extract_reg(text):
    if not text:
        return None
    m = _REG_RE.search(str(text))
    return m.group(1) if m else None


def strip_reg(text):
    """Return the name part with the ' - REG' suffix removed."""
    if not text:
        return text
    return _REG_RE.sub('', str(text)).rstrip(' -').strip()


def resolve_reg(idx, reg_val, name_val, email_val, mobile_val):
    by_reg, by_name, by_email, by_mobile = idx
    # 1) Reg embedded in the candidate-name string (most reliable).
    emb = extract_reg(name_val) or extract_reg(reg_val)
    if emb:
        if norm(emb) in by_reg:
            return emb
        # also try a couple of common reg-format variants
        for cand in {emb, emb.replace('/20', '/'), emb}:
            if norm(cand) in by_reg:
                return cand
        return emb  # trust the Zoho reg even if exact row not pre-indexed
    # 2) Explicit reg column value
    if reg_val and norm(reg_val) in by_reg:
        return reg_val
    # 3) Name (with any reg suffix stripped)
    nm = norm(strip_reg(name_val))
    if nm and nm in by_name:
        return by_name[nm]
    # 4) Email
    if email_val and norm(email_val) in by_email:
        return by_email[norm(email_val)]
    # 5) Mobile
    if mobile_val:
        dm = digits(mobile_val)[-10:]
        if dm and dm in by_mobile:
            return by_mobile[dm]
    return reg_val


# ── Per-section config ─────────────────────────────────────────────
# (file, table, reg_col, name_col, email_col, mobile_col, mapping)
# mapping: table_col -> (excel_header, kind)  kind in {s,d,num}
SECTIONS = {
 'academic': ("Academic Details (R).xlsx", "ops_academic_details", None, "Enter Candidate Name", None, None, {
    'img_fmg':('IMG / FMG','s'),'img_medical_college':('IMG Medical College Name','s'),
    'fmg_medical_college':('FMG Medical College Name','s'),'country':('Country','s'),
    'mbbs_status':('MBBS Status','s'),'mbbs_start_date':('MBBS Start Date','d'),'mbbs_end_date':('MBBS End Date','d'),
    'speciality_interest_1':('Speciality Interest 1','s'),'speciality_interest_2':('Speciality Interest 2','s'),
    'internship_status':('Internship Status','s'),'internship_hospital':('Internship Hospital','s'),
    'internship_location':('Internship Location (State / Country)','s'),
    'internship_hospital_2':('Internship Hospital 2','s'),
    'internship_location_2':('Internship Location 2 (State / Country)','s'),
    'internship_start_date':('Internship Start Date','d'),'internship_end_date':('Internship End Date','d'),
    'internship_gap':('Internship Gap','s'),'gap_in_months':('Gap in Months','s'),'gap_reason':('Gap Reason','s'),
    'working_status':('Working Status','s'),'working_hospital_name':('Working Hospital Name','s'),
    'additional_info':('Additional Info','s'),
 }),
 'gmc': ("All Gmc Registrations.xlsx", "ops_gmc_registration", None, "Enter Candidate Name", "Candidate Email", "Mobile Number", {
    'gmc_reference_number':('GMC Reference Number','s'),'login_pwd':('Login Pwd','s'),
    'secret_question':('Secret Question 1','s'),'secret_answer':('Secret Answer','s'),'gmc_setup':('GMC Setup','s'),
    'registration_date':('Registration Date','d'),'english_exam':('English Exam','s'),'exam_date':('Exam Date','d'),
    'english_result_expiry_date':('English Result Expiry Date','d'),'license':('License','s'),
    'license_received_date':('License Received Date','d'),
    'candidate_email':('Candidate Email','s'),'mobile_number':('Mobile Number','s'),
 }),
 'epic': ("Epic Registration Status (R).xlsx", "ops_epic_registration", None, "Enter Candidate Name", "Candidate Email", "Mobile Number", {
    'epic_registration':('EPIC Registration','s'),'epic_status':('EPIC Status','s'),'notary_camp':('Notary Camp','s'),
    'registration_date':('Registration Date','d'),'documents_stage':('Documents Stage','s'),
    'document_stage_status':('Document Stage Status','s'),'login_id':('Login ID','s'),'login_pwd':('Login Pwd','s'),
    'secret_question_1':('Secret Question 1','s'),'secret_answer_1':('Secret Answer 1','s'),
    'secret_question_2':('Secret Question 2','s'),'secret_answer_2':('Secret Answer 2','s'),
    'secret_question_3':('Secret Question 3','s'),'secret_answer_3':('Secret Answer 3','s'),
    'secret_question_4':('Secret Question 4','s'),'secret_answer_4':('Secret Answer 4','s'),
    'epic_id_number':('EPIC ID Number','s'),'notary_camp_login':('Notary Camp Login','s'),
    'notary_camp_password':('Notary Camp Password','s'),
 }),
 'test_bookings': ("Test Bookings (R).xlsx", "ops_test_bookings", None, "Enter Candidate Name", "Candidate Email", "Mobile Number", {
    'exam':('Exam','s'),'exam_type':('Exam Method','s'),'booking_date':('Booking Date','d'),'exam_date':('Exam Date','d'),
    'exam_status':('Exam Status','s'),'exam_result':('Exam Result','s'),'exam_result_date':('Exam Result Date','d'),
    'score':('Test Score','s'),'test_center':('Exam Center Name','s'),'city_state':('City / State','s'),
    'country':('Country','s'),'booked_by':('Exam Booked By','s'),
    'revaluation':('Are we applying for revaluation','s'),'reval_applied_date':('Reval Applied Date','d'),
    'reval_result':('Reval Result','s'),'reval_score':('Reval Score','s'),'notes':('Notes','s'),
 }),
 'coaching': ("ICP Training Report List.xlsx", "ops_coaching", None, "Enter Candidate Name", "Candidate Email", "Mobile Number", {
    'english_training':('Training Program','s'),'course_type':('Course Type','s'),
    'coaching_method':('Coaching Method','s'),'coaching_status':('Coaching Status','s'),
    'batch_month':('Batch (Month)','s'),'batch_year':('Batch (year)','s'),
    'start_date':('Start Date','d'),'end_date':('End Date','d'),'attendance':('Training Attendance','s'),
    # New export collapsed the 5 per-exam vendor columns into one "Training Vendor".
    'vendor_provider':('Training Vendor','s'),
 }),
 'online_courses': ("Online Courses (R).xlsx", "ops_online_courses", None, "Candidate Name", "Candidate Email", "Mobile Number", {
    'courses_name':('Courses Name','s'),'course_type':('Course Type','s'),'start_date':('Start Date','d'),
    'end_date':('End Date','d'),'validity_months':('Validity in Months','s'),'course_provider':('Course Provider','s'),
    'certification_body':('Certification Body','s'),'subject_name':('Subject Name','s'),
    'course_status':('Course Status','s'),'booked_by':('Booked By','s'),'client_email':('Client Email','s'),
    'mobile_number':('Mobile Number','s'),'candidate_email':('Candidate Email','s'),'notes':('Notes','s'),
 }),
 'subscriptions': ("Online Subscriptions.xlsx", "ops_online_subscriptions", None, "Candidate Name", "Candidate Email", "Mobile Number", {
    'online_subscription':('Online Subscription','s'),'subscription_provider':('Subscription Providers','s'),
    'issued_date':('Issued Date','d'),
    'activation_type':('Activation type','s'),'notes':('Notes','s'),'client_email':('Client Email','s'),
    'login_id':('Login Id','s'),'password':('Password','s'),'booked_by':('Booked By','s'),
    'mobile_number':('Mobile Number','s'),'candidate_email':('Candidate Email','s'),
 }),
 'research': ("Research and Publication (R).xlsx", "ops_research_publication", None, "Enter Candidate Name", "Candidate Email", "Mobile Number", {
    'research_status':('Research Status','s'),'research_topic':('Research Topic','s'),'service':('Service','s'),
    'published_copy':('Published Copy','s'),'research_start_date':('Research Start Date','d'),
    'research_end_date':('Research End Date','d'),'research_provider':('Research Provider','s'),
    'published_journal_name':('Published Journal Name','s'),'author_position':('Author Position','s'),
    'research_batch':('Research Batch','s'),'mobile_number':('Mobile Number','s'),
    'candidate_email':('Candidate Email','s'),'additional_notes':('Additional Notes','s'),
 }),
 'webinars': ("Webinar and Conferences.xlsx", "ops_webinars_conferences", None, "Candidate Name", "Candidate Email", "Mobile Number", {
    'event_type':('Event Type','s'),'start_date':('Start Date','d'),'end_date':('End Date','d'),
    'duration_days':('Duration (Days)','s'),'event_value':('Event Value','s'),
    'cpd_points':('CPD Points','s'),'event_name':('Webinar and Event Name','s'),
    'provider':('Provider','s'),
    'participation_type':('Participation Type','s'),'notes':('Notes','s'),
    'mobile_number':('Mobile Number','s'),'candidate_email':('Candidate Email','s'),
 }),
 'call_notes': ("Client Call Notes Report (1).xlsx", "ops_call_notes", None, "Candidate Name", None, None, {
    'call_date':('Call Date','d'),'call_note':('Call Note','s'),'added_by':('Added User','s'),
 }),
 'english_logins': ("All English Login Details.xlsx", "ops_english_logins", None, "Enter Candidate Name", None, None, {
    'ielts_login_id':('IELTS Login IID','s'),'ielts_password':('IELTS Password','s'),
    'ielts_hint_question':('IELTS Hint Question','s'),'ielts_security_answer':('IELTS Security Answer','s'),
    'oet_login_id':('OET Login ID','s'),'oet_password':('OET Password','s'),
    'oet_hint_question':('OET Hint Question','s'),'oet_security_answer':('OET Security Answer','s'),
 }),
 'ngo': ("All Ngo Activities.xlsx", "ops_ngo_activities", None, "Enter Candidate Name", "Candidate Email", "Mobile Number", {
    'ngo_vendor_name':('NGO Vendor Name','s'),'activity_type':('Activity Type','s'),
    'batch_name_no':('Batch Name / No','s'),'activity_start_date':('Activity Start Date','d'),
 }),
 'uk_cab': ("All Uk Cab Bookings.xlsx", "ops_uk_cab_bookings", None, "Enter Candidate Name", "Candidate Email", "Contact Number", {
    'candidate_email':('Candidate Email','s'),'contact_number':('Contact Number','s'),
    'whatsapp_number':('Whats App Number','s'),'whatsapp_number_2':('Whats App Number 2','s'),
    'vendor':('Vendor','s'),'services':('Services','s'),'booking_date':('Booking Date','d'),
    'invoice_number':('Invoice Number','s'),'additional_instructions':('Additional Instructions','s'),
    'pick_up_date':('Pick Up Date','d'),'pick_up_location':('Pick Up Loaction','s'),
    'drop_location':('Drop Location','s'),'additional_notes':('Additional Notes','s'),
 }),
 'observerships': ("UK Observership Report.xlsx", "ops_uk_observerships", None, "Enter Candidate Name", None, None, {
    'hospital_name':('Hospital Name','s'),'hospital_location':('Hospital Location','s'),
    'start_date':('Start Date','d'),'end_date':('End Date','d'),'speciality':('Speciality','s'),
    'payment':('Payment','s'),'mobile_number':('Mobile Number','s'),'candidate_email':('Candidate Email','s'),
 }),
 'visa': ("Visa and Travel (R).xlsx", "ops_uk_visa_travel", None, "Enter Candidate Name", "Candidate Email", "Mobile Number", {
    'visa_application_status':('Visa Application Status','s'),'documents_collected':('Documents Collected','s'),
    'note_on_documents':('Note on Documents','s'),'documents_status':('Documents Status','s'),
    'visa_interview_date':('Visa Interview Date','d'),'visa_status':('Visa Status','s'),
    'note_on_visa_status':('Note on Visa status','s'),'visa_issued_date':('Visa Issued Date','d'),
    'visa_period':('Visa Period','s'),'visa_expiry_date':('Visa Expiry Date','d'),'visa_type':('Visa Type','s'),
    'departure_date':('Departure Date','d'),'quarantine':('Quarantine','s'),'quarantine_days':('Quarantine Days','s'),
    'quarantine_hotel_name':('Quarantine Hotel Name','s'),'lodging_name':('Lodging Name','s'),
    'lodging_type':('Lodging Type','s'),'note':('Note','s'),'arrival_date':('Arrival Date','d'),
    'uk_phone_number':('UK Phone Number','s'),'mobile_number':('Mobile Number','s'),
    'candidate_email':('Candidate Email','s'),
 }),
 'mentorship': ("Mentorship Report.xlsx", "ops_mentorship", None, "Candidate Name", "Candidate Email", "Mobile Number", {
    'session_date':('Date','d'),'start_time':('Start Time','s'),'end_time':('End Time','s'),
    'duration_minutes':('Duration (Minutes)','s'),
    # New export: Amount Paid -> Fee Paid (+ currency); Services Type is the
    # Paid/Not Paid flag; Service Description is free text.
    'amount_paid':('Fee Paid','num'),'fee_currency':('Fee Currency','s'),
    'payment_status':('Services Type','s'),'service_description':('Service Description','s'),
    'candidate_attendance':('Candidate Attendance','s'),'additional_notes':('Additional Notes','s'),
    'session_confirmation':('Session Confirmation','s'),'program_provider':('Program Provider','s'),
    'mentor_attendance':('Mentor Attendance','s'),
 }),
}

CONV = {'s': s, 'd': d, 'num': num}


def _pg():
    """Fresh psycopg2 connection with TCP keepalives so long imports
    survive idle gaps. Direct (not db.py) so we can batch with
    execute_values."""
    import psycopg2
    return psycopg2.connect(
        os.environ['DATABASE_URL'], connect_timeout=20,
        keepalives=1, keepalives_idle=30, keepalives_interval=10, keepalives_count=5,
    )


def run():
    import openpyxl
    from psycopg2.extras import execute_values
    only = os.environ.get('SECTIONS')
    only = set(only.split(',')) if only else None

    # Build the client index once on its own connection.
    conn = _pg()
    cur = conn.cursor()
    cur.execute("SELECT registration_number, prefix, first_name, last_name, email, mobile "
                "FROM plab_clients WHERE COALESCE(pathway,'plab')='plab'")
    by_reg, by_name, by_email, by_mobile = set(), {}, {}, {}
    for reg, prefix, fn, ln, email, mob in cur.fetchall():
        if not reg:
            continue
        by_reg.add(norm(reg))
        full = norm(f"{fn or ''} {ln or ''}")
        if full:
            by_name.setdefault(full, reg)
        if email:
            by_email.setdefault(norm(email), reg)
        if mob:
            dm = digits(mob)[-10:]
            if dm:
                by_mobile.setdefault(dm, reg)
    idx = (by_reg, by_name, by_email, by_mobile)
    cur.close(); conn.close()
    print(f"Client index built: {len(by_name)} names, {len(by_email)} emails, {len(by_mobile)} mobiles\n")

    summary = []
    for key, cfg in SECTIONS.items():
        if only and key not in only:
            continue
        fname, table, reg_col, name_col, email_col, mobile_col, mapping = cfg
        # New Zoho exports live in UK_DIR with a "PLAB - " prefix (the GMC file
        # has no space: "PLAB -..."). Resolve whichever variant exists so the
        # legacy bare filename in SECTIONS still maps to the new file.
        path = None
        for cand in (fname, f"PLAB - {fname}", f"PLAB -{fname}"):
            p = os.path.join(DL, cand)
            if os.path.exists(p):
                path = p
                break
        if not path:
            print(f"[{key}] MISSING FILE {fname}, skip")
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]   # data sheet (sheet 0); sheet 1 is "Drop Downs"
        headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        hidx = {h: i for i, h in enumerate(headers)}

        cols = ['registration_number'] + list(mapping.keys()) + ['pathway']

        # Collect all rows in memory first (no DB held open during parse).
        batch = []
        unmatched = blank = 0
        for rcells in ws.iter_rows(min_row=2):
            vals = [c.value for c in rcells]
            if not any(v not in (None, '') for v in vals):
                continue
            def g(h):
                i = hidx.get(h)
                return vals[i] if (i is not None and i < len(vals)) else None
            reg_val = s(g(reg_col)) if reg_col else None
            name_val = s(g(name_col)) if name_col else None
            email_val = s(g(email_col)) if email_col else None
            mobile_val = s(g(mobile_col)) if mobile_col else None
            if not (reg_val or name_val or email_val or mobile_val):
                blank += 1
                continue
            reg = resolve_reg(idx, reg_val, name_val, email_val, mobile_val)
            if not reg:
                unmatched += 1
                continue
            row_vals = [reg] + [CONV[kind](g(xh)) for (xh, kind) in mapping.values()] + ['plab']
            batch.append(tuple(row_vals))
        wb.close()

        # Fresh connection per section; wipe + batch insert; commit; close.
        conn = _pg()
        conn.autocommit = False
        cur = conn.cursor()
        try:
            cur.execute(f"DELETE FROM {table} WHERE COALESCE(pathway,'plab')='plab'")
            if batch:
                tmpl = "(" + ",".join(["%s"] * len(cols)) + ")"
                execute_values(
                    cur,
                    f"INSERT INTO {table} ({','.join(cols)}) VALUES %s",
                    batch, template=tmpl, page_size=500,
                )
            conn.commit()
            print(f"[{key:14s}] {table:26s} inserted={len(batch):5d}  unmatched={unmatched:4d}  blank={blank}")
            summary.append((key, len(batch), unmatched))
        except Exception as e:
            conn.rollback()
            print(f"[{key}] FAILED: {e}")
            summary.append((key, -1, unmatched))
        finally:
            cur.close(); conn.close()

    print("\n=== SUMMARY ===")
    for k, ins, un in summary:
        print(f"  {k:16s} {ins:6d} inserted, {un} unmatched")


if __name__ == '__main__':
    run()
