"""
One-time import: Australia research & publication records from Excel into
ops_research_publication.

Each Excel row goes into ops_research_publication with pathway='australia'.

Linking strategy:
  The Excel's "Enter Candidate Name" cell contains both the candidate name
  AND the registration number, e.g.:
      " Dr. Mohammad  Shibil  -GCAUSIP/2023/044"
      " Dr. Shiva Teja Akoju -GCAUSIP/24-25/039"
  We parse out the GCAUSIP-prefixed reg number and store it in
  registration_number so existing list/edit pages that JOIN against
  plab_clients keep working unchanged.

  If the reg number is missing from the cell we fall back to a fuzzy name
  match against plab_clients WHERE pathway='australia' (LOWER(TRIM(name))
  match on prefix+first+last).

Idempotent: keyed on (registration_number, research_topic,
research_start_date, research_provider). Re-running with no marker change
is a no-op. Bumping IMPORT_VERSION re-runs the upsert. Bad rows (missing
reg number, no plab_clients match) are skipped with a warning so one bad
row never blocks the whole import.

Hooked into app boot from app.py the same way as
import_australia_test_bookings.py.
"""

import logging
import os
import re
from datetime import datetime, date


EXCEL_FILENAME = 'All_Australia_Research.xlsx'
IMPORT_VERSION = 'au_research_v1_initial_17'


H = {
    'cand_name':       'Enter Candidate Name',
    'status':          'Research Status',
    'start_date':      'Research Start Date',
    'topic':           'Research Topic',
    'batch':           'Research Batch',
    'end_date':        'Research End Date',
    'provider':        'Research Provider',
    'journal':         'Published Journal Name',
    'author_position': 'Author Position',
    'published_copy':  'Upload Published Copy',
}


_REG_RE = re.compile(r'(GCAUSIP/\S+)', re.IGNORECASE)
_DR_PREFIX_RE = re.compile(r'^\s*(dr\.?|mr\.?|mrs\.?|ms\.?)\s+', re.IGNORECASE)


def _extract_reg_number(candidate_cell):
    """Pull the GCAUSIP/... reg number out of the combined name+reg cell.

    Returns '' if none found. The cell can have stray spaces and dashes
    around the reg number, so we use a generous regex and strip noise.
    """
    if not candidate_cell:
        return ''
    s = str(candidate_cell)
    m = _REG_RE.search(s)
    if not m:
        return ''
    raw = m.group(1).rstrip('.,;:-/ \t')
    return raw.upper().replace('GCAUSIP', 'GCAUSIP')


def _extract_name(candidate_cell):
    """Pull the candidate name (without prefix and without reg-no) from the cell."""
    if not candidate_cell:
        return ''
    s = str(candidate_cell)
    # Strip the trailing reg-number chunk (after the last '-' before GCAUSIP)
    s = re.sub(r'\s*-?\s*GCAUSIP/\S+\s*$', '', s, flags=re.IGNORECASE)
    s = _DR_PREFIX_RE.sub('', s.strip())
    # Collapse internal whitespace
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _fuzzy_match_reg(conn, candidate_cell):
    """Fuzzy match a candidate name (no reg no in cell) against plab_clients."""
    name = _extract_name(candidate_cell)
    if not name:
        return ''
    # Try first+last full-name match first
    parts = name.split()
    if not parts:
        return ''
    lname = name.lower()
    try:
        # Exact full-name (case-insensitive) match
        row = conn.execute(
            """SELECT registration_number FROM plab_clients
                WHERE pathway = 'australia'
                  AND LOWER(TRIM(COALESCE(first_name,'') || ' ' || COALESCE(last_name,''))) = ?
                LIMIT 1""",
            (lname,),
        ).fetchone()
        if row:
            return row['registration_number'] or ''
        # First-name + last-name two-side match
        if len(parts) >= 2:
            first = parts[0].lower()
            last = parts[-1].lower()
            row = conn.execute(
                """SELECT registration_number FROM plab_clients
                    WHERE pathway = 'australia'
                      AND LOWER(TRIM(first_name)) = ?
                      AND LOWER(TRIM(last_name)) = ?
                    LIMIT 1""",
                (first, last),
            ).fetchone()
            if row:
                return row['registration_number'] or ''
    except Exception:
        pass
    return ''


def _ss(v):
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _sd(v):
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def run_import_australia_research_once(get_db_fn):
    """Import Australia research records into ops_research_publication (pathway='australia')."""
    import openpyxl

    excel_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        EXCEL_FILENAME,
    )
    if not os.path.exists(excel_path):
        logging.info("Australia research import: no Excel found, skipping")
        return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

    conn = get_db_fn()
    try:
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS _import_markers (key TEXT PRIMARY KEY, value TEXT)")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        marker = conn.execute(
            "SELECT value FROM _import_markers WHERE key = 'au_research_import'"
        ).fetchone()
        if marker and marker['value'] == IMPORT_VERSION:
            logging.info(f"Australia research import: already at {IMPORT_VERSION}, skipping")
            return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

        logging.info(f"Australia research import: starting {IMPORT_VERSION}...")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h is not None}

        def cell(row, key):
            i = idx.get(H[key])
            return row[i] if i is not None and i < len(row) else None

        inserted = 0
        updated = 0
        skipped = 0
        errors = 0

        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row_cells):
                continue
            cand_cell = cell(row_cells, 'cand_name')
            reg_num = _extract_reg_number(cand_cell)
            if not reg_num:
                # Fuzzy name fallback
                reg_num = _fuzzy_match_reg(conn, cand_cell)
            if not reg_num:
                skipped += 1
                logging.warning(
                    f"Australia research: skipping row, no reg no and no name match "
                    f"(cell={cand_cell!r})"
                )
                continue

            topic = _ss(cell(row_cells, 'topic'))
            start_date = _sd(cell(row_cells, 'start_date'))
            provider = _ss(cell(row_cells, 'provider'))

            data = {
                'registration_number':     reg_num,
                'research_status':         _ss(cell(row_cells, 'status')),
                'research_start_date':     start_date,
                'research_topic':          topic,
                'research_batch':          _ss(cell(row_cells, 'batch')),
                'research_end_date':       _sd(cell(row_cells, 'end_date')),
                'research_provider':       provider,
                'published_journal_name':  _ss(cell(row_cells, 'journal')),
                'author_position':         _ss(cell(row_cells, 'author_position')),
                'published_copy':          _ss(cell(row_cells, 'published_copy')),
                'pathway':                 'australia',
            }

            try:
                existing = conn.execute(
                    "SELECT id FROM ops_research_publication "
                    "WHERE registration_number = ? "
                    "  AND COALESCE(research_topic,'') = ? "
                    "  AND COALESCE(research_start_date,'') = ? "
                    "  AND COALESCE(research_provider,'') = ? "
                    "  AND pathway = 'australia'",
                    (reg_num, topic, start_date, provider),
                ).fetchone()
                if existing:
                    sets = ', '.join(f"{k} = ?" for k in data.keys())
                    conn.execute(
                        f"UPDATE ops_research_publication SET {sets} WHERE id = ?",
                        list(data.values()) + [existing['id']],
                    )
                    updated += 1
                else:
                    cols = ', '.join(data.keys())
                    placeholders = ', '.join(['?'] * len(data))
                    conn.execute(
                        f"INSERT INTO ops_research_publication ({cols}) VALUES ({placeholders})",
                        list(data.values()),
                    )
                    inserted += 1
                conn.commit()
            except Exception as e:
                errors += 1
                logging.warning(f"Australia research row error ({reg_num} / {topic}): {e}")
                try:
                    conn.rollback()
                except Exception:
                    pass

        # Marker
        try:
            conn.execute(
                "INSERT INTO _import_markers (key, value) VALUES ('au_research_import', ?) "
                "ON CONFLICT (key) DO UPDATE SET value = excluded.value",
                (IMPORT_VERSION,),
            )
            conn.commit()
        except Exception:
            try:
                conn.execute("DELETE FROM _import_markers WHERE key = 'au_research_import'")
                conn.execute(
                    "INSERT INTO _import_markers (key, value) VALUES ('au_research_import', ?)",
                    (IMPORT_VERSION,),
                )
                conn.commit()
            except Exception as e:
                logging.warning(f"Australia research: could not write marker: {e}")

        logging.info(
            f"Australia research {IMPORT_VERSION} complete — "
            f"inserted={inserted} updated={updated} skipped={skipped} errors={errors}"
        )
        wb.close()
        return {'inserted': inserted, 'updated': updated, 'skipped': skipped, 'errors': errors}
    finally:
        try:
            conn.close()
        except Exception:
            pass
