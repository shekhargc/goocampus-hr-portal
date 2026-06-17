"""
One-time import script: Load PLAB clients from Excel into DB.
Upserts by registration_number: updates existing, inserts new.
Run from the project directory: python import_excel_clients.py
"""
import os, sys, re
from datetime import datetime, date

# Add project dir to path so we can import db module
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from db import get_db

EXCEL_PATH = os.environ.get('EXCEL_PATH', '/sessions/nice-serene-galileo/mnt/uploads/GC UK Reg-List.xlsx')

def safe_float(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(',', '').replace('₹', '').replace('$', '')
    try: return float(s)
    except: return 0

def safe_str(v):
    if v is None: return ''
    return str(v).strip()

def safe_date(v):
    if v is None: return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()

def run_import():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    status_fix = {'Oh Hold': 'On Hold'}
    conn = get_db()

    # Find highest existing sequence number
    result = conn.execute(
        "SELECT registration_number FROM plab_clients WHERE registration_number LIKE ? ORDER BY id DESC LIMIT 1",
        ('GCUKIP/%',)
    ).fetchone()
    if result:
        m = re.search(r'/(\d+)$', result['registration_number'])
        next_seq = int(m.group(1)) + 1 if m else 1
    else:
        next_seq = 1

    def next_reg():
        nonlocal next_seq
        reg = f"GCUKIP/25-26/{next_seq:03d}"
        next_seq += 1
        return reg

    inserted = 0
    updated = 0
    skipped = 0

    for row_cells in ws.iter_rows(min_row=2):
        vals = [cell.value for cell in row_cells]
        if not any(vals):
            continue
        row = dict(zip(headers, vals))

        reg_num = safe_str(row.get('Registration Number'))
        candidate_name = safe_str(row.get('Candidate Name'))

        # Parse name
        first_name = ''
        last_name = ''
        prefix = 'Dr.'
        if candidate_name:
            parts = candidate_name.strip().split()
            if parts and parts[0].rstrip('.') in ('Dr', 'Mr', 'Mrs', 'Ms', 'Prof'):
                prefix = parts[0] if parts[0].endswith('.') else parts[0] + '.'
                parts = parts[1:]
            if len(parts) >= 2:
                first_name = parts[0]
                last_name = ' '.join(parts[1:])
            elif parts:
                first_name = parts[0]

        if not first_name and not reg_num:
            skipped += 1
            continue

        pkg = safe_float(row.get('Package (Mention Plab Actual Package)'))
        disc = safe_float(row.get('Discount Allowed (Discount Offer + Stage Discount)'))
        final_raw = safe_float(row.get('Finalised Package')) or safe_float(row.get('Final Package'))
        final = final_raw or (pkg - disc)
        i1 = safe_float(row.get('1st Installment'))
        i2 = safe_float(row.get('2nd Installment'))
        i3 = safe_float(row.get('3rd Installment'))
        i4 = safe_float(row.get('4th Installment'))
        total_paid = i1 + i2 + i3 + i4

        raw_status = safe_str(row.get('Account Status')) or 'In Process'
        account_status = status_fix.get(raw_status, raw_status)

        # Check if exists
        existing = None
        if reg_num:
            existing = conn.execute(
                "SELECT id FROM plab_clients WHERE registration_number = ?", (reg_num,)
            ).fetchone()

        if existing:
            # UPDATE
            conn.execute('''UPDATE plab_clients SET
                registration_date=?, customer_id=?,
                prefix=?, first_name=?, last_name=?,
                mobile=?, whatsapp1=?, whatsapp2=?, email=?, dob=?, city=?, state=?,
                instagram=?, facebook=?, linkedin=?,
                father_name=?, father_phone=?, mother_name=?, mother_phone=?, parents_email=?,
                joined_stage=?, plan_type=?,
                account_status=?, current_stage=?, switched_program=?,
                counsellor=?, counsellor_email=?, counsellor_number=?,
                lead_source=?, referral_type=?, operations_referral=?,
                package_amount=?, discount_allowed=?, additional_package_notes=?,
                final_package=?, total_paid=?,
                inst1_amount=?, inst1_date=?, inst1_note=?,
                inst2_amount=?, inst2_date=?, inst2_note=?,
                inst3_amount=?, inst3_date=?, inst3_note=?,
                inst4_amount=?, inst4_date=?, inst4_note=?,
                dropped_date=?, upgraded_to=?,
                additional_notes=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            ''', (
                safe_date(row.get('Registration Date (Payment Date)')),
                safe_str(row.get('Customer ID')),
                prefix, first_name, last_name,
                safe_str(row.get('Mobile Number')),
                safe_str(row.get('Whats App Number')),
                safe_str(row.get('Whats App Number 2')),
                safe_str(row.get('Candidate Email')),
                safe_date(row.get('D.O.B')),
                safe_str(row.get('CITY')),
                safe_str(row.get('STATE')),
                safe_str(row.get('Instagram Name')),
                safe_str(row.get('Facebook Name')),
                safe_str(row.get('LinkedIn Name')),
                safe_str(row.get('Fathers Name')),
                safe_str(row.get('Fathers Mobile Number')),
                safe_str(row.get('Mothers Name')),
                safe_str(row.get('Mothers Mobile Number')),
                safe_str(row.get('Parents Email ID')),
                safe_str(row.get('Joined Stage')),
                safe_str(row.get('Plan Type')),
                account_status,
                safe_str(row.get('PLAB Stage (Current Status)')),
                safe_str(row.get('Switched Program')),
                safe_str(row.get('Counsellor')),
                safe_str(row.get('Counsellor Email')),
                safe_str(row.get('Counsellor Number')),
                safe_str(row.get('Lead Source')),
                safe_str(row.get('UK Client Referral')),
                safe_str(row.get('Operations Team Referral')),
                pkg, disc,
                safe_str(row.get('Additional Notes (Discount Given, Package reduced reason)')),
                final, total_paid,
                i1, safe_date(row.get('1st Instalment Date')),
                safe_str(row.get('Ist Installment Note')),
                i2, safe_date(row.get('2nd Instalment Date')),
                safe_str(row.get('2nd Installment Note')),
                i3, safe_date(row.get('3rd Instalment Date')),
                safe_str(row.get('3rd Installment Note')),
                i4, safe_date(row.get('4th Instalment Date')),
                safe_str(row.get('4th Installment Note')),
                safe_date(row.get('Dropped Out / Switched Program Date')),
                safe_str(row.get('Upgraded To')),
                safe_str(row.get('Additional Notes')),
                existing['id']
            ))
            updated += 1
        else:
            # INSERT
            conn.execute('''INSERT INTO plab_clients (
                registration_number, registration_date, customer_id,
                prefix, first_name, last_name,
                mobile, whatsapp1, whatsapp2, email, dob, city, state,
                instagram, facebook, linkedin,
                father_name, father_phone, mother_name, mother_phone, parents_email,
                joined_stage, plan_type,
                account_status, current_stage, switched_program,
                counsellor, counsellor_email, counsellor_number,
                lead_source, referral_type, operations_referral,
                package_amount, discount_allowed, additional_package_notes,
                final_package, total_paid,
                inst1_amount, inst1_date, inst1_note,
                inst2_amount, inst2_date, inst2_note,
                inst3_amount, inst3_date, inst3_note,
                inst4_amount, inst4_date, inst4_note,
                dropped_date, upgraded_to,
                additional_notes, pathway
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'plab')''', (
                reg_num or next_reg(),
                safe_date(row.get('Registration Date (Payment Date)')),
                safe_str(row.get('Customer ID')),
                prefix, first_name, last_name,
                safe_str(row.get('Mobile Number')),
                safe_str(row.get('Whats App Number')),
                safe_str(row.get('Whats App Number 2')),
                safe_str(row.get('Candidate Email')),
                safe_date(row.get('D.O.B')),
                safe_str(row.get('CITY')),
                safe_str(row.get('STATE')),
                safe_str(row.get('Instagram Name')),
                safe_str(row.get('Facebook Name')),
                safe_str(row.get('LinkedIn Name')),
                safe_str(row.get('Fathers Name')),
                safe_str(row.get('Fathers Mobile Number')),
                safe_str(row.get('Mothers Name')),
                safe_str(row.get('Mothers Mobile Number')),
                safe_str(row.get('Parents Email ID')),
                safe_str(row.get('Joined Stage')),
                safe_str(row.get('Plan Type')),
                account_status,
                safe_str(row.get('PLAB Stage (Current Status)')),
                safe_str(row.get('Switched Program')),
                safe_str(row.get('Counsellor')),
                safe_str(row.get('Counsellor Email')),
                safe_str(row.get('Counsellor Number')),
                safe_str(row.get('Lead Source')),
                safe_str(row.get('UK Client Referral')),
                safe_str(row.get('Operations Team Referral')),
                pkg, disc,
                safe_str(row.get('Additional Notes (Discount Given, Package reduced reason)')),
                final, total_paid,
                i1, safe_date(row.get('1st Instalment Date')),
                safe_str(row.get('Ist Installment Note')),
                i2, safe_date(row.get('2nd Instalment Date')),
                safe_str(row.get('2nd Installment Note')),
                i3, safe_date(row.get('3rd Instalment Date')),
                safe_str(row.get('3rd Installment Note')),
                i4, safe_date(row.get('4th Instalment Date')),
                safe_str(row.get('4th Installment Note')),
                safe_date(row.get('Dropped Out / Switched Program Date')),
                safe_str(row.get('Upgraded To')),
                safe_str(row.get('Additional Notes')),
            ))
            inserted += 1

    conn.commit()
    conn.close()
    wb.close()
    print(f"Done! Inserted: {inserted}, Updated: {updated}, Skipped: {skipped}")

if __name__ == '__main__':
    run_import()
