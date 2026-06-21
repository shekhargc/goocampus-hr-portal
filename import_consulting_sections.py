"""
import_consulting_sections.py — refresh Standard Consulting (pathway='consulting')
from the new two-sheet Zoho exports. GCCSS reg numbers.

  1. step_clients()  — upsert plab_clients (pathway='consulting'); ~62 already exist.
  2. step_sections() — replace-per-section: Academic / AMC Reg / EPIC / Mentorship /
     Payments / Call Notes, reg-matched.

Run:
  DATABASE_URL=...  DRY_RUN=1  python3 import_consulting_sections.py   # plan only
  DATABASE_URL=...               python3 import_consulting_sections.py   # live
"""
import os
from import_amc_sections import (
    s, d, num, norm, digits, extract_reg, strip_reg, reg_variants,
    split_name, _strip_prefix, CONV, _pg,
)

DL = os.environ.get('CS_DIR', '/Users/Santosh/Desktop/Zoho Data/Standard Consulting')
PATHWAY = 'consulting'
DRY = os.environ.get('DRY_RUN') in ('1', 'true', 'True')


def _resolve(fname):
    return os.path.join(DL, fname)


CLIENT_FILE = "Standard Consulting -  Registration list.xlsx"
CLIENT_MAP = {
    'customer_id': ('Customer ID', 's'),
    'registration_date': ('Registration Date (Payment Date)', 'd'),
    'mobile': ('Mobile Number', 's'), 'email': ('Candidate Email', 's'),
    'whatsapp1': ('Whats App Number', 's'),
    'plan_type': ('Plan Type', 's'),
    'package_amount': ('Package (Mention Actual Package)', 'num'),
    'final_package': ('Final Package', 'num'),
    'discount_allowed': ('Discount Allowed (Discount Offer + Stage Discount)', 'num'),
    'account_status': ('Account Status', 's'), 'counsellor': ('Counsellor Name', 's'),
    'current_stage': ('Stage (Current Status)', 's'), 'joined_stage': ('Joined Stage', 's'),
    'lead_source': ('Lead Source', 's'),
    'dob': ('D.O.B', 'd'), 'city': ('CITY', 's'), 'state': ('STATE', 's'),
    'instagram': ('Instgram Account Name', 's'), 'facebook': ('Facebook Account Name', 's'),
    'linkedin': ('LinkedIn Account Name', 's'),
    'father_name': ('Fathers Name', 's'), 'father_phone': ('Fathers Mobile Number', 's'),
    'mother_name': ('Mothers Name', 's'), 'mother_phone': ('Mothers Mobile Number', 's'),
    'parents_email': ('Parents Email ID', 's'),
    'counsellor_email': ('Counsellor Email', 's'), 'counsellor_number': ('Counsellor Number', 's'),
    'portfolio_referral': ('Portfolio Client Referral', 's'),
    'australia_referral': ('Australia Client Referral:', 's'),
    'additional_package_notes': ('Discount Notes', 's'),
    'inst1_amount': ('1st Installment', 'num'), 'inst1_date': ('1st Installment Date', 'd'), 'inst1_note': ('1st Installment Note', 's'),
    'inst2_amount': ('2nd Installment', 'num'), 'inst2_date': ('2nd Installment Date', 'd'), 'inst2_note': ('2nd Installment Note', 's'),
    'additional_notes': ('Notes', 's'),
}


def step_clients():
    import openpyxl
    path = _resolve(CLIENT_FILE)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else '' for h in next(it)]
    hi = {h: i for i, h in enumerate(headers)}
    def g(row, h):
        i = hi.get(h)
        return row[i] if (i is not None and i < len(row)) else None
    rows = [r for r in it if any(v not in (None, '') for v in r)]
    wb.close()

    conn = _pg(); cur = conn.cursor()
    cur.execute("SELECT registration_number FROM plab_clients WHERE COALESCE(pathway,'plab')='consulting'")
    existing = set((r[0] or '').strip() for r in cur.fetchall())
    ins = upd = skip = 0
    for r in rows:
        reg = extract_reg(g(r, 'Registration Number')) or s(g(r, 'Registration Number'))
        if not reg:
            skip += 1; continue
        prefix, first, last = split_name(g(r, 'Candidate Name'))
        data = {'prefix': prefix, 'first_name': first, 'last_name': last}
        for col, (h, k) in CLIENT_MAP.items():
            data[col] = CONV[k](g(r, h))
        if DRY:
            ins += 0 if reg in existing else 1
            upd += 1 if reg in existing else 0
            continue
        if reg in existing:
            sets = ', '.join(f"{c}=%s" for c in data)
            cur.execute(f"UPDATE plab_clients SET {sets}, updated_at=NOW() "
                        f"WHERE registration_number=%s AND COALESCE(pathway,'plab')='consulting'",
                        list(data.values()) + [reg]); upd += 1
        else:
            cols = ['registration_number', 'pathway'] + list(data.keys())
            vals = [reg, 'consulting'] + list(data.values())
            cur.execute(f"INSERT INTO plab_clients ({','.join(cols)}) VALUES ({','.join(['%s']*len(cols))})", vals)
            existing.add(reg); ins += 1
    if not DRY:
        conn.commit()
    cur.close(); conn.close()
    print(f"[clients] inserted={ins} updated={upd} skipped(no reg)={skip}  (DRY={DRY})")


def build_index():
    conn = _pg(); cur = conn.cursor()
    cur.execute("SELECT registration_number, first_name, last_name, email, mobile "
                "FROM plab_clients WHERE COALESCE(pathway,'plab')='consulting'")
    by_reg, by_name, by_email, by_mobile = {}, {}, {}, {}
    for reg, fn, ln, email, mob in cur.fetchall():
        if not reg:
            continue
        canon = reg.strip()
        for v in reg_variants(canon):
            by_reg.setdefault(norm(v), canon)
        full = norm(f"{fn or ''} {ln or ''}")
        if full:
            by_name.setdefault(full, canon)
        if email:
            by_email.setdefault(norm(email), canon)
        if mob:
            dm = digits(mob)[-10:]
            if dm:
                by_mobile.setdefault(dm, canon)
    cur.close(); conn.close()
    return by_reg, by_name, by_email, by_mobile


def resolve_reg(idx, reg_val, name_val):
    by_reg, by_name, by_email, by_mobile = idx
    emb = extract_reg(name_val) or extract_reg(reg_val) or reg_val
    if emb:
        for v in reg_variants(emb):
            cc = by_reg.get(norm(v))
            if cc:
                return cc
    nm = norm(_strip_prefix(strip_reg(name_val)))
    if nm and nm in by_name:
        return by_name[nm]
    return None


SECTIONS = {
 'academic': ("Standard Consulting - Academic Details Report.xlsx", "ops_academic_details", "Registration Number", "Enter Candidate name", {
    'img_fmg': ('IMG / FMG', 's'), 'img_medical_college': ('IMG Medical College Name', 's'),
    'fmg_medical_college': ('FMG Medical College Name', 's'), 'mbbs_status': ('MBBS Status', 's'),
    'speciality_interest_1': ('Speciality Interest 1', 's'), 'speciality_interest_2': ('Speciality Interest 2', 's'),
    'internship_status': ('Internship Status', 's'), 'internship_gap': ('Internship Gap', 's'),
    'mbbs_start_date': ('MBBS Start Date', 'd'), 'mbbs_end_date': ('MBBS End Date', 'd'),
    'internship_hospital': ('Internship Hospital', 's'), 'internship_location': ('Internship Location (State/Country)', 's'),
    'internship_hospital_2': ('Internship Hospital 2', 's'), 'internship_location_2': ('Internship Location 2 (State / Country)', 's'),
    'internship_start_date': ('Internship Start Date', 'd'), 'internship_end_date': ('Internship End Date', 'd'),
    'gap_in_months': ('Internship Gap in Months', 's'), 'gap_reason': ('Intership Gap Reason', 's'),
    'working_status': ('Working Status', 's'), 'working_hospital_name': ('Working Hospital Name', 's'),
    'additional_info': ('Additional Info', 's'),
 }),
 'amc_reg': ("Standard Consulting - All Amc Registrations.xlsx", "ops_amc_registration", None, "Enter Candidate Name", {
    'amc_reference_number': ('AMC Reference Number', 's'), 'login_pwd': ('Login Password', 's'),
    'amc_setup': ('AMC Setup Status', 's'), 'registration_date': ('Registration Date', 'd'),
 }),
 'epic': ("Standard Consulting - All Epic Verifications.xlsx", "ops_epic_registration", None, "Enter Candidate Name", {
    'login_id': ('Login ID', 's'), 'login_pwd': ('Login Password', 's'),
    'secret_question_1': ('Secret Question 1', 's'), 'secret_answer_1': ('Secret Answer 1', 's'),
    'secret_question_2': ('Secret Question 2', 's'), 'secret_answer_2': ('Secret Answer 2', 's'),
    'secret_question_3': ('Secret Question 3', 's'), 'secret_answer_3': ('Secret Answer 3', 's'),
    'secret_question_4': ('Secret Question 4', 's'), 'secret_answer_4': ('Secret Answer 4', 's'),
    'notary_camp_login': ('Notary Cam Login', 's'), 'notary_camp_password': ('Notary Cam Password', 's'),
    'epic_registration': ('Epic Registration Status', 's'), 'registration_date': ('Registration Date', 'd'),
    'epic_id_number': ('EPIC ID Number', 's'), 'epic_status': ('EPIC Status', 's'),
    'notary_camp': ('Notary Cam Status', 's'), 'documents_stage': ('Document Stage', 's'),
    'document_stage_status': ('Document Stage Status', 's'),
 }),
 'mentorship': ("Standard Consulting - All Mentorship Sessions.xlsx", "ops_mentorship", None, "Enter Candidate Name", {
    'session_date': ('Date', 'd'), 'service_description': ('Session Name', 's'),
    'duration_minutes': ('Duration (In Hours)', 's'), 'amount_paid': ('Amount Paid', 'num'),
    'program_provider': ('Mentor Name', 's'), 'additional_notes': ('Additional Notes', 's'),
 }),
 'payments': ("Standard Consutling - All Clients Payments.xlsx", "ops_payments", "Registration Number", "Enter Candidate Name", {
    'total_package': ('Total Package', 'num'), 'instalment': ('Instalment', 's'),
    'total_amount_paid': ('Total Amount Paid', 'num'), 'payment_date': ('Payment Date', 'd'),
    'amount_paid': ('Amount Paid', 'num'), 'gst_paid': ('GST Paid', 'num'),
    'payment_method': ('Payment Method', 's'), 'notes': ('Notes', 's'),
 }),
 'call_notes': ("Standard Consulting - Client Call Notes Report (1).xlsx", "ops_call_notes", None, "Candidate Name", {
    'call_date': ('Call Date', 'd'), 'call_note': ('Call Notes', 's'), 'added_by': ('Added User', 's'),
 }),
}


def step_sections():
    import openpyxl
    only = os.environ.get('SECTIONS')
    only = set(only.split(',')) if only else None
    idx = build_index()
    print(f"[index] {len(idx[1])} names, {len(idx[0])} reg-variants")
    for key, (fname, table, reg_col, name_col, mapping) in SECTIONS.items():
        if only and key not in only:
            continue
        path = _resolve(fname)
        if not os.path.exists(path):
            print(f"[{key}] MISSING {fname}"); continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else '' for h in next(it)]
        hi = {h: i for i, h in enumerate(headers)}
        def g(row, h):
            i = hi.get(h)
            return row[i] if (i is not None and i < len(row)) else None
        batch = []; unmatched = 0
        for r in it:
            if not any(v not in (None, '') for v in r):
                continue
            reg_val = s(g(r, reg_col)) if reg_col else None
            name_val = s(g(r, name_col)) if name_col else None
            if not (reg_val or name_val):
                continue
            reg = resolve_reg(idx, reg_val, name_val)
            if not reg:
                unmatched += 1; continue
            row_vals = [reg] + [CONV[k](g(r, h)) for (h, k) in mapping.values()] + ['consulting']
            batch.append(tuple(row_vals))
        wb.close()
        cols = ['registration_number'] + list(mapping.keys()) + ['pathway']
        if DRY:
            print(f"[{key:10}] {table:26} would insert={len(batch):4} unmatched={unmatched}")
            continue
        conn = _pg(); conn.autocommit = False; cur = conn.cursor()
        try:
            cur.execute(f"DELETE FROM {table} WHERE COALESCE(pathway,'plab')='consulting'")
            if batch:
                from psycopg2.extras import execute_values
                tmpl = "(" + ",".join(["%s"] * len(cols)) + ")"
                execute_values(cur, f"INSERT INTO {table} ({','.join(cols)}) VALUES %s",
                               batch, template=tmpl, page_size=500)
            conn.commit()
            print(f"[{key:10}] {table:26} inserted={len(batch):4} unmatched={unmatched}")
        except Exception as e:
            conn.rollback(); print(f"[{key}] FAILED: {e}")
        finally:
            cur.close(); conn.close()


def run():
    print(f"=== Consulting import  DRY={DRY} ===")
    step_clients()
    step_sections()


if __name__ == '__main__':
    run()
