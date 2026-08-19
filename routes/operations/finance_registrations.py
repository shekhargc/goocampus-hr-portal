"""
routes/operations/finance_registrations.py

READ-ONLY "Finance — Client Registrations" panel: an audit view of client
registrations across EVERY pathway (plab / australia [AMC] / consulting / uae /
portfolio / training) with each client's payments. SELECT only — never writes.

Access: gated by Access Master  finance → registrations → view  (admins bypass).
Routes are @login_required and mapped in app.py's ACCESS_ROUTE_MAP so the
before_request access hook enforces the grant.
"""
import logging
from datetime import datetime, date

from flask import render_template, request, jsonify, redirect, flash, session

from core.auth import login_required
from core.users import get_user
from db import get_db

_PW_LABELS = {
    'plab': 'UK / PLAB', 'australia': 'AMC (Australia)', 'consulting': 'Consulting',
    'uae': 'UAE', 'portfolio': 'Portfolio', 'training': 'Training',
}

# registration_date is free-text in mixed formats — parse tolerantly in Python
# (never cast to ::date in SQL: one bad row would error the whole query).
_DATE_FMTS = ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%d-%b-%Y', '%d %b %Y',
              '%d-%B-%Y', '%d %B %Y', '%m/%d/%Y', '%d.%m.%Y', '%Y-%m-%d %H:%M:%S')


def _parse_date_loose(s):
    if not s:
        return None
    s = str(s).strip()
    if not s:
        return None
    for fmt in _DATE_FMTS:
        try:
            return datetime.strptime(s[:19] if fmt.endswith('%S') else s, fmt).date()
        except Exception:
            continue
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d').date()
    except Exception:
        return None


def _full_name(r):
    from core.helpers import format_dr_name
    raw = ' '.join(p for p in [(r.get('prefix') or '').strip(),
                               (r.get('first_name') or '').strip(),
                               (r.get('last_name') or '').strip()] if p).strip()
    return format_dr_name(raw)


def _reg_key(reg):
    return (reg or '').strip().upper()


def _fetch_finance_rows(conn, f_pathway, f_from, f_to, q):
    """Shared filter/sort so the list page and the Excel export return exactly
    the same set for the same filters."""
    rows = []
    try:
        where = ["registration_number IS NOT NULL AND registration_number <> ''"]
        params = []
        if f_pathway in _PW_LABELS:
            where.append("COALESCE(pathway,'plab') = ?"); params.append(f_pathway)
        if q:
            where.append("(first_name ILIKE ? OR last_name ILIKE ? OR registration_number ILIKE ? "
                         "OR mobile ILIKE ? OR email ILIKE ? OR whatsapp1 ILIKE ?)")
            params += [f'%{q}%'] * 6
        recs = [dict(r) for r in conn.execute(
            "SELECT id, registration_number, registration_date, prefix, first_name, last_name, "
            "mobile, whatsapp1, email, city, state, COALESCE(pathway,'plab') AS pathway, "
            # Resolve the counsellor from the best available source so ops-team
            # counsellors show too (founder 2026-08-14). NOTE: plab_clients has no
            # counsellor_name column — only counsellor (name) + counsellor_id.
            "COALESCE(NULLIF(TRIM(counsellor),''), "
            "  (SELECT e.name FROM employees e WHERE e.id = plab_clients.counsellor_id)) AS counsellor, "
            "account_status, final_package "
            f"FROM plab_clients WHERE {' AND '.join(where)}", params).fetchall()]
        # total paid per registration (SUM of GST-inclusive totals)
        paid_map = {}
        try:
            for pr in conn.execute("SELECT UPPER(TRIM(registration_number)) AS k, "
                                   "COALESCE(SUM(total_amount_paid),0) AS tp "
                                   "FROM ops_payments GROUP BY 1").fetchall():
                paid_map[pr['k']] = float(pr['tp'] or 0)
        except Exception as e:
            logging.error(f"finance_registrations paid_map: {e}")
        d_from = _parse_date_loose(f_from) if f_from else None
        d_to = _parse_date_loose(f_to) if f_to else None
        for r in recs:
            r['full_name'] = _full_name(r) or '—'
            r['pathway_label'] = _PW_LABELS.get(r['pathway'], r['pathway'])
            r['_pd'] = _parse_date_loose(r.get('registration_date'))
            r['total_paid'] = paid_map.get(_reg_key(r.get('registration_number')), 0.0)
        if d_from:
            recs = [r for r in recs if r['_pd'] and r['_pd'] >= d_from]
        if d_to:
            recs = [r for r in recs if r['_pd'] and r['_pd'] <= d_to]
        # newest registration date first; undated rows sink to the bottom
        recs.sort(key=lambda r: (r['_pd'] or date.min, r['id']), reverse=True)
        rows = recs
    except Exception as e:
        logging.error(f"_fetch_finance_rows: {e}")
        try: conn.rollback()
        except Exception: pass
    return rows


@login_required
def finance_registrations_list():
    user = get_user()
    conn = get_db()
    f_pathway = (request.args.get('pathway') or '').strip().lower()
    f_from = (request.args.get('from') or '').strip()
    f_to = (request.args.get('to') or '').strip()
    q = (request.args.get('q') or '').strip()
    rows = _fetch_finance_rows(conn, f_pathway, f_from, f_to, q)
    conn.close()
    return render_template('finance_registrations.html', user=user, rows=rows,
                           pathways=_PW_LABELS, f_pathway=f_pathway,
                           f_from=f_from, f_to=f_to, q=q, total=len(rows),
                           active_section='finance')


@login_required
def finance_registrations_export():
    """Export the CURRENT filtered view to Excel (.xlsx). Read-only."""
    import io, re as _re
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from core.helpers import format_reg
    from flask import Response

    conn = get_db()
    f_pathway = (request.args.get('pathway') or '').strip().lower()
    f_from = (request.args.get('from') or '').strip()
    f_to = (request.args.get('to') or '').strip()
    q = (request.args.get('q') or '').strip()
    rows = _fetch_finance_rows(conn, f_pathway, f_from, f_to, q)
    conn.close()

    _ILLEGAL = _re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')  # openpyxl rejects these
    def _c(v):
        if v is None:
            return ''
        return _ILLEGAL.sub('', v) if isinstance(v, str) else v

    wb = Workbook(); ws = wb.active; ws.title = 'Registrations'
    headers = ['#', 'Registration Date', 'Registration Number', 'Full Name', 'Pathway',
               'State', 'City', 'Mobile', 'WhatsApp', 'Email', 'Counsellor', 'Status', 'Total Paid (INR)']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1E3A5F')
    for i, r in enumerate(rows, start=1):
        ws.append([
            i, _c(r.get('registration_date')), _c(format_reg(r.get('registration_number'))),
            _c(r.get('full_name')), _c(r.get('pathway_label')), _c(r.get('state')), _c(r.get('city')),
            _c(r.get('mobile')), _c(r.get('whatsapp1')), _c(r.get('email')), _c(r.get('counsellor')),
            _c(r.get('account_status')), round(float(r.get('total_paid') or 0), 2),
        ])
    if not rows:
        ws.append(['No registrations match the current filters.'])
    # readable column widths
    for col, w in zip('ABCDEFGHIJKLM', (5, 16, 22, 26, 16, 16, 14, 14, 14, 26, 18, 16, 16)):
        ws.column_dimensions[col].width = w

    out = io.BytesIO(); wb.save(out); out.seek(0)
    parts = ['client_registrations']
    if f_pathway:
        parts.append(f_pathway)
    fname = '_'.join(parts) + '.xlsx'
    return Response(out.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{fname}"'})


@login_required
def finance_registration_detail(client_id):
    conn = get_db()
    c = conn.execute("SELECT * FROM plab_clients WHERE id = ?", (client_id,)).fetchone()
    if not c:
        conn.close(); return jsonify({'error': 'not found'}), 404
    c = dict(c)
    reg = (c.get('registration_number') or '').strip()
    pays = []
    try:
        pays = [dict(p) for p in conn.execute(
            "SELECT payment_date, amount_paid, gst_paid, total_amount_paid, instalment, "
            "payment_method, notes FROM ops_payments "
            "WHERE UPPER(TRIM(registration_number)) = ? ORDER BY id", (_reg_key(reg),)).fetchall()]
    except Exception as e:
        logging.error(f"finance_registration_detail payments: {e}")
    conn.close()
    profile = {
        'name': _full_name(c) or '—', 'registration_number': reg,
        'registration_date': c.get('registration_date') or '',
        'pathway': _PW_LABELS.get(c.get('pathway') or 'plab', c.get('pathway') or 'plab'),
        'mobile': c.get('mobile') or '', 'whatsapp': c.get('whatsapp1') or '',
        'email': c.get('email') or '', 'city': c.get('city') or '', 'state': c.get('state') or '',
        'counsellor': c.get('counsellor') or '', 'status': c.get('account_status') or '',
        'package': c.get('final_package') or '',
        'total_paid': sum(float(p.get('total_amount_paid') or 0) for p in pays),
    }
    payments = [{
        'date': p.get('payment_date') or '', 'amount': float(p.get('amount_paid') or 0),
        'gst': float(p.get('gst_paid') or 0), 'total': float(p.get('total_amount_paid') or 0),
        'instalment': p.get('instalment') or '', 'method': p.get('payment_method') or '',
        'notes': p.get('notes') or '',
    } for p in pays]
    return jsonify({'profile': profile, 'payments': payments})


# Smart Finance landing: the top-nav "Finance" link points here so a finance
# staff member lands on the FIRST finance page they actually have access to
# (previously it always went to /finance/budget, which a registrations-only
# grant can't open).
_FIN_LANDING = [
    ('budget', '/finance/budget'), ('revenue', '/finance/budget/revenue'),
    ('expenses', '/finance/budget/expenses'), ('salary', '/finance/salaries'),
    ('reports', '/finance/budget/report'), ('registrations', '/finance/registrations'),
    ('settings', '/finance/budget/settings'),
]


@login_required
def finance_index():
    if session.get('is_admin'):
        return redirect('/finance/budget')
    user = get_user()
    from app import has_section_permission
    for sub, url in _FIN_LANDING:
        try:
            if has_section_permission(user, 'finance', sub, 'view'):
                return redirect(url)
        except Exception:
            continue
    flash('You do not have access to the Finance section.', 'error')
    return redirect('/dashboard')


def register_routes(app):
    app.add_url_rule('/finance', endpoint='finance_index',
                     view_func=finance_index, methods=['GET'])
    app.add_url_rule('/finance/registrations', endpoint='finance_registrations_list',
                     view_func=finance_registrations_list, methods=['GET'])
    app.add_url_rule('/finance/registrations/export.xlsx', endpoint='finance_registrations_export',
                     view_func=finance_registrations_export, methods=['GET'])
    app.add_url_rule('/finance/registrations/<int:client_id>/detail.json',
                     endpoint='finance_registration_detail',
                     view_func=finance_registration_detail, methods=['GET'])
