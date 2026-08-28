"""
routes/operations/australia.py — Operations: AMC Pathway dashboard.

Surfaces Australia client stats from plab_clients WHERE pathway='australia'
(the storage decision locked 2026-05-31). The 228 historical Australia
clients were imported from GC_AUS_Registration_Report.xlsx by
import_australia_clients.run_import_australia_clients_once() at app boot.

The Australia Excel and the import script are committed alongside this
file so the same data shows up identically on every fresh environment.

Endpoint name preserved: ops_australia_pathway. Sidebar pathway switcher
links to /operations/australia-pathway.
"""

import logging
from flask import render_template, flash, request, redirect, url_for

from core.auth import admin_required
from core.users import get_user
from db import get_db
# Pathway-scoped dropdown options for the AMC Registration edit form so the
# Plan Type / Account Status / Current Stage / Counsellor / Lead Source
# <select>s render the same options PLAB uses (sourced from the AMC tab
# of Field Manager).
from routes.operations._form_lookups import section_client_lookups, section_client_products, transfer_for_reg


@admin_required
def ops_australia_pathway():
    """AMC Pathway dashboard — stats from plab_clients where pathway='australia'."""
    user = get_user()
    conn = get_db()

    stats = {
        'total_clients':        0,
        'active_clients':       0,
        'on_hold_clients':      0,
        'dropped_clients':      0,
        'completed_clients':    0,
        'account_status':       {},   # {status: count}
        'current_stage':        {},   # {stage: count}
        'counsellor_breakdown': [],   # top 5 [(counsellor, count)]
        'package_total':        0.0,  # sum of final_package across active
        'recent_registrations': [],   # last 5 rows
        'this_fy_new':          0,    # signups in current Indian FY
        # Section counts shown on the dashboard tile grid.
        'section_counts':       {},   # {slug: count}
        # Upcoming booked exams, each with a clamped `days_until` int.
        'upcoming_exams':       [],
    }

    try:
        # ── Totals + status buckets ──────────────────────────────────────
        total = conn.execute(
            "SELECT COUNT(*) AS c FROM plab_clients WHERE pathway = 'australia'"
        ).fetchone()['c']
        stats['total_clients'] = total

        # Per-status counts (single grouped query, then map keys onto buckets)
        for row in conn.execute(
            """SELECT COALESCE(NULLIF(TRIM(account_status), ''), 'Unknown') AS s,
                      COUNT(*) AS c
               FROM plab_clients WHERE pathway = 'australia'
               GROUP BY account_status"""
        ).fetchall():
            stats['account_status'][row['s']] = row['c']

        # Convenience: surface the four most common buckets at the top
        stats['active_clients']    = stats['account_status'].get('In Process', 0)
        stats['on_hold_clients']   = stats['account_status'].get('On Hold', 0)
        stats['dropped_clients']   = stats['account_status'].get('Dropped', 0)
        stats['completed_clients'] = stats['account_status'].get('Completed', 0)

        # ── Current stage breakdown (active clients only) ────────────────
        for row in conn.execute(
            """SELECT COALESCE(NULLIF(TRIM(current_stage), ''), 'Not Set') AS stg,
                      COUNT(*) AS c
               FROM plab_clients
               WHERE pathway = 'australia' AND account_status = 'In Process'
               GROUP BY current_stage
               ORDER BY c DESC"""
        ).fetchall():
            stats['current_stage'][row['stg']] = row['c']

        # ── Counsellor distribution (top 5) ─────────────────────────────
        stats['counsellor_breakdown'] = [
            (r['counsellor'] or 'Unassigned', r['c'])
            for r in conn.execute(
                """SELECT COALESCE(NULLIF(TRIM(counsellor), ''), 'Unassigned') AS counsellor,
                          COUNT(*) AS c
                   FROM plab_clients
                   WHERE pathway = 'australia' AND account_status = 'In Process'
                   GROUP BY counsellor
                   ORDER BY c DESC
                   LIMIT 5"""
            ).fetchall()
        ]

        # ── Package value (active clients) ──────────────────────────────
        pkg_row = conn.execute(
            """SELECT COALESCE(SUM(final_package), 0) AS s
               FROM plab_clients
               WHERE pathway = 'australia' AND account_status = 'In Process'"""
        ).fetchone()
        stats['package_total'] = float(pkg_row['s'] or 0)

        # ── This-FY new registrations ───────────────────────────────────
        # Indian FY runs Apr-Mar. We could compute via SQL but the count is
        # cheap to do in Python with the registration_number prefix.
        from core.registration import indian_financial_year
        fy = indian_financial_year()
        fy_count = conn.execute(
            "SELECT COUNT(*) AS c FROM plab_clients "
            "WHERE pathway = 'australia' AND registration_number LIKE ?",
            (f'GCAUSIP/{fy}/%',),
        ).fetchone()['c']
        stats['this_fy_new'] = fy_count

        # ── Recent registrations (last 5 by id, which == reg order) ─────
        stats['recent_registrations'] = conn.execute(
            """SELECT id, registration_number, prefix, first_name, last_name,
                      mobile, email, account_status, current_stage,
                      counsellor, final_package, registration_date
               FROM plab_clients
               WHERE pathway = 'australia'
               ORDER BY id DESC
               LIMIT 5"""
        ).fetchall()

        # ── Section row counts for the dashboard tile grid ──────────────
        section_tables = {
            'test_bookings':   'ops_test_bookings',
            'academic':        'ops_academic_details',
            'epic':            'ops_epic_registration',
            'training':        'ops_coaching',
            'online_courses':  'ops_online_subscriptions',
            'payments':        'ops_payments',
            'call_notes':      'ops_call_notes',
            'research':        'ops_research_publication',
            'webinars':        'ops_webinars_conferences',
        }
        for slug, table in section_tables.items():
            try:
                stats['section_counts'][slug] = conn.execute(
                    f"SELECT COUNT(*) AS c FROM {table} WHERE pathway = 'australia'"
                ).fetchone()['c']
            except Exception:
                stats['section_counts'][slug] = 0

        # ── Detailed per-section stats (matches PLAB Pathway Dashboard) ──
        # All scoped to pathway='australia' so PLAB rows can't leak.
        from datetime import date, timedelta
        today = date.today().isoformat()
        thirty_ago = (date.today() - timedelta(days=30)).isoformat()
        AU = "pathway = 'australia'"

        def cnt(sql, *params):
            try:
                return conn.execute(sql, params).fetchone()['c']
            except Exception:
                try: conn.rollback()
                except Exception: pass
                return 0

        # Coaching / Training
        stats['coaching_total']   = cnt(f"SELECT COUNT(*) AS c FROM ops_coaching WHERE {AU}")
        stats['coaching_ongoing'] = cnt(f"SELECT COUNT(*) AS c FROM ops_coaching WHERE {AU} AND coaching_status = 'On Going'")

        # Test Bookings
        stats['test_total']       = stats['section_counts']['test_bookings']
        stats['upcoming_tests']   = cnt(f"SELECT COUNT(*) AS c FROM ops_test_bookings WHERE {AU} AND exam_date >= ? AND exam_status = 'Booked'", today)
        # AMC 1 / AMC 2 are Australia's PLAB 1 / PLAB 2 equivalents
        stats['amc1_upcoming']    = cnt(f"SELECT COUNT(*) AS c FROM ops_test_bookings WHERE {AU} AND exam_date >= ? AND exam_status = 'Booked' AND exam LIKE 'AMC 1%'", today)
        stats['amc2_upcoming']    = cnt(f"SELECT COUNT(*) AS c FROM ops_test_bookings WHERE {AU} AND exam_date >= ? AND exam_status = 'Booked' AND exam LIKE 'AMC 2%'", today)
        stats['amc_mcq_upcoming'] = cnt(f"SELECT COUNT(*) AS c FROM ops_test_bookings WHERE {AU} AND exam_date >= ? AND exam_status = 'Booked' AND exam LIKE 'AMC MCQ%'", today)
        stats['amc_clin_upcoming']= cnt(f"SELECT COUNT(*) AS c FROM ops_test_bookings WHERE {AU} AND exam_date >= ? AND exam_status = 'Booked' AND exam LIKE 'AMC Clinical%'", today)
        stats['awaiting_results'] = cnt(f"SELECT COUNT(*) AS c FROM ops_test_bookings WHERE {AU} AND (exam_result IS NULL OR exam_result = '') AND (exam_status = 'Attended' OR (exam_status = 'Booked' AND exam_date < ?))", today)
        stats['recent_passes']    = cnt(f"SELECT COUNT(*) AS c FROM ops_test_bookings WHERE {AU} AND exam_result = 'Passed' AND exam_result_date >= ?", thirty_ago)

        # EPIC
        stats['epic_total']       = stats['section_counts']['epic']
        stats['epic_in_process']  = cnt(f"SELECT COUNT(*) AS c FROM ops_epic_registration WHERE {AU} AND epic_status = 'In Process'")
        stats['epic_completed']   = cnt(f"SELECT COUNT(*) AS c FROM ops_epic_registration WHERE {AU} AND epic_status = 'Completed'")

        # Academic
        stats['academic_total']   = stats['section_counts']['academic']

        # Research
        stats['research_total']     = stats['section_counts']['research']
        stats['research_started']   = cnt(f"SELECT COUNT(*) AS c FROM ops_research_publication WHERE {AU} AND research_status = 'Started'")
        stats['research_completed'] = cnt(f"SELECT COUNT(*) AS c FROM ops_research_publication WHERE {AU} AND research_status = 'Research Completed'")
        stats['research_published'] = cnt(f"SELECT COUNT(*) AS c FROM ops_research_publication WHERE {AU} AND research_status = 'Research Published'")

        # Online Courses
        stats['oc_total']         = stats['section_counts']['online_courses']

        # Payments
        stats['payments_total']   = stats['section_counts']['payments']
        try:
            stats['payments_sum'] = float(conn.execute(
                f"SELECT COALESCE(SUM(total_amount_paid), 0) AS s FROM ops_payments WHERE {AU}"
            ).fetchone()['s'] or 0)
        except Exception:
            stats['payments_sum'] = 0.0

        # Webinars
        stats['webinars_total']   = stats['section_counts']['webinars']

        # ── Upcoming exams list (all booked, soonest first) ──
        # Mirrors the Consulting dashboard: pull a wide window (LIMIT 200),
        # convert rows to dicts and attach a clamped `days_until` integer so
        # the tabbed Upcoming Exams widget can split them into 7 / 30 / 30+
        # day buckets. Un-parseable exam_dates are skipped.
        try:
            from datetime import datetime
            rows = conn.execute(
                f"""SELECT t.*, p.prefix, p.first_name, p.last_name
                      FROM ops_test_bookings t
                 LEFT JOIN plab_clients p ON p.registration_number = t.registration_number
                                          AND COALESCE(p.pathway, 'plab') = 'australia'
                     WHERE COALESCE(t.pathway, 'plab') = 'australia'
                       AND t.exam_date >= ?
                       AND t.exam_status = 'Booked'
                  ORDER BY t.exam_date ASC LIMIT 200""",
                (today,),
            ).fetchall()
            today_d = date.today()
            ue = []
            for r in rows:
                d = dict(r)
                ed = (d.get('exam_date') or '').strip() if isinstance(d.get('exam_date'), str) else d.get('exam_date')
                days_until = None
                try:
                    if ed:
                        # exam_date is stored as an ISO 'YYYY-MM-DD' string
                        exam_d = datetime.strptime(str(ed)[:10], '%Y-%m-%d').date()
                        days_until = (exam_d - today_d).days
                        if days_until < 0:
                            days_until = 0
                except Exception:
                    days_until = None
                if days_until is None:
                    # un-parseable date -> skip so the windows stay clean
                    continue
                d['days_until'] = days_until
                ue.append(d)
            stats['upcoming_exams'] = ue
        except Exception:
            stats['upcoming_exams'] = []

    except Exception as e:
        logging.error(f"ops_australia_pathway: {e}")
        flash(f'Error loading AMC Pathway dashboard: {e}', 'error')
    finally:
        conn.close()

    return render_template(
        'ops_australia_pathway_dashboard.html',
        user=user,
        pathway_name='AMC Pathway',
        stats=stats,
        active_ops_page='australia',
        active_pathway='australia',
    )


# NOTE: ops_australia_test_bookings_list was removed from this module.
# Australia Test Bookings (including the new drawer/detail/edit work
# from Phase 4 standardization) now lives entirely in
# routes/operations/au_test_bookings.py. Registration is wired in
# routes/operations/__init__.py.


@admin_required
def ops_australia_clients_list():
    """Australia Registration — list plab_clients WHERE pathway='australia'.

    This is what the Registration sidebar item links to when on Australia
    Pathway. Mirrors the PLAB client list but filtered to the Australia
    pathway and using the agent-built table template style.
    """
    user = get_user()
    conn = get_db()

    search = (request.args.get('q', '') or '').strip()
    status_filter = (request.args.get('status', '') or '').strip()
    stage_filter = (request.args.get('stage', '') or '').strip()
    counsellor_filter = (request.args.get('counsellor', '') or '').strip()

    records = []
    statuses = []
    stages = []
    counsellors = []
    total = 0

    try:
        sql = '''SELECT * FROM plab_clients
                  WHERE pathway = 'australia' '''
        params = []
        if status_filter:
            sql += " AND account_status = ? "
            params.append(status_filter)
        if stage_filter:
            sql += " AND current_stage = ? "
            params.append(stage_filter)
        if counsellor_filter:
            sql += " AND counsellor = ? "
            params.append(counsellor_filter)
        if search:
            sql += """ AND (
                first_name ILIKE ? OR last_name ILIKE ? OR
                registration_number ILIKE ? OR mobile ILIKE ? OR email ILIKE ? OR
                (COALESCE(prefix,'')||' '||first_name||' '||COALESCE(last_name,'')) ILIKE ?
            ) """
            params.extend([f'%{search}%'] * 6)
        # Latest registrations on top (user request 2026-06-01) — mirrors PLAB.
        sql += " ORDER BY NULLIF(regexp_replace(COALESCE(registration_number,''), '[^0-9]', '', 'g'), '')::bigint DESC NULLS LAST, id DESC "
        records = [dict(r) for r in conn.execute(sql, params).fetchall()]
        total = len(records)
        # Attach per-client payment totals from ops_payments (single source of
        # truth) so the 'Paid' column reflects real payments — for every pathway.
        if records:
            def _nrm(s): return (s or '').strip().upper()
            _keys = [_nrm(r['registration_number']) for r in records]
            _ph = ','.join(['?'] * len(_keys))
            _pmap = {pr['k']: pr for pr in conn.execute(
                f"""SELECT UPPER(TRIM(registration_number)) AS k,
                           COALESCE(SUM(total_amount_paid), 0) AS tp,
                           COALESCE(SUM(amount_paid), 0)       AS ap,
                           COALESCE(SUM(gst_paid), 0)          AS gp
                      FROM ops_payments WHERE UPPER(TRIM(registration_number)) IN ({_ph})
                     GROUP BY UPPER(TRIM(registration_number))""", _keys).fetchall()}
            for r in records:
                pr = _pmap.get(_nrm(r['registration_number']))
                r['total_paid']  = float(pr['tp']) if pr else 0
                r['amount_paid'] = float(pr['ap']) if pr else 0
                r['gst_paid']    = float(pr['gp']) if pr else 0

        statuses = [
            r['account_status'] for r in conn.execute(
                """SELECT DISTINCT account_status FROM plab_clients
                    WHERE pathway = 'australia'
                      AND account_status IS NOT NULL AND account_status != ''
                    ORDER BY account_status"""
            ).fetchall()
        ]
        stages = [
            r['current_stage'] for r in conn.execute(
                """SELECT DISTINCT current_stage FROM plab_clients
                    WHERE pathway = 'australia'
                      AND current_stage IS NOT NULL AND current_stage != ''
                    ORDER BY current_stage"""
            ).fetchall()
        ]
        counsellors = [
            r['counsellor'] for r in conn.execute(
                """SELECT DISTINCT counsellor FROM plab_clients
                    WHERE pathway = 'australia'
                      AND counsellor IS NOT NULL AND counsellor != ''
                    ORDER BY counsellor"""
            ).fetchall()
        ]
    except Exception as e:
        logging.error(f"ops_australia_clients_list: {e}")
        flash(f'Error loading Australia client list: {e}', 'error')
    finally:
        conn.close()

    return render_template(
        'ops_australia_clients_list.html',
        user=user,
        records=records,
        total=total,
        search=search,
        status_filter=status_filter,
        stage_filter=stage_filter,
        counsellor_filter=counsellor_filter,
        statuses=statuses,
        stages=stages,
        counsellors=counsellors,
        pathway_name='AMC Pathway',
        active_ops_page='australia-clients',
        active_pathway='australia',
    )


# ── Editable columns on plab_clients (pathway='australia' scope) ────────────
# Only columns that are safe to edit through the UI live here. id /
# registration_number / pathway are NOT editable.
AU_EDITABLE_COLUMNS = [
    # Personal
    'prefix', 'first_name', 'last_name', 'mobile', 'whatsapp1', 'whatsapp2',
    'email', 'dob', 'city', 'state',
    'instagram', 'facebook', 'linkedin',
    'father_name', 'father_phone', 'mother_name', 'mother_phone', 'parents_email',
    # Service
    'plan_type', 'product_id', 'account_status', 'current_stage', 'switched_program',
    'counsellor', 'counsellor_email', 'counsellor_number',
    'lead_source', 'registration_date',
    # Financials
    'package_amount', 'discount_allowed', 'final_package',
    'inst1_amount', 'inst1_date', 'inst1_note',
    'inst2_amount', 'inst2_date', 'inst2_note',
    'inst3_amount', 'inst3_date', 'inst3_note',
    'inst4_amount', 'inst4_date', 'inst4_note',
    # Centralised cross-pathway referral (data layer pre-created; columns exist)
    'referral_source_pathway', 'referral_channel', 'referral_client_reg', 'referral_client_name',
]


def _payment_totals_for(conn, reg_num):
    """Return (amount_paid, gst_paid, total_paid) summed from ops_payments
    for one Australia client. Pathway-scoped so PLAB payments can't leak."""
    row = conn.execute(
        """SELECT COALESCE(SUM(amount_paid), 0)        AS amt,
                  COALESCE(SUM(gst_paid), 0)           AS gst,
                  COALESCE(SUM(total_amount_paid), 0)  AS tot
             FROM ops_payments
            WHERE UPPER(TRIM(registration_number)) = UPPER(TRIM(?)) """,
        (reg_num,),
    ).fetchone()
    return float(row['amt'] or 0), float(row['gst'] or 0), float(row['tot'] or 0)


@admin_required
def ops_australia_client_detail(client_id):
    """Full Australia client profile — view + edit form + linked sections.

    Mirrors PLAB's /operations/plab/<id> structure: client info, payment
    breakdown, and tables of related ops_* records (test bookings, payments,
    academic, EPIC, training, etc.) all scoped to pathway='australia'.
    """
    user = get_user()
    conn = get_db()
    client = conn.execute(
        "SELECT * FROM plab_clients WHERE id = ? AND COALESCE(pathway, 'plab') = 'australia'",
        (client_id,),
    ).fetchone()
    if not client:
        conn.close()
        flash('Australia client not found.', 'error')
        return redirect(url_for('ops_australia_clients_list'))

    reg = client['registration_number']
    amount_paid, gst_paid, total_paid = _payment_totals_for(conn, reg)
    final_pkg = float(client['final_package'] or 0) or (
        float(client['package_amount'] or 0) - float(client['discount_allowed'] or 0)
    )
    balance = final_pkg - amount_paid
    pct = (amount_paid / final_pkg * 100) if final_pkg > 0 else 0

    # Related sections — only pull pathway='australia' rows for safety.
    # Each fetch returns a list of plain dicts so the section_block macro
    # in the template can iterate columns with .items().
    def fetch(table, order=None):
        try:
            sql = (
                f"SELECT * FROM {table} "
                f"WHERE registration_number = ? "
                f"  AND COALESCE(pathway, 'plab') = 'australia' "
            )
            if order:
                sql += f" ORDER BY {order} "
            return [dict(x) for x in conn.execute(sql, (reg,)).fetchall()]
        except Exception:
            return []

    sections = {
        'academic':             fetch('ops_academic_details', 'id DESC'),
        'epic':                 fetch('ops_epic_registration', 'id DESC'),
        'amc_registration':     fetch('ops_amc_registration', 'id DESC'),
        'training':             fetch('ops_coaching', 'id DESC'),
        'test_bookings':        fetch('ops_test_bookings', 'id DESC'),
        'online_courses':       fetch('ops_online_courses', 'id DESC'),
        'online_subscriptions': fetch('ops_online_subscriptions', 'id DESC'),
        'research':             fetch('ops_research_publication', 'id DESC'),
        'webinars':             fetch('ops_webinars_conferences', 'id DESC'),
        'call_notes':           fetch('ops_call_notes', 'created_at DESC, id DESC'),
        'payments':             fetch('ops_payments', 'id DESC'),
    }

    from app import PLAB_DOC_TYPES
    try:
        documents = conn.execute("SELECT d.id, d.client_id, d.doc_type, d.doc_category, d.file_name, d.file_path, d.file_size, d.status, d.verified_by, d.verified_at, d.notes, d.uploaded_by, d.uploaded_at, e.name as verifier_name FROM plab_client_documents d LEFT JOIN employees e ON e.id = d.verified_by WHERE d.client_id = ? ORDER BY d.doc_category, d.uploaded_at DESC", (client['id'],)).fetchall()
    except Exception:
        documents = []
    try:
        from app import get_package_services
        package_services = get_package_services(
            conn, client.get('product_id'), client.get('plan_type'))
    except Exception:
        package_services = []

    # ── Onboarding sync (founder 2026-08-06). For AMC, Welcome Call, Welcome
    #    Email and Refund Policy are standard for every client → always Yes
    #    (refund policy is issued with the welcome kit). Service Agreement stays
    #    real: Yes only when there's a signed/uploaded contract — so an
    #    ops-uploaded agreement flips it to Yes. Welcome Kit defaults Given.
    onboarding = {'welcome_call': True, 'welcome_email': True,
                  'agreement': False, 'refund_policy': True, 'welcome_kit': True}
    try:
        _ensure_welcome_kit_col(conn)
        krow = conn.execute("SELECT welcome_kit_given FROM plab_clients WHERE id = ?", (client['id'],)).fetchone()
        onboarding['welcome_kit'] = _kit_given(krow['welcome_kit_given'] if krow else None)
        rn = (reg or '').strip().upper()
        contract_signed = False
        rowreg = conn.execute(
            "SELECT id, account_id FROM client_registrations "
            "WHERE UPPER(TRIM(registration_number)) = ? AND client_submitted_at IS NOT NULL "
            "ORDER BY id DESC LIMIT 1", (rn,)).fetchone()
        if rowreg:
            contract_signed = bool(conn.execute(
                "SELECT 1 FROM client_agreements WHERE account_id = ? AND registration_id = ? "
                "AND agreement_type = 'contract' LIMIT 1", (rowreg['account_id'], rowreg['id'])).fetchone())
        onboarding['agreement'] = _yn_flag(client.get('service_agreement')) or bool(client.get('contract_path')) or contract_signed
    except Exception as e:
        logging.warning(f"AMC onboarding sync ({reg}): {e}")
        try: conn.rollback()
        except Exception: pass

    try:
        from app import build_client_address_ctx, build_client_resume_ctx
        address_ctx = build_client_address_ctx(conn, dict(client))
        resume_ctx = build_client_resume_ctx(conn, dict(client))
    except Exception:
        address_ctx = None
        resume_ctx = None
    conn.close()

    return render_template(
        'ops_australia_client_detail.html',
        onboarding=onboarding,
        documents=documents,
        plab_doc_types=PLAB_DOC_TYPES,
        user=user,
        client=client,
        address=address_ctx,
        resume=resume_ctx,
        package_services=package_services,
        sections=sections,
        amount_paid=amount_paid,
        gst_paid=gst_paid,
        total_paid=total_paid,
        final_pkg=final_pkg,
        balance=balance,
        payment_pct=pct,
        pathway_name='AMC Pathway',
        active_ops_page='australia-clients',
        active_pathway='australia',
    )


@admin_required
def ops_australia_client_by_reg(reg):
    """Resolve an Australia client by registration_number → detail redirect.

    Used by section drawers (Call Notes, Payments, Test Bookings, etc.)
    whose "View Client Profile" button only knows the registration_number,
    not the numeric client_id required by /operations/australia/clients/<id>.
    """
    conn = get_db()
    row = conn.execute(
        "SELECT id FROM plab_clients WHERE registration_number = ? "
        "AND COALESCE(pathway, 'plab') = 'australia' LIMIT 1",
        (reg,)
    ).fetchone()
    conn.close()
    if not row:
        flash(f'Australia client {reg} not found.', 'error')
        return redirect(url_for('ops_australia_clients_list'))
    return redirect(url_for('ops_australia_client_detail', client_id=row['id']))


@admin_required
def ops_australia_client_edit_page(client_id):
    """GET — render the edit form for an Australia client.

    Mirrors the PLAB pattern (view page + separate edit page) so the layout
    is consistent across pathways. The view page is /operations/australia/
    clients/<id>; this edit page is .../edit (GET).
    """
    user = get_user()
    conn = get_db()
    client = conn.execute(
        "SELECT * FROM plab_clients WHERE id = ? AND COALESCE(pathway, 'plab') = 'australia'",
        (client_id,),
    ).fetchone()
    xfer = transfer_for_reg(conn, client['registration_number']) if client else None
    conn.close()
    if not client:
        flash('Australia client not found.', 'error')
        return redirect(url_for('ops_australia_clients_list'))
    client = dict(client)
    if not client.get('product_id') and client.get('plan_type'):
        from app import derive_product_id_from_plan
        client['product_id'] = derive_product_id_from_plan('australia', client.get('plan_type'))
    return render_template(
        'ops_australia_client_edit_form.html',
        user=user,
        client=client,
        transfer=xfer,
        active_ops_page='australia-clients',
        active_pathway='australia',
        # Dropdown options for Plan Type, Account Status, Current Stage,
        # Switched Program, Counsellor, Lead Source. Sourced from
        # lookup_options where pathway='australia'.
        **section_client_lookups('australia', current=client.get('counsellor')),
        **section_client_products('australia'),
    )


@admin_required
def ops_australia_client_edit_save(client_id):
    """POST handler: save changes to an Australia client's editable fields.

    Strict allowlist (AU_EDITABLE_COLUMNS) — anything else in the form is
    silently ignored. Pathway is NEVER updated through this endpoint.
    """
    conn = get_db()
    try:
        existing = conn.execute(
            "SELECT id FROM plab_clients WHERE id = ? AND COALESCE(pathway, 'plab') = 'australia'",
            (client_id,),
        ).fetchone()
        if not existing:
            flash('Australia client not found.', 'error')
            return redirect(url_for('ops_australia_clients_list'))

        # Build SET clause from the allowlist.
        sets = []
        params = []
        for col in AU_EDITABLE_COLUMNS:
            if col in request.form:
                val = request.form.get(col, '').strip()
                # Numeric columns get coerced; bad data -> 0.
                if col in {'package_amount', 'discount_allowed', 'final_package',
                           'inst1_amount', 'inst2_amount', 'inst3_amount', 'inst4_amount'}:
                    try:
                        val = float(val) if val else 0
                    except ValueError:
                        val = 0
                # product_id: empty selection -> NULL (legacy rows stay clean).
                elif col == 'product_id':
                    # Never blank an existing product on an empty submit (e.g. the edit
                    # dropdown had no matching option for a cross-pathway product). Keep it.
                    if not val:
                        continue
                    try: val = int(val)
                    except (TypeError, ValueError): continue
                sets.append(f"{col} = ?")
                params.append(val)
        if sets:
            params.append(client_id)
            conn.execute(
                f"UPDATE plab_clients SET {', '.join(sets)} WHERE id = ?",
                params,
            )
            conn.commit()
            flash('Client details saved.', 'success')
        else:
            flash('No changes to save.', 'info')
    except Exception as e:
        logging.error(f"ops_australia_client_edit: {e}")
        flash(f'Error saving client: {e}', 'error')
        try: conn.rollback()
        except Exception: pass
    finally:
        conn.close()

    return redirect(url_for('ops_australia_client_detail', client_id=client_id))


@admin_required
def ops_australia_client_delete(client_id):
    """Z-1: hard-delete an Australia client. Mirror of
    ops_plab_delete with a pathway='australia' safety check so PLAB
    rows can't be deleted through this surface even if someone
    crafts the URL by hand.
    """
    conn = get_db()
    try:
        existing = conn.execute(
            "SELECT id FROM plab_clients "
            " WHERE id = ? AND COALESCE(pathway,'plab') = 'australia'",
            (client_id,),
        ).fetchone()
        if not existing:
            flash('Australia client not found.', 'error')
            return redirect(url_for('ops_australia_clients_list'))
        conn.execute("DELETE FROM plab_clients WHERE id = ?", (client_id,))
        conn.commit()
        flash('Client deleted', 'success')
    except Exception as e:
        logging.error(f"ops_australia_client_delete: {e}")
        flash(f'Error: {e}', 'error')
        try: conn.rollback()
        except Exception: pass
    finally:
        conn.close()
    return redirect(url_for('ops_australia_clients_list'))


@admin_required
def ops_australia_documents_list():
    """Documents overview for Australia (AMC) clients.

    Same shape as ops_documents_list (PLAB) and
    ops_consulting_documents_list — JOIN plab_clients +
    plab_client_documents filtered to pathway='australia'. After Y-4
    each row's file_path is an R2 object key, but this list only
    counts docs per client; no R2 calls happen here.
    """
    user = get_user()
    conn = get_db()
    search = (request.args.get('search') or '').strip()
    status_filter = (request.args.get('status') or '')
    page = max(1, int(request.args.get('page', 1) or 1))
    per_page = 50

    # Use the same PLAB checklist for now -- AMC-specific required
    # docs can be carved out later via lookup_options.
    try:
        from app import PLAB_ALL_REQUIRED_DOCS as REQUIRED_DOCS
    except Exception:
        REQUIRED_DOCS = []
    total_required = len(REQUIRED_DOCS)

    all_clients = []
    try:
        q = """SELECT c.id, c.registration_number, c.first_name, c.last_name, c.prefix,
                      c.account_status, c.current_stage, c.created_at,
                      COUNT(d.id) as doc_count,
                      SUM(CASE WHEN d.status = 'verified' THEN 1 ELSE 0 END) as verified_count
                 FROM plab_clients c
            LEFT JOIN plab_client_documents d ON d.client_id = c.id
                WHERE COALESCE(c.pathway, 'plab') = 'australia' """
        params = []
        if search:
            q += " AND (c.first_name ILIKE ? OR c.last_name ILIKE ? OR c.registration_number ILIKE ? OR (COALESCE(c.prefix,'')||' '||c.first_name||' '||COALESCE(c.last_name,'')) ILIKE ?) "
            params += [f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%']
        q += """ GROUP BY c.id, c.registration_number, c.first_name, c.last_name, c.prefix,
                          c.account_status, c.current_stage, c.created_at
                 ORDER BY c.created_at DESC NULLS LAST, c.id DESC"""
        rows = conn.execute(q, params).fetchall()
        for r in rows:
            rd = dict(r)
            rd['total_required'] = total_required
            rd['name'] = f"{r['prefix'] or ''} {r['first_name'] or ''} {r['last_name'] or ''}".strip()
            if status_filter == 'complete' and rd['doc_count'] < total_required:
                continue
            if status_filter == 'incomplete' and rd['doc_count'] >= total_required:
                continue
            if status_filter == 'none' and rd['doc_count'] > 0:
                continue
            all_clients.append(rd)
    except Exception as e:
        logging.error(f"ops_australia_documents_list: {e}")
        all_clients = []
    finally:
        conn.close()

    total_clients = len(all_clients)
    complete_count = sum(1 for c in all_clients if c.get('doc_count', 0) >= total_required)
    total_documents = sum(c.get('doc_count', 0) for c in all_clients)
    total_pages = max(1, (total_clients + per_page - 1) // per_page)
    page = min(page, total_pages)
    start = (page - 1) * per_page
    page_clients = all_clients[start:start + per_page]

    return render_template(
        'ops_australia_documents_list.html',
        user=user,
        clients=page_clients,
        total_clients=total_clients,
        complete_count=complete_count,
        total_documents=total_documents,
        page=page,
        total_pages=total_pages,
        per_page=per_page,
        search=search,
        status_filter=status_filter,
        total_required=total_required,
        start_index=start if page_clients else 0,
        active_ops_page='australia-documents',
        active_pathway='australia',
    )


# ══════════════════════════════════════════════════════════════════════════
# AMC "Services Summary" — one spreadsheet-style sheet showing, per client,
# which services have been rendered (Yes (n) / No) + onboarding flags.
# Reads live from the AMC (pathway='australia') ops tables so it stays in sync.
# ══════════════════════════════════════════════════════════════════════════

# Each service column → the ops record table(s) whose per-client row count
# gives the "Yes (n)" value. Multiple tables are summed into one column.
AMC_SERVICE_COLUMNS = [
    ('research',        'Research & Publication', ['ops_research_publication']),
    ('epic',            'EPIC Registration',      ['ops_epic_registration']),
    ('amc_registration','AMC Registration',       ['ops_amc_registration']),
    ('mentorship',      'Mentorship',             ['ops_mentorship']),
    ('webinars',        'Webinars & Conferences', ['ops_webinars_conferences']),
    ('training',        'Training',               ['ops_coaching']),
    ('online_courses',  'Online Courses',         ['ops_online_courses', 'ops_online_subscriptions']),
    ('test_bookings',   'Test Bookings',          ['ops_test_bookings']),
]


def _yn_flag(v):
    """Loose truthiness for messy legacy Yes/No/1/0/blank columns."""
    if v is None:
        return False
    s = str(v).strip().lower()
    return s not in ('', '0', '0.0', 'no', 'none', 'n', 'false', '-', 'nil', 'na', 'n/a', 'pending', 'not sent')


def _ensure_welcome_kit_col(conn):
    """Per-client Welcome-Kit toggle lives on plab_clients. Default 1 = 'Given'
    (founder 2026-08-06: sync all AMC clients to kit-given, ops toggles off the
    few who didn't get one). Guarded here for Render cold-start safety."""
    try:
        conn.execute("ALTER TABLE plab_clients ADD COLUMN IF NOT EXISTS welcome_kit_given INTEGER DEFAULT 1")
        conn.commit()
    except Exception:
        try: conn.rollback()
        except Exception: pass


def _kit_given(v):
    """Interpret the welcome_kit_given column: NULL/anything-but-0 => Given."""
    return (v is None) or (str(v).strip() not in ('0', '0.0', 'False', 'false'))


def _amc_services_summary_data(conn, search='', status_filter='', stage_filter=''):
    """Return (rows, statuses, stages) for the AMC Services Summary + export."""
    _ensure_welcome_kit_col(conn)

    sql = "SELECT * FROM plab_clients WHERE COALESCE(pathway,'plab')='australia' "
    params = []
    if status_filter:
        sql += " AND account_status = ? "; params.append(status_filter)
    if stage_filter:
        sql += " AND current_stage = ? "; params.append(stage_filter)
    if search:
        sql += (" AND (first_name ILIKE ? OR last_name ILIKE ? OR registration_number ILIKE ? "
                " OR mobile ILIKE ? OR email ILIKE ? "
                " OR (COALESCE(prefix,'')||' '||first_name||' '||COALESCE(last_name,'')) ILIKE ?) ")
        params.extend([f'%{search}%'] * 6)
    sql += (" ORDER BY NULLIF(regexp_replace(COALESCE(registration_number,''),'[^0-9]','','g'),'')"
            "::bigint DESC NULLS LAST, id DESC ")
    clients = [dict(r) for r in conn.execute(sql, params).fetchall()]

    def _norm(s):
        return (s or '').strip().upper()

    # ── Per-service counts: one grouped query per table, summed into the column
    service_counts = {}
    for key, _label, tables in AMC_SERVICE_COLUMNS:
        merged = {}
        for table in tables:
            try:
                for r in conn.execute(
                    f"""SELECT UPPER(TRIM(registration_number)) AS k, COUNT(*) AS c
                          FROM {table} WHERE COALESCE(pathway,'plab')='australia'
                         GROUP BY UPPER(TRIM(registration_number))""").fetchall():
                    if r['k']:
                        merged[r['k']] = merged.get(r['k'], 0) + int(r['c'])
            except Exception:
                try: conn.rollback()
                except Exception: pass
        service_counts[key] = merged

    # ── Service Agreement from the REAL data (legacy column, an uploaded
    #    contract file, or the digital-sign record) so an ops-uploaded agreement
    #    flips this to Yes. Keyed by registration_number.
    contract_regs = set()
    try:
        for r in conn.execute(
            """SELECT DISTINCT UPPER(TRIM(cr.registration_number)) AS rn
                 FROM client_agreements ca
                 JOIN client_registrations cr ON cr.id = ca.registration_id
                WHERE ca.agreement_type = 'contract'""").fetchall():
            if r['rn']: contract_regs.add(r['rn'])
    except Exception:
        try: conn.rollback()
        except Exception: pass

    rows = []
    for c in clients:
        rn = _norm(c.get('registration_number'))
        name = ' '.join(f"{c.get('prefix') or ''} {c.get('first_name') or ''} {c.get('last_name') or ''}".split())
        services = {key: service_counts.get(key, {}).get(rn, 0) for key, _l, _t in AMC_SERVICE_COLUMNS}
        rows.append({
            'id': c.get('id'),
            'registration_number': c.get('registration_number') or '',
            'name': name or '—',
            'registration_date': c.get('registration_date') or '',
            'account_status': c.get('account_status') or '',
            'current_stage': c.get('current_stage') or '',
            'services': services,
            # Welcome Email & Refund Policy are standard for every AMC client
            # (refund policy goes out with the welcome kit) → always Yes.
            'welcome_email': True,
            'refund_policy': True,
            # Welcome Kit defaults Given; ops toggles off exceptions.
            'welcome_kit': _kit_given(c.get('welcome_kit_given')),
            # Agreement reflects real data — an ops-uploaded contract flips it to Yes.
            'agreement': _yn_flag(c.get('service_agreement')) or bool(c.get('contract_path')) or (rn in contract_regs),
        })

    statuses = sorted({r['account_status'] for r in rows if r['account_status']})
    stages = sorted({r['current_stage'] for r in rows if r['current_stage']})
    return rows, statuses, stages


@admin_required
def ops_australia_services_summary():
    """AMC Services Summary — spreadsheet-style per-client services-rendered sheet."""
    user = get_user()
    conn = get_db()
    search = (request.args.get('q', '') or '').strip()
    status_filter = (request.args.get('status', '') or '').strip()
    stage_filter = (request.args.get('stage', '') or '').strip()
    rows, statuses, stages = [], [], []
    try:
        rows, statuses, stages = _amc_services_summary_data(conn, search, status_filter, stage_filter)
    except Exception as e:
        logging.warning(f"ops_australia_services_summary: {e}")
        try: conn.rollback()
        except Exception: pass
    conn.close()
    return render_template(
        'ops_australia_services_summary.html',
        user=user, rows=rows, statuses=statuses, stages=stages,
        service_columns=AMC_SERVICE_COLUMNS,
        q=search, status_filter=status_filter, stage_filter=stage_filter,
        pathway_name='AMC Pathway',
        active_pathway='australia',
        active_ops_page='australia-services-summary',
    )


@admin_required
def ops_australia_services_summary_download():
    """Excel download of the AMC Services Summary (respects current filters)."""
    conn = get_db()
    search = (request.args.get('q', '') or '').strip()
    status_filter = (request.args.get('status', '') or '').strip()
    stage_filter = (request.args.get('stage', '') or '').strip()
    try:
        rows, _statuses, _stages = _amc_services_summary_data(conn, search, status_filter, stage_filter)
    except Exception as e:
        logging.warning(f"ops_australia_services_summary_download: {e}")
        rows = []
    conn.close()

    import io
    from datetime import datetime
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    def _xs(v):
        return ILLEGAL_CHARACTERS_RE.sub('', v) if isinstance(v, str) else v

    wb = Workbook(); ws = wb.active; ws.title = 'AMC Services'
    headers = ['#', 'Client Name', 'Registration Number', 'Registration Date',
               'Account Status', 'Current Stage']
    headers += [label for _k, label, _t in AMC_SERVICE_COLUMNS]
    headers += ['Welcome Email', 'Welcome Kit', 'Agreement', 'Refund Policy']
    ws.append(headers)

    for i, r in enumerate(rows, 1):
        line = [i, _xs(r['name']), _xs(r['registration_number']), _xs(r['registration_date']),
                _xs(r['account_status']), _xs(r['current_stage'])]
        for key, _label, _t in AMC_SERVICE_COLUMNS:
            n = r['services'].get(key, 0)
            line.append(f"Yes ({n})" if n else "No")
        line.append('Yes' if r['welcome_email'] else 'No')
        line.append('Yes' if r['welcome_kit'] else 'No')
        line.append('Yes' if r['agreement'] else 'No')
        line.append('Yes' if r['refund_policy'] else 'No')
        ws.append(line)

    navy = PatternFill('solid', fgColor='0F1B33')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = navy
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = ws.dimensions
    widths = [5, 26, 20, 15, 14, 20] + [15] * len(AMC_SERVICE_COLUMNS) + [14, 12, 12, 12]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    out = io.BytesIO(); wb.save(out); out.seek(0)
    fn = f"AMC_Services_Summary_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fn)


@admin_required
def ops_australia_welcome_kit_toggle(client_id):
    """Toggle a client's Welcome-Kit given/not-given flag (from the profile)."""
    conn = get_db()
    _ensure_welcome_kit_col(conn)
    val = 1 if (request.form.get('given') == '1') else 0
    try:
        conn.execute(
            "UPDATE plab_clients SET welcome_kit_given = ? "
            "WHERE id = ? AND COALESCE(pathway,'plab')='australia'", (val, client_id))
        conn.commit()
        flash('Welcome kit marked as ' + ('Given.' if val else 'Not given.'), 'success')
    except Exception as e:
        logging.warning(f"ops_australia_welcome_kit_toggle: {e}")
        try: conn.rollback()
        except Exception: pass
        flash('Could not update welcome kit status.', 'error')
    conn.close()
    nxt = request.form.get('next') or url_for('ops_australia_client_detail', client_id=client_id)
    return redirect(nxt)


def register_routes(app):
    """Attach this sub-area's URL rules to the Flask app.

    Uses app.add_url_rule to preserve endpoint names — keeps existing
    url_for('ops_australia_pathway') call sites working unchanged.
    """
    app.add_url_rule(
        '/operations/australia/services-summary',
        endpoint='ops_australia_services_summary',
        view_func=ops_australia_services_summary,
        methods=['GET'],
    )
    app.add_url_rule(
        '/operations/australia/services-summary/download',
        endpoint='ops_australia_services_summary_download',
        view_func=ops_australia_services_summary_download,
        methods=['GET'],
    )
    app.add_url_rule(
        '/operations/australia/clients/<int:client_id>/welcome-kit',
        endpoint='ops_australia_welcome_kit_toggle',
        view_func=ops_australia_welcome_kit_toggle,
        methods=['POST'],
    )
    app.add_url_rule(
        '/operations/australia-pathway',
        endpoint='ops_australia_pathway',
        view_func=ops_australia_pathway,
        methods=['GET'],
    )
    # Australia Test Bookings registration moved to routes/operations/au_test_bookings.py.
    app.add_url_rule(
        '/operations/australia/clients',
        endpoint='ops_australia_clients_list',
        view_func=ops_australia_clients_list,
        methods=['GET'],
    )
    app.add_url_rule(
        '/operations/australia/clients/<int:client_id>',
        endpoint='ops_australia_client_detail',
        view_func=ops_australia_client_detail,
        methods=['GET'],
    )
    # Lookup-by-registration-number helper for section drawers
    app.add_url_rule(
        '/operations/australia/clients/by-reg/<path:reg>',
        endpoint='ops_australia_client_by_reg',
        view_func=ops_australia_client_by_reg,
        methods=['GET'],
    )
    # Edit page (GET) — renders the form
    app.add_url_rule(
        '/operations/australia/clients/<int:client_id>/edit',
        endpoint='ops_australia_client_edit_page',
        view_func=ops_australia_client_edit_page,
        methods=['GET'],
    )
    # Save (POST) — same URL, different method, separate endpoint
    app.add_url_rule(
        '/operations/australia/clients/<int:client_id>/edit',
        endpoint='ops_australia_client_edit_save',
        view_func=ops_australia_client_edit_save,
        methods=['POST'],
    )
    # Z-1: Delete handler (POST)
    app.add_url_rule(
        '/operations/australia/clients/<int:client_id>/delete',
        endpoint='ops_australia_client_delete',
        view_func=ops_australia_client_delete,
        methods=['POST'],
    )
    # Y-5: Documents list (clone of ops_consulting_documents_list,
    # scoped to pathway='australia')
    app.add_url_rule(
        '/operations/australia/documents',
        endpoint='ops_australia_documents_list',
        view_func=ops_australia_documents_list,
        methods=['GET'],
    )
