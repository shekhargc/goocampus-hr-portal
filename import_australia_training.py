"""
One-time import: Australia training records from Excel into ops_coaching.

Each Excel row goes into ops_coaching with pathway='australia'.

Linking strategy:
  Unlike the test-bookings sheet, the "Enter Candidate Name" cell on this
  Excel rarely contains a registration number — it is usually just the
  doctor's name, e.g. " Dr. Arjun C.s.  ". We therefore:
    1) Try to parse out a GCAUSIP/... reg number if it happens to be in
       the cell (cheap, exact, prefer this).
    2) Otherwise fuzzy-match the cleaned-up name against plab_clients
       WHERE pathway='australia'. Match strategy is LOWER(TRIM(name))
       comparison against "first last" (and "last first" fallback) so
       both name orderings work.

  When neither yields a match we increment skipped, log a warning, and
  carry on so a few bad rows never block the whole import.

Idempotent: keyed on (registration_number, course_type, start_date). Bumping
IMPORT_VERSION re-runs the upsert. Marker key: au_training_import.

Hooked into app boot from app.py the same way as
import_australia_test_bookings.py.
"""

import logging
import os
import re
from datetime import datetime, date


EXCEL_FILENAME = 'All_Australia_Trainings.xlsx'
IMPORT_VERSION = 'au_training_v1_initial_355'


H = {
    'auto_no':       'Auto Number',
    'cand_name':     'Enter Candidate Name',
    'cand_email':    'Candidate Email',
    'training_prog': 'Training Program',
    'course_type':   'Course Type',
    'train_method':  'Training Method',
    'vendor':        'Training Vendor Name',
    'booked_by':     'Booked By',
    'batch_year':    'Batch Year',
    'batch_month':   'Batch (Month)',
    'start_date':    'Start Date',
    'end_date':      'End Date',
    'train_status':  'Training Status',
}


_REG_RE = re.compile(r'(GCAUSIP/\S+)', re.IGNORECASE)
# Strip "Dr." style prefixes and trailing punctuation when we need the
# bare name to fuzzy-match. Keep it permissive.
_PREFIX_RE = re.compile(r'^\s*(dr\.?|mr\.?|mrs\.?|ms\.?|miss|prof\.?)\s+', re.IGNORECASE)


def _extract_reg_number(candidate_cell):
    """Pull the GCAUSIP/... reg number out of the candidate-name cell.

    Most Training rows do NOT contain one — that's fine, caller will
    fall back to a fuzzy name match.
    """
    if not candidate_cell:
        return ''
    s = str(candidate_cell)
    m = _REG_RE.search(s)
    if not m:
        return ''
    raw = m.group(1).rstrip('.,;:-/ \t')
    return raw.upper()


def _clean_name(candidate_cell):
    """Normalize the candidate-name cell to a comparable form.

    Strips honorifics, trailing punctuation, collapses whitespace, and
    lowercases. Returns '' if there is nothing usable.
    """
    if not candidate_cell:
        return ''
    s = str(candidate_cell).strip()
    # Drop a trailing "-GCAUSIP/..." chunk if present so it doesn't
    # pollute the name comparison.
    s = _REG_RE.sub('', s)
    s = s.replace('-', ' ').strip(' .,-')
    s = _PREFIX_RE.sub('', s)
    s = re.sub(r'\s+', ' ', s).strip().lower()
    return s


def _fuzzy_match_reg(conn, name_clean, cache):
    """Look up a registration_number by cleaned candidate name.

    `cache` is a dict {clean_name: reg_number} populated lazily from
    plab_clients WHERE pathway='australia' so we don't run a query per
    row. Returns '' on no match.
    """
    if not name_clean:
        return ''
    if cache:  # already built
        return cache.get(name_clean, '')

    # Build the lookup once.
    rows = conn.execute(
        """SELECT registration_number, prefix, first_name, last_name
             FROM plab_clients
            WHERE pathway = 'australia'"""
    ).fetchall()
    for r in rows:
        first = (r['first_name'] or '').strip()
        last = (r['last_name'] or '').strip()
        if not (first or last):
            continue
        # Forward: "first last"
        fwd = re.sub(r'\s+', ' ', f"{first} {last}").strip().lower()
        # Reverse: "last first" (sometimes data has it flipped)
        rev = re.sub(r'\s+', ' ', f"{last} {first}").strip().lower()
        if fwd and fwd not in cache:
            cache[fwd] = r['registration_number']
        if rev and rev not in cache:
            # Don't clobber a forward match
            cache.setdefault(rev, r['registration_number'])
        # Also index by first-name-only and last-name-only as a last
        # resort (low-precision but the data is messy enough that this
        # rescues a useful chunk of rows).
        if first:
            cache.setdefault(first.lower(), r['registration_number'])
        if last:
            cache.setdefault(last.lower(), r['registration_number'])

    return cache.get(name_clean, '')


def _ss(v):
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _sd(v):
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def run_import_australia_training_once(get_db_fn):
    """Import Australia training records into ops_coaching (pathway='australia')."""
    import openpyxl

    excel_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        EXCEL_FILENAME,
    )
    if not os.path.exists(excel_path):
        logging.info("Australia training import: no Excel found, skipping")
        return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

    conn = get_db_fn()
    try:
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS _import_markers (key TEXT PRIMARY KEY, value TEXT)")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        marker = conn.execute(
            "SELECT value FROM _import_markers WHERE key = 'au_training_import'"
        ).fetchone()
        if marker and marker['value'] == IMPORT_VERSION:
            logging.info(f"Australia training import: already at {IMPORT_VERSION}, skipping")
            return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

        logging.info(f"Australia training import: starting {IMPORT_VERSION}...")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h is not None}

        def cell(row, key):
            i = idx.get(H[key])
            return row[i] if i is not None and i < len(row) else None

        inserted = 0
        updated = 0
        skipped = 0
        errors = 0
        name_cache = {}

        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row_cells):
                continue

            cand_cell = cell(row_cells, 'cand_name')
            reg_num = _extract_reg_number(cand_cell)
            if not reg_num:
                # Fall back to fuzzy name match.
                reg_num = _fuzzy_match_reg(conn, _clean_name(cand_cell), name_cache)

            if not reg_num:
                skipped += 1
                logging.warning(
                    f"Australia training row skipped (no reg, no name match): "
                    f"{(_ss(cand_cell) or '?')[:80]}"
                )
                continue

            course_type = _ss(cell(row_cells, 'course_type'))
            start_date = _sd(cell(row_cells, 'start_date'))
            training_prog = _ss(cell(row_cells, 'training_prog'))

            # Notes capture the email + the original training program label,
            # since ops_coaching's columns don't have direct slots for them.
            note_bits = []
            cand_email = _ss(cell(row_cells, 'cand_email'))
            if training_prog:
                note_bits.append(f"program: {training_prog}")
            if cand_email:
                note_bits.append(f"email: {cand_email}")
            notes = ' | '.join(note_bits)

            data = {
                'registration_number': reg_num,
                'course_type':         course_type,
                'coaching_method':     _ss(cell(row_cells, 'train_method')),
                'coaching_status':     _ss(cell(row_cells, 'train_status')),
                'batch_month':         _ss(cell(row_cells, 'batch_month')),
                'batch_year':          _ss(cell(row_cells, 'batch_year')),
                'start_date':          start_date,
                'end_date':            _sd(cell(row_cells, 'end_date')),
                'vendor_provider':     _ss(cell(row_cells, 'vendor')),
                'other_vendor':        _ss(cell(row_cells, 'booked_by')),
                'english_training':    training_prog,
                'pathway':             'australia',
            }

            try:
                existing = conn.execute(
                    "SELECT id FROM ops_coaching "
                    "WHERE registration_number = ? AND COALESCE(course_type,'') = ? "
                    "  AND COALESCE(start_date,'') = ? AND pathway = 'australia'",
                    (reg_num, course_type, start_date),
                ).fetchone()
                if existing:
                    sets = ', '.join(f"{k} = ?" for k in data.keys())
                    conn.execute(
                        f"UPDATE ops_coaching SET {sets} WHERE id = ?",
                        list(data.values()) + [existing['id']],
                    )
                    updated += 1
                else:
                    cols = ', '.join(data.keys())
                    placeholders = ', '.join(['?'] * len(data))
                    conn.execute(
                        f"INSERT INTO ops_coaching ({cols}) VALUES ({placeholders})",
                        list(data.values()),
                    )
                    inserted += 1
                conn.commit()
            except Exception as e:
                errors += 1
                logging.warning(
                    f"Australia training row error ({reg_num} / {course_type}): {e}"
                )
                try:
                    conn.rollback()
                except Exception:
                    pass

        # Marker
        try:
            conn.execute(
                "INSERT INTO _import_markers (key, value) VALUES ('au_training_import', ?) "
                "ON CONFLICT (key) DO UPDATE SET value = excluded.value",
                (IMPORT_VERSION,),
            )
            conn.commit()
        except Exception:
            try:
                conn.execute("DELETE FROM _import_markers WHERE key = 'au_training_import'")
                conn.execute(
                    "INSERT INTO _import_markers (key, value) VALUES ('au_training_import', ?)",
                    (IMPORT_VERSION,),
                )
                conn.commit()
            except Exception as e:
                logging.warning(f"Australia training: could not write marker: {e}")

        logging.info(
            f"Australia training {IMPORT_VERSION} complete — "
            f"inserted={inserted} updated={updated} skipped={skipped} errors={errors}"
        )
        wb.close()
        return {'inserted': inserted, 'updated': updated, 'skipped': skipped, 'errors': errors}
    finally:
        try:
            conn.close()
        except Exception:
            pass
