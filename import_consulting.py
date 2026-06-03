"""
One-time imports for Standard Consulting data (S-3).

Loads the 7 user-supplied Excel files from imports/ into the appropriate
tables on first boot after this commit. Idempotent via per-file
_import_markers entries. Each importer matches the same shape as the
existing Australia importers (import_australia_*.py):
  * open workbook
  * read headers row 1, build index map (when headers present)
  * iterate data rows, build typed record dict
  * INSERT or UPSERT per row, scoped to pathway='consulting'

Files imported (all under imports/):
  GC CSS Registration Report.xlsx      -> plab_clients          (56 rows)
  GC Consulting Academic Details ...    -> ops_academic_details   (56)
  Client Call Notes Report (1).xlsx     -> ops_call_notes         (199)
  All Epic Verifications.xlsx           -> ops_epic_registration  (15)
  All Clients Payments.xlsx             -> ops_payments           (38)
  All Amc Registrations.xlsx            -> ops_amc_registration   (14)
  All Mentorship Sessions.xlsx          -> ops_mentorship         (5)

Most files reference clients by registration_number (format
GCCSS/YY-YY/NN) which we extract from a candidate-name string where
no dedicated column exists. The clients importer runs FIRST so the
other importers can FK-reference the resulting plab_clients rows.

Invoked from app.py at boot via run_all_consulting_imports_once().
"""

import logging
import os
import re
from datetime import datetime, date


IMPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'imports')


# ── Filenames ──────────────────────────────────────────────────────────
FILE_CLIENTS    = 'GC CSS Registration Report.xlsx'
FILE_ACADEMIC   = 'GC Consulting Academic Details Report.xlsx'
FILE_CALL_NOTES = 'Client Call Notes Report (1).xlsx'
FILE_EPIC       = 'All Epic Verifications.xlsx'
FILE_PAYMENTS   = 'All Clients Payments.xlsx'
FILE_AMC        = 'All Amc Registrations.xlsx'
FILE_MENTORSHIP = 'All Mentorship Sessions.xlsx'


# ── Marker keys (bumped together when re-import is needed) ─────────────
MARK_CLIENTS    = ('cs_clients_seeded',     'v2_fixed_columns')
# Dependent sections all bumped to v2 too -- on the v1 run nothing was
# imported (no clients existed yet, so the FK gate skipped every row).
# v2 re-runs them after the clients importer above has done its work.
MARK_ACADEMIC   = ('cs_academic_seeded',    'v2_after_clients')
MARK_CALL_NOTES = ('cs_call_notes_seeded',  'v2_after_clients')
MARK_EPIC       = ('cs_epic_seeded',        'v2_after_clients')
MARK_PAYMENTS   = ('cs_payments_seeded',    'v4_dedupe_clean')
MARK_AMC        = ('cs_amc_seeded',         'v3_table_created')
MARK_MENTORSHIP = ('cs_mentorship_seeded',  'v2_after_clients')


# ── Shared cell-value helpers ──────────────────────────────────────────
def _ss(v):
    """Safe string: strip, drop None, collapse stray int-floats to int."""
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _sf(v):
    """Safe float -- 0 on parse failure / None."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '').replace('₹', '').replace('$', '')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _sd(v):
    """Safe date: ISO YYYY-MM-DD or empty."""
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip().split(' ')[0]


_PREFIX_TOKENS = {'Dr', 'Dr.', 'Mr', 'Mr.', 'Mrs', 'Mrs.', 'Ms', 'Ms.', 'Prof', 'Prof.'}


def _split_name(cname):
    """Split 'Dr. First Last' into (prefix, first, last)."""
    cname = (cname or '').strip()
    if not cname:
        return 'Dr.', '', ''
    parts = cname.split()
    prefix = 'Dr.'
    if parts and parts[0].rstrip('.') in {'Dr', 'Mr', 'Mrs', 'Ms', 'Prof'}:
        prefix = parts[0] if parts[0].endswith('.') else parts[0] + '.'
        parts = parts[1:]
    if not parts:
        return prefix, '', ''
    if len(parts) == 1:
        return prefix, parts[0], ''
    return prefix, parts[0], ' '.join(parts[1:])


_REG_RE = re.compile(r'GCCSS/\d{1,4}(?:-\d{2,4})?/\d{1,4}', re.IGNORECASE)


def _extract_reg(value):
    """Extract a GCCSS/YY-YY/NN registration number from a string.
    Returns the matched substring upper-cased, or None.

    Handles inputs like:
        ' Dr. Sheethal mariat  Bobby  -GCCSS/25-26/042'
        ' Dr. Lithika Rajan - GCCSS/26-27/08'
        'GCCSS/24-25/06'
    """
    if not value:
        return None
    m = _REG_RE.search(str(value))
    return m.group(0).upper() if m else None


# ── Marker helpers ────────────────────────────────────────────────────
def _ensure_marker_table(conn):
    try:
        conn.execute(
            "CREATE TABLE IF NOT EXISTS _import_markers "
            "(key TEXT PRIMARY KEY, value TEXT)"
        )
        conn.commit()
    except Exception:
        try: conn.rollback()
        except Exception: pass


def _has_marker(conn, key, value):
    try:
        row = conn.execute(
            "SELECT value FROM _import_markers WHERE key = ?", (key,)
        ).fetchone()
        return bool(row) and row['value'] == value
    except Exception:
        return False


def _set_marker(conn, key, value):
    try:
        conn.execute(
            "INSERT INTO _import_markers (key, value) VALUES (?, ?) "
            "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            (key, value),
        )
        conn.commit()
    except Exception:
        try: conn.rollback()
        except Exception: pass
        # Fall back to DELETE + INSERT for SQLite quirks.
        try:
            conn.execute("DELETE FROM _import_markers WHERE key = ?", (key,))
            conn.execute(
                "INSERT INTO _import_markers (key, value) VALUES (?, ?)",
                (key, value),
            )
            conn.commit()
        except Exception:
            try: conn.rollback()
            except Exception: pass


def _open_workbook(filename):
    fpath = os.path.join(IMPORTS_DIR, filename)
    if not os.path.isfile(fpath):
        return None
    try:
        import openpyxl
        return openpyxl.load_workbook(
            fpath, read_only=True, data_only=True)
    except Exception as e:
        logging.warning(f"_open_workbook({filename}): {e}")
        return None


# ── 1) CLIENTS ────────────────────────────────────────────────────────
H_CLIENTS = {
    'customer_id':      'Customer ID',
    'reg_number':       'Registration Number',
    'candidate_name':   'Candidate Name',
    'mobile':           'Mobile Number',
    'email':            'Candidate Email',
    'reg_date':         'Registration Date (Payment Date)',
    'plan_type':        'Plan Type',
    'package':          'Package (Mention Actual Package)',
    'final_package':    'Final Package',
    'discount':         'Discount Allowed (Discount Offer + Stage Discount)',
    'account_status':   'Account Status',
    'counsellor':       'Counsellor Name',
    'current_stage':    'Stage (Current Status)',
    'dob':              'D.O.B',
    'city':             'CITY',
    'joined_stage':     'Joined Stage',
    'lead_source':      'Lead Source',
    'state':            'STATE',
    'whatsapp':         'Whats App Number',
    'registration_id':  'Registration ID',
    'instagram':        'Instgram Account Name',
    'facebook':         'Facebook Account Name',
    'linkedin':         'LinkedIn Account Name',
    'father_name':      'Fathers Name',
    'father_phone':     'Fathers Mobile Number',
    'mother_name':      'Mothers Name',
    'mother_phone':     'Mothers Mobile Number',
    'parents_email':    'Parents Email ID',
    'counsellor_number':'Counsellor Number',
    'counsellor_email': 'Counsellor Email',
    'uk_referral':      'UK Client Referral',
    'portfolio_referral':'Portfolio Client Referral',
    'australia_referral':'Australia Client Referral:',
    'discount_notes':   'Discount Notes',
    'inst1_amt':        '1st Installment',
    'inst1_date':       '1st Installment Date',
    'inst1_note':       '1st Installment Note',
    'inst2_amt':        '2nd Installment',
    'inst2_date':       '2nd Installment Date',
    'inst2_note':       '2nd Installment Note',
    'notes':            'Notes',
}


def _plan_to_product_id(plan_type, prod_map):
    """Map free-text plan_type to one of the 4 consulting product IDs.
    prod_map is {name_lower: id} for pathway='consulting' products.

    Matching logic (keyword-based, case-insensitive):
        Australia / AUS / AMC -> 'amc consulting'
        UAE                   -> 'uae consulting'
        USA / US              -> 'usa consulting'
        UK / Britain          -> 'uk consulting'
    Unknown / empty -> None.
    """
    p = (plan_type or '').lower()
    if not p:
        return None
    if 'australia' in p or 'amc' in p or 'aus' in p.split():
        return prod_map.get('amc consulting')
    if 'uae' in p:
        return prod_map.get('uae consulting')
    if 'usa' in p or ' us ' in f' {p} ':
        return prod_map.get('usa consulting')
    if 'uk' in p or 'britain' in p:
        return prod_map.get('uk consulting')
    return None


def run_import_consulting_clients_once(get_db_fn):
    if _open_workbook(FILE_CLIENTS) is None:
        logging.info(f"Consulting clients import: no {FILE_CLIENTS} found, skipping")
        return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}
    conn = get_db_fn()
    try:
        _ensure_marker_table(conn)
        if _has_marker(conn, *MARK_CLIENTS):
            logging.info(
                f"Consulting clients import: already at {MARK_CLIENTS[1]}, skipping")
            return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

        logging.info(f"Consulting clients import: starting {MARK_CLIENTS[1]}...")
        wb = _open_workbook(FILE_CLIENTS)
        ws = wb.active

        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h is not None}

        def cell(row, key):
            i = idx.get(H_CLIENTS.get(key, key))
            return row[i] if i is not None and i < len(row) else None

        # Build product-name -> id map
        prod_map = {}
        try:
            for r in conn.execute(
                "SELECT id, name FROM products_services "
                " WHERE COALESCE(pathway,'') = 'consulting'"
            ).fetchall():
                prod_map[(r['name'] or '').lower()] = r['id']
        except Exception as e:
            logging.warning(f"clients: prod_map fetch: {e}")

        inserted = updated = skipped = errors = 0
        seen = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row):
                continue
            reg = _ss(cell(row, 'reg_number'))
            if not reg:
                # Fallback: try extracting from candidate_name
                reg = _extract_reg(_ss(cell(row, 'candidate_name'))) or ''
            if not reg:
                skipped += 1
                continue
            if reg in seen:
                skipped += 1
                continue
            seen.add(reg)

            prefix, fname, lname = _split_name(_ss(cell(row, 'candidate_name')))
            plan_type = _ss(cell(row, 'plan_type'))
            product_id = _plan_to_product_id(plan_type, prod_map)

            data = {
                'registration_number': reg,
                'prefix':            prefix,
                'first_name':        fname or '(missing)',
                'last_name':         lname,
                'mobile':            _ss(cell(row, 'mobile')),
                'email':             _ss(cell(row, 'email')),
                'dob':               _sd(cell(row, 'dob')),
                'city':              _ss(cell(row, 'city')),
                'state':             _ss(cell(row, 'state')),
                'registration_date': _sd(cell(row, 'reg_date')),
                'plan_type':         plan_type,
                'package_amount':    _sf(cell(row, 'package')),
                'final_package':     _sf(cell(row, 'final_package')),
                'discount_allowed':  _sf(cell(row, 'discount')),
                'account_status':    _ss(cell(row, 'account_status')) or 'In Process',
                'counsellor':        _ss(cell(row, 'counsellor')),
                'current_stage':     _ss(cell(row, 'current_stage')),
                'joined_stage':      _ss(cell(row, 'joined_stage')),
                'lead_source':       _ss(cell(row, 'lead_source')),
                'instagram':         _ss(cell(row, 'instagram')),
                'facebook':          _ss(cell(row, 'facebook')),
                'linkedin':          _ss(cell(row, 'linkedin')),
                'father_name':       _ss(cell(row, 'father_name')),
                'father_phone':      _ss(cell(row, 'father_phone')),
                'mother_name':       _ss(cell(row, 'mother_name')),
                'mother_phone':      _ss(cell(row, 'mother_phone')),
                'parents_email':     _ss(cell(row, 'parents_email')),
                'counsellor_number': _ss(cell(row, 'counsellor_number')),
                'counsellor_email':  _ss(cell(row, 'counsellor_email')),
                'inst1_amount':      _sf(cell(row, 'inst1_amt')),
                'inst1_date':        _sd(cell(row, 'inst1_date')),
                'inst1_note':        _ss(cell(row, 'inst1_note')),
                'inst2_amount':      _sf(cell(row, 'inst2_amt')),
                'inst2_date':        _sd(cell(row, 'inst2_date')),
                'inst2_note':        _ss(cell(row, 'inst2_note')),
                'additional_notes':  _ss(cell(row, 'notes')),
                'pathway':           'consulting',
                'product_id':        product_id,
            }

            try:
                exists = conn.execute(
                    "SELECT id FROM plab_clients WHERE registration_number = ?",
                    (reg,),
                ).fetchone()
                if exists:
                    conn.execute("""
                        UPDATE plab_clients SET
                            prefix            = ?,
                            first_name        = ?,
                            last_name         = ?,
                            mobile            = ?,
                            email             = ?,
                            dob               = ?,
                            city              = ?,
                            state             = ?,
                            registration_date = ?,
                            plan_type         = ?,
                            package_amount    = ?,
                            final_package     = ?,
                            discount_allowed  = ?,
                            account_status    = ?,
                            counsellor        = ?,
                            current_stage     = ?,
                            joined_stage      = ?,
                            lead_source       = ?,
                            instagram         = ?,
                            facebook          = ?,
                            linkedin          = ?,
                            father_name       = ?,
                            father_phone      = ?,
                            mother_name       = ?,
                            mother_phone      = ?,
                            parents_email     = ?,
                            counsellor_number = ?,
                            counsellor_email  = ?,
                            inst1_amount      = ?,
                            inst1_date        = ?,
                            inst1_note        = ?,
                            inst2_amount      = ?,
                            inst2_date        = ?,
                            inst2_note        = ?,
                            additional_notes  = ?,
                            pathway           = 'consulting',
                            product_id        = COALESCE(?, product_id)
                        WHERE id = ?
                    """, (
                        data['prefix'], data['first_name'], data['last_name'],
                        data['mobile'], data['email'],
                        data['dob'], data['city'], data['state'],
                        data['registration_date'], data['plan_type'],
                        data['package_amount'], data['final_package'],
                        data['discount_allowed'], data['account_status'],
                        data['counsellor'], data['current_stage'],
                        data['joined_stage'], data['lead_source'],
                        data['instagram'], data['facebook'], data['linkedin'],
                        data['father_name'], data['father_phone'],
                        data['mother_name'], data['mother_phone'],
                        data['parents_email'], data['counsellor_number'],
                        data['counsellor_email'],
                        data['inst1_amount'], data['inst1_date'], data['inst1_note'],
                        data['inst2_amount'], data['inst2_date'], data['inst2_note'],
                        data['additional_notes'],
                        data['product_id'], exists['id'],
                    ))
                    updated += 1
                else:
                    conn.execute("""
                        INSERT INTO plab_clients (
                            registration_number, prefix, first_name, last_name,
                            mobile, email, dob, city, state,
                            registration_date, plan_type, package_amount, final_package,
                            discount_allowed, account_status, counsellor,
                            current_stage, joined_stage, lead_source,
                            instagram, facebook, linkedin,
                            father_name, father_phone, mother_name, mother_phone,
                            parents_email, counsellor_number, counsellor_email,
                            inst1_amount, inst1_date, inst1_note,
                            inst2_amount, inst2_date, inst2_note,
                            additional_notes, pathway, product_id
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'consulting',?)
                    """, (
                        data['registration_number'], data['prefix'],
                        data['first_name'], data['last_name'],
                        data['mobile'], data['email'],
                        data['dob'], data['city'], data['state'],
                        data['registration_date'], data['plan_type'],
                        data['package_amount'], data['final_package'],
                        data['discount_allowed'], data['account_status'],
                        data['counsellor'], data['current_stage'],
                        data['joined_stage'], data['lead_source'],
                        data['instagram'], data['facebook'], data['linkedin'],
                        data['father_name'], data['father_phone'],
                        data['mother_name'], data['mother_phone'],
                        data['parents_email'], data['counsellor_number'],
                        data['counsellor_email'],
                        data['inst1_amount'], data['inst1_date'], data['inst1_note'],
                        data['inst2_amount'], data['inst2_date'], data['inst2_note'],
                        data['additional_notes'],
                        data['product_id'],
                    ))
                    inserted += 1
            except Exception as e:
                logging.warning(f"clients upsert {reg}: {e}")
                try: conn.rollback()
                except Exception: pass
                errors += 1

        conn.commit()
        _set_marker(conn, *MARK_CLIENTS)
        logging.info(
            f"Consulting clients import: inserted {inserted}, updated {updated}, "
            f"skipped {skipped}, errors {errors}"
        )
        return {'inserted': inserted, 'updated': updated,
                'skipped': skipped, 'errors': errors}
    except Exception as e:
        logging.error(f"run_import_consulting_clients_once: {e}")
        try: conn.rollback()
        except Exception: pass
        return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 1}
    finally:
        try: conn.close()
        except Exception: pass


# ── 2) ACADEMIC DETAILS ───────────────────────────────────────────────
# Header row is messy (col 2 has no label). Use positional indices.
ACADEMIC_COL = {
    'reg':               1,   # col 2 (0-indexed=1)
    'img_fmg':           2,
    'full_name':         3,
    'img_college':       4,
    'fmg_college':       5,
    'mbbs_status':       6,
    'speciality_1':      7,
    'speciality_2':      8,
    'internship_status': 9,
    'internship_gap':    10,
    'mbbs_start':        11,
    'address':           12,
    'mbbs_end':          13,
    'intern_hospital':   14,
    'intern_location':   15,
    'intern_hospital_2': 16,
    'intern_location_2': 17,
    'intern_start':      18,
    'intern_end':        19,
    'gap_months':        20,
    'gap_reason':        21,
    'working_status':    22,
    'working_hospital':  23,
    'additional_info':   24,
}


def run_import_consulting_academic_once(get_db_fn):
    if _open_workbook(FILE_ACADEMIC) is None:
        logging.info(f"Consulting academic import: no {FILE_ACADEMIC}, skipping")
        return {'inserted': 0, 'skipped': 0, 'errors': 0}
    conn = get_db_fn()
    try:
        _ensure_marker_table(conn)
        if _has_marker(conn, *MARK_ACADEMIC):
            logging.info(f"Consulting academic: already at {MARK_ACADEMIC[1]}, skipping")
            return {'inserted': 0, 'skipped': 0, 'errors': 0}

        wb = _open_workbook(FILE_ACADEMIC)
        ws = wb.active
        inserted = skipped = errors = 0
        seen = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row):
                continue
            reg = _extract_reg(_ss(row[ACADEMIC_COL['reg']]))
            if not reg:
                skipped += 1
                continue
            if reg in seen:
                skipped += 1
                continue
            seen.add(reg)
            # Skip if no client row exists -- the FK would fail.
            try:
                client = conn.execute(
                    "SELECT id FROM plab_clients WHERE registration_number = ?",
                    (reg,),
                ).fetchone()
            except Exception:
                client = None
            if not client:
                skipped += 1
                continue
            try:
                conn.execute("""
                    INSERT INTO ops_academic_details (
                        registration_number, img_fmg,
                        img_medical_college, fmg_medical_college,
                        mbbs_status, mbbs_start_date, mbbs_end_date,
                        speciality_interest_1, speciality_interest_2,
                        internship_status, internship_hospital,
                        internship_location, internship_hospital_2,
                        internship_location_2,
                        internship_start_date, internship_end_date,
                        internship_gap, gap_in_months, gap_reason,
                        working_status, working_hospital_name,
                        additional_info, pathway
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'consulting')
                """, (
                    reg,
                    _ss(row[ACADEMIC_COL['img_fmg']]),
                    _ss(row[ACADEMIC_COL['img_college']]),
                    _ss(row[ACADEMIC_COL['fmg_college']]),
                    _ss(row[ACADEMIC_COL['mbbs_status']]),
                    _sd(row[ACADEMIC_COL['mbbs_start']]),
                    _sd(row[ACADEMIC_COL['mbbs_end']]),
                    _ss(row[ACADEMIC_COL['speciality_1']]),
                    _ss(row[ACADEMIC_COL['speciality_2']]),
                    _ss(row[ACADEMIC_COL['internship_status']]),
                    _ss(row[ACADEMIC_COL['intern_hospital']]),
                    _ss(row[ACADEMIC_COL['intern_location']]),
                    _ss(row[ACADEMIC_COL['intern_hospital_2']]),
                    _ss(row[ACADEMIC_COL['intern_location_2']]),
                    _sd(row[ACADEMIC_COL['intern_start']]),
                    _sd(row[ACADEMIC_COL['intern_end']]),
                    _ss(row[ACADEMIC_COL['internship_gap']]),
                    _ss(row[ACADEMIC_COL['gap_months']]),
                    _ss(row[ACADEMIC_COL['gap_reason']]),
                    _ss(row[ACADEMIC_COL['working_status']]),
                    _ss(row[ACADEMIC_COL['working_hospital']]),
                    _ss(row[ACADEMIC_COL['additional_info']]),
                ))
                inserted += 1
            except Exception as e:
                logging.warning(f"academic insert {reg}: {e}")
                try: conn.rollback()
                except Exception: pass
                errors += 1
        conn.commit()
        _set_marker(conn, *MARK_ACADEMIC)
        logging.info(
            f"Consulting academic import: inserted {inserted}, "
            f"skipped {skipped}, errors {errors}"
        )
        return {'inserted': inserted, 'skipped': skipped, 'errors': errors}
    except Exception as e:
        logging.error(f"run_import_consulting_academic_once: {e}")
        try: conn.rollback()
        except Exception: pass
        return {'inserted': 0, 'skipped': 0, 'errors': 1}
    finally:
        try: conn.close()
        except Exception: pass


# ── 3) CALL NOTES ─────────────────────────────────────────────────────
def run_import_consulting_call_notes_once(get_db_fn):
    if _open_workbook(FILE_CALL_NOTES) is None:
        logging.info(f"Consulting call notes: no {FILE_CALL_NOTES}, skipping")
        return {'inserted': 0, 'skipped': 0, 'errors': 0}
    conn = get_db_fn()
    try:
        _ensure_marker_table(conn)
        if _has_marker(conn, *MARK_CALL_NOTES):
            logging.info(f"Consulting call notes: already at {MARK_CALL_NOTES[1]}, skipping")
            return {'inserted': 0, 'skipped': 0, 'errors': 0}

        wb = _open_workbook(FILE_CALL_NOTES)
        ws = wb.active
        inserted = skipped = errors = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(c is not None and c != '' for c in row):
                continue
            cname = _ss(row[0]) if len(row) > 0 else ''
            reg = _extract_reg(cname)
            if not reg:
                skipped += 1
                continue
            try:
                client = conn.execute(
                    "SELECT id FROM plab_clients WHERE registration_number = ?",
                    (reg,),
                ).fetchone()
            except Exception:
                client = None
            if not client:
                skipped += 1
                continue
            call_date = _sd(row[1]) if len(row) > 1 else ''
            note = _ss(row[2]) if len(row) > 2 else ''
            added_by = _ss(row[3]) if len(row) > 3 else ''
            try:
                conn.execute("""
                    INSERT INTO ops_call_notes (
                        registration_number, call_date, call_note,
                        contacted_status, contact_type, added_by, pathway
                    ) VALUES (?, ?, ?, 'Yes', 'Call', ?, 'consulting')
                """, (reg, call_date, note, added_by))
                inserted += 1
            except Exception as e:
                logging.warning(f"call_notes insert {reg}: {e}")
                try: conn.rollback()
                except Exception: pass
                errors += 1
        conn.commit()
        _set_marker(conn, *MARK_CALL_NOTES)
        logging.info(
            f"Consulting call notes import: inserted {inserted}, "
            f"skipped {skipped}, errors {errors}"
        )
        return {'inserted': inserted, 'skipped': skipped, 'errors': errors}
    except Exception as e:
        logging.error(f"run_import_consulting_call_notes_once: {e}")
        try: conn.rollback()
        except Exception: pass
        return {'inserted': 0, 'skipped': 0, 'errors': 1}
    finally:
        try: conn.close()
        except Exception: pass


# ── 4) EPIC ───────────────────────────────────────────────────────────
EPIC_COL = {
    'name':           0,
    'login_id':       1,
    'login_pwd':      2,
    'sq1':            3, 'sa1': 4,
    'sq2':            5, 'sa2': 6,
    'sq3':            7,
    'sq4':            8,
    'sa3':            9, 'sa4': 10,
    'notary_login':   11,
    'notary_pwd':     12,
    'epic_reg_status':13,
    'reg_date':       14,
    'epic_id_num':    15,
    'epic_status':    16,
    'notary_status':  17,
    'document_stage': 18,
    'document_stage_status': 19,
}


def run_import_consulting_epic_once(get_db_fn):
    if _open_workbook(FILE_EPIC) is None:
        logging.info(f"Consulting EPIC: no {FILE_EPIC}, skipping")
        return {'inserted': 0, 'skipped': 0, 'errors': 0}
    conn = get_db_fn()
    try:
        _ensure_marker_table(conn)
        if _has_marker(conn, *MARK_EPIC):
            logging.info(f"Consulting EPIC: already at {MARK_EPIC[1]}, skipping")
            return {'inserted': 0, 'skipped': 0, 'errors': 0}

        wb = _open_workbook(FILE_EPIC)
        ws = wb.active
        inserted = skipped = errors = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row):
                continue
            reg = _extract_reg(_ss(row[EPIC_COL['name']]))
            if not reg:
                skipped += 1
                continue
            try:
                client = conn.execute(
                    "SELECT id FROM plab_clients WHERE registration_number = ?",
                    (reg,),
                ).fetchone()
            except Exception:
                client = None
            if not client:
                skipped += 1
                continue
            try:
                conn.execute("""
                    INSERT INTO ops_epic_registration (
                        registration_number,
                        login_id, login_pwd,
                        secret_question_1, secret_answer_1,
                        secret_question_2, secret_answer_2,
                        secret_question_3, secret_answer_3,
                        secret_question_4, secret_answer_4,
                        notary_camp_login, notary_camp_password,
                        registration_date, epic_id_number,
                        epic_status, notary_camp,
                        documents_stage, document_stage_status,
                        pathway
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'consulting')
                """, (
                    reg,
                    _ss(row[EPIC_COL['login_id']]),
                    _ss(row[EPIC_COL['login_pwd']]),
                    _ss(row[EPIC_COL['sq1']]), _ss(row[EPIC_COL['sa1']]),
                    _ss(row[EPIC_COL['sq2']]), _ss(row[EPIC_COL['sa2']]),
                    _ss(row[EPIC_COL['sq3']]), _ss(row[EPIC_COL['sa3']]),
                    _ss(row[EPIC_COL['sq4']]), _ss(row[EPIC_COL['sa4']]),
                    _ss(row[EPIC_COL['notary_login']]),
                    _ss(row[EPIC_COL['notary_pwd']]),
                    _sd(row[EPIC_COL['reg_date']]),
                    _ss(row[EPIC_COL['epic_id_num']]),
                    _ss(row[EPIC_COL['epic_status']]),
                    _ss(row[EPIC_COL['notary_status']]),
                    _ss(row[EPIC_COL['document_stage']]),
                    _ss(row[EPIC_COL['document_stage_status']]),
                ))
                inserted += 1
            except Exception as e:
                logging.warning(f"epic insert {reg}: {e}")
                try: conn.rollback()
                except Exception: pass
                errors += 1
        conn.commit()
        _set_marker(conn, *MARK_EPIC)
        logging.info(
            f"Consulting EPIC import: inserted {inserted}, "
            f"skipped {skipped}, errors {errors}"
        )
        return {'inserted': inserted, 'skipped': skipped, 'errors': errors}
    except Exception as e:
        logging.error(f"run_import_consulting_epic_once: {e}")
        try: conn.rollback()
        except Exception: pass
        return {'inserted': 0, 'skipped': 0, 'errors': 1}
    finally:
        try: conn.close()
        except Exception: pass


# ── 5) PAYMENTS ───────────────────────────────────────────────────────
PAYMENTS_COL = {
    'name':               0,
    'reg_number':         1,
    'total_package':      2,
    'instalment':         3,
    'total_amount_paid':  4,
    'payment_date':       5,
    'amount_paid':        6,
    'gst_paid':           7,
    'payment_method':     8,
    'notes':              9,
}


def run_import_consulting_payments_once(get_db_fn):
    if _open_workbook(FILE_PAYMENTS) is None:
        logging.info(f"Consulting payments: no {FILE_PAYMENTS}, skipping")
        return {'inserted': 0, 'skipped': 0, 'errors': 0}
    conn = get_db_fn()
    try:
        _ensure_marker_table(conn)
        if _has_marker(conn, *MARK_PAYMENTS):
            logging.info(f"Consulting payments: already at {MARK_PAYMENTS[1]}, skipping")
            return {'inserted': 0, 'skipped': 0, 'errors': 0}

        wb = _open_workbook(FILE_PAYMENTS)
        ws = wb.active
        inserted = skipped = errors = recovered = 0
        # S-3 fix 3: since this is a marker-bump re-run, the previous
        # 33 rows from v2 are already in ops_payments. Wipe pathway=
        # 'consulting' rows first so the re-insert lands a clean
        # 38-row set rather than duplicating the previous 33.
        try:
            del_result = conn.execute(
                "DELETE FROM ops_payments "
                " WHERE COALESCE(pathway,'plab') = 'consulting'"
            )
            conn.commit()
            logging.info(
                "payments fix3: cleared prior consulting payment rows "
                "before re-import"
            )
        except Exception as e:
            logging.warning(f"payments fix3 wipe: {e}")
            try: conn.rollback()
            except Exception: pass
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row):
                continue
            col_b = _ss(row[PAYMENTS_COL['reg_number']])
            name_reg = _extract_reg(_ss(row[PAYMENTS_COL['name']]))

            # Try col B first (the explicit Registration Number column).
            reg = col_b if col_b and col_b.upper().startswith('GCCSS') else None
            client = None
            if reg:
                try:
                    client = conn.execute(
                        "SELECT id FROM plab_clients WHERE registration_number = ?",
                        (reg,),
                    ).fetchone()
                except Exception:
                    client = None

            # Fallback: if col B didn't match a client, try the reg# embedded
            # in the candidate name. Several rows in the source Excel have
            # typos in col B (missing leading zero, off-by-one digit) where
            # the name suffix is correct. Recover those rather than skipping.
            if not client and name_reg and name_reg != reg:
                try:
                    client = conn.execute(
                        "SELECT id FROM plab_clients WHERE registration_number = ?",
                        (name_reg,),
                    ).fetchone()
                except Exception:
                    client = None
                if client:
                    if reg:
                        logging.info(
                            f"payments fallback: col B '{reg}' had no match, "
                            f"using name-suffix '{name_reg}' instead")
                    reg = name_reg
                    recovered += 1

            if not reg or not client:
                skipped += 1
                continue
            try:
                conn.execute("""
                    INSERT INTO ops_payments (
                        registration_number, payment_date,
                        amount_paid, gst_paid, total_amount_paid,
                        instalment, payment_method,
                        total_package, notes, pathway
                    ) VALUES (?,?,?,?,?,?,?,?,?,'consulting')
                """, (
                    reg,
                    _sd(row[PAYMENTS_COL['payment_date']]),
                    _sf(row[PAYMENTS_COL['amount_paid']]),
                    _sf(row[PAYMENTS_COL['gst_paid']]),
                    _sf(row[PAYMENTS_COL['total_amount_paid']]),
                    _ss(row[PAYMENTS_COL['instalment']]),
                    _ss(row[PAYMENTS_COL['payment_method']]),
                    _sf(row[PAYMENTS_COL['total_package']]),
                    _ss(row[PAYMENTS_COL['notes']]),
                ))
                inserted += 1
            except Exception as e:
                logging.warning(f"payments insert {reg}: {e}")
                try: conn.rollback()
                except Exception: pass
                errors += 1
        conn.commit()
        _set_marker(conn, *MARK_PAYMENTS)
        logging.info(
            f"Consulting payments import: inserted {inserted} "
            f"({recovered} via name-fallback), "
            f"skipped {skipped}, errors {errors}"
        )
        return {'inserted': inserted, 'skipped': skipped,
                'errors': errors, 'recovered': recovered}
    except Exception as e:
        logging.error(f"run_import_consulting_payments_once: {e}")
        try: conn.rollback()
        except Exception: pass
        return {'inserted': 0, 'skipped': 0, 'errors': 1}
    finally:
        try: conn.close()
        except Exception: pass


# ── 6) AMC REGISTRATIONS ──────────────────────────────────────────────
AMC_COL = {
    'name':       0,
    'amc_ref':    1,
    'login_pwd':  2,
    'amc_setup':  3,
    'reg_date':   4,
}


def run_import_consulting_amc_once(get_db_fn):
    if _open_workbook(FILE_AMC) is None:
        logging.info(f"Consulting AMC: no {FILE_AMC}, skipping")
        return {'inserted': 0, 'skipped': 0, 'errors': 0}
    conn = get_db_fn()
    try:
        _ensure_marker_table(conn)
        if _has_marker(conn, *MARK_AMC):
            logging.info(f"Consulting AMC: already at {MARK_AMC[1]}, skipping")
            return {'inserted': 0, 'skipped': 0, 'errors': 0}

        wb = _open_workbook(FILE_AMC)
        ws = wb.active
        inserted = skipped = errors = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row):
                continue
            reg = _extract_reg(_ss(row[AMC_COL['name']]))
            if not reg:
                skipped += 1
                continue
            try:
                client = conn.execute(
                    "SELECT id FROM plab_clients WHERE registration_number = ?",
                    (reg,),
                ).fetchone()
            except Exception:
                client = None
            if not client:
                skipped += 1
                continue
            try:
                conn.execute("""
                    INSERT INTO ops_amc_registration (
                        registration_number, amc_reference_number,
                        login_pwd, amc_setup, registration_date, pathway
                    ) VALUES (?, ?, ?, ?, ?, 'consulting')
                """, (
                    reg,
                    _ss(row[AMC_COL['amc_ref']]),
                    _ss(row[AMC_COL['login_pwd']]),
                    _ss(row[AMC_COL['amc_setup']]),
                    _sd(row[AMC_COL['reg_date']]),
                ))
                inserted += 1
            except Exception as e:
                logging.warning(f"amc insert {reg}: {e}")
                try: conn.rollback()
                except Exception: pass
                errors += 1
        conn.commit()
        _set_marker(conn, *MARK_AMC)
        logging.info(
            f"Consulting AMC import: inserted {inserted}, "
            f"skipped {skipped}, errors {errors}"
        )
        return {'inserted': inserted, 'skipped': skipped, 'errors': errors}
    except Exception as e:
        logging.error(f"run_import_consulting_amc_once: {e}")
        try: conn.rollback()
        except Exception: pass
        return {'inserted': 0, 'skipped': 0, 'errors': 1}
    finally:
        try: conn.close()
        except Exception: pass


# ── 7) MENTORSHIP ─────────────────────────────────────────────────────
MENTORSHIP_COL = {
    'name':     0,
    'date':     1,
    'session':  2,
    'duration': 3,
    'amount':   4,
    'mentor':   5,
    'notes':    6,
}


def run_import_consulting_mentorship_once(get_db_fn):
    if _open_workbook(FILE_MENTORSHIP) is None:
        logging.info(f"Consulting mentorship: no {FILE_MENTORSHIP}, skipping")
        return {'inserted': 0, 'skipped': 0, 'errors': 0}
    conn = get_db_fn()
    try:
        _ensure_marker_table(conn)
        if _has_marker(conn, *MARK_MENTORSHIP):
            logging.info(f"Consulting mentorship: already at {MARK_MENTORSHIP[1]}, skipping")
            return {'inserted': 0, 'skipped': 0, 'errors': 0}

        wb = _open_workbook(FILE_MENTORSHIP)
        ws = wb.active
        inserted = skipped = errors = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row):
                continue
            reg = _extract_reg(_ss(row[MENTORSHIP_COL['name']]))
            if not reg:
                skipped += 1
                continue
            try:
                client = conn.execute(
                    "SELECT id FROM plab_clients WHERE registration_number = ?",
                    (reg,),
                ).fetchone()
            except Exception:
                client = None
            if not client:
                skipped += 1
                continue
            # Duration -- Excel has "3.0" (hours). Convert to minutes string.
            duration_val = row[MENTORSHIP_COL['duration']]
            duration_str = ''
            try:
                if duration_val not in (None, ''):
                    duration_str = str(int(float(duration_val) * 60))
            except Exception:
                duration_str = _ss(duration_val)
            try:
                conn.execute("""
                    INSERT INTO ops_mentorship (
                        registration_number, session_date,
                        duration_minutes, amount_paid,
                        program_provider, additional_notes, pathway
                    ) VALUES (?, ?, ?, ?, ?, ?, 'consulting')
                """, (
                    reg,
                    _sd(row[MENTORSHIP_COL['date']]),
                    duration_str,
                    _sf(row[MENTORSHIP_COL['amount']]),
                    _ss(row[MENTORSHIP_COL['mentor']]) or _ss(row[MENTORSHIP_COL['session']]),
                    _ss(row[MENTORSHIP_COL['notes']]),
                ))
                inserted += 1
            except Exception as e:
                logging.warning(f"mentorship insert {reg}: {e}")
                try: conn.rollback()
                except Exception: pass
                errors += 1
        conn.commit()
        _set_marker(conn, *MARK_MENTORSHIP)
        logging.info(
            f"Consulting mentorship import: inserted {inserted}, "
            f"skipped {skipped}, errors {errors}"
        )
        return {'inserted': inserted, 'skipped': skipped, 'errors': errors}
    except Exception as e:
        logging.error(f"run_import_consulting_mentorship_once: {e}")
        try: conn.rollback()
        except Exception: pass
        return {'inserted': 0, 'skipped': 0, 'errors': 1}
    finally:
        try: conn.close()
        except Exception: pass


# ── ORCHESTRATOR ──────────────────────────────────────────────────────
def run_all_consulting_imports_once(get_db_fn):
    """Run all 7 consulting imports in dependency order.
    Clients first -- the others FK-reference plab_clients rows.
    """
    results = {}
    results['clients']    = run_import_consulting_clients_once(get_db_fn)
    results['academic']   = run_import_consulting_academic_once(get_db_fn)
    results['call_notes'] = run_import_consulting_call_notes_once(get_db_fn)
    results['epic']       = run_import_consulting_epic_once(get_db_fn)
    results['payments']   = run_import_consulting_payments_once(get_db_fn)
    results['amc']        = run_import_consulting_amc_once(get_db_fn)
    results['mentorship'] = run_import_consulting_mentorship_once(get_db_fn)
    return results
