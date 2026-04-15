import os
import hashlib
import logging
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import calendar
from db import get_db
from email_utils import send_birthday_reminder, send_anniversary_reminder, send_announcement_email, send_happy_birthday_email

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'goocampus-leave-2026')
app.config['DEBUG'] = False

PHOTO_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'photos')
os.makedirs(PHOTO_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        conn = get_db()
        user = conn.execute('SELECT is_admin FROM employees WHERE id = ?', (session['user_id'],)).fetchone()
        conn.close()
        if not user or user['is_admin'] != 1:
            flash('Admin access required', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# Management codes that can post announcements (along with admin)
MANAGEMENT_CODES = ['GC001', 'GC002', 'GC003']

def can_post_announcements(user):
    """Check if user is admin or management team."""
    return user['is_admin'] == 1 or user['emp_code'] in MANAGEMENT_CODES

def can_approve_leave(approver, leave_employee, conn):
    """Determine if approver can approve/reject a leave for leave_employee.
    Returns (allowed: bool, reason: str).
    Rules:
      1. Never self-approve
      2. Management (GC001-GC003) leaves → other management members can approve
      3. Employee with reporting_to → reporting manager OR any admin can approve
      4. Employee without reporting_to → only admin can approve
    """
    if approver['id'] == leave_employee['id']:
        return False, 'You cannot approve your own leave request'

    emp_code = leave_employee.get('emp_code', '')
    approver_code = approver.get('emp_code', '')

    # Management cross-approval: any management member can approve another's leave
    if emp_code in MANAGEMENT_CODES and approver_code in MANAGEMENT_CODES:
        return True, 'management_peer'

    # Admin can approve anyone else's leave
    if approver['is_admin'] == 1:
        return True, 'admin'

    # Reporting manager can approve their direct report's leave
    reporting_to = leave_employee.get('reporting_to')
    if reporting_to and reporting_to == approver['id']:
        return True, 'manager'

    return False, 'Not authorized to approve this leave'

def has_module_access(user, module):
    """Check if user has access to a CRM module (sales, projects, b2b_meetings).
    Admin and management always have access. Others need explicit grant."""
    if user['is_admin'] == 1 or user['emp_code'] in MANAGEMENT_CODES:
        return True
    conn = get_db()
    access = conn.execute(
        'SELECT id FROM module_access WHERE employee_id = ? AND module = ? AND is_active = 1',
        (user['id'], module)
    ).fetchone()
    conn.close()
    return access is not None

def sales_access_required(f):
    """Decorator requiring sales module access."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = get_user()
        if not has_module_access(user, 'sales'):
            flash('Sales module access required', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def get_user():
    if 'user_id' not in session:
        return None
    conn = get_db()
    user = conn.execute('SELECT * FROM employees WHERE id = ?', (session['user_id'],)).fetchone()
    conn.close()
    return user

def get_monthly_alloc(month_num):
    """Get monthly leave allocation for a given month number (1-12).
    April (month 4) gets 3 days, all other months get 2 days. Base total = 25/year."""
    return 3 if month_num == 4 else 2

def is_manager(user_id):
    """Check if user has any direct reports"""
    conn = get_db()
    count = conn.execute('SELECT COUNT(*) as cnt FROM employees WHERE reporting_to = ? AND is_active = 1', (user_id,)).fetchone()
    conn.close()
    return count['cnt'] > 0

def get_pending_team_count(user_id):
    """Get count of pending leave requests from direct reports"""
    conn = get_db()
    count = conn.execute('''
        SELECT COUNT(*) as cnt FROM leave_records lr
        JOIN employees e ON lr.employee_id = e.id
        WHERE lr.status = 'pending' AND e.reporting_to = ?
    ''', (user_id,)).fetchone()
    conn.close()
    return count['cnt']

@app.context_processor
def inject_manager_status():
    """Make is_manager, pending_team_count, and has_sales_access available in all templates"""
    if 'user_id' in session:
        user_id = session['user_id']
        user = get_user()
        mgr = is_manager(user_id)
        pending_team = get_pending_team_count(user_id) if mgr else 0
        sales_access = has_module_access(user, 'sales') if user else False
        return {'is_manager': mgr, 'pending_team_count': pending_team, 'has_sales_access': sales_access}
    return {'is_manager': False, 'pending_team_count': 0, 'has_sales_access': False}

def calculate_monthly_balance(employee_id, year, month):
    """Calculate running balance for a given month"""
    conn = get_db()

    # Get carry forward
    emp = conn.execute('SELECT carry_forward FROM employees WHERE id = ?', (employee_id,)).fetchone()
    carry_forward = emp['carry_forward'] if emp else 0

    # Total annual allocation = 25 + carry_forward
    total_allocation = 25 + carry_forward

    # Calculate balance up to and including the given month
    # Carry forward is added as starting balance; monthly alloc: April=3, others=2
    balance = carry_forward
    for m in range(4, month + 1):  # FY starts April (month 4)
        balance += get_monthly_alloc(m)
        # Subtract approved leaves in this month
        leaves = conn.execute('''
            SELECT SUM(days) as total_days FROM leave_records
            WHERE employee_id = ? AND status = 'approved'
            AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
        ''', (employee_id, str(year if m >= 4 else year - 1), str(m).zfill(2))).fetchone()

        if leaves['total_days']:
            balance -= leaves['total_days']

    conn.close()
    return max(0, balance)

def get_available_balance(employee_id, year, month):
    """Get available balance at the start of a month"""
    conn = get_db()

    emp = conn.execute('SELECT carry_forward FROM employees WHERE id = ?', (employee_id,)).fetchone()
    carry_forward = emp['carry_forward'] if emp else 0

    total_allocation = 25 + carry_forward

    # Get balance from start of FY to end of previous month
    # Carry forward added as starting balance; monthly alloc: April=3, others=2
    balance = carry_forward
    for m in range(4, month):
        balance += get_monthly_alloc(m)
        leaves = conn.execute('''
            SELECT SUM(days) as total_days FROM leave_records
            WHERE employee_id = ? AND status = 'approved'
            AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
        ''', (employee_id, str(year if m >= 4 else year - 1), str(m).zfill(2))).fetchone()

        if leaves['total_days']:
            balance -= leaves['total_days']

    conn.close()
    return max(0, balance)

def get_leaves_for_month(employee_id, year, month):
    """Get all approved leaves for a specific month"""
    conn = get_db()
    leaves = conn.execute('''
        SELECT * FROM leave_records
        WHERE employee_id = ? AND status = 'approved'
        AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
        ORDER BY leave_date
    ''', (employee_id, str(year), str(month).zfill(2))).fetchall()
    conn.close()
    return leaves

def get_all_holidays():
    """Get all holidays"""
    conn = get_db()
    holidays = conn.execute('SELECT * FROM holidays ORDER BY holiday_date').fetchall()
    conn.close()
    return {row['holiday_date']: row for row in holidays}

@app.route('/')
@login_required
def index():
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        emp_code = request.form.get('emp_code', '').strip()
        password = request.form.get('password', '').strip()

        if not emp_code or not password:
            flash('Employee code and password required', 'error')
            return render_template('login.html')

        conn = get_db()
        user = conn.execute('SELECT * FROM employees WHERE emp_code = ? AND is_active = 1', (emp_code,)).fetchone()
        conn.close()

        if user and user['password'] == hash_password(password):
            session['user_id'] = user['id']
            session['is_admin'] = user['is_admin']
            session['emp_code'] = user['emp_code']
            # Check if password is still the default (first name lowercase)
            default_pw = user['name'].split()[0].lower()
            if password == default_pw:
                session['show_welcome'] = True
            flash('Login successful', 'success')
            return redirect(url_for('index'))
        else:
            flash('Invalid credentials or account inactive', 'error')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully', 'success')
    return redirect(url_for('login'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    user = get_user()
    if request.method == 'POST':
        old_password = request.form.get('old_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not old_password or not new_password or not confirm_password:
            flash('All fields required', 'error')
            return render_template('change_password.html', user=user)

        if user['password'] != hash_password(old_password):
            flash('Current password is incorrect', 'error')
            return render_template('change_password.html', user=user)

        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return render_template('change_password.html', user=user)

        if len(new_password) < 4:
            flash('Password must be at least 4 characters', 'error')
            return render_template('change_password.html', user=user)

        conn = get_db()
        conn.execute('UPDATE employees SET password = ? WHERE id = ?', (hash_password(new_password), user['id']))
        conn.commit()
        conn.close()

        flash('Password changed successfully', 'success')
        return redirect(url_for('dashboard'))

    return render_template('change_password.html', user=user)

@app.route('/dashboard')
@login_required
def dashboard():
    """Main CRM dashboard for all users (admin and employees)"""
    user = get_user()
    conn = get_db()
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')
    fy_year = now.year if now.month >= 4 else now.year - 1

    # Total employees
    total_employees = conn.execute("SELECT COUNT(*) as count FROM employees WHERE is_active = 1 AND emp_code != 'admin'").fetchone()
    total_employees_count = total_employees['count'] if total_employees else 0

    # Leaves today
    leaves_today = conn.execute("SELECT COUNT(DISTINCT employee_id) as count FROM leave_records WHERE leave_date = ? AND status = 'approved'", (today,)).fetchone()
    leaves_today_count = leaves_today['count'] if leaves_today else 0

    # Pending leave approvals
    pending_leaves = conn.execute("SELECT COUNT(*) as count FROM leave_records WHERE status = 'pending'").fetchone()
    pending_leaves_count = pending_leaves['count'] if pending_leaves else 0

    # Departments
    departments = conn.execute("SELECT department, COUNT(*) as count FROM employees WHERE is_active = 1 AND emp_code != 'admin' GROUP BY department ORDER BY department").fetchall()
    total_departments = len(departments) if departments else 0

    # Active projects (try/except in case projects table doesn't exist)
    active_projects = 0
    try:
        ap_result = conn.execute("SELECT COUNT(*) as count FROM projects WHERE status = 'active'").fetchone()
        active_projects = ap_result['count'] if ap_result else 0
    except:
        active_projects = 0

    # Meetings this month
    month_start = now.strftime('%Y-%m-01')
    month_end = now.strftime('%Y-%m-') + str(calendar.monthrange(now.year, now.month)[1])
    total_meetings = 0
    try:
        tm_result = conn.execute("SELECT COUNT(*) as count FROM b2b_trips WHERE from_date >= ? AND from_date <= ?", (month_start, month_end)).fetchone()
        total_meetings = tm_result['count'] if tm_result else 0
    except Exception as e:
        logging.error(f"Dashboard total_meetings query error: {e}")
        total_meetings = 0
        try:
            conn.rollback()
        except:
            pass

    # Meetings this week
    week_start = (now - timedelta(days=now.weekday())).strftime('%Y-%m-%d')
    week_end = (now + timedelta(days=6-now.weekday())).strftime('%Y-%m-%d')
    meetings_this_week = 0
    try:
        mtw_result = conn.execute("SELECT COUNT(*) as count FROM b2b_trips WHERE from_date >= ? AND from_date <= ?", (week_start, week_end)).fetchone()
        meetings_this_week = mtw_result['count'] if mtw_result else 0
    except Exception as e:
        logging.error(f"Dashboard meetings_this_week query error: {e}")
        meetings_this_week = 0
        try:
            conn.rollback()
        except:
            pass

    # Sales news count
    total_news = 0
    try:
        tn_result = conn.execute("SELECT COUNT(*) as count FROM sales_news").fetchone()
        total_news = tn_result['count'] if tn_result else 0
    except Exception as e:
        logging.error(f"Dashboard total_news query error: {e}")
        total_news = 0
        try:
            conn.rollback()
        except:
            pass

    # On leave today list
    on_leave_today = conn.execute('''
        SELECT e.name, e.photo_url, e.department, lr.day_portion, lr.leave_type
        FROM leave_records lr JOIN employees e ON lr.employee_id = e.id
        WHERE lr.leave_date = ? AND lr.status = 'approved'
        ORDER BY e.name
    ''', (today,)).fetchall()

    # Recent announcements
    announcements = conn.execute('''
        SELECT a.*, e.name as posted_by_name FROM announcements a
        JOIN employees e ON a.posted_by = e.id
        WHERE a.is_active = 1
        ORDER BY a.created_at DESC LIMIT 5
    ''', ()).fetchall()

    # Upcoming holidays
    upcoming_holidays = conn.execute('''
        SELECT * FROM holidays
        WHERE holiday_date >= ?
        ORDER BY holiday_date LIMIT 5
    ''', (today,)).fetchall()

    # Recent activities (combine from different tables)
    recent_activities = []

    # Recent leaves
    try:
        recent_leaves_act = conn.execute('''
            SELECT lr.leave_date as date, e.name, lr.leave_type, lr.status, lr.created_at
            FROM leave_records lr JOIN employees e ON lr.employee_id = e.id
            ORDER BY lr.created_at DESC LIMIT 5
        ''').fetchall()
        for l in recent_leaves_act:
            recent_activities.append({
                'type': 'leave',
                'title': f"{l['name']} - {l['leave_type'].capitalize()} Leave",
                'subtitle': f"Status: {l['status'].capitalize()}",
                'time_ago': str(l['created_at'])[:10] if l['created_at'] else '',
                'icon_color': 'var(--orange)'
            })
    except Exception as e:
        logging.error(f"Dashboard recent_leaves_act query error: {e}")
        try:
            conn.rollback()
        except:
            pass

    # Recent meetings
    try:
        recent_meetings = conn.execute('''
            SELECT t.from_date, e.name, t.trip_type, t.created_at
            FROM b2b_trips t JOIN employees e ON t.employee_id = e.id
            ORDER BY t.created_at DESC LIMIT 5
        ''').fetchall()
        for m in recent_meetings:
            recent_activities.append({
                'type': 'meeting',
                'title': f"{m['name']} - {(m['trip_type'] or 'Meeting').capitalize()} Meeting",
                'subtitle': f"From: {m['from_date']}",
                'time_ago': str(m['created_at'])[:10] if m['created_at'] else '',
                'icon_color': 'var(--blue)'
            })
    except Exception as e:
        logging.error(f"Dashboard recent_meetings query error: {e}")
        try:
            conn.rollback()
        except:
            pass

    # Sort by time_ago desc and limit to 8
    recent_activities.sort(key=lambda x: x.get('time_ago', ''), reverse=True)
    recent_activities = recent_activities[:8]

    # Management flag
    is_management = user['emp_code'] in MANAGEMENT_CODES

    # Leave data for ALL users (shown in banner pills)
    my_leave_data = {}
    mgmt_leave_data = {}
    if user['emp_code'] != 'admin':
        carry_forward = user['carry_forward'] or 0
        total_allocation = 25 + carry_forward
        leaves_taken = conn.execute('''
            SELECT SUM(days) as total_days FROM leave_records
            WHERE employee_id = ? AND status = 'approved'
            AND ((strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) >= '04')
                 OR (strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) < '04'))
        ''', (user['id'], str(fy_year), str(fy_year + 1))).fetchone()
        days_taken = leaves_taken['total_days'] if leaves_taken['total_days'] else 0
        my_pending = conn.execute(
            "SELECT COUNT(*) as cnt FROM leave_records WHERE employee_id = ? AND status = 'pending'",
            (user['id'],)
        ).fetchone()['cnt']
        my_leave_data = {
            'total_allocation': total_allocation,
            'available_balance': round(total_allocation - days_taken, 2),
            'days_taken': days_taken,
            'pending_count': my_pending
        }
        # Keep mgmt_leave_data for backward compatibility
        if is_management or user['is_admin']:
            mgmt_leave_data = my_leave_data

    # ── Employee-specific data ──

    # My pending leave applications (employee's own)
    my_pending_leaves = []
    if user['emp_code'] != 'admin':
        my_pending_leaves = conn.execute('''
            SELECT lr.*, e.name, e.emp_code
            FROM leave_records lr JOIN employees e ON lr.employee_id = e.id
            WHERE lr.employee_id = ? AND lr.status = 'pending'
            ORDER BY lr.leave_date ASC
        ''', (user['id'],)).fetchall()

    # My recent activity (employee's own leaves and meetings)
    my_recent_activity = []
    if user['emp_code'] != 'admin':
        try:
            my_leaves_recent = conn.execute('''
                SELECT lr.leave_date as date, lr.leave_type, lr.status, lr.created_at, lr.day_portion
                FROM leave_records lr
                WHERE lr.employee_id = ?
                ORDER BY lr.created_at DESC LIMIT 6
            ''', (user['id'],)).fetchall()
            for l in my_leaves_recent:
                my_recent_activity.append({
                    'type': 'leave',
                    'title': f"{l['leave_type'].capitalize()} Leave - {l['date']}",
                    'subtitle': f"Status: {l['status'].capitalize()} | {'Full Day' if l['day_portion'] == 'full_day' else 'Half Day'}",
                    'time_ago': str(l['created_at'])[:10] if l['created_at'] else '',
                    'icon_color': '#10B981' if l['status'] == 'approved' else '#F59E0B' if l['status'] == 'pending' else '#EF4444'
                })
        except Exception as e:
            logging.error(f"Dashboard my_leaves_recent query error: {e}")
            try:
                conn.rollback()
            except:
                pass
        try:
            my_meetings_recent = conn.execute('''
                SELECT t.from_date, t.trip_type, t.created_at
                FROM b2b_trips t
                WHERE t.employee_id = ?
                ORDER BY t.created_at DESC LIMIT 4
            ''', (user['id'],)).fetchall()
            for m in my_meetings_recent:
                my_recent_activity.append({
                    'type': 'meeting',
                    'title': f"{(m['trip_type'] or 'Meeting').capitalize()} Meeting",
                    'subtitle': f"Date: {m['from_date']}",
                    'time_ago': str(m['created_at'])[:10] if m['created_at'] else '',
                    'icon_color': '#4A7AB5'
                })
        except Exception as e:
            logging.error(f"Dashboard my_meetings_recent query error: {e}")
            try:
                conn.rollback()
            except:
                pass
        my_recent_activity.sort(key=lambda x: x.get('time_ago', ''), reverse=True)
        my_recent_activity = my_recent_activity[:8]

    # Birthdays this month
    birthdays_this_month = []
    try:
        month_str = now.strftime('%m')
        birthdays_this_month = conn.execute('''
            SELECT id, name, photo_url, dob, department
            FROM employees
            WHERE is_active = 1 AND emp_code != 'admin'
            AND dob IS NOT NULL AND dob != ''
            AND strftime('%m', dob) = ?
            ORDER BY strftime('%d', dob)
        ''', (month_str,)).fetchall()
    except Exception as e:
        logging.error(f"Dashboard birthdays query error: {e}")
        try:
            conn.rollback()
        except:
            pass

    # Work anniversaries this month
    anniversaries_this_month = []
    try:
        current_year_str = str(now.year)
        anniversaries_this_month = conn.execute('''
            SELECT id, name, photo_url, joining_date, department
            FROM employees
            WHERE is_active = 1 AND emp_code != 'admin'
            AND joining_date IS NOT NULL AND joining_date != ''
            AND strftime('%m', joining_date) = ?
            AND strftime('%Y', joining_date) != ?
            ORDER BY strftime('%d', joining_date)
        ''', (month_str, current_year_str)).fetchall()
    except Exception as e:
        logging.error(f"Dashboard anniversaries query error: {e}")
        try:
            conn.rollback()
        except:
            pass

    # Unread notification count for this employee
    my_unread_notifs = 0
    try:
        nr = conn.execute(
            'SELECT COUNT(*) as cnt FROM notifications WHERE employee_id = ? AND is_read = 0',
            (user['id'],)
        ).fetchone()
        my_unread_notifs = nr['cnt'] if nr else 0
    except Exception as e:
        logging.error(f"Dashboard unread_notifs query error: {e}")
        try:
            conn.rollback()
        except:
            pass

    # My KRA score (if assigned)
    my_kra_score = None
    try:
        current_month = now.month
        current_year_val = now.year
        kra_row = conn.execute('''
            SELECT AVG(kr.manager_rating) as final_score
            FROM kra_monthly_ratings kr
            JOIN kra_assignments ka ON kr.assignment_id = ka.id
            WHERE ka.employee_id = ? AND kr.month = ? AND kr.year = ?
            AND kr.manager_submitted = 1
        ''', (user['id'], current_month, current_year_val)).fetchone()
        if kra_row:
            my_kra_score = kra_row['final_score']
    except Exception as e:
        logging.error(f"Dashboard kra_score query error: {e}")
        try:
            conn.rollback()
        except:
            pass

    # Approvals I need to action (for managers) - their team members' pending leaves
    team_pending_approvals = []
    if user['emp_code'] != 'admin':
        try:
            team_pending_approvals = conn.execute('''
                SELECT lr.id, lr.leave_date, lr.leave_type, lr.day_portion, lr.reason, lr.status,
                       e.name, e.emp_code, e.photo_url
                FROM leave_records lr
                JOIN employees e ON lr.employee_id = e.id
                WHERE e.reporting_to = ? AND lr.status = 'pending'
                ORDER BY lr.leave_date ASC
            ''', (user['id'],)).fetchall()
        except Exception as e:
            logging.error(f"Dashboard team_pending_approvals query error: {e}")
            try:
                conn.rollback()
            except:
                pass
        # If management, also include pending from other management members
        if user['emp_code'] in MANAGEMENT_CODES:
            try:
                mgmt_pending = conn.execute('''
                    SELECT lr.id, lr.leave_date, lr.leave_type, lr.day_portion, lr.reason, lr.status,
                           e.name, e.emp_code, e.photo_url
                    FROM leave_records lr
                    JOIN employees e ON lr.employee_id = e.id
                    WHERE lr.status = 'pending' AND e.emp_code IN (?, ?, ?)
                    AND e.id != ?
                    ORDER BY lr.leave_date ASC
                ''', (*MANAGEMENT_CODES, user['id'])).fetchall()
                existing_ids = {l['id'] for l in team_pending_approvals}
                for leave in mgmt_pending:
                    if leave['id'] not in existing_ids:
                        team_pending_approvals.append(leave)
            except Exception as e:
                logging.error(f"Dashboard mgmt_pending_approvals query error: {e}")
                try:
                    conn.rollback()
                except:
                    pass

    conn.close()

    return render_template('main_dashboard.html',
                         user=user,
                         total_employees=total_employees_count,
                         leaves_today=leaves_today_count,
                         pending_leaves=pending_leaves_count,
                         pending_count=pending_leaves_count,
                         departments=departments,
                         total_departments=total_departments,
                         active_projects=active_projects,
                         total_meetings=total_meetings,
                         meetings_this_week=meetings_this_week,
                         total_news=total_news,
                         on_leave_today=on_leave_today,
                         announcements=announcements,
                         upcoming_holidays=upcoming_holidays,
                         recent_activities=recent_activities,
                         current_month_name=calendar.month_name[now.month],
                         current_year=now.year,
                         can_announce=can_post_announcements(user),
                         is_management=is_management,
                         mgmt_leave_data=mgmt_leave_data,
                         my_leave_data=my_leave_data,
                         my_pending_leaves=my_pending_leaves,
                         my_recent_activity=my_recent_activity,
                         birthdays_this_month=birthdays_this_month,
                         anniversaries_this_month=anniversaries_this_month,
                         my_unread_notifs=my_unread_notifs,
                         my_kra_score=my_kra_score,
                         team_pending_approvals=team_pending_approvals,
                         today=today)

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    user = get_user()
    conn = get_db()

    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        dob = request.form.get('dob', '').strip()
        address = request.form.get('address', '').strip()
        emergency_contact_name = request.form.get('emergency_contact_name', '').strip()
        emergency_contact_phone = request.form.get('emergency_contact_phone', '').strip()
        emergency_contact_relation = request.form.get('emergency_contact_relation', '').strip()

        # Handle lock-once fields: designation and joining_date
        # Only update if the field is currently empty (not yet set)
        designation = request.form.get('designation', '').strip()
        joining_date = request.form.get('joining_date', '').strip()

        update_fields = '''
            UPDATE employees
            SET email = ?, phone = ?, dob = ?, address = ?,
                emergency_contact_name = ?, emergency_contact_phone = ?, emergency_contact_relation = ?
        '''
        params = [email, phone, dob, address, emergency_contact_name, emergency_contact_phone, emergency_contact_relation]

        # Only allow setting designation if not already set
        if designation and not user.get('designation'):
            update_fields += ', designation = ?'
            params.append(designation)

        # Only allow setting joining_date if not already set
        if joining_date and not user.get('joining_date'):
            update_fields += ', joining_date = ?'
            params.append(joining_date)

        update_fields += ' WHERE id = ?'
        params.append(user['id'])

        conn.execute(update_fields, tuple(params))
        conn.commit()
        conn.close()

        flash('Profile updated successfully', 'success')
        return redirect(url_for('profile'))

    # Get reporting manager name
    reporting_manager = None
    if user.get('reporting_to'):
        mgr = conn.execute('SELECT name FROM employees WHERE id = ?', (user['reporting_to'],)).fetchone()
        if mgr:
            reporting_manager = mgr['name']

    conn.close()
    return render_template('employee_profile.html', user=user, reporting_manager=reporting_manager)


@app.route('/profile/upload-photo', methods=['POST'])
@login_required
def profile_upload_photo():
    """Allow employees to upload their own profile photo."""
    user = get_user()

    if 'photo' not in request.files:
        flash('No photo selected', 'error')
        return redirect(url_for('profile'))

    file = request.files['photo']
    if file.filename == '':
        flash('No photo selected', 'error')
        return redirect(url_for('profile'))

    if file and allowed_file(file.filename):
        ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"{user['emp_code']}.{ext}"
        filepath = os.path.join(PHOTO_FOLDER, filename)

        # Remove old photos with different extensions
        for old_ext in ALLOWED_EXTENSIONS:
            old_path = os.path.join(PHOTO_FOLDER, f"{user['emp_code']}.{old_ext}")
            if os.path.exists(old_path) and old_path != filepath:
                os.remove(old_path)

        file.save(filepath)

        # Update photo_url in database
        conn = get_db()
        conn.execute('UPDATE employees SET photo_url = ? WHERE id = ?', (filename, user['id']))
        conn.commit()
        conn.close()

        flash('Photo updated successfully', 'success')
    else:
        flash('Invalid file type. Allowed: png, jpg, jpeg, gif, webp', 'error')

    return redirect(url_for('profile'))

@app.route('/apply-leave', methods=['GET', 'POST'])
@login_required
def apply_leave():
    user = get_user()

    # Admin control account cannot apply for personal leave
    if user['emp_code'] == 'admin':
        flash('Admin account cannot apply for leave. Use "Add Leave (Employee)" to apply on behalf of an employee.', 'error')
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        leave_type = request.form.get('leave_type', '').strip()
        from_date = request.form.get('from_date', '').strip()
        to_date = request.form.get('to_date', '').strip()
        day_portion = request.form.get('day_portion', 'full').strip()  # for single day
        last_day_portion = request.form.get('last_day_portion', 'full').strip()  # for multi-day last day
        reason = request.form.get('reason', '').strip()

        if not leave_type or not from_date or not reason:
            flash('All fields required', 'error')
            return render_template('apply_leave.html', user=user)

        if leave_type not in ['annual', 'sick', 'casual']:
            flash('Invalid leave type', 'error')
            return render_template('apply_leave.html', user=user)

        # If no to_date, treat as single day
        if not to_date:
            to_date = from_date

        from_dt = datetime.strptime(from_date, '%Y-%m-%d')
        to_dt = datetime.strptime(to_date, '%Y-%m-%d')

        if to_dt < from_dt:
            flash('To date cannot be before from date', 'error')
            return render_template('apply_leave.html', user=user)

        conn = get_db()
        holidays_db = {row['holiday_date'] for row in conn.execute('SELECT holiday_date FROM holidays').fetchall()}

        # Build list of working days in the range
        leave_days = []
        current = from_dt
        while current <= to_dt:
            date_str = current.strftime('%Y-%m-%d')
            if current.weekday() < 5 and date_str not in holidays_db:
                leave_days.append(current)
            current += timedelta(days=1)

        if not leave_days:
            flash('No working days in the selected date range (weekends and holidays are excluded)', 'error')
            conn.close()
            return render_template('apply_leave.html', user=user)

        # Determine day portions for each day
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        total_days = 0

        for i, day in enumerate(leave_days):
            date_str = day.strftime('%Y-%m-%d')
            is_single = (len(leave_days) == 1)
            is_last = (i == len(leave_days) - 1 and len(leave_days) > 1)

            if is_single:
                portion = day_portion if day_portion in ['full', 'first_half', 'second_half'] else 'full'
            elif is_last:
                portion = last_day_portion if last_day_portion in ['full', 'first_half', 'second_half'] else 'full'
            else:
                portion = 'full'

            day_val = 0.5 if portion in ['first_half', 'second_half'] else 1.0
            total_days += day_val

            # Check for duplicate
            existing = conn.execute('SELECT id FROM leave_records WHERE employee_id = ? AND leave_date = ? AND status != ?',
                                   (user['id'], date_str, 'rejected')).fetchone()
            if existing:
                flash(f'You already have a leave record for {date_str}', 'error')
                conn.close()
                return render_template('apply_leave.html', user=user)

            conn.execute('''
                INSERT INTO leave_records (employee_id, leave_type, leave_date, days, day_portion, reason, status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, 'pending', ?)
            ''', (user['id'], leave_type, date_str, day_val, portion, reason, now_str))

        conn.commit()
        conn.close()

        conn2 = get_db()
        # Notify the reporting manager
        reporting_mgr = conn2.execute('SELECT reporting_to FROM employees WHERE id = ?', (user['id'],)).fetchone()
        if reporting_mgr and reporting_mgr['reporting_to']:
            try:
                create_notification(conn2, reporting_mgr['reporting_to'],
                    f"Leave Request from {user['name']}",
                    f"{user['name']} has applied for {total_days:.1f} day(s) of {leave_type} leave ({from_date} to {to_date})",
                    'leave_request', '/approvals')
                conn2.commit()
            except Exception as e:
                logging.error(f"Failed to create leave notification for manager: {e}")

        # Notify admin
        admins = conn2.execute('SELECT id FROM employees WHERE is_admin = 1 AND is_active = 1', ()).fetchall()
        for adm in admins:
            if not reporting_mgr or adm['id'] != reporting_mgr.get('reporting_to'):
                try:
                    create_notification(conn2, adm['id'],
                        f"Leave Request from {user['name']}",
                        f"{user['name']} has applied for {total_days:.1f} day(s) of {leave_type} leave ({from_date} to {to_date})",
                        'leave_request', '/admin/pending-approvals')
                    conn2.commit()
                except Exception as e:
                    logging.error(f"Failed to create leave notification for admin: {e}")
        conn2.close()

        day_label = 'day' if total_days == 1 else 'days'
        flash(f'Leave request submitted for {total_days:.1f} {day_label} ({len(leave_days)} working day{"s" if len(leave_days) > 1 else ""})', 'success')
        return redirect(url_for('dashboard'))

    return render_template('apply_leave.html', user=user)


@app.route('/apply-late-leave', methods=['GET', 'POST'])
@login_required
def apply_late_leave():
    """Late leave application — for past dates up to 15 days ago"""
    user = get_user()

    if user['emp_code'] == 'admin':
        flash('Admin account cannot apply for leave.', 'error')
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        leave_type = request.form.get('leave_type', '').strip()
        from_date = request.form.get('from_date', '').strip()
        to_date = request.form.get('to_date', '').strip()
        day_portion = request.form.get('day_portion', 'full').strip()
        last_day_portion = request.form.get('last_day_portion', 'full').strip()
        reason = request.form.get('reason', '').strip()
        late_reason = request.form.get('late_reason', '').strip()

        if not leave_type or not from_date or not reason or not late_reason:
            flash('All fields are required, including reason for late application', 'error')
            conn = get_db()
            late_count = conn.execute('SELECT late_leave_count FROM employees WHERE id = ?', (user['id'],)).fetchone()
            conn.close()
            return render_template('apply_late_leave.html', user=user,
                                 late_leave_count=late_count['late_leave_count'] if late_count else 0)

        if leave_type not in ['annual', 'sick', 'casual']:
            flash('Invalid leave type', 'error')
            conn = get_db()
            late_count = conn.execute('SELECT late_leave_count FROM employees WHERE id = ?', (user['id'],)).fetchone()
            conn.close()
            return render_template('apply_late_leave.html', user=user,
                                 late_leave_count=late_count['late_leave_count'] if late_count else 0)

        if not to_date:
            to_date = from_date

        from_dt = datetime.strptime(from_date, '%Y-%m-%d')
        to_dt = datetime.strptime(to_date, '%Y-%m-%d')
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        min_date = today - timedelta(days=15)

        # Validate: dates must be in the past
        if from_dt >= today:
            flash('Late leave is for past dates only. Use regular leave application for today or future dates.', 'error')
            return redirect(url_for('apply_late_leave'))

        # Validate: not older than 15 days
        if from_dt < min_date:
            flash('Cannot apply late leave for dates older than 15 days', 'error')
            return redirect(url_for('apply_late_leave'))

        if to_dt < from_dt:
            flash('To date cannot be before from date', 'error')
            return redirect(url_for('apply_late_leave'))

        if to_dt >= today:
            flash('To date must also be a past date for late leave', 'error')
            return redirect(url_for('apply_late_leave'))

        conn = get_db()
        holidays_db = {row['holiday_date'] for row in conn.execute('SELECT holiday_date FROM holidays').fetchall()}

        # Build list of working days
        leave_days = []
        current = from_dt
        while current <= to_dt:
            date_str = current.strftime('%Y-%m-%d')
            if current.weekday() < 5 and date_str not in holidays_db:
                leave_days.append(current)
            current += timedelta(days=1)

        if not leave_days:
            flash('No working days in the selected date range (weekends and holidays are excluded)', 'error')
            conn.close()
            return redirect(url_for('apply_late_leave'))

        # Insert leave records with is_late = 1
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        total_days = 0
        combined_reason = f"{reason}\n\n[Late Application Reason]: {late_reason}"

        for i, day in enumerate(leave_days):
            date_str = day.strftime('%Y-%m-%d')
            is_single = (len(leave_days) == 1)
            is_last = (i == len(leave_days) - 1 and len(leave_days) > 1)

            if is_single:
                portion = day_portion if day_portion in ['full', 'first_half', 'second_half'] else 'full'
            elif is_last:
                portion = last_day_portion if last_day_portion in ['full', 'first_half', 'second_half'] else 'full'
            else:
                portion = 'full'

            day_val = 0.5 if portion in ['first_half', 'second_half'] else 1.0
            total_days += day_val

            # Check for duplicate
            existing = conn.execute('SELECT id FROM leave_records WHERE employee_id = ? AND leave_date = ? AND status != ?',
                                   (user['id'], date_str, 'rejected')).fetchone()
            if existing:
                flash(f'You already have a leave record for {date_str}', 'error')
                conn.close()
                return redirect(url_for('apply_late_leave'))

            conn.execute('''
                INSERT INTO leave_records (employee_id, leave_type, leave_date, days, day_portion, reason, status, is_late, created_at)
                VALUES (?, ?, ?, ?, ?, ?, 'pending', 1, ?)
            ''', (user['id'], leave_type, date_str, day_val, portion, combined_reason, now_str))

        # Increment late leave count
        conn.execute('UPDATE employees SET late_leave_count = late_leave_count + 1 WHERE id = ?', (user['id'],))
        conn.commit()
        conn.close()

        # Notify reporting manager with LATE flag
        conn2 = get_db()
        reporting_mgr = conn2.execute('SELECT reporting_to FROM employees WHERE id = ?', (user['id'],)).fetchone()
        if reporting_mgr and reporting_mgr['reporting_to']:
            try:
                create_notification(conn2, reporting_mgr['reporting_to'],
                    f"Late Leave Request from {user['name']}",
                    f"{user['name']} has applied LATE for {total_days:.1f} day(s) of {leave_type} leave ({from_date} to {to_date}). Late reason: {late_reason}",
                    'leave_request', '/approvals')
                conn2.commit()
            except Exception as e:
                logging.error(f"Failed to create late leave notification for manager: {e}")

        # Notify admin
        admins = conn2.execute('SELECT id FROM employees WHERE is_admin = 1 AND is_active = 1', ()).fetchall()
        for adm in admins:
            if not reporting_mgr or adm['id'] != reporting_mgr.get('reporting_to'):
                try:
                    create_notification(conn2, adm['id'],
                        f"Late Leave Request from {user['name']}",
                        f"{user['name']} has applied LATE for {total_days:.1f} day(s) of {leave_type} leave ({from_date} to {to_date}). Late reason: {late_reason}",
                        'leave_request', '/admin/pending-approvals')
                    conn2.commit()
                except Exception as e:
                    logging.error(f"Failed to create late leave notification for admin: {e}")
        conn2.close()

        day_label = 'day' if total_days == 1 else 'days'
        flash(f'Late leave request submitted for {total_days:.1f} {day_label}. Your late leave count has been updated.', 'success')
        return redirect(url_for('dashboard'))

    # GET request
    conn = get_db()
    late_count = conn.execute('SELECT late_leave_count FROM employees WHERE id = ?', (user['id'],)).fetchone()
    conn.close()
    return render_template('apply_late_leave.html', user=user,
                         late_leave_count=late_count['late_leave_count'] if late_count else 0)


@app.route('/cancel-leave/<int:leave_id>', methods=['POST'])
@login_required
def cancel_leave(leave_id):
    user = get_user()
    conn = get_db()

    leave = conn.execute('SELECT * FROM leave_records WHERE id = ?', (leave_id,)).fetchone()

    if not leave or leave['employee_id'] != user['id']:
        flash('Leave not found or unauthorized', 'error')
        conn.close()
        return redirect(url_for('dashboard'))

    if leave['status'] != 'pending':
        flash('Only pending leaves can be cancelled', 'error')
        conn.close()
        return redirect(url_for('dashboard'))

    conn.execute('DELETE FROM leave_records WHERE id = ?', (leave_id,))
    conn.commit()
    conn.close()

    flash('Leave request cancelled', 'success')
    return redirect(url_for('dashboard'))


@app.route('/retrieve-leave/<int:leave_id>', methods=['POST'])
@login_required
def retrieve_leave(leave_id):
    """Retrieve (recall) a leave — works for any status, before leave date."""
    user = get_user()
    conn = get_db()

    leave = conn.execute('SELECT * FROM leave_records WHERE id = ?', (leave_id,)).fetchone()
    if not leave or leave['employee_id'] != user['id']:
        flash('Leave not found or unauthorized', 'error')
        conn.close()
        return redirect(url_for('my_leave_report'))

    if leave['status'] == 'retrieved':
        flash('This leave has already been retrieved', 'error')
        conn.close()
        return redirect(url_for('my_leave_report'))

    # Check leave date is today or in the future
    try:
        leave_dt = datetime.strptime(leave['leave_date'], '%Y-%m-%d')
    except (ValueError, TypeError):
        leave_dt = leave['leave_date'] if hasattr(leave['leave_date'], 'strftime') else datetime.now()
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    retrieve_reason = request.form.get('retrieve_reason', '').strip()
    if not retrieve_reason:
        flash('Please provide a reason for retrieving this leave', 'error')
        conn.close()
        return redirect(url_for('my_leave_report'))

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute('''
        UPDATE leave_records
        SET status = 'retrieved', modification_reason = ?, approved_at = ?
        WHERE id = ?
    ''', (retrieve_reason, now_str, leave_id))
    conn.commit()
    conn.close()

    flash('Leave retrieved successfully', 'success')
    return redirect(url_for('my_leave_report'))


@app.route('/modify-leave/<int:leave_id>', methods=['GET', 'POST'])
@login_required
def modify_leave(leave_id):
    """Modify an existing leave — creates new record, marks old as modified, goes to pending."""
    user = get_user()
    conn = get_db()

    leave = conn.execute('SELECT * FROM leave_records WHERE id = ?', (leave_id,)).fetchone()
    if not leave or leave['employee_id'] != user['id']:
        flash('Leave not found or unauthorized', 'error')
        conn.close()
        return redirect(url_for('my_leave_report'))

    if leave['status'] in ('retrieved', 'modified'):
        flash('This leave has already been retrieved or modified', 'error')
        conn.close()
        return redirect(url_for('my_leave_report'))

    if request.method == 'POST':
        new_leave_type = request.form.get('leave_type', '').strip()
        new_from_date = request.form.get('from_date', '').strip()
        new_to_date = request.form.get('to_date', '').strip() or new_from_date
        new_day_portion = request.form.get('day_portion', 'full').strip()
        new_last_day_portion = request.form.get('last_day_portion', 'full').strip()
        new_reason = request.form.get('reason', '').strip()
        mod_reason = request.form.get('modification_reason', '').strip()

        if not all([new_leave_type, new_from_date, new_reason, mod_reason]):
            flash('All fields including modification reason are required', 'error')
            conn.close()
            return redirect(url_for('modify_leave', leave_id=leave_id))

        try:
            from_dt = datetime.strptime(new_from_date, '%Y-%m-%d')
            to_dt = datetime.strptime(new_to_date, '%Y-%m-%d')
        except ValueError:
            flash('Invalid date format', 'error')
            conn.close()
            return redirect(url_for('modify_leave', leave_id=leave_id))

        if to_dt < from_dt:
            flash('To date cannot be before from date', 'error')
            conn.close()
            return redirect(url_for('modify_leave', leave_id=leave_id))

        # Check holidays
        holidays_db = {row['holiday_date'] for row in conn.execute('SELECT holiday_date FROM holidays').fetchall()}

        # Determine if this is a late leave modification
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        is_late = 1 if leave.get('is_late') else 0

        # Build working days
        leave_days = []
        current = from_dt
        while current <= to_dt:
            date_str = current.strftime('%Y-%m-%d')
            if current.weekday() < 5 and date_str not in holidays_db:
                leave_days.append(current)
            current += timedelta(days=1)

        if not leave_days:
            flash('No working days in the selected date range', 'error')
            conn.close()
            return redirect(url_for('modify_leave', leave_id=leave_id))

        # Mark original leave as 'modified'
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        original_reason = leave['reason'] or ''
        conn.execute('''
            UPDATE leave_records SET status = 'modified', modification_reason = ? WHERE id = ?
        ''', (mod_reason, leave_id))

        # Insert new leave records (one per working day)
        total_days = 0
        for i, day in enumerate(leave_days):
            date_str = day.strftime('%Y-%m-%d')
            is_single = (len(leave_days) == 1)
            is_first = (i == 0)
            is_last = (i == len(leave_days) - 1)

            if is_single:
                portion = new_day_portion
            elif is_first:
                portion = new_day_portion
            elif is_last:
                portion = new_last_day_portion
            else:
                portion = 'full'

            day_val = 0.5 if portion in ('first_half', 'second_half') else 1.0
            total_days += day_val

            conn.execute('''
                INSERT INTO leave_records (employee_id, leave_type, leave_date, days, day_portion,
                    reason, status, is_late, original_id, original_reason, modification_reason, created_at)
                VALUES (?, ?, ?, ?, ?, ?, 'pending', ?, ?, ?, ?, ?)
            ''', (user['id'], new_leave_type, date_str, day_val, portion,
                  new_reason, is_late, leave_id, original_reason, mod_reason, now_str))

        conn.commit()
        conn.close()

        flash(f'Leave modified ({total_days:.1f} day(s)). Re-approval is required.', 'success')
        return redirect(url_for('my_leave_report'))

    # GET: show modification form
    conn.close()
    today_str = datetime.now().strftime('%Y-%m-%d')
    return render_template('modify_leave.html', user=user, leave=leave, today=today_str)


# ─── Edit Leave (in-place update for future leaves) ───
@app.route('/edit-leave/<int:leave_id>', methods=['GET', 'POST'])
@login_required
def edit_leave(leave_id):
    """Edit a leave record in-place. Allowed for future-dated leaves only.
       Authorised users: the applicant, their reporting manager, or admin."""
    user = get_user()
    conn = get_db()

    leave = conn.execute('''
        SELECT lr.*, e.name, e.emp_code, e.reporting_to
        FROM leave_records lr JOIN employees e ON lr.employee_id = e.id
        WHERE lr.id = ?
    ''', (leave_id,)).fetchone()

    if not leave:
        flash('Leave not found', 'error')
        conn.close()
        return redirect(request.referrer or url_for('dashboard'))

    # Authorization: applicant, reporting manager, or admin
    is_owner = (leave['employee_id'] == user['id'])
    is_reporting = (leave['reporting_to'] == user['id'])
    is_admin_user = bool(user.get('is_admin'))
    if not (is_owner or is_reporting or is_admin_user):
        flash('You are not authorised to edit this leave', 'error')
        conn.close()
        return redirect(request.referrer or url_for('dashboard'))

    # Only future leaves can be edited
    today = datetime.now().strftime('%Y-%m-%d')
    if leave['leave_date'] < today:
        flash('Past leaves cannot be edited', 'error')
        conn.close()
        return redirect(request.referrer or url_for('dashboard'))

    if leave['status'] in ('retrieved', 'modified'):
        flash('This leave has already been retrieved or modified', 'error')
        conn.close()
        return redirect(request.referrer or url_for('dashboard'))

    if request.method == 'POST':
        new_leave_type = request.form.get('leave_type', '').strip()
        new_day_portion = request.form.get('day_portion', 'full').strip()
        new_reason = request.form.get('reason', '').strip()

        if not all([new_leave_type, new_reason]):
            flash('Leave type and reason are required', 'error')
            conn.close()
            return redirect(url_for('edit_leave', leave_id=leave_id))

        day_val = 0.5 if new_day_portion in ('first_half', 'second_half') else 1.0

        conn.execute('''
            UPDATE leave_records SET leave_type = ?, day_portion = ?, days = ?, reason = ?
            WHERE id = ?
        ''', (new_leave_type, new_day_portion, day_val, new_reason, leave_id))
        conn.commit()
        conn.close()

        flash('Leave updated successfully', 'success')
        return redirect(request.referrer or url_for('my_leave_applications'))

    # GET: show edit form
    conn.close()
    return render_template('edit_leave.html', user=user, leave=leave)


@app.route('/approvals')
@login_required
def employee_approvals():
    """Approvals page for managers and management peers"""
    user = get_user()
    conn = get_db()

    # Get pending leave requests from direct reports (late leaves first)
    pending = conn.execute('''
        SELECT lr.*, e.name, e.emp_code, e.department FROM leave_records lr
        JOIN employees e ON lr.employee_id = e.id
        WHERE lr.status = 'pending' AND e.reporting_to = ?
        ORDER BY lr.is_late DESC, lr.created_at DESC
    ''', (user['id'],)).fetchall()
    pending = list(pending)

    # If current user is management, also show pending from other management members
    if user['emp_code'] in MANAGEMENT_CODES:
        mgmt_pending = conn.execute('''
            SELECT lr.*, e.name, e.emp_code, e.department FROM leave_records lr
            JOIN employees e ON lr.employee_id = e.id
            WHERE lr.status = 'pending' AND e.emp_code IN (?, ?, ?)
            AND e.id != ?
            ORDER BY lr.is_late DESC, lr.created_at DESC
        ''', (*MANAGEMENT_CODES, user['id'])).fetchall()
        # Merge, avoiding duplicates
        existing_ids = {l['id'] for l in pending}
        for leave in mgmt_pending:
            if leave['id'] not in existing_ids:
                pending.append(leave)

    # Get recently approved/rejected by this manager
    recent = conn.execute('''
        SELECT lr.*, e.name, e.emp_code, e.department FROM leave_records lr
        JOIN employees e ON lr.employee_id = e.id
        WHERE lr.approved_by = ? AND lr.status IN ('approved', 'rejected')
        ORDER BY lr.approved_at DESC LIMIT 10
    ''', (user['id'],)).fetchall()

    # Get direct reports list (include late leave count)
    direct_reports = conn.execute('''
        SELECT id, name, emp_code, department, designation, late_leave_count FROM employees
        WHERE reporting_to = ? AND is_active = 1 ORDER BY name
    ''', (user['id'],)).fetchall()

    conn.close()

    return render_template('employee_approvals.html',
                         user=user,
                         pending_leaves=pending,
                         recent_actions=recent,
                         direct_reports=direct_reports)

@app.route('/approve/<int:leave_id>', methods=['POST'])
@login_required
def employee_approve_leave(leave_id):
    """Approve leave - for non-admin managers"""
    user = get_user()
    conn = get_db()

    leave = conn.execute('SELECT * FROM leave_records WHERE id = ?', (leave_id,)).fetchone()
    if not leave:
        flash('Leave not found', 'error')
        conn.close()
        return redirect(url_for('employee_approvals'))

    leave_emp = conn.execute('SELECT * FROM employees WHERE id = ?', (leave['employee_id'],)).fetchone()
    allowed, reason = can_approve_leave(user, leave_emp, conn)
    if not allowed:
        flash(reason, 'error')
        conn.close()
        return redirect(url_for('employee_approvals'))

    conn.execute('''
        UPDATE leave_records SET status = 'approved', approved_by = ?, approved_at = ? WHERE id = ?
    ''', (user['id'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'), leave_id))
    # Notify employee
    create_notification(conn, leave['employee_id'],
        'Leave Approved',
        f'Your {leave["leave_type"]} leave on {leave["leave_date"]} has been approved by {user["name"]}.',
        'success', '/my-leaves')
    conn.commit()
    conn.close()

    flash('Leave approved', 'success')
    return redirect(url_for('employee_approvals'))

@app.route('/reject/<int:leave_id>', methods=['POST'])
@login_required
def employee_reject_leave(leave_id):
    """Reject leave - for managers and management peers"""
    user = get_user()
    conn = get_db()

    leave = conn.execute('SELECT * FROM leave_records WHERE id = ?', (leave_id,)).fetchone()
    if not leave:
        flash('Leave not found', 'error')
        conn.close()
        return redirect(url_for('employee_approvals'))

    leave_emp = conn.execute('SELECT * FROM employees WHERE id = ?', (leave['employee_id'],)).fetchone()
    allowed, reason = can_approve_leave(user, leave_emp, conn)
    if not allowed:
        flash(reason, 'error')
        conn.close()
        return redirect(url_for('employee_approvals'))

    conn.execute('''
        UPDATE leave_records SET status = 'rejected', approved_by = ?, approved_at = ? WHERE id = ?
    ''', (user['id'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'), leave_id))
    # Notify employee
    create_notification(conn, leave['employee_id'],
        'Leave Rejected',
        f'Your {leave["leave_type"]} leave on {leave["leave_date"]} has been rejected by {user["name"]}.',
        'danger', '/my-leaves')
    conn.commit()
    conn.close()

    flash('Leave rejected', 'success')
    return redirect(url_for('employee_approvals'))

@app.route('/calendar')
@login_required
def calendar_view():
    user = get_user()
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))

    # Get calendar days
    month_days = calendar.monthcalendar(year, month)
    month_name = calendar.month_name[month]

    # Get holidays and leaves
    holidays = get_all_holidays()
    leaves = get_leaves_for_month(user['id'], year, month)
    leave_dates = {}
    for leave in leaves:
        if leave['leave_date'] not in leave_dates:
            leave_dates[leave['leave_date']] = []
        leave_dates[leave['leave_date']].append(leave)

    # Navigation
    prev_month = month - 1
    prev_year = year
    if prev_month < 1:
        prev_month = 12
        prev_year -= 1

    next_month = month + 1
    next_year = year
    if next_month > 12:
        next_month = 1
        next_year += 1

    today = datetime.now().strftime('%Y-%m-%d')

    return render_template('employee_calendar.html',
                         user=user,
                         year=year,
                         month=month,
                         month_name=month_name,
                         month_days=month_days,
                         holidays=holidays,
                         leave_dates=leave_dates,
                         prev_month=prev_month,
                         prev_year=prev_year,
                         next_month=next_month,
                         next_year=next_year,
                         today=today)

@app.route('/holidays')
@login_required
def employee_holidays():
    user = get_user()
    conn = get_db()
    holidays = conn.execute('SELECT * FROM holidays ORDER BY holiday_date').fetchall()
    conn.close()
    today_date = datetime.now().strftime('%Y-%m-%d')
    return render_template('employee_holidays.html', user=user, holidays=holidays, today_date=today_date)

@app.route('/org-chart')
@login_required
def employee_org_chart():
    user = get_user()
    conn = get_db()
    employees = conn.execute('''
        SELECT e.*, m.name as manager_name, m.photo_url as manager_photo
        FROM employees e
        LEFT JOIN employees m ON e.reporting_to = m.id
        WHERE e.is_active = 1 AND e.emp_code != 'admin'
        ORDER BY e.department, e.name
    ''').fetchall()
    departments = conn.execute('''
        SELECT department, COUNT(*) as count FROM employees
        WHERE is_active = 1 AND emp_code != 'admin'
        GROUP BY department ORDER BY department
    ''').fetchall()
    conn.close()
    return render_template('employee_org_chart.html', user=user, employees=employees, departments=departments)

@app.route('/leave-info')
@login_required
def leave_info():
    """Leave policy and structure information page."""
    user = get_user()
    today = datetime.now()
    fy_year = today.year if today.month >= 4 else today.year - 1

    # Get employee's carry forward
    conn = get_db()
    emp = conn.execute('SELECT carry_forward FROM employees WHERE id = ?', (user['id'],)).fetchone()
    carry_forward = emp['carry_forward'] if emp else 0
    total_allocation = 25 + carry_forward
    conn.close()

    return render_template('leave_info.html',
                         user=user,
                         fy_year=fy_year,
                         my_carry_forward=carry_forward,
                         my_total_allocation=total_allocation)


@app.route('/my-leave-report')
@login_required
def my_leave_report():
    user = get_user()

    # Admin control account has no personal leave report
    if user['emp_code'] == 'admin':
        return redirect(url_for('admin_employee_leave_report'))
    conn = get_db()

    today = datetime.now()
    current_month = today.month
    fy_year = today.year if current_month >= 4 else today.year - 1

    # Total balance and carry forward
    emp = conn.execute('SELECT carry_forward FROM employees WHERE id = ?', (user['id'],)).fetchone()
    carry_forward = emp['carry_forward'] if emp else 0
    total_allocation = 25 + carry_forward

    # Pending requests count
    pending_count = conn.execute(
        'SELECT COUNT(*) as cnt FROM leave_records WHERE employee_id = ? AND status = ?',
        (user['id'], 'pending')
    ).fetchone()['cnt']

    # Month-wise leave report for the full FY with running balance
    monthly_leave_data = []
    running_balance = carry_forward  # Start with carry forward
    for m in range(12):
        report_month = ((m + 3) % 12) + 1
        report_year = fy_year if report_month >= 4 else fy_year + 1
        m_alloc = get_monthly_alloc(report_month)

        month_data = conn.execute('''
            SELECT SUM(days) as total_days,
                   SUM(CASE WHEN leave_type = 'annual' THEN days ELSE 0 END) as annual,
                   SUM(CASE WHEN leave_type = 'sick' THEN days ELSE 0 END) as sick,
                   SUM(CASE WHEN leave_type = 'casual' THEN days ELSE 0 END) as casual,
                   COUNT(*) as count
            FROM leave_records
            WHERE employee_id = ? AND status = 'approved'
            AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
        ''', (user['id'], str(report_year), str(report_month).zfill(2))).fetchone()

        month_total = month_data['total_days'] or 0
        running_balance = round(running_balance + m_alloc - month_total, 2)

        monthly_leave_data.append({
            'month': calendar.month_name[report_month],
            'month_short': calendar.month_abbr[report_month],
            'year': report_year,
            'total': month_total,
            'annual': month_data['annual'] or 0,
            'sick': month_data['sick'] or 0,
            'casual': month_data['casual'] or 0,
            'count': month_data['count'] or 0,
            'monthly_alloc': m_alloc,
            'balance': running_balance
        })

    # All leave records for the FY
    all_leaves = conn.execute('''
        SELECT * FROM leave_records
        WHERE employee_id = ?
        AND ((strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) >= '04')
             OR (strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) < '04'))
        ORDER BY leave_date DESC
    ''', (user['id'], str(fy_year), str(fy_year + 1))).fetchall()

    # Days taken - with type breakdown
    total_taken = sum(m['total'] for m in monthly_leave_data)
    annual_total = sum(m['annual'] for m in monthly_leave_data)
    sick_total = sum(m['sick'] for m in monthly_leave_data)
    casual_total = sum(m['casual'] for m in monthly_leave_data)
    available_balance = total_allocation - total_taken

    conn.close()

    return render_template('employee_leave_report.html',
                         user=user,
                         monthly_leave_data=monthly_leave_data,
                         all_leaves=all_leaves,
                         fy_year=fy_year,
                         total_allocation=total_allocation,
                         total_taken=total_taken,
                         annual_total=annual_total,
                         sick_total=sick_total,
                         casual_total=casual_total,
                         available_balance=round(available_balance, 2),
                         pending_count=pending_count,
                         carry_forward=carry_forward,
                         monthly_alloc='Apr=3, Others=2')

@app.route('/hr-dashboard')
@login_required
def hr_dashboard():
    """HR-specific dashboard - admins see full view, employees see limited view"""
    user = get_user()

    # Redirect non-admin employees to main dashboard
    if not user['is_admin']:
        return redirect(url_for('dashboard'))

    conn = get_db()

    # Summary stats
    total_employees = conn.execute("SELECT COUNT(*) as count FROM employees WHERE is_active = 1 AND emp_code != 'admin'").fetchone()

    today = datetime.now().strftime('%Y-%m-%d')
    leaves_today = conn.execute('''
        SELECT COUNT(DISTINCT employee_id) as count FROM leave_records
        WHERE leave_date = ? AND status = 'approved'
    ''', (today,)).fetchone()

    pending = conn.execute("SELECT COUNT(*) as count FROM leave_records WHERE status = 'pending'").fetchone()

    # Department-wise count (include management as employees, exclude admin login)
    departments = conn.execute('''
        SELECT department, COUNT(*) as count FROM employees
        WHERE is_active = 1 AND emp_code != 'admin'
        GROUP BY department ORDER BY department
    ''').fetchall()

    # Who's on leave today
    on_leave_today = conn.execute('''
        SELECT e.name, e.photo_url, e.department, lr.day_portion, lr.leave_type
        FROM leave_records lr JOIN employees e ON lr.employee_id = e.id
        WHERE lr.leave_date = ? AND lr.status = 'approved'
        ORDER BY e.name
    ''', (today,)).fetchall()

    # Recent leave applications
    recent = conn.execute('''
        SELECT lr.*, e.name, e.emp_code, e.photo_url FROM leave_records lr
        JOIN employees e ON lr.employee_id = e.id
        ORDER BY lr.created_at DESC LIMIT 10
    ''').fetchall()

    # Total departments
    total_depts = len(departments)

    # Upcoming holidays this month
    now = datetime.now()
    month_start = now.strftime('%Y-%m-01')
    month_end = now.strftime('%Y-%m-') + str(calendar.monthrange(now.year, now.month)[1])
    upcoming_holidays = conn.execute('''
        SELECT * FROM holidays
        WHERE holiday_date >= ? AND holiday_date <= ?
        ORDER BY holiday_date
    ''', (now.strftime('%Y-%m-%d'), month_end)).fetchall()

    # Birthdays this month
    current_mm = str(now.month).zfill(2)
    birthdays_this_month = conn.execute('''
        SELECT name, dob, photo_url, department FROM employees
        WHERE is_active = 1 AND emp_code != 'admin' AND dob IS NOT NULL AND dob != ''
        AND strftime('%m', dob) = ?
        ORDER BY strftime('%d', dob)
    ''', (current_mm,)).fetchall()

    # Work anniversaries this month
    anniversaries_this_month = conn.execute('''
        SELECT name, joining_date, photo_url, department FROM employees
        WHERE is_active = 1 AND emp_code != 'admin' AND joining_date IS NOT NULL AND joining_date != ''
        AND strftime('%m', joining_date) = ?
        AND strftime('%Y', joining_date) != ?
        ORDER BY strftime('%d', joining_date)
    ''', (current_mm, str(now.year))).fetchall()

    # Recent announcements
    announcements = conn.execute('''
        SELECT a.*, e.name as posted_by_name FROM announcements a
        JOIN employees e ON a.posted_by = e.id
        WHERE a.is_active = 1
        ORDER BY a.created_at DESC LIMIT 5
    ''', ()).fetchall()

    # Management user leave balance (for welcome banner)
    is_management = user['emp_code'] in MANAGEMENT_CODES
    mgmt_leave_data = {}
    if is_management:
        fy_year = now.year if now.month >= 4 else now.year - 1
        carry_forward = user['carry_forward'] or 0
        total_allocation = 25 + carry_forward
        leaves_taken = conn.execute('''
            SELECT SUM(days) as total_days FROM leave_records
            WHERE employee_id = ? AND status = 'approved'
            AND ((strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) >= '04')
                 OR (strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) < '04'))
        ''', (user['id'], str(fy_year), str(fy_year + 1))).fetchone()
        days_taken = leaves_taken['total_days'] if leaves_taken['total_days'] else 0
        my_pending = conn.execute(
            "SELECT COUNT(*) as cnt FROM leave_records WHERE employee_id = ? AND status = 'pending'",
            (user['id'],)
        ).fetchone()['cnt']
        mgmt_leave_data = {
            'total_allocation': total_allocation,
            'available_balance': round(total_allocation - days_taken, 2),
            'days_taken': days_taken,
            'pending_count': my_pending
        }

    conn.close()

    return render_template('admin_dashboard.html',
                         user=user,
                         total_employees=total_employees['count'],
                         leaves_today=leaves_today['count'],
                         pending_count=pending['count'],
                         departments=departments,
                         total_departments=total_depts,
                         on_leave_today=on_leave_today,
                         recent_leaves=recent,
                         upcoming_holidays=upcoming_holidays,
                         current_month_name=calendar.month_name[now.month],
                         current_year=now.year,
                         birthdays_this_month=birthdays_this_month,
                         anniversaries_this_month=anniversaries_this_month,
                         announcements=announcements,
                         can_announce=can_post_announcements(user),
                         is_management=is_management,
                         mgmt_leave_data=mgmt_leave_data)

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    """Redirect to the new HR dashboard"""
    return redirect(url_for('hr_dashboard'))

@app.route('/admin/org-chart')
@admin_required
def org_chart():
    user = get_user()
    conn = get_db()
    # Get admin user (top of hierarchy)
    admin_user = conn.execute('SELECT * FROM employees WHERE is_admin = 1 LIMIT 1').fetchone()
    employees = conn.execute('''
        SELECT e.*, m.name as manager_name, m.photo_url as manager_photo
        FROM employees e
        LEFT JOIN employees m ON e.reporting_to = m.id
        WHERE e.is_active = 1 AND e.emp_code != 'admin'
        ORDER BY e.department, e.name
    ''').fetchall()
    departments = conn.execute('''
        SELECT department, COUNT(*) as count FROM employees
        WHERE is_active = 1 AND emp_code != 'admin'
        GROUP BY department ORDER BY department
    ''').fetchall()
    conn.close()
    return render_template('org_chart.html', user=user, admin_user=admin_user, employees=employees, departments=departments)

@app.route('/admin/calendar')
@admin_required
def admin_calendar():
    user = get_user()
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))

    # Get calendar days
    month_days = calendar.monthcalendar(year, month)
    month_name = calendar.month_name[month]

    # Get all employees
    conn = get_db()
    employees = conn.execute("SELECT id, name FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name").fetchall()

    # Get all leaves for the month
    all_leaves = conn.execute('''
        SELECT lr.*, e.name, e.emp_code FROM leave_records lr
        JOIN employees e ON lr.employee_id = e.id
        WHERE lr.status = 'approved'
        AND strftime('%Y', lr.leave_date) = ? AND strftime('%m', lr.leave_date) = ?
        ORDER BY lr.leave_date, e.name
    ''', (str(year), str(month).zfill(2))).fetchall()

    conn.close()

    # Organize leaves by date
    leave_dates = {}
    for leave in all_leaves:
        if leave['leave_date'] not in leave_dates:
            leave_dates[leave['leave_date']] = []
        leave_dates[leave['leave_date']].append(leave)

    # Get holidays
    holidays = get_all_holidays()

    # Navigation
    prev_month = month - 1
    prev_year = year
    if prev_month < 1:
        prev_month = 12
        prev_year -= 1

    next_month = month + 1
    next_year = year
    if next_month > 12:
        next_month = 1
        next_year += 1

    today = datetime.now().strftime('%Y-%m-%d')

    return render_template('admin_calendar.html',
                         user=user,
                         year=year,
                         month=month,
                         month_name=month_name,
                         month_days=month_days,
                         holidays=holidays,
                         leave_dates=leave_dates,
                         employees=employees,
                         prev_month=prev_month,
                         prev_year=prev_year,
                         next_month=next_month,
                         next_year=next_year,
                         today=today)

@app.route('/admin/employee/<int:emp_id>')
@admin_required
def admin_employee_detail(emp_id):
    user = get_user()
    conn = get_db()

    employee = conn.execute('SELECT * FROM employees WHERE id = ?', (emp_id,)).fetchone()
    if not employee:
        flash('Employee not found', 'error')
        conn.close()
        return redirect(url_for('admin_dashboard'))

    leaves = conn.execute('''
        SELECT * FROM leave_records
        WHERE employee_id = ?
        ORDER BY leave_date DESC
    ''', (emp_id,)).fetchall()

    # Calculate balance
    emp_data = conn.execute('SELECT carry_forward FROM employees WHERE id = ?', (emp_id,)).fetchone()
    carry_forward = emp_data['carry_forward'] if emp_data else 0
    total_allocation = 25 + carry_forward

    approved_leaves = conn.execute('''
        SELECT SUM(days) as total_days FROM leave_records
        WHERE employee_id = ? AND status = 'approved'
    ''', (emp_id,)).fetchone()

    days_taken = approved_leaves['total_days'] if approved_leaves['total_days'] else 0
    available = total_allocation - days_taken

    conn.close()

    return render_template('admin_employee_detail.html',
                         user=user,
                         employee=employee,
                         leaves=leaves,
                         total_allocation=total_allocation,
                         days_taken=days_taken,
                         available_balance=max(0, available))

@app.route('/admin/employee/<int:emp_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_employee_edit(emp_id):
    user = get_user()
    conn = get_db()

    employee = conn.execute('SELECT * FROM employees WHERE id = ?', (emp_id,)).fetchone()
    if not employee:
        flash('Employee not found', 'error')
        conn.close()
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        dob = request.form.get('dob', '').strip()
        address = request.form.get('address', '').strip()
        department = request.form.get('department', '').strip()
        designation = request.form.get('designation', '').strip()
        joining_date = request.form.get('joining_date', '').strip()
        carry_forward = request.form.get('carry_forward', '0').strip()

        try:
            carry_forward = float(carry_forward)
        except ValueError:
            carry_forward = 0

        reporting_to = request.form.get('reporting_to', '').strip()
        reporting_to = int(reporting_to) if reporting_to else None

        emergency_contact_name = request.form.get('emergency_contact_name', '').strip()
        emergency_contact_phone = request.form.get('emergency_contact_phone', '').strip()
        emergency_contact_relation = request.form.get('emergency_contact_relation', '').strip()

        conn.execute('''
            UPDATE employees
            SET name = ?, email = ?, phone = ?, dob = ?, address = ?, department = ?,
                designation = ?, joining_date = ?, carry_forward = ?, reporting_to = ?,
                emergency_contact_name = ?, emergency_contact_phone = ?, emergency_contact_relation = ?
            WHERE id = ?
        ''', (name, email, phone, dob, address, department, designation, joining_date, carry_forward,
              reporting_to, emergency_contact_name, emergency_contact_phone, emergency_contact_relation, emp_id))
        conn.commit()
        conn.close()

        flash('Employee updated successfully', 'success')
        return redirect(url_for('admin_employee_detail', emp_id=emp_id))

    # Get all active employees as potential managers (exclude the employee being edited)
    managers = conn.execute('SELECT id, name, emp_code FROM employees WHERE is_active = 1 AND id != ? ORDER BY name', (emp_id,)).fetchall()

    conn.close()

    departments = ['Sales', 'Operations', 'Marketing', 'Admin', 'Senior Management', 'HR', 'Management']

    return render_template('admin_employee_profile.html',
                         user=user,
                         employee=employee,
                         departments=departments,
                         managers=managers)

@app.route('/admin/employee/<int:emp_id>/upload-photo', methods=['POST'])
@admin_required
def admin_upload_photo(emp_id):
    conn = get_db()
    employee = conn.execute('SELECT * FROM employees WHERE id = ?', (emp_id,)).fetchone()
    if not employee:
        flash('Employee not found', 'error')
        conn.close()
        return redirect(url_for('admin_dashboard'))

    if 'photo' not in request.files:
        flash('No photo selected', 'error')
        conn.close()
        return redirect(url_for('admin_employee_edit', emp_id=emp_id))

    file = request.files['photo']
    if file.filename == '':
        flash('No photo selected', 'error')
        conn.close()
        return redirect(url_for('admin_employee_edit', emp_id=emp_id))

    if file and allowed_file(file.filename):
        ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"{employee['emp_code']}.{ext}"
        filepath = os.path.join(PHOTO_FOLDER, filename)

        # Remove old photos with different extensions
        for old_ext in ALLOWED_EXTENSIONS:
            old_path = os.path.join(PHOTO_FOLDER, f"{employee['emp_code']}.{old_ext}")
            if os.path.exists(old_path) and old_path != filepath:
                os.remove(old_path)

        file.save(filepath)

        # Update photo_url in database
        conn.execute('UPDATE employees SET photo_url = ? WHERE id = ?', (filename, emp_id))
        conn.commit()
        conn.close()

        flash('Photo uploaded successfully', 'success')
        return redirect(url_for('admin_employee_edit', emp_id=emp_id))
    else:
        flash('Invalid file type. Allowed: png, jpg, jpeg, gif, webp', 'error')
        conn.close()
        return redirect(url_for('admin_employee_edit', emp_id=emp_id))

@app.route('/admin/add-leave', methods=['GET', 'POST'])
@admin_required
def admin_add_leave():
    user = get_user()
    conn = get_db()

    employees = conn.execute("SELECT id, name FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name").fetchall()

    if request.method == 'POST':
        employee_id = request.form.get('employee_id', '').strip()
        leave_type = request.form.get('leave_type', '').strip()
        leave_date = request.form.get('leave_date', '').strip()
        day_portion = request.form.get('day_portion', 'full').strip()
        reason = request.form.get('reason', '').strip()

        if not employee_id or not leave_type or not leave_date:
            flash('Required fields missing', 'error')
            return render_template('admin_add_leave.html', user=user, employees=employees)

        try:
            employee_id = int(employee_id)
        except ValueError:
            flash('Invalid input', 'error')
            return render_template('admin_add_leave.html', user=user, employees=employees)

        if leave_type not in ['annual', 'sick', 'casual']:
            flash('Invalid leave type', 'error')
            return render_template('admin_add_leave.html', user=user, employees=employees)

        if day_portion not in ['full', 'first_half', 'second_half']:
            flash('Invalid duration', 'error')
            return render_template('admin_add_leave.html', user=user, employees=employees)

        days = 0.5 if day_portion in ['first_half', 'second_half'] else 1.0

        conn.execute('''
            INSERT INTO leave_records (employee_id, leave_type, leave_date, days, day_portion, reason, status, approved_by, approved_at, created_at)
            VALUES (?, ?, ?, ?, ?, ?, 'approved', ?, ?, ?)
        ''', (employee_id, leave_type, leave_date, days, day_portion, reason or '', user['id'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'), datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        conn.commit()
        conn.close()

        flash('Leave added successfully', 'success')
        return redirect(url_for('admin_dashboard'))

    conn.close()

    return render_template('admin_add_leave.html', user=user, employees=employees)

@app.route('/admin/bulk-leave', methods=['GET', 'POST'])
@admin_required
def admin_bulk_leave():
    user = get_user()
    conn = get_db()

    employees = conn.execute("SELECT id, name FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name").fetchall()

    if request.method == 'POST':
        department = request.form.get('department', '').strip()
        leave_type = request.form.get('leave_type', '').strip()
        leave_date = request.form.get('leave_date', '').strip()
        days = request.form.get('days', '').strip()
        reason = request.form.get('reason', '').strip()

        if not department or not leave_type or not leave_date or not days:
            flash('All fields required', 'error')
            return render_template('admin_bulk_leave.html', user=user, employees=employees)

        try:
            days = float(days)
            if days <= 0:
                flash('Days must be positive', 'error')
                return render_template('admin_bulk_leave.html', user=user, employees=employees)
        except ValueError:
            flash('Invalid days value', 'error')
            return render_template('admin_bulk_leave.html', user=user, employees=employees)

        # Get all employees in department
        dept_employees = conn.execute('SELECT id FROM employees WHERE department = ? AND is_active = 1', (department,)).fetchall()

        count = 0
        for emp in dept_employees:
            conn.execute('''
                INSERT INTO leave_records (employee_id, leave_type, leave_date, days, reason, status, approved_by, approved_at, created_at)
                VALUES (?, ?, ?, ?, ?, 'approved', ?, ?, ?)
            ''', (emp['id'], leave_type, leave_date, days, reason or '', user['id'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'), datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            count += 1

        conn.commit()
        conn.close()

        flash(f'Added leave for {count} employees', 'success')
        return redirect(url_for('admin_dashboard'))

    conn.close()

    departments = ['Sales', 'Operations', 'Marketing', 'Admin', 'Senior Management', 'HR', 'Management']

    return render_template('admin_bulk_leave.html', user=user, employees=employees, departments=departments)

@app.route('/admin/manage-employees', methods=['GET', 'POST'])
@admin_required
def manage_employees():
    user = get_user()
    conn = get_db()

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        emp_code = request.form.get('emp_code', '').strip()
        email = request.form.get('email', '').strip()
        department = request.form.get('department', '').strip()

        if not name or not emp_code or not department:
            flash('Required fields missing', 'error')
            return redirect(url_for('manage_employees'))

        # Check if emp_code already exists
        existing = conn.execute('SELECT id FROM employees WHERE emp_code = ?', (emp_code,)).fetchone()
        if existing:
            flash('Employee code already exists', 'error')
            return redirect(url_for('manage_employees'))

        conn.execute('''
            INSERT INTO employees (name, emp_code, password, department, is_active, joining_date, created_at)
            VALUES (?, ?, ?, ?, 1, ?, ?)
        ''', (name, emp_code, hash_password(name.split()[0].lower()), department, datetime.now().strftime('%Y-%m-%d'), datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        conn.commit()

        flash('Employee added successfully', 'success')
        return redirect(url_for('manage_employees'))

    employees = conn.execute("SELECT * FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY department, name").fetchall()

    # Group employees by department
    dept_employees = {}
    for emp in employees:
        dept = emp['department'] or 'Unassigned'
        if dept not in dept_employees:
            dept_employees[dept] = []
        dept_employees[dept].append(emp)

    conn.close()

    departments = ['Sales', 'Operations', 'Marketing', 'Admin', 'Senior Management', 'HR', 'Management']

    return render_template('manage_employees.html',
                         user=user,
                         employees=employees,
                         dept_employees=dept_employees,
                         departments=departments)

@app.route('/admin/delete-employee/<int:emp_id>', methods=['POST'])
@admin_required
def delete_employee(emp_id):
    conn = get_db()

    employee = conn.execute('SELECT * FROM employees WHERE id = ?', (emp_id,)).fetchone()
    if not employee:
        flash('Employee not found', 'error')
        conn.close()
        return redirect(url_for('manage_employees'))

    # Delete all leave records for this employee
    conn.execute('DELETE FROM leave_records WHERE employee_id = ?', (emp_id,))

    # Delete employee
    conn.execute('DELETE FROM employees WHERE id = ?', (emp_id,))
    conn.commit()
    conn.close()

    flash(f'Employee {employee["name"]} and all their records deleted', 'success')
    return redirect(url_for('manage_employees'))


@app.route('/admin/reset-passwords', methods=['GET', 'POST'])
@admin_required
def admin_reset_passwords():
    """Admin-only password reset for any user. Restricted to emp_code='admin'."""
    user = get_user()
    # Extra check: only the admin account, not management
    if user['emp_code'] != 'admin':
        flash('Only the admin account can reset passwords', 'error')
        return redirect(url_for('dashboard'))

    conn = get_db()

    if request.method == 'POST':
        emp_id = request.form.get('emp_id', type=int)
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not emp_id or not new_password:
            flash('Please select an employee and enter a new password', 'error')
        elif len(new_password) < 4:
            flash('Password must be at least 4 characters', 'error')
        elif new_password != confirm_password:
            flash('Passwords do not match', 'error')
        else:
            emp = conn.execute('SELECT id, name, emp_code FROM employees WHERE id = ?', (emp_id,)).fetchone()
            if not emp:
                flash('Employee not found', 'error')
            else:
                conn.execute('UPDATE employees SET password = ? WHERE id = ?',
                             (hash_password(new_password), emp_id))
                conn.commit()
                flash(f'Password reset successfully for {emp["name"]} ({emp["emp_code"]})', 'success')

    employees = conn.execute(
        "SELECT id, name, emp_code, department, is_active FROM employees WHERE emp_code != 'admin' ORDER BY is_active DESC, department, name"
    ).fetchall()
    conn.close()

    return render_template('admin_reset_passwords.html', user=user, employees=employees)


@app.route('/admin/delete-leave/<int:leave_id>', methods=['POST'])
@admin_required
def delete_leave(leave_id):
    conn = get_db()

    leave = conn.execute('SELECT * FROM leave_records WHERE id = ?', (leave_id,)).fetchone()
    if not leave:
        flash('Leave not found', 'error')
        conn.close()
        return redirect(url_for('admin_dashboard'))

    conn.execute('DELETE FROM leave_records WHERE id = ?', (leave_id,))
    conn.commit()
    conn.close()

    flash('Leave record deleted', 'success')
    return redirect(request.referrer or url_for('admin_dashboard'))

@app.route('/admin/monthly-report')
@admin_required
def admin_monthly_report():
    user = get_user()
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))

    conn = get_db()
    employees = conn.execute("SELECT * FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name").fetchall()

    report_data = []
    m_alloc = get_monthly_alloc(month)
    for emp in employees:
        carry_forward = emp['carry_forward']
        total_allocation = 25 + carry_forward

        # Get leaves taken in this month
        month_leaves = conn.execute('''
            SELECT SUM(days) as total_days FROM leave_records
            WHERE employee_id = ? AND status = 'approved'
            AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
        ''', (emp['id'], str(year), str(month).zfill(2))).fetchone()

        days_taken_month = month_leaves['total_days'] if month_leaves['total_days'] else 0

        # Get balance at start of month
        balance_start = get_available_balance(emp['id'], year, month)
        balance_available = balance_start + m_alloc

        # Calculate deduction
        deduction = max(0, days_taken_month - balance_available)

        report_data.append({
            'name': emp['name'],
            'emp_code': emp['emp_code'],
            'department': emp['department'],
            'monthly_allocation': m_alloc,
            'balance_start': round(balance_start, 2),
            'balance_available': round(balance_available, 2),
            'days_taken': round(days_taken_month, 2),
            'deduction': round(deduction, 2)
        })

    conn.close()

    month_name = calendar.month_name[month]

    return render_template('admin_monthly_report.html',
                         user=user,
                         year=year,
                         month=month,
                         month_name=month_name,
                         report_data=report_data)

@app.route('/admin/monthly-report/download')
@admin_required
def admin_monthly_report_download():
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    file_format = request.args.get('format', 'xlsx')

    conn = get_db()
    employees = conn.execute("SELECT * FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name").fetchall()

    if file_format == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = f"Leave Report {month}-{year}"

        # Headers
        headers = ['Name', 'Employee Code', 'Department', 'Monthly Allocation', 'Balance Start', 'Balance Available', 'Days Taken', 'Salary Deduction']
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data
        dl_m_alloc = get_monthly_alloc(month)
        for emp in employees:
            carry_forward = emp['carry_forward']
            total_allocation = 25 + carry_forward

            month_leaves = conn.execute('''
                SELECT SUM(days) as total_days FROM leave_records
                WHERE employee_id = ? AND status = 'approved'
                AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
            ''', (emp['id'], str(year), str(month).zfill(2))).fetchone()

            days_taken_month = month_leaves['total_days'] if month_leaves['total_days'] else 0
            balance_start = get_available_balance(emp['id'], year, month)
            balance_available = balance_start + dl_m_alloc
            deduction = max(0, days_taken_month - balance_available)

            ws.append([
                emp['name'],
                emp['emp_code'],
                emp['department'],
                dl_m_alloc,
                round(balance_start, 2),
                round(balance_available, 2),
                round(days_taken_month, 2),
                round(deduction, 2)
            ])

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 17
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 17

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        conn.close()

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Leave_Report_{month}_{year}.xlsx'
        )

    conn.close()
    flash('Invalid format', 'error')
    return redirect(url_for('admin_monthly_report', year=year, month=month))

@app.route('/admin/holidays', methods=['GET', 'POST'])
@admin_required
def admin_holidays():
    user = get_user()
    conn = get_db()

    if request.method == 'POST':
        holiday_date = request.form.get('holiday_date', '').strip()
        name = request.form.get('name', '').strip()
        holiday_type = request.form.get('holiday_type', '').strip()

        if not holiday_date or not name:
            flash('Required fields missing', 'error')
            return redirect(url_for('admin_holidays'))

        # Check if holiday already exists
        existing = conn.execute('SELECT id FROM holidays WHERE holiday_date = ?', (holiday_date,)).fetchone()
        if existing:
            flash('Holiday already exists for this date', 'error')
            return redirect(url_for('admin_holidays'))

        conn.execute('''
            INSERT INTO holidays (holiday_date, name, holiday_type, created_at)
            VALUES (?, ?, ?, ?)
        ''', (holiday_date, name, holiday_type, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        conn.commit()

        flash('Holiday added successfully', 'success')
        return redirect(url_for('admin_holidays'))

    holidays = conn.execute('SELECT * FROM holidays ORDER BY holiday_date').fetchall()
    conn.close()

    return render_template('admin_holidays.html', user=user, holidays=holidays)

@app.route('/admin/delete-holiday/<int:holiday_id>', methods=['POST'])
@admin_required
def delete_holiday(holiday_id):
    conn = get_db()

    conn.execute('DELETE FROM holidays WHERE id = ?', (holiday_id,))
    conn.commit()
    conn.close()

    flash('Holiday deleted', 'success')
    return redirect(url_for('admin_holidays'))

@app.route('/admin/pending-approvals')
@login_required
def pending_approvals():
    user = get_user()
    conn = get_db()

    # Filters
    f_employee = (request.args.get('employee') or '').strip()
    f_department = (request.args.get('department') or '').strip()
    f_type = (request.args.get('leave_type') or '').strip()

    base_select = '''
        SELECT lr.id, lr.employee_id, lr.leave_type, lr.leave_date, lr.days,
               lr.day_portion, lr.reason, lr.status, lr.is_late,
               lr.original_id, lr.modification_reason, lr.original_reason,
               lr.approved_by, lr.approved_at, lr.created_at,
               e.name        AS employee_name,
               e.emp_code    AS emp_code,
               e.department  AS department,
               e.designation AS designation,
               e.email       AS email,
               e.phone       AS mobile,
               e.photo_url   AS photo_url,
               e.late_leave_count AS late_leave_count
        FROM leave_records lr
        JOIN employees e ON lr.employee_id = e.id
    '''
    where = ["lr.status = 'pending'"]
    params = []

    # If admin, show all pending; else only direct reports
    if not user['is_admin']:
        where.append('e.reporting_to = ?')
        params.append(user['id'])

    if f_employee:
        where.append('(LOWER(e.name) LIKE ? OR LOWER(e.emp_code) LIKE ?)')
        like = f'%{f_employee.lower()}%'
        params.extend([like, like])
    if f_department:
        where.append('LOWER(e.department) LIKE ?')
        params.append(f'%{f_department.lower()}%')
    if f_type:
        where.append('lr.leave_type = ?')
        params.append(f_type)

    sql = base_select + ' WHERE ' + ' AND '.join(where) + ' ORDER BY lr.is_late DESC, lr.created_at DESC'
    pending_rows = conn.execute(sql, tuple(params)).fetchall()

    # Enrich rows: photo_src + parsed reason fields
    pending = []
    for r in pending_rows:
        d = dict(r)
        p = d.get('photo_url') or ''
        d['photo_src'] = (p if p.startswith('http') else f'/static/photos/{p}') if p else ''
        # Split out late reason if embedded
        raw = d.get('reason') or ''
        if '[Late Application Reason]:' in raw:
            parts = raw.split('\n\n[Late Application Reason]:', 1)
            d['reason_main'] = parts[0].strip()
            d['late_reason'] = parts[1].strip() if len(parts) > 1 else ''
        else:
            d['reason_main'] = raw
            d['late_reason'] = ''
        pending.append(d)

    # Late leave summary
    if user['is_admin']:
        late_summary = conn.execute('''
            SELECT name, emp_code, late_leave_count FROM employees
            WHERE late_leave_count > 0 AND is_active = 1 AND emp_code != 'admin'
            ORDER BY late_leave_count DESC
        ''').fetchall()
    else:
        late_summary = conn.execute('''
            SELECT name, emp_code, late_leave_count FROM employees
            WHERE late_leave_count > 0 AND reporting_to = ? AND is_active = 1
            ORDER BY late_leave_count DESC
        ''', (user['id'],)).fetchall()

    total_count = len(pending)
    late_count = sum(1 for p in pending if p.get('is_late'))
    reapproval_count = sum(1 for p in pending if p.get('original_id'))

    conn.close()

    return render_template('pending_approvals.html', user=user, pending_leaves=pending,
                         late_leave_summary=late_summary,
                         f_employee=f_employee, f_department=f_department, f_type=f_type,
                         total_count=total_count, late_count=late_count,
                         reapproval_count=reapproval_count)

@app.route('/admin/approve-leave/<int:leave_id>', methods=['POST'])
@login_required
def approve_leave(leave_id):
    user = get_user()
    conn = get_db()

    leave = conn.execute('SELECT * FROM leave_records WHERE id = ?', (leave_id,)).fetchone()
    if not leave:
        flash('Leave not found', 'error')
        conn.close()
        return redirect(url_for('pending_approvals'))

    leave_emp = conn.execute('SELECT * FROM employees WHERE id = ?', (leave['employee_id'],)).fetchone()
    allowed, reason = can_approve_leave(user, leave_emp, conn)
    if not allowed:
        flash(reason, 'error')
        conn.close()
        return redirect(url_for('pending_approvals'))

    conn.execute('''
        UPDATE leave_records
        SET status = 'approved', approved_by = ?, approved_at = ?
        WHERE id = ?
    ''', (user['id'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'), leave_id))
    # Notify employee
    create_notification(conn, leave['employee_id'],
        'Leave Approved',
        f'Your {leave["leave_type"]} leave on {leave["leave_date"]} has been approved by {user["name"]}.',
        'success', '/my-leaves')
    conn.commit()
    conn.close()

    flash('Leave approved', 'success')
    return redirect(request.referrer or url_for('pending_approvals'))

@app.route('/admin/reject-leave/<int:leave_id>', methods=['POST'])
@login_required
def reject_leave(leave_id):
    user = get_user()
    conn = get_db()

    leave = conn.execute('SELECT * FROM leave_records WHERE id = ?', (leave_id,)).fetchone()
    if not leave:
        flash('Leave not found', 'error')
        conn.close()
        return redirect(url_for('pending_approvals'))

    leave_emp = conn.execute('SELECT * FROM employees WHERE id = ?', (leave['employee_id'],)).fetchone()
    allowed, reason = can_approve_leave(user, leave_emp, conn)
    if not allowed:
        flash(reason, 'error')
        conn.close()
        return redirect(url_for('pending_approvals'))

    conn.execute('''
        UPDATE leave_records
        SET status = 'rejected', approved_by = ?, approved_at = ?
        WHERE id = ?
    ''', (user['id'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'), leave_id))
    # Notify employee
    create_notification(conn, leave['employee_id'],
        'Leave Rejected',
        f'Your {leave["leave_type"]} leave on {leave["leave_date"]} has been rejected by {user["name"]}.',
        'danger', '/my-leaves')
    conn.commit()
    conn.close()

    flash('Leave rejected', 'success')
    return redirect(request.referrer or url_for('pending_approvals'))

@app.route('/api/balance')
@login_required
def api_balance():
    emp_id = request.args.get('emp_id', type=int)
    if not emp_id:
        return jsonify({'error': 'emp_id required'}), 400

    user = get_user()
    if not user['is_admin'] and user['id'] != emp_id:
        return jsonify({'error': 'Unauthorized'}), 403

    conn = get_db()
    emp = conn.execute('SELECT carry_forward FROM employees WHERE id = ?', (emp_id,)).fetchone()

    if not emp:
        conn.close()
        return jsonify({'error': 'Employee not found'}), 404

    carry_forward = emp['carry_forward']
    total_allocation = 25 + carry_forward

    approved_leaves = conn.execute('''
        SELECT SUM(days) as total_days FROM leave_records
        WHERE employee_id = ? AND status = 'approved'
    ''', (emp_id,)).fetchone()

    days_taken = approved_leaves['total_days'] if approved_leaves['total_days'] else 0
    available = total_allocation - days_taken

    conn.close()

    return jsonify({
        'total_allocation': total_allocation,
        'days_taken': days_taken,
        'available_balance': max(0, available),
        'monthly_allocation': 'Apr=3, Others=2'
    })

@app.route('/api/employee-name')
@login_required
def api_employee_name():
    emp_id = request.args.get('emp_id', type=int)
    if not emp_id:
        return jsonify({'error': 'emp_id required'}), 400

    conn = get_db()
    emp = conn.execute('SELECT name FROM employees WHERE id = ?', (emp_id,)).fetchone()
    conn.close()

    if not emp:
        return jsonify({'error': 'Employee not found'}), 404

    return jsonify({'name': emp['name']})

# ===== NOTIFICATION HELPERS =====

def get_unread_count(user_id, is_admin=False):
    """Get unread notification count for a user."""
    conn = get_db()
    count = conn.execute(
        'SELECT COUNT(*) as cnt FROM notifications WHERE employee_id = ? AND is_read = 0',
        (user_id,)
    ).fetchone()
    conn.close()
    return count['cnt'] if count else 0


@app.context_processor
def inject_notification_count():
    """Make unread notification count available in all templates."""
    if 'user_id' in session:
        is_admin = session.get('is_admin', False)
        count = get_unread_count(session['user_id'], is_admin)
        return {'unread_notification_count': count}
    return {'unread_notification_count': 0}


@app.route('/api/notifications')
@login_required
def get_notifications():
    """Get notifications for the current user."""
    user = get_user()
    conn = get_db()
    notifications = conn.execute('''
        SELECT * FROM notifications
        WHERE employee_id = ?
        ORDER BY created_at DESC LIMIT 30
    ''', (user['id'],)).fetchall()
    conn.close()

    result = []
    for n in notifications:
        result.append({
            'id': n['id'],
            'type': n['type'],
            'title': n['title'],
            'message': n['message'],
            'is_read': n['is_read'],
            'created_at': str(n['created_at'])
        })

    return jsonify(result)


@app.route('/api/notifications/mark-read', methods=['POST'])
@login_required
def mark_notifications_read():
    """Mark all notifications as read for the current user."""
    user = get_user()
    conn = get_db()
    conn.execute(
        'UPDATE notifications SET is_read = 1 WHERE is_read = 0 AND employee_id = ?',
        (user['id'],)
    )
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/notifications/<int:notif_id>/read', methods=['POST'])
@login_required
def mark_single_notification_read(notif_id):
    """Mark a single notification as read."""
    conn = get_db()
    conn.execute('UPDATE notifications SET is_read = 1 WHERE id = ?', (notif_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/dismiss-welcome', methods=['POST'])
@login_required
def dismiss_welcome():
    """Dismiss the first-login welcome popup."""
    session.pop('show_welcome', None)
    return jsonify({'status': 'ok'})


@app.route('/announcements')
@login_required
def announcements():
    user = get_user()
    if not can_post_announcements(user):
        flash('You do not have permission to manage announcements', 'error')
        return redirect(url_for('dashboard'))

    conn = get_db()
    all_announcements = conn.execute('''
        SELECT a.*, e.name as posted_by_name FROM announcements a
        JOIN employees e ON a.posted_by = e.id
        ORDER BY a.created_at DESC
    ''', ()).fetchall()
    conn.close()

    return render_template('announcements.html', user=user, announcements=all_announcements)


@app.route('/announcements/create', methods=['POST'])
@login_required
def create_announcement():
    user = get_user()
    if not can_post_announcements(user):
        flash('You do not have permission to post announcements', 'error')
        return redirect(url_for('dashboard'))

    title = request.form.get('title', '').strip()
    message = request.form.get('message', '').strip()

    if not title or not message:
        flash('Title and message are required', 'error')
        return redirect(url_for('announcements'))

    conn = get_db()
    conn.execute('''
        INSERT INTO announcements (title, message, posted_by, is_active, created_at)
        VALUES (?, ?, ?, 1, ?)
    ''', (title, message, user['id'], datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit()

    # Create notification for all active employees
    try:
        all_emps = conn.execute("SELECT id FROM employees WHERE is_active = 1 AND emp_code != 'admin'", ()).fetchall()
        for emp in all_emps:
            create_notification(conn, emp['id'], f"Announcement: {title}",
                              f"New announcement by {user['name']}: {message[:100]}",
                              'announcement', '/notifications')
        conn.commit()
    except Exception as e:
        logging.error(f"Failed to create announcement notification: {e}")

    # Send email to all active employees with email addresses
    recipients = conn.execute('''
        SELECT email FROM employees WHERE is_active = 1 AND email IS NOT NULL AND email != ''
    ''', ()).fetchall()
    conn.close()

    recipient_emails = [r['email'] for r in recipients]
    if recipient_emails:
        try:
            send_announcement_email(title, message, user['name'], recipient_emails)
        except Exception as e:
            logging.error(f"Failed to send announcement email: {e}")

    flash('Announcement posted successfully', 'success')
    return redirect(url_for('announcements'))


@app.route('/announcements/delete/<int:ann_id>', methods=['POST'])
@login_required
def delete_announcement(ann_id):
    user = get_user()
    if not can_post_announcements(user):
        flash('You do not have permission to delete announcements', 'error')
        return redirect(url_for('dashboard'))

    conn = get_db()
    conn.execute('UPDATE announcements SET is_active = 0 WHERE id = ?', (ann_id,))
    conn.commit()
    conn.close()

    flash('Announcement removed', 'success')
    return redirect(url_for('announcements'))


@app.route('/api/daily-notifications')
def daily_notifications():
    """
    API endpoint to send birthday and anniversary reminder emails.
    Call this daily via cron job. Sends reminders for tomorrow's events.
    Also sends a happy birthday email directly to the birthday person.
    Creates in-app notifications for all users.
    """
    conn = get_db()
    tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
    tomorrow_mm = tomorrow[5:7]
    tomorrow_dd = tomorrow[8:10]
    current_year = datetime.now().year

    # Get all active employee emails for recipients
    all_employees = conn.execute('''
        SELECT email FROM employees WHERE is_active = 1 AND email IS NOT NULL AND email != ''
    ''', ()).fetchall()
    recipient_emails = [e['email'] for e in all_employees]

    results = {'birthdays_sent': 0, 'birthday_wishes_sent': 0, 'anniversaries_sent': 0, 'notifications_created': 0, 'errors': []}

    if not recipient_emails:
        conn.close()
        return jsonify({'status': 'ok', 'message': 'No recipients with email addresses', **results})

    # Birthday reminders for tomorrow
    birthday_people = conn.execute('''
        SELECT name, dob, email FROM employees
        WHERE is_active = 1 AND emp_code != 'admin' AND dob IS NOT NULL AND dob != ''
        AND strftime('%m', dob) = ? AND strftime('%d', dob) = ?
    ''', (tomorrow_mm, tomorrow_dd)).fetchall()

    for person in birthday_people:
        # Send team reminder email
        try:
            send_birthday_reminder(person['name'], tomorrow, recipient_emails)
            results['birthdays_sent'] += 1
        except Exception as e:
            results['errors'].append(f"Birthday reminder for {person['name']}: {str(e)}")

        # Send happy birthday email directly to the person
        if person['email']:
            try:
                send_happy_birthday_email(person['name'], person['email'])
                results['birthday_wishes_sent'] += 1
            except Exception as e:
                results['errors'].append(f"Happy birthday email for {person['name']}: {str(e)}")

        # Create in-app notification for everyone
        try:
            all_emps = conn.execute("SELECT id FROM employees WHERE is_active = 1 AND emp_code != 'admin'", ()).fetchall()
            for emp in all_emps:
                create_notification(conn, emp['id'],
                    f"{person['name']}'s Birthday Tomorrow!",
                    f"Tomorrow is {person['name']}'s birthday! Join us in wishing them a wonderful day.",
                    'birthday')
            results['notifications_created'] += 1
        except Exception as e:
            results['errors'].append(f"Birthday notification for {person['name']}: {str(e)}")

    # Anniversary reminders for tomorrow
    anniversary_people = conn.execute('''
        SELECT name, joining_date, email FROM employees
        WHERE is_active = 1 AND emp_code != 'admin' AND joining_date IS NOT NULL AND joining_date != ''
        AND strftime('%m', joining_date) = ? AND strftime('%d', joining_date) = ?
        AND strftime('%Y', joining_date) != ?
    ''', (tomorrow_mm, tomorrow_dd, str(current_year))).fetchall()

    for person in anniversary_people:
        try:
            join_year = int(person['joining_date'][:4])
            years = current_year - join_year
            send_anniversary_reminder(person['name'], years, tomorrow, recipient_emails)
            results['anniversaries_sent'] += 1
        except Exception as e:
            results['errors'].append(f"Anniversary email for {person['name']}: {str(e)}")

        # Create in-app notification for everyone
        try:
            join_year = int(person['joining_date'][:4])
            years = current_year - join_year
            yr_text = f"{years} year{'s' if years != 1 else ''}"
            all_emps = conn.execute("SELECT id FROM employees WHERE is_active = 1 AND emp_code != 'admin'", ()).fetchall()
            for emp in all_emps:
                create_notification(conn, emp['id'],
                    f"{person['name']}'s Work Anniversary!",
                    f"Tomorrow marks {person['name']}'s {yr_text} at GooCampus! Congratulations!",
                    'anniversary')
            results['notifications_created'] += 1
        except Exception as e:
            results['errors'].append(f"Anniversary notification for {person['name']}: {str(e)}")

    # Create notifications for upcoming holidays (next 2 days)
    day_after = (datetime.now() + timedelta(days=2)).strftime('%Y-%m-%d')
    upcoming = conn.execute('''
        SELECT holiday_name, holiday_date, holiday_type FROM holidays
        WHERE holiday_date = ? OR holiday_date = ?
    ''', (tomorrow, day_after)).fetchall()

    for holiday in upcoming:
        try:
            all_emps = conn.execute("SELECT id FROM employees WHERE is_active = 1 AND emp_code != 'admin'", ()).fetchall()
            for emp in all_emps:
                create_notification(conn, emp['id'],
                    f"Holiday: {holiday['holiday_name']}",
                    f"Upcoming holiday on {holiday['holiday_date']} — {holiday['holiday_name']} ({holiday['holiday_type']})",
                    'holiday')
            results['notifications_created'] += 1
        except Exception as e:
            results['errors'].append(f"Holiday notification: {str(e)}")

    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', **results})


# ─── Quarterly Report Routes ───
@app.route('/admin/quarterly-report')
@admin_required
def admin_quarterly_report():
    user = get_user()
    year = int(request.args.get('year', datetime.now().year))
    quarter = int(request.args.get('quarter', 1))

    # Q1=Apr-Jun, Q2=Jul-Sep, Q3=Oct-Dec, Q4=Jan-Mar
    quarter_months = {
        1: [4, 5, 6],
        2: [7, 8, 9],
        3: [10, 11, 12],
        4: [1, 2, 3]
    }
    months = quarter_months.get(quarter, [4, 5, 6])

    conn = get_db()
    employees = conn.execute("SELECT * FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name").fetchall()

    report_data = []
    for emp in employees:
        # Get leaves for each month in quarter
        quarterly_days = 0
        leave_breakdown = {'annual': 0, 'sick': 0, 'casual': 0}

        for m in months:
            # Adjust year for Q4 (Jan-Mar of next year)
            check_year = year + 1 if m < 4 else year
            leaves = conn.execute('''
                SELECT SUM(days) as total_days, leave_type FROM leave_records
                WHERE employee_id = ? AND status = 'approved'
                AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
                GROUP BY leave_type
            ''', (emp['id'], str(check_year), str(m).zfill(2))).fetchall()

            for row in leaves:
                if row['total_days']:
                    quarterly_days += row['total_days']
                    leave_type = (row['leave_type'] or 'casual').lower()
                    if leave_type in leave_breakdown:
                        leave_breakdown[leave_type] += row['total_days']

        report_data.append({
            'name': emp['name'],
            'emp_code': emp['emp_code'],
            'department': emp['department'],
            'total_days': quarterly_days,
            'annual': leave_breakdown['annual'],
            'sick': leave_breakdown['sick'],
            'casual': leave_breakdown['casual']
        })

    conn.close()
    quarter_name = f"Q{quarter} (FY {year}-{year+1})"

    return render_template('admin_quarterly_report.html',
                         user=user,
                         year=year,
                         quarter=quarter,
                         quarter_name=quarter_name,
                         report_data=report_data)


@app.route('/admin/quarterly-report/download')
@admin_required
def admin_quarterly_report_download():
    year = int(request.args.get('year', datetime.now().year))
    quarter = int(request.args.get('quarter', 1))

    quarter_months = {1: [4, 5, 6], 2: [7, 8, 9], 3: [10, 11, 12], 4: [1, 2, 3]}
    months = quarter_months.get(quarter, [4, 5, 6])

    conn = get_db()
    employees = conn.execute("SELECT * FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Q{quarter} Report {year}"

    headers = ['Name', 'Employee Code', 'Department', 'Total Days', 'Annual', 'Sick', 'Casual']
    ws.append(headers)

    header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for emp in employees:
        quarterly_days = 0
        leave_breakdown = {'annual': 0, 'sick': 0, 'casual': 0}

        for m in months:
            check_year = year + 1 if m < 4 else year
            leaves = conn.execute('''
                SELECT SUM(days) as total_days, leave_type FROM leave_records
                WHERE employee_id = ? AND status = 'approved'
                AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
                GROUP BY leave_type
            ''', (emp['id'], str(check_year), str(m).zfill(2))).fetchall()

            for row in leaves:
                if row['total_days']:
                    quarterly_days += row['total_days']
                    leave_type = (row['leave_type'] or 'casual').lower()
                    if leave_type in leave_breakdown:
                        leave_breakdown[leave_type] += row['total_days']

        ws.append([
            emp['name'],
            emp['emp_code'],
            emp['department'],
            quarterly_days,
            leave_breakdown['annual'],
            leave_breakdown['sick'],
            leave_breakdown['casual']
        ])

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    conn.close()

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Quarterly_Report_Q{quarter}_{year}.xlsx'
    )


# ─── Annual Report Routes ───
@app.route('/admin/annual-report')
@admin_required
def admin_annual_report():
    user = get_user()
    # year parameter is the FY start year (e.g. 2025 for FY 2025-26)
    year = int(request.args.get('year', datetime.now().year))

    conn = get_db()
    employees = conn.execute("SELECT * FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name").fetchall()

    report_data = []
    for emp in employees:
        carry_forward = emp['carry_forward']
        total_allocation = 25 + carry_forward

        # Get approved leaves for the entire FY (Apr year to Mar year+1)
        fy_leaves = conn.execute('''
            SELECT SUM(days) as total_days FROM leave_records
            WHERE employee_id = ? AND status = 'approved'
            AND ((strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) >= '04')
                 OR (strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) <= '03'))
        ''', (emp['id'], str(year), str(year + 1))).fetchone()

        total_taken = fy_leaves['total_days'] if fy_leaves['total_days'] else 0
        remaining_balance = total_allocation - total_taken

        # Get monthly breakdown
        monthly_breakdown = []
        for m in range(1, 13):
            check_year = year if m >= 4 else year + 1
            month_leaves = conn.execute('''
                SELECT SUM(days) as total_days FROM leave_records
                WHERE employee_id = ? AND status = 'approved'
                AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
            ''', (emp['id'], str(check_year), str(m).zfill(2))).fetchone()
            month_days = month_leaves['total_days'] if month_leaves['total_days'] else 0
            monthly_breakdown.append(month_days)

        report_data.append({
            'name': emp['name'],
            'emp_code': emp['emp_code'],
            'department': emp['department'],
            'total_allocation': total_allocation,
            'total_taken': total_taken,
            'remaining_balance': remaining_balance,
            'monthly_breakdown': monthly_breakdown
        })

    conn.close()
    fy_label = f"FY {year}-{year+1}"

    return render_template('admin_annual_report.html',
                         user=user,
                         year=year,
                         fy_label=fy_label,
                         report_data=report_data)


@app.route('/admin/annual-report/download')
@admin_required
def admin_annual_report_download():
    year = int(request.args.get('year', datetime.now().year))

    conn = get_db()
    employees = conn.execute("SELECT * FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Annual Report {year}"

    headers = ['Name', 'Employee Code', 'Department', 'Total Allocation', 'Days Taken', 'Remaining Balance']
    # Add month headers
    month_names = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']
    headers.extend(month_names)
    ws.append(headers)

    header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for emp in employees:
        carry_forward = emp['carry_forward']
        total_allocation = 25 + carry_forward

        fy_leaves = conn.execute('''
            SELECT SUM(days) as total_days FROM leave_records
            WHERE employee_id = ? AND status = 'approved'
            AND ((strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) >= '04')
                 OR (strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) <= '03'))
        ''', (emp['id'], str(year), str(year + 1))).fetchone()

        total_taken = fy_leaves['total_days'] if fy_leaves['total_days'] else 0
        remaining_balance = total_allocation - total_taken

        row = [emp['name'], emp['emp_code'], emp['department'], total_allocation, total_taken, remaining_balance]

        # Add monthly breakdown
        for m in range(1, 13):
            check_year = year if m >= 4 else year + 1
            month_leaves = conn.execute('''
                SELECT SUM(days) as total_days FROM leave_records
                WHERE employee_id = ? AND status = 'approved'
                AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
            ''', (emp['id'], str(check_year), str(m).zfill(2))).fetchone()
            month_days = month_leaves['total_days'] if month_leaves['total_days'] else 0
            row.append(month_days)

        ws.append(row)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 17
    for i, col in enumerate(['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']):
        ws.column_dimensions[col].width = 10

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    conn.close()

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Annual_Report_FY_{year}_{year+1}.xlsx'
    )



# ─── Admin Employee Leave Report (select any employee) ───
@app.route('/admin/employee-leave-report')
@admin_required
def admin_employee_leave_report():
    user = get_user()
    conn = get_db()

    employees = conn.execute("SELECT id, name, emp_code, department FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name").fetchall()

    selected_emp_id = request.args.get('employee_id', type=int)
    selected_emp = None
    monthly_leave_data = []
    all_leaves = []
    total_allocation = 0
    total_taken = 0
    annual_total = 0
    sick_total = 0
    casual_total = 0
    available_balance = 0
    pending_count = 0
    carry_forward = 0
    monthly_alloc = 'Apr=3, Others=2'

    today = datetime.now()
    current_month = today.month
    fy_year = today.year if current_month >= 4 else today.year - 1

    if selected_emp_id:
        selected_emp = conn.execute('SELECT * FROM employees WHERE id = ?', (selected_emp_id,)).fetchone()
        if selected_emp:
            carry_forward = selected_emp['carry_forward'] or 0
            total_allocation = 25 + carry_forward

            pending_count = conn.execute(
                "SELECT COUNT(*) as cnt FROM leave_records WHERE employee_id = ? AND status = 'pending'",
                (selected_emp_id,)
            ).fetchone()['cnt']

            running_balance = carry_forward  # Start with carry forward
            for m in range(12):
                report_month = ((m + 3) % 12) + 1
                report_year = fy_year if report_month >= 4 else fy_year + 1
                m_alloc = get_monthly_alloc(report_month)

                month_data = conn.execute('''
                    SELECT SUM(days) as total_days,
                           SUM(CASE WHEN leave_type = 'annual' THEN days ELSE 0 END) as annual,
                           SUM(CASE WHEN leave_type = 'sick' THEN days ELSE 0 END) as sick,
                           SUM(CASE WHEN leave_type = 'casual' THEN days ELSE 0 END) as casual,
                           COUNT(*) as count
                    FROM leave_records
                    WHERE employee_id = ? AND status = 'approved'
                    AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
                ''', (selected_emp_id, str(report_year), str(report_month).zfill(2))).fetchone()

                month_total = month_data['total_days'] or 0
                running_balance = round(running_balance + m_alloc - month_total, 2)

                monthly_leave_data.append({
                    'month': calendar.month_name[report_month],
                    'month_short': calendar.month_abbr[report_month],
                    'year': report_year,
                    'total': month_total,
                    'annual': month_data['annual'] or 0,
                    'sick': month_data['sick'] or 0,
                    'casual': month_data['casual'] or 0,
                    'count': month_data['count'] or 0,
                    'monthly_alloc': m_alloc,
                    'balance': running_balance
                })

            all_leaves = conn.execute('''
                SELECT * FROM leave_records
                WHERE employee_id = ?
                AND ((strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) >= '04')
                     OR (strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) < '04'))
                ORDER BY leave_date DESC
            ''', (selected_emp_id, str(fy_year), str(fy_year + 1))).fetchall()

            total_taken = sum(m['total'] for m in monthly_leave_data)
            annual_total = sum(m['annual'] for m in monthly_leave_data)
            sick_total = sum(m['sick'] for m in monthly_leave_data)
            casual_total = sum(m['casual'] for m in monthly_leave_data)
            available_balance = round(total_allocation - total_taken, 2)

    conn.close()

    return render_template('admin_employee_leave_report.html',
                         user=user,
                         employees=employees,
                         selected_emp=selected_emp,
                         selected_emp_id=selected_emp_id,
                         monthly_leave_data=monthly_leave_data,
                         all_leaves=all_leaves,
                         fy_year=fy_year,
                         total_allocation=total_allocation,
                         total_taken=total_taken,
                         annual_total=annual_total,
                         sick_total=sick_total,
                         casual_total=casual_total,
                         available_balance=available_balance,
                         pending_count=pending_count,
                         carry_forward=carry_forward,
                         monthly_alloc=monthly_alloc)


# ─── Team Leave Report Routes ───
@app.route('/reports/team')
@login_required
def team_leave_report():
    user = get_user()
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))

    # Check if user is a manager
    conn = get_db()
    direct_reports = conn.execute(
        'SELECT * FROM employees WHERE reporting_to = ? AND is_active = 1 ORDER BY name',
        (user['id'],)
    ).fetchall()

    if not direct_reports:
        conn.close()
        flash('You have no direct reports', 'info')
        return redirect(url_for('dashboard'))

    report_data = []
    for emp in direct_reports:
        carry_forward = emp['carry_forward']
        total_allocation = 25 + carry_forward
        monthly_allocation = get_monthly_alloc(month)

        # Get leaves for this month
        month_leaves = conn.execute('''
            SELECT SUM(days) as total_days FROM leave_records
            WHERE employee_id = ? AND status = 'approved'
            AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
        ''', (emp['id'], str(year), str(month).zfill(2))).fetchone()

        days_taken = month_leaves['total_days'] if month_leaves['total_days'] else 0
        balance_start = get_available_balance(emp['id'], year, month)
        balance_available = balance_start + monthly_allocation

        report_data.append({
            'name': emp['name'],
            'emp_code': emp['emp_code'],
            'allocation': monthly_allocation,
            'balance_start': round(balance_start, 2),
            'balance_available': round(balance_available, 2),
            'days_taken': days_taken
        })

    conn.close()
    month_name = calendar.month_name[month]

    return render_template('team_leave_report.html',
                         user=user,
                         year=year,
                         month=month,
                         month_name=month_name,
                         report_data=report_data)


@app.route('/reports/team/download')
@login_required
def team_leave_report_download():
    user = get_user()
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))

    conn = get_db()
    direct_reports = conn.execute(
        'SELECT * FROM employees WHERE reporting_to = ? AND is_active = 1 ORDER BY name',
        (user['id'],)
    ).fetchall()

    if not direct_reports:
        conn.close()
        flash('You have no direct reports', 'info')
        return redirect(url_for('team_leave_report', year=year, month=month))

    wb = Workbook()
    ws = wb.active
    ws.title = f"Team Report {month}-{year}"

    headers = ['Name', 'Employee Code', 'Monthly Allocation', 'Balance Start', 'Balance Available', 'Days Taken']
    ws.append(headers)

    header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for emp in direct_reports:
        carry_forward = emp['carry_forward']
        total_allocation = 25 + carry_forward
        monthly_allocation = get_monthly_alloc(month)

        month_leaves = conn.execute('''
            SELECT SUM(days) as total_days FROM leave_records
            WHERE employee_id = ? AND status = 'approved'
            AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
        ''', (emp['id'], str(year), str(month).zfill(2))).fetchone()

        days_taken = month_leaves['total_days'] if month_leaves['total_days'] else 0
        balance_start = get_available_balance(emp['id'], year, month)
        balance_available = balance_start + monthly_allocation

        ws.append([
            emp['name'],
            emp['emp_code'],
            monthly_allocation,
            round(balance_start, 2),
            round(balance_available, 2),
            days_taken
        ])

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 14

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    conn.close()

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Team_Report_{month}_{year}.xlsx'
    )


# ─── Redirect Routes for Report Navigation ───
@app.route('/reports/monthly')
@login_required
def reports_monthly():
    """Monthly calendar view of leaves for the employee."""
    user = get_user()
    if user['emp_code'] == 'admin':
        return redirect(url_for('admin_employee_leave_report'))

    today = datetime.now()
    view_year = int(request.args.get('year', today.year))
    view_month = int(request.args.get('month', today.month))

    # Clamp month
    if view_month < 1: view_month = 12; view_year -= 1
    if view_month > 12: view_month = 1; view_year += 1

    fy_year = view_year if view_month >= 4 else view_year - 1

    conn = get_db()

    # Total allocation
    emp = conn.execute('SELECT carry_forward FROM employees WHERE id = ?', (user['id'],)).fetchone()
    carry_forward = emp['carry_forward'] if emp else 0
    total_allocation = 25 + carry_forward

    # FY total taken
    fy_leaves = conn.execute('''
        SELECT SUM(days) as total FROM leave_records
        WHERE employee_id = ? AND status = 'approved'
        AND ((strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) >= '04')
             OR (strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) < '04'))
    ''', (user['id'], str(fy_year), str(fy_year + 1))).fetchone()
    total_taken = fy_leaves['total'] or 0
    available_balance = round(total_allocation - total_taken, 2)

    # Leaves for this month
    month_leaves_raw = conn.execute('''
        SELECT * FROM leave_records
        WHERE employee_id = ? AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
        AND status IN ('approved', 'pending')
        ORDER BY leave_date
    ''', (user['id'], str(view_year), str(view_month).zfill(2))).fetchall()

    # Holidays for this month
    holidays_raw = conn.execute('''
        SELECT * FROM holidays
        WHERE strftime('%Y', holiday_date) = ? AND strftime('%m', holiday_date) = ?
    ''', (str(view_year), str(view_month).zfill(2))).fetchall()

    holidays_map = {}
    for h in holidays_raw:
        day = int(h['holiday_date'][8:10])
        holidays_map[day] = h['name']

    # Build leaves map by day
    leaves_map = {}
    for lv in month_leaves_raw:
        day = int(lv['leave_date'][8:10])
        leaves_map[day] = lv

    conn.close()

    # Build calendar days
    import calendar as cal_module
    days_in_month = cal_module.monthrange(view_year, view_month)[1]
    first_day_offset = cal_module.monthrange(view_year, view_month)[0]  # 0=Mon
    # Convert to Sunday-start: Mon=0 -> offset=1, Sun=6 -> offset=0
    first_day_offset = (first_day_offset + 1) % 7

    calendar_days = []
    month_leaves_count = 0
    working_days = 0
    for d in range(1, days_in_month + 1):
        day_of_week = cal_module.weekday(view_year, view_month, d)  # 0=Mon, 6=Sun
        is_weekend = day_of_week in (5, 6)  # Sat, Sun
        is_today = (view_year == today.year and view_month == today.month and d == today.day)
        is_holiday = d in holidays_map

        if not is_weekend and not is_holiday:
            working_days += 1

        leave = leaves_map.get(d)
        if leave:
            month_leaves_count += 1

        calendar_days.append({
            'date': d,
            'is_today': is_today,
            'is_weekend': is_weekend,
            'is_holiday': is_holiday,
            'holiday_name': holidays_map.get(d, ''),
            'leave': leave
        })

    # Trailing empty cells
    total_cells = first_day_offset + days_in_month
    trailing_empty = (7 - (total_cells % 7)) % 7

    # Month leave records with day names for the table
    month_leaves = []
    day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    for lv in month_leaves_raw:
        d = int(lv['leave_date'][8:10])
        dow = cal_module.weekday(view_year, view_month, d)
        leave_dict = dict(lv)
        leave_dict['day_name'] = day_names[dow]
        month_leaves.append(leave_dict)

    # Prev/next month
    if view_month == 1:
        prev_month, prev_year = 12, view_year - 1
    else:
        prev_month, prev_year = view_month - 1, view_year
    if view_month == 12:
        next_month, next_year = 1, view_year + 1
    else:
        next_month, next_year = view_month + 1, view_year

    month_name = cal_module.month_name[view_month]

    return render_template('report_monthly.html',
                         user=user, fy_year=fy_year,
                         view_year=view_year, view_month=view_month,
                         month_name=month_name, days_in_month=days_in_month,
                         first_day_offset=first_day_offset,
                         calendar_days=calendar_days,
                         trailing_empty=trailing_empty,
                         month_leaves=month_leaves,
                         month_leaves_count=month_leaves_count,
                         working_days=working_days,
                         total_allocation=total_allocation,
                         available_balance=available_balance,
                         prev_month=prev_month, prev_year=prev_year,
                         next_month=next_month, next_year=next_year)


@app.route('/reports/quarterly')
@login_required
def reports_quarterly():
    """Quarterly breakdown of leaves for the employee."""
    user = get_user()
    if user['emp_code'] == 'admin':
        return redirect(url_for('admin_employee_leave_report'))

    today = datetime.now()
    fy_year = int(request.args.get('fy', today.year if today.month >= 4 else today.year - 1))

    conn = get_db()
    emp = conn.execute('SELECT carry_forward FROM employees WHERE id = ?', (user['id'],)).fetchone()
    carry_forward = emp['carry_forward'] if emp else 0
    total_allocation = 25 + carry_forward

    pending_count = conn.execute(
        'SELECT COUNT(*) as cnt FROM leave_records WHERE employee_id = ? AND status = ?',
        (user['id'], 'pending')
    ).fetchone()['cnt']

    import calendar as cal_module
    # Q1: Apr-Jun, Q2: Jul-Sep, Q3: Oct-Dec, Q4: Jan-Mar
    quarter_defs = [
        {'label': 'Q1 (Apr — Jun)', 'months_label': 'April, May, June', 'months': [4, 5, 6]},
        {'label': 'Q2 (Jul — Sep)', 'months_label': 'July, August, September', 'months': [7, 8, 9]},
        {'label': 'Q3 (Oct — Dec)', 'months_label': 'October, November, December', 'months': [10, 11, 12]},
        {'label': 'Q4 (Jan — Mar)', 'months_label': 'January, February, March', 'months': [1, 2, 3]},
    ]

    quarters = []
    running_balance = carry_forward
    annual_total = sick_total = casual_total = 0

    for qdef in quarter_defs:
        q_annual = q_sick = q_casual = q_total = 0
        q_alloc = sum(get_monthly_alloc(m) for m in qdef['months'])
        month_details = []

        for m in qdef['months']:
            yr = fy_year if m >= 4 else fy_year + 1
            mdata = conn.execute('''
                SELECT SUM(days) as total_days,
                       SUM(CASE WHEN leave_type = 'annual' THEN days ELSE 0 END) as annual,
                       SUM(CASE WHEN leave_type = 'sick' THEN days ELSE 0 END) as sick,
                       SUM(CASE WHEN leave_type = 'casual' THEN days ELSE 0 END) as casual
                FROM leave_records
                WHERE employee_id = ? AND status = 'approved'
                AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
            ''', (user['id'], str(yr), str(m).zfill(2))).fetchone()

            m_total = mdata['total_days'] or 0
            m_annual = mdata['annual'] or 0
            m_sick = mdata['sick'] or 0
            m_casual = mdata['casual'] or 0

            q_annual += m_annual; q_sick += m_sick; q_casual += m_casual; q_total += m_total

            month_details.append({
                'name': cal_module.month_name[m],
                'annual': m_annual, 'sick': m_sick, 'casual': m_casual, 'total': m_total
            })

        running_balance += q_alloc - q_total
        annual_total += q_annual; sick_total += q_sick; casual_total += q_casual

        quarters.append({
            'label': qdef['label'],
            'months_label': qdef['months_label'],
            'months': month_details,
            'annual': q_annual, 'sick': q_sick, 'casual': q_casual,
            'total': q_total, 'alloc': q_alloc,
            'balance': round(running_balance, 2)
        })

    total_taken = annual_total + sick_total + casual_total
    available_balance = round(total_allocation - total_taken, 2)
    conn.close()

    return render_template('report_quarterly.html',
                         user=user, fy_year=fy_year,
                         quarters=quarters,
                         total_allocation=total_allocation,
                         total_taken=total_taken,
                         annual_total=annual_total,
                         sick_total=sick_total,
                         casual_total=casual_total,
                         available_balance=available_balance,
                         pending_count=pending_count)


@app.route('/reports/annual')
@login_required
def reports_annual():
    """Annual month-wise leave report for the employee."""
    user = get_user()
    if user['emp_code'] == 'admin':
        return redirect(url_for('admin_employee_leave_report'))

    today = datetime.now()
    fy_year = int(request.args.get('fy', today.year if today.month >= 4 else today.year - 1))

    conn = get_db()
    emp = conn.execute('SELECT carry_forward FROM employees WHERE id = ?', (user['id'],)).fetchone()
    carry_forward = emp['carry_forward'] if emp else 0
    total_allocation = 25 + carry_forward

    pending_count = conn.execute(
        'SELECT COUNT(*) as cnt FROM leave_records WHERE employee_id = ? AND status = ?',
        (user['id'], 'pending')
    ).fetchone()['cnt']

    import calendar as cal_module
    monthly_leave_data = []
    running_balance = carry_forward
    for m_idx in range(12):
        report_month = ((m_idx + 3) % 12) + 1
        report_year = fy_year if report_month >= 4 else fy_year + 1
        m_alloc = get_monthly_alloc(report_month)

        month_data = conn.execute('''
            SELECT SUM(days) as total_days,
                   SUM(CASE WHEN leave_type = 'annual' THEN days ELSE 0 END) as annual,
                   SUM(CASE WHEN leave_type = 'sick' THEN days ELSE 0 END) as sick,
                   SUM(CASE WHEN leave_type = 'casual' THEN days ELSE 0 END) as casual
            FROM leave_records
            WHERE employee_id = ? AND status = 'approved'
            AND strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) = ?
        ''', (user['id'], str(report_year), str(report_month).zfill(2))).fetchone()

        month_total = month_data['total_days'] or 0
        running_balance = round(running_balance + m_alloc - month_total, 2)

        monthly_leave_data.append({
            'month': cal_module.month_name[report_month],
            'year': report_year,
            'total': month_total,
            'annual': month_data['annual'] or 0,
            'sick': month_data['sick'] or 0,
            'casual': month_data['casual'] or 0,
            'monthly_alloc': m_alloc,
            'balance': running_balance
        })

    all_leaves = conn.execute('''
        SELECT * FROM leave_records
        WHERE employee_id = ?
        AND ((strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) >= '04')
             OR (strftime('%Y', leave_date) = ? AND strftime('%m', leave_date) < '04'))
        ORDER BY leave_date DESC
    ''', (user['id'], str(fy_year), str(fy_year + 1))).fetchall()

    total_taken = sum(m['total'] for m in monthly_leave_data)
    annual_total = sum(m['annual'] for m in monthly_leave_data)
    sick_total = sum(m['sick'] for m in monthly_leave_data)
    casual_total = sum(m['casual'] for m in monthly_leave_data)
    available_balance = round(total_allocation - total_taken, 2)

    conn.close()

    return render_template('report_annual.html',
                         user=user, fy_year=fy_year,
                         monthly_leave_data=monthly_leave_data,
                         all_leaves=all_leaves,
                         total_allocation=total_allocation,
                         total_taken=total_taken,
                         annual_total=annual_total,
                         sick_total=sick_total,
                         casual_total=casual_total,
                         available_balance=available_balance,
                         pending_count=pending_count,
                         carry_forward=carry_forward,
                         monthly_alloc='Apr=3, Others=2')


# ─── Forex / Currency Conversion ───
# Cached FX rates (1 unit of currency → INR). Refreshed every 6 hours.
SUPPORTED_CURRENCIES = ['INR', 'USD', 'EUR', 'GBP', 'AUD', 'AED', 'RUB']
_fx_cache = {'rates': None, 'fetched_at': 0, 'source': 'fallback', 'updated_iso': None}

# Reasonable fallback rates in case the API is unreachable (as of early 2026)
_fx_fallback = {
    'INR': 1.0, 'USD': 83.50, 'EUR': 90.20, 'GBP': 106.40,
    'AUD': 55.10, 'AED': 22.73, 'RUB': 0.91
}

def get_fx_rates_inr(force_refresh=False):
    """Returns dict mapping currency code to INR per 1 unit. Cached 6 hours.
    Falls back to hardcoded rates if network fails."""
    import time as _time
    now = _time.time()
    if (not force_refresh) and _fx_cache['rates'] and (now - _fx_cache['fetched_at']) < 6 * 3600:
        return _fx_cache['rates']
    try:
        import urllib.request as _ur
        import json as _json
        with _ur.urlopen('https://open.er-api.com/v6/latest/USD', timeout=6) as r:
            data = _json.loads(r.read().decode())
        usd_rates = data.get('rates', {}) or {}
        inr_per_usd = usd_rates.get('INR')
        if not inr_per_usd:
            raise ValueError('No INR rate in response')
        rates = {'INR': 1.0}
        for code in SUPPORTED_CURRENCIES:
            if code == 'INR':
                continue
            r_code = usd_rates.get(code)
            if r_code and r_code > 0:
                # 1 unit of <code> in INR = inr_per_usd / (USD→code rate)
                rates[code] = round(float(inr_per_usd) / float(r_code), 4)
            else:
                rates[code] = _fx_fallback.get(code, 1.0)
        _fx_cache['rates'] = rates
        _fx_cache['fetched_at'] = now
        _fx_cache['source'] = 'open.er-api.com'
        _fx_cache['updated_iso'] = data.get('time_last_update_utc', '')
        return rates
    except Exception as e:
        logging.warning(f"FX fetch failed, using fallback: {e}")
        if _fx_cache['rates']:
            return _fx_cache['rates']
        _fx_cache['rates'] = dict(_fx_fallback)
        _fx_cache['fetched_at'] = now
        _fx_cache['source'] = 'fallback'
        return _fx_cache['rates']

def to_inr(amount, currency):
    """Convert amount in given currency to INR using cached rates."""
    try:
        amt = float(amount or 0)
    except (ValueError, TypeError):
        amt = 0.0
    cur = (currency or 'INR').upper()
    if cur == 'INR':
        return amt
    rates = get_fx_rates_inr()
    return amt * rates.get(cur, 1.0)


@app.route('/api/forex-rates')
def api_forex_rates():
    """Public endpoint returning current FX→INR rates (JSON). Used by frontend for live conversion preview."""
    rates = get_fx_rates_inr()
    return jsonify({
        'rates': rates,
        'source': _fx_cache.get('source', 'fallback'),
        'updated': _fx_cache.get('updated_iso', ''),
        'currencies': SUPPORTED_CURRENCIES
    })


def ensure_crm_tables():
    """Create CRM tables if they don't exist (safe to run repeatedly)."""
    try:
        conn = get_db()
        tables_sql = [
            '''CREATE TABLE IF NOT EXISTS wfh_requests (
                id SERIAL PRIMARY KEY,
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                from_date TEXT NOT NULL,
                to_date TEXT NOT NULL,
                reason TEXT NOT NULL,
                status TEXT DEFAULT 'pending',
                approved_by INTEGER REFERENCES employees(id),
                approved_at TEXT,
                rejection_reason TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS projects (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                description TEXT,
                status TEXT DEFAULT 'active',
                created_by INTEGER REFERENCES employees(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS revenue_streams (
                id SERIAL PRIMARY KEY,
                project_id INTEGER REFERENCES projects(id),
                name TEXT NOT NULL,
                description TEXT,
                is_active INTEGER DEFAULT 1,
                created_by INTEGER REFERENCES employees(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS products_services (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                description TEXT,
                type TEXT NOT NULL DEFAULT 'product',
                project_id INTEGER REFERENCES projects(id),
                revenue_stream_id INTEGER REFERENCES revenue_streams(id),
                status TEXT DEFAULT 'active',
                product_cost NUMERIC(14,2) DEFAULT 0,
                sale_price NUMERIC(14,2) DEFAULT 0,
                created_by INTEGER REFERENCES employees(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS sales_news (
                id SERIAL PRIMARY KEY,
                title TEXT NOT NULL,
                content TEXT NOT NULL,
                category TEXT DEFAULT 'general',
                posted_by INTEGER NOT NULL REFERENCES employees(id),
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS meeting_types (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS b2b_trips (
                id SERIAL PRIMARY KEY,
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                trip_type TEXT NOT NULL,
                meeting_category TEXT DEFAULT 'face_to_face',
                from_date TEXT NOT NULL,
                to_date TEXT NOT NULL,
                travel_date TEXT,
                project_id INTEGER REFERENCES projects(id),
                notes TEXT,
                status TEXT DEFAULT 'planned',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS b2b_meetings (
                id SERIAL PRIMARY KEY,
                trip_id INTEGER NOT NULL REFERENCES b2b_trips(id) ON DELETE CASCADE,
                meeting_type_id INTEGER REFERENCES meeting_types(id),
                meeting_with TEXT NOT NULL,
                meeting_date TEXT NOT NULL,
                project_id INTEGER REFERENCES projects(id),
                location TEXT,
                contact_person TEXT,
                contact_phone TEXT,
                outcome TEXT,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS module_access (
                id SERIAL PRIMARY KEY,
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                module TEXT NOT NULL,
                granted_by INTEGER REFERENCES employees(id),
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(employee_id, module)
            )''',
        ]
        for sql in tables_sql:
            conn.execute(sql)
        conn.commit()
        # Add is_active column if not exists (migration for existing deployments)
        try:
            conn.execute('ALTER TABLE module_access ADD COLUMN is_active INTEGER DEFAULT 1')
            conn.commit()
        except Exception:
            conn.rollback()  # Required for PostgreSQL to reset aborted transaction state
        # Add meeting_category column to b2b_trips (migration for existing deployments)
        try:
            conn.execute("ALTER TABLE b2b_trips ADD COLUMN meeting_category TEXT DEFAULT 'face_to_face'")
            conn.commit()
        except Exception:
            conn.rollback()
        # Add product_cost column to products_services (migration)
        try:
            conn.execute("ALTER TABLE products_services ADD COLUMN product_cost NUMERIC(14,2) DEFAULT 0")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        # Add sale_price column to products_services (migration)
        try:
            conn.execute("ALTER TABLE products_services ADD COLUMN sale_price NUMERIC(14,2) DEFAULT 0")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        # Add cost_currency column (migration)
        try:
            conn.execute("ALTER TABLE products_services ADD COLUMN cost_currency TEXT DEFAULT 'INR'")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        # Add sale_currency column (migration)
        try:
            conn.execute("ALTER TABLE products_services ADD COLUMN sale_currency TEXT DEFAULT 'INR'")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        # Add revenue_stream_id column to products_services (migration)
        try:
            conn.execute("ALTER TABLE products_services ADD COLUMN revenue_stream_id INTEGER REFERENCES revenue_streams(id)")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        # Drop NOT NULL on revenue_streams.project_id so streams become global (migration)
        try:
            conn.execute("ALTER TABLE revenue_streams ALTER COLUMN project_id DROP NOT NULL")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        # One-time cleanup: null out legacy revenue_streams.project_id so the
        # column stops implying a home project. Streams are now truly global.
        # Guarded by a flags row so it only runs once.
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS app_migrations (key TEXT PRIMARY KEY, applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
            conn.commit()
            already = conn.execute(
                "SELECT 1 FROM app_migrations WHERE key = ?",
                ('null_revenue_streams_project_id_v1',)
            ).fetchone()
            if not already:
                conn.execute("UPDATE revenue_streams SET project_id = NULL WHERE project_id IS NOT NULL")
                conn.execute(
                    "INSERT INTO app_migrations (key) VALUES (?)",
                    ('null_revenue_streams_project_id_v1',)
                )
                conn.commit()
                logging.info("Migration: nulled legacy revenue_streams.project_id")
        except Exception as _mig_err:
            try:
                conn.rollback()
            except Exception:
                pass
            logging.warning(f"stream project_id cleanup skipped: {_mig_err}")
        conn.close()
        logging.info("CRM tables ensured.")
    except Exception as e:
        logging.error(f"ensure_crm_tables: {e}")


def ensure_kra_tables():
    """Create KRA tables if they don't exist."""
    try:
        conn = get_db()
        kra_sql = [
            '''CREATE TABLE IF NOT EXISTS kra_templates (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                department TEXT,
                role_title TEXT,
                fy_year INTEGER NOT NULL,
                is_active INTEGER DEFAULT 1,
                created_by INTEGER REFERENCES employees(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS kra_categories (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                sort_order INTEGER DEFAULT 0,
                is_common INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS kra_template_items (
                id SERIAL PRIMARY KEY,
                template_id INTEGER NOT NULL REFERENCES kra_templates(id) ON DELETE CASCADE,
                category_id INTEGER NOT NULL REFERENCES kra_categories(id),
                kpi_code TEXT NOT NULL,
                measure_description TEXT NOT NULL,
                target_value REAL,
                percentage REAL DEFAULT 0,
                percentage_sharing REAL NOT NULL DEFAULT 0,
                is_target_based INTEGER DEFAULT 0,
                sort_order INTEGER DEFAULT 0
            )''',
            '''CREATE TABLE IF NOT EXISTS kra_assignments (
                id SERIAL PRIMARY KEY,
                template_id INTEGER NOT NULL REFERENCES kra_templates(id),
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                fy_year INTEGER NOT NULL,
                assigned_by INTEGER REFERENCES employees(id),
                assigned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_active INTEGER DEFAULT 1
            )''',
            '''CREATE TABLE IF NOT EXISTS kra_monthly_ratings (
                id SERIAL PRIMARY KEY,
                assignment_id INTEGER NOT NULL REFERENCES kra_assignments(id),
                template_item_id INTEGER NOT NULL REFERENCES kra_template_items(id),
                month INTEGER NOT NULL,
                year INTEGER NOT NULL,
                achieved_value REAL,
                employee_rating REAL,
                manager_rating REAL,
                employee_result REAL,
                manager_result REAL,
                employee_submitted INTEGER DEFAULT 0,
                manager_submitted INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS kra_monthly_notes (
                id SERIAL PRIMARY KEY,
                assignment_id INTEGER NOT NULL REFERENCES kra_assignments(id),
                month INTEGER NOT NULL,
                year INTEGER NOT NULL,
                notes TEXT,
                created_by INTEGER REFERENCES employees(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
        ]
        for sql in kra_sql:
            conn.execute(sql)
        conn.commit()
        conn.close()
        logging.info("KRA tables ensured.")
    except Exception as e:
        logging.error(f"ensure_kra_tables: {e}")


def ensure_notification_tables():
    """Create or migrate notifications table to use employee_id schema."""
    try:
        conn = get_db()
        # Check if the table exists and has the old schema (target_user_id)
        if conn.is_postgres:
            col_check = conn.execute(
                "SELECT column_name FROM information_schema.columns WHERE table_name = 'notifications' AND column_name = 'target_user_id'",
                ()
            ).fetchone()
            if col_check:
                # Old schema detected — drop and recreate
                logging.info("Old notifications schema detected (target_user_id). Migrating to employee_id schema...")
                conn.execute('DROP TABLE notifications', ())
                conn.commit()

        conn.execute('''CREATE TABLE IF NOT EXISTS notifications (
            id SERIAL PRIMARY KEY,
            employee_id INTEGER NOT NULL REFERENCES employees(id),
            title TEXT NOT NULL,
            message TEXT NOT NULL,
            type TEXT DEFAULT 'info',
            link TEXT,
            is_read INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        conn.commit()
        conn.close()
        logging.info("Notification tables ensured.")
    except Exception as e:
        logging.error(f"ensure_notification_tables: {e}")


def create_notification(conn, employee_id, title, message, notif_type='info', link=None):
    """Helper to create a notification for an employee."""
    conn.execute(
        'INSERT INTO notifications (employee_id, title, message, type, link) VALUES (?, ?, ?, ?, ?)',
        (employee_id, title, message, notif_type, link)
    )


def seed_kra_categories():
    """Seed the 6 default KRA categories if not present."""
    try:
        conn = get_db()
        count = conn.execute('SELECT COUNT(*) as cnt FROM kra_categories').fetchone()
        if count['cnt'] == 0:
            cats = [
                ('Target', 1, 0),
                ('Knowledge', 2, 1),
                ('Customer Handling', 3, 1),
                ('HR', 4, 1),
                ('Extra Mile', 5, 1),
                ('Interpersonal Skill', 6, 1),
            ]
            for name, sort_order, is_common in cats:
                conn.execute(
                    'INSERT INTO kra_categories (name, sort_order, is_common) VALUES (?, ?, ?)',
                    (name, sort_order, is_common)
                )
            conn.commit()
        conn.close()
    except Exception as e:
        logging.error(f"seed_kra_categories: {e}")


def seed_default_meeting_types():
    """Seed default meeting client types if table is empty."""
    try:
        conn = get_db()
        count = conn.execute('SELECT COUNT(*) as cnt FROM meeting_types').fetchone()
        if count['cnt'] == 0:
            for mt in ['School', 'College', 'Partner', 'Branch Partner', 'Agent']:
                conn.execute('INSERT INTO meeting_types (name) VALUES (?)', (mt,))
            conn.commit()
        conn.close()
    except Exception as e:
        logging.error(f"seed_default_meeting_types: {e}")


# ─── WFH (Work from Home) Routes ───

@app.route('/wfh/apply', methods=['GET', 'POST'])
@login_required
def apply_wfh():
    user = get_user()

    # Admin control account cannot apply for WFH
    if user['emp_code'] == 'admin':
        flash('Admin account cannot apply for WFH.', 'error')
        return redirect(url_for('wfh_approvals'))

    if request.method == 'POST':
        from_date = request.form.get('from_date')
        to_date = request.form.get('to_date')
        reason = request.form.get('reason', '').strip()

        if not from_date or not to_date or not reason:
            flash('All fields are required', 'error')
            return redirect(url_for('apply_wfh'))

        if to_date < from_date:
            flash('To date cannot be before from date', 'error')
            return redirect(url_for('apply_wfh'))

        conn = get_db()
        conn.execute('''
            INSERT INTO wfh_requests (employee_id, from_date, to_date, reason)
            VALUES (?, ?, ?, ?)
        ''', (user['id'], from_date, to_date, reason))
        conn.commit()

        # Notify reporting manager and admins
        admins = conn.execute('SELECT id FROM employees WHERE is_admin = 1 AND is_active = 1').fetchall()
        for admin in admins:
            if admin['id'] != user['id']:
                create_notification(conn, admin['id'],
                    f"WFH Request from {user['name']}",
                    f"{user['name']} has requested WFH from {from_date} to {to_date}. Reason: {reason}",
                    'leave_request', '/wfh/approvals')
        conn.commit()
        conn.close()

        flash('WFH request submitted successfully', 'success')
        return redirect(url_for('my_wfh_requests'))

    return render_template('apply_wfh.html', user=user)


@app.route('/wfh/my-requests')
@login_required
def my_wfh_requests():
    user = get_user()

    if user['emp_code'] == 'admin':
        return redirect(url_for('wfh_approvals'))

    conn = get_db()
    requests_list = conn.execute('''
        SELECT w.*, a.name as approver_name
        FROM wfh_requests w
        LEFT JOIN employees a ON w.approved_by = a.id
        WHERE w.employee_id = ?
        ORDER BY w.created_at DESC
    ''', (user['id'],)).fetchall()
    conn.close()
    return render_template('my_wfh_requests.html', user=user, requests=requests_list)


@app.route('/wfh/approvals')
@login_required
def wfh_approvals():
    user = get_user()
    conn = get_db()

    is_mgmt = user['emp_code'] in MANAGEMENT_CODES
    is_admin_user = user['is_admin'] == 1

    # Filters
    f_employee = (request.args.get('employee') or '').strip()
    f_department = (request.args.get('department') or '').strip()
    f_status = (request.args.get('status') or '').strip()

    base_select = '''
        SELECT w.id, w.employee_id, w.from_date, w.to_date, w.reason, w.status,
               w.approved_by, w.approved_at, w.rejection_reason, w.created_at,
               e.name          AS employee_name,
               e.emp_code      AS emp_code,
               e.department    AS department,
               e.designation   AS designation,
               e.email         AS email,
               e.phone         AS mobile,
               e.photo_url     AS photo_url,
               a.name          AS approver_name
        FROM wfh_requests w
        JOIN employees e ON w.employee_id = e.id
        LEFT JOIN employees a ON w.approved_by = a.id
    '''

    where_clauses = []
    params = []

    if not is_admin_user and not is_mgmt:
        direct_report_ids = [r['id'] for r in conn.execute(
            'SELECT id FROM employees WHERE reporting_to = ? AND is_active = 1', (user['id'],)
        ).fetchall()]
        if not direct_report_ids:
            flash('No team members to manage', 'error')
            conn.close()
            return redirect(url_for('dashboard'))
        placeholders = ','.join('?' * len(direct_report_ids))
        where_clauses.append(f'w.employee_id IN ({placeholders})')
        params.extend(direct_report_ids)

    if f_employee:
        where_clauses.append('(LOWER(e.name) LIKE ? OR LOWER(e.emp_code) LIKE ?)')
        like = f'%{f_employee.lower()}%'
        params.extend([like, like])
    if f_department:
        where_clauses.append('LOWER(e.department) LIKE ?')
        params.append(f'%{f_department.lower()}%')
    if f_status:
        where_clauses.append('w.status = ?')
        params.append(f_status)

    sql = base_select
    if where_clauses:
        sql += ' WHERE ' + ' AND '.join(where_clauses)
    sql += ' ORDER BY w.created_at DESC'

    requests_list = conn.execute(sql, tuple(params)).fetchall()

    # Normalize rows to dicts and add photo_src (ready-to-use img URL)
    enriched = []
    for r in requests_list:
        d = dict(r)
        p = d.get('photo_url') or ''
        if p:
            d['photo_src'] = p if p.startswith('http') else f'/static/photos/{p}'
        else:
            d['photo_src'] = ''
        enriched.append(d)

    # Stat counts (respect filters but not status, so pills stay meaningful)
    count_sql_base = '''
        SELECT w.status AS st, COUNT(*) AS c
        FROM wfh_requests w
        JOIN employees e ON w.employee_id = e.id
    '''
    count_where = []
    count_params = []
    if not is_admin_user and not is_mgmt:
        placeholders = ','.join('?' * len(direct_report_ids))
        count_where.append(f'w.employee_id IN ({placeholders})')
        count_params.extend(direct_report_ids)
    if f_employee:
        count_where.append('(LOWER(e.name) LIKE ? OR LOWER(e.emp_code) LIKE ?)')
        like = f'%{f_employee.lower()}%'
        count_params.extend([like, like])
    if f_department:
        count_where.append('LOWER(e.department) LIKE ?')
        count_params.append(f'%{f_department.lower()}%')
    count_sql = count_sql_base
    if count_where:
        count_sql += ' WHERE ' + ' AND '.join(count_where)
    count_sql += ' GROUP BY w.status'
    rows = conn.execute(count_sql, tuple(count_params)).fetchall()
    counts_by_status = {r['st']: r['c'] for r in rows}
    pending_count = counts_by_status.get('pending', 0)
    approved_count = counts_by_status.get('approved', 0)
    rejected_count = counts_by_status.get('rejected', 0)
    total_count = pending_count + approved_count + rejected_count

    conn.close()
    return render_template('wfh_approvals.html', user=user, requests=enriched,
                           f_employee=f_employee, f_department=f_department, f_status=f_status,
                           total_count=total_count, pending_count=pending_count,
                           approved_count=approved_count, rejected_count=rejected_count)


@app.route('/wfh/approve/<int:wfh_id>', methods=['POST'])
@login_required
def approve_wfh(wfh_id):
    user = get_user()
    action = request.form.get('action')  # 'approve' or 'reject'
    rejection_reason = request.form.get('rejection_reason', '').strip()

    conn = get_db()
    wfh = conn.execute('SELECT * FROM wfh_requests WHERE id = ?', (wfh_id,)).fetchone()
    if not wfh:
        flash('WFH request not found', 'error')
        conn.close()
        return redirect(url_for('wfh_approvals'))

    if wfh['status'] != 'pending':
        flash('This request has already been processed', 'error')
        conn.close()
        return redirect(url_for('wfh_approvals'))

    new_status = 'approved' if action == 'approve' else 'rejected'
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    conn.execute('''
        UPDATE wfh_requests SET status = ?, approved_by = ?, approved_at = ?, rejection_reason = ?
        WHERE id = ?
    ''', (new_status, user['id'], now, rejection_reason if action == 'reject' else None, wfh_id))

    # Notify the employee
    emp = conn.execute('SELECT name FROM employees WHERE id = ?', (wfh['employee_id'],)).fetchone()
    status_text = 'approved' if action == 'approve' else 'rejected'
    msg = f"Your WFH request ({wfh['from_date']} to {wfh['to_date']}) has been {status_text} by {user['name']}."
    if action == 'reject' and rejection_reason:
        msg += f" Reason: {rejection_reason}"
    create_notification(conn, wfh['employee_id'],
        f"WFH Request {status_text.title()}", msg,
        'success' if action == 'approve' else 'danger', '/wfh/my-requests')

    conn.commit()
    conn.close()

    flash(f'WFH request {status_text}', 'success')
    return redirect(url_for('wfh_approvals'))


# ─── Projects Routes ───

@app.route('/projects')
@login_required
def projects_list():
    user = get_user()
    if not has_module_access(user, 'projects') and not user['is_admin']:
        # All employees can view projects, but only certain can manage
        pass
    conn = get_db()
    projects_raw = conn.execute('''
        SELECT p.*, e.name as created_by_name
        FROM projects p
        LEFT JOIN employees e ON p.created_by = e.id
        ORDER BY p.status, p.name
    ''').fetchall()
    products_raw = conn.execute('''
        SELECT project_id, product_cost, sale_price, cost_currency, sale_currency
        FROM products_services
    ''').fetchall()
    conn.close()
    # Aggregate in Python (multi-currency aware)
    agg = {}
    for r in products_raw:
        pid = r['project_id']
        if pid is None:
            continue
        a = agg.setdefault(pid, {'count': 0, 'cost': 0.0, 'rev': 0.0})
        a['count'] += 1
        a['cost'] += to_inr(r['product_cost'], r['cost_currency'] or 'INR')
        a['rev'] += to_inr(r['sale_price'], r['sale_currency'] or 'INR')
    projects = []
    for p in projects_raw:
        d = dict(p)
        pid = d['id']
        a = agg.get(pid, {'count': 0, 'cost': 0.0, 'rev': 0.0})
        d['product_count'] = a['count']
        d['total_cost'] = a['cost']
        d['total_revenue'] = a['rev']
        projects.append(d)
    return render_template('projects.html', user=user, projects=projects,
                         can_manage=bool(user['is_admin']),
                         is_admin=bool(user['is_admin']))


@app.route('/projects/add', methods=['GET', 'POST'])
@admin_required
def add_project():
    user = get_user()
    if not has_module_access(user, 'projects') and not user['is_admin']:
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        if not name:
            flash('Project name is required', 'error')
            return redirect(url_for('add_project'))

        conn = get_db()
        conn.execute('INSERT INTO projects (name, description, created_by) VALUES (?, ?, ?)',
                    (name, description, user['id']))
        conn.commit()
        conn.close()
        flash('Project added successfully', 'success')
        return redirect(url_for('projects_list'))

    return render_template('add_project.html', user=user)


@app.route('/projects/<int:project_id>')
@login_required
def project_detail(project_id):
    user = get_user()
    conn = get_db()
    project = conn.execute('SELECT p.*, e.name as created_by_name FROM projects p LEFT JOIN employees e ON p.created_by = e.id WHERE p.id = ?', (project_id,)).fetchone()
    if not project:
        flash('Project not found', 'error')
        conn.close()
        return redirect(url_for('projects_list'))

    # Global streams list (for edit dropdowns etc.)
    all_streams = conn.execute(
        "SELECT * FROM revenue_streams ORDER BY is_active DESC, name"
    ).fetchall()

    products_raw = conn.execute('''
        SELECT ps.*, e.name as created_by_name
        FROM products_services ps
        LEFT JOIN employees e ON ps.created_by = e.id
        WHERE ps.project_id = ?
        ORDER BY ps.type, ps.name
    ''', (project_id,)).fetchall()

    conn.close()
    # Only include streams that have at least one product in this project
    used_stream_ids = {p['revenue_stream_id'] for p in products_raw if p['revenue_stream_id']}
    streams_raw = [s for s in all_streams if s['id'] in used_stream_ids]
    # Attach live-converted INR fields to each product
    products = []
    for p in products_raw:
        d = dict(p)
        d['cost_currency'] = d.get('cost_currency') or 'INR'
        d['sale_currency'] = d.get('sale_currency') or 'INR'
        d['cost_inr'] = to_inr(d.get('product_cost'), d['cost_currency'])
        d['revenue_inr'] = to_inr(d.get('sale_price'), d['sale_currency'])
        d['margin_inr'] = d['revenue_inr'] - d['cost_inr']
        products.append(d)

    # Group products by revenue stream id (None = unassigned)
    streams = []
    for s in streams_raw:
        sd = dict(s)
        sd['products'] = [p for p in products if p.get('revenue_stream_id') == sd['id']]
        sd['total_cost'] = sum(p['cost_inr'] for p in sd['products'])
        sd['total_revenue'] = sum(p['revenue_inr'] for p in sd['products'])
        sd['total_margin'] = sd['total_revenue'] - sd['total_cost']
        streams.append(sd)
    unassigned = [p for p in products if not p.get('revenue_stream_id')]

    return render_template('project_detail.html', user=user, project=project, products=products,
                         streams=streams, unassigned=unassigned,
                         all_streams=all_streams,
                         can_manage=bool(user['is_admin']),
                         supported_currencies=SUPPORTED_CURRENCIES,
                         fx_rates=get_fx_rates_inr())


@app.route('/projects/<int:project_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_project(project_id):
    user = get_user()
    if not has_module_access(user, 'projects') and not user['is_admin']:
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))

    conn = get_db()
    project = conn.execute('SELECT * FROM projects WHERE id = ?', (project_id,)).fetchone()
    if not project:
        conn.close()
        flash('Project not found', 'error')
        return redirect(url_for('projects_list'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        status = request.form.get('status', 'active')
        conn.execute('UPDATE projects SET name = ?, description = ?, status = ? WHERE id = ?',
                    (name, description, status, project_id))
        conn.commit()
        conn.close()
        flash('Project updated', 'success')
        return redirect(url_for('project_detail', project_id=project_id))

    conn.close()
    return render_template('project_form.html', user=user, project=project, mode='edit')


@app.route('/projects/<int:project_id>/delete', methods=['POST'])
@admin_required
def delete_project(project_id):
    """Admin-only hard delete of a project + its products."""
    conn = get_db()
    try:
        conn.execute('DELETE FROM products_services WHERE project_id = ?', (project_id,))
        conn.execute('DELETE FROM revenue_streams WHERE project_id = ?', (project_id,))
        conn.execute('DELETE FROM projects WHERE id = ?', (project_id,))
        conn.commit()
        flash('Project, streams and products deleted', 'success')
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        flash(f'Could not delete project: {e}', 'error')
    finally:
        conn.close()
    return redirect(url_for('projects_list'))


# ─── Revenue Stream Routes (GLOBAL — not tied to a single project) ───

@app.route('/streams')
@login_required
def streams_list():
    """Global listing of all revenue streams with usage stats."""
    user = get_user()
    conn = get_db()
    streams_raw = conn.execute(
        '''SELECT rs.*, e.name as created_by_name
           FROM revenue_streams rs
           LEFT JOIN employees e ON rs.created_by = e.id
           ORDER BY rs.is_active DESC, rs.name'''
    ).fetchall()
    # Usage stats per stream: product count, distinct project count, total INR cost/rev
    products_raw = conn.execute(
        'SELECT revenue_stream_id, project_id, product_cost, sale_price, cost_currency, sale_currency FROM products_services'
    ).fetchall()
    conn.close()
    usage = {}
    for p in products_raw:
        sid = p['revenue_stream_id']
        if not sid:
            continue
        u = usage.setdefault(sid, {'count': 0, 'projects': set(), 'cost': 0.0, 'rev': 0.0})
        u['count'] += 1
        if p['project_id']:
            u['projects'].add(p['project_id'])
        u['cost'] += to_inr(p['product_cost'], p['cost_currency'] or 'INR')
        u['rev'] += to_inr(p['sale_price'], p['sale_currency'] or 'INR')
    streams = []
    for s in streams_raw:
        d = dict(s)
        u = usage.get(d['id'], {'count': 0, 'projects': set(), 'cost': 0.0, 'rev': 0.0})
        d['product_count'] = u['count']
        d['project_count'] = len(u['projects'])
        d['total_cost'] = u['cost']
        d['total_revenue'] = u['rev']
        d['total_margin'] = u['rev'] - u['cost']
        streams.append(d)
    return render_template('streams.html', user=user, streams=streams,
                           can_manage=bool(user['is_admin']),
                           is_admin=bool(user['is_admin']))


@app.route('/streams/add', methods=['GET', 'POST'])
@admin_required
def add_revenue_stream():
    user = get_user()
    if not has_module_access(user, 'projects') and not user['is_admin']:
        flash('Access denied', 'error')
        return redirect(url_for('streams_list'))
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        if not name:
            flash('Stream name is required', 'error')
            return redirect(url_for('add_revenue_stream'))
        conn = get_db()
        try:
            conn.execute(
                'INSERT INTO revenue_streams (name, description, created_by) VALUES (?, ?, ?)',
                (name, description, user['id'])
            )
            conn.commit()
            flash('Revenue stream created', 'success')
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            flash(f'Could not add stream: {e}', 'error')
        finally:
            conn.close()
        return redirect(url_for('streams_list'))
    return render_template('stream_form.html', user=user, stream=None, mode='add')


@app.route('/streams/<int:stream_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_revenue_stream(stream_id):
    user = get_user()
    if not has_module_access(user, 'projects') and not user['is_admin']:
        flash('Access denied', 'error')
        return redirect(url_for('streams_list'))
    conn = get_db()
    stream = conn.execute('SELECT * FROM revenue_streams WHERE id = ?', (stream_id,)).fetchone()
    if not stream:
        conn.close()
        flash('Stream not found', 'error')
        return redirect(url_for('streams_list'))
    if request.method == 'POST':
        name = request.form.get('name', '').strip() or stream['name']
        description = request.form.get('description', '').strip()
        is_active = 1 if request.form.get('is_active', '1') == '1' else 0
        try:
            conn.execute(
                'UPDATE revenue_streams SET name = ?, description = ?, is_active = ? WHERE id = ?',
                (name, description, is_active, stream_id)
            )
            conn.commit()
            flash('Stream updated', 'success')
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            flash(f'Could not update stream: {e}', 'error')
        conn.close()
        return redirect(url_for('streams_list'))
    conn.close()
    return render_template('stream_form.html', user=user, stream=stream, mode='edit')


@app.route('/streams/<int:stream_id>/delete', methods=['POST'])
@admin_required
def delete_revenue_stream(stream_id):
    """Admin-only: delete a revenue stream. Products under it become unassigned."""
    conn = get_db()
    stream = conn.execute('SELECT * FROM revenue_streams WHERE id = ?', (stream_id,)).fetchone()
    if not stream:
        conn.close()
        flash('Stream not found', 'error')
        return redirect(url_for('streams_list'))
    try:
        # Unlink products from this stream
        conn.execute('UPDATE products_services SET revenue_stream_id = NULL WHERE revenue_stream_id = ?', (stream_id,))
        # Cascade delete linked finance budget_category (and its entries) so Finance Settings stays in sync
        linked_cats = conn.execute(
            "SELECT id FROM budget_categories WHERE stream_id = ? AND cat_type = 'revenue'",
            (stream_id,)
        ).fetchall()
        for cat in linked_cats:
            cat_id = cat['id']
            try:
                conn.execute('DELETE FROM budget_entries WHERE category_id = ?', (cat_id,))
            except Exception:
                pass
            conn.execute('DELETE FROM budget_categories WHERE id = ?', (cat_id,))
        conn.execute('DELETE FROM revenue_streams WHERE id = ?', (stream_id,))
        conn.commit()
        flash('Revenue stream deleted (and removed from Finance Settings)', 'success')
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        flash(f'Could not delete stream: {e}', 'error')
    finally:
        conn.close()
    return redirect(url_for('streams_list'))


# ─── Products & Services Routes ───

@app.route('/products')
@login_required
def products_list():
    """Dedicated products listing across all projects."""
    user = get_user()
    conn = get_db()
    products_raw = conn.execute('''
        SELECT ps.*,
               p.name as project_name,
               rs.name as stream_name,
               e.name as created_by_name
        FROM products_services ps
        LEFT JOIN projects p ON ps.project_id = p.id
        LEFT JOIN revenue_streams rs ON ps.revenue_stream_id = rs.id
        LEFT JOIN employees e ON ps.created_by = e.id
        ORDER BY ps.status, p.name, rs.name, ps.type, ps.name
    ''').fetchall()
    projects = conn.execute(
        "SELECT id, name FROM projects WHERE status = 'active' ORDER BY name"
    ).fetchall()
    conn.close()
    # Attach live-converted INR fields to each product
    products = []
    total_cost = 0.0
    total_rev = 0.0
    for p in products_raw:
        d = dict(p)
        d['cost_currency'] = d.get('cost_currency') or 'INR'
        d['sale_currency'] = d.get('sale_currency') or 'INR'
        d['cost_inr'] = to_inr(d.get('product_cost'), d['cost_currency'])
        d['revenue_inr'] = to_inr(d.get('sale_price'), d['sale_currency'])
        d['margin_inr'] = d['revenue_inr'] - d['cost_inr']
        total_cost += d['cost_inr']
        total_rev += d['revenue_inr']
        products.append(d)
    return render_template('products.html', user=user, products=products, projects=projects,
                           can_manage=bool(user['is_admin']),
                           is_admin=bool(user['is_admin']),
                           supported_currencies=SUPPORTED_CURRENCIES,
                           fx_rates=get_fx_rates_inr(),
                           total_cost=total_cost,
                           total_rev=total_rev,
                           total_margin=total_rev - total_cost)


@app.route('/products/add/<int:project_id>', methods=['GET', 'POST'])
@admin_required
def add_product(project_id):
    user = get_user()
    if not has_module_access(user, 'projects') and not user['is_admin']:
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))

    conn = get_db()
    project = conn.execute('SELECT * FROM projects WHERE id = ?', (project_id,)).fetchone()
    if not project:
        flash('Project not found', 'error')
        conn.close()
        return redirect(url_for('projects_list'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        ps_type = request.form.get('type', 'product')
        try:
            product_cost = float(request.form.get('product_cost') or 0)
        except ValueError:
            product_cost = 0
        try:
            sale_price = float(request.form.get('sale_price') or 0)
        except ValueError:
            sale_price = 0
        # Unified currency for the product (applies to both cost and revenue)
        currency = (request.form.get('currency') or 'INR').upper()
        if currency not in SUPPORTED_CURRENCIES:
            currency = 'INR'
        cost_currency = currency
        sale_currency = currency
        # Revenue stream. Streams are global — the product's picked project
        # always wins. We only null the stream out if it doesn't exist or is
        # inactive. The legacy revenue_streams.project_id column is ignored.
        stream_raw = request.form.get('revenue_stream_id', '').strip()
        revenue_stream_id = int(stream_raw) if stream_raw.isdigit() else None
        effective_project_id = project_id
        if revenue_stream_id:
            chk = conn.execute(
                'SELECT id FROM revenue_streams WHERE id = ? AND is_active = 1',
                (revenue_stream_id,)
            ).fetchone()
            if not chk:
                revenue_stream_id = None

        if not name:
            flash('Name is required', 'error')
            conn.close()
            return redirect(url_for('add_product', project_id=project_id))

        conn.execute('''
            INSERT INTO products_services
                (name, description, type, project_id, revenue_stream_id, product_cost, sale_price,
                 cost_currency, sale_currency, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (name, description, ps_type, effective_project_id, revenue_stream_id, product_cost, sale_price,
              cost_currency, sale_currency, user['id']))
        conn.commit()
        conn.close()
        flash('Product/Service added', 'success')
        src = request.form.get('source', '')
        if src == 'products':
            return redirect(url_for('products_list'))
        return redirect(url_for('project_detail', project_id=effective_project_id))

    streams = conn.execute(
        "SELECT id, name FROM revenue_streams WHERE is_active = 1 ORDER BY name"
    ).fetchall()
    conn.close()
    return render_template('add_product.html', user=user, project=project,
                           streams=streams,
                           supported_currencies=SUPPORTED_CURRENCIES,
                           fx_rates=get_fx_rates_inr())


@app.route('/products/<int:ps_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_product(ps_id):
    user = get_user()
    if not has_module_access(user, 'projects') and not user['is_admin']:
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))

    conn = get_db()
    ps = conn.execute('SELECT * FROM products_services WHERE id = ?', (ps_id,)).fetchone()
    if not ps:
        flash('Not found', 'error')
        conn.close()
        return redirect(url_for('projects_list'))

    if request.method == 'GET':
        projects_all = conn.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
        streams_all = conn.execute(
            "SELECT id, name FROM revenue_streams WHERE is_active = 1 ORDER BY name"
        ).fetchall()
        conn.close()
        product = dict(ps)
        product['cost_currency'] = product.get('cost_currency') or 'INR'
        return render_template('product_form.html', user=user, product=product,
                               projects_all=projects_all, streams_all=streams_all,
                               supported_currencies=SUPPORTED_CURRENCIES,
                               fx_rates=get_fx_rates_inr(), mode='edit')

    name = request.form.get('name', '').strip()
    description = request.form.get('description', '').strip()
    ps_type = request.form.get('type', ps['type'])
    status = request.form.get('status', ps['status'])
    try:
        product_cost = float(request.form.get('product_cost') or 0)
    except ValueError:
        product_cost = 0
    try:
        sale_price = float(request.form.get('sale_price') or 0)
    except ValueError:
        sale_price = 0
    # Unified currency for the product
    currency = (request.form.get('currency') or 'INR').upper()
    if currency not in SUPPORTED_CURRENCIES:
        currency = 'INR'
    cost_currency = currency
    sale_currency = currency
    # Optional reassignment of project
    new_project_id = request.form.get('project_id')
    if new_project_id and new_project_id.isdigit():
        project_id_val = int(new_project_id)
    else:
        project_id_val = ps['project_id']

    # Revenue stream. Streams are global — the product's picked project
    # always wins. We only null the stream out if it doesn't exist or is
    # inactive. The legacy revenue_streams.project_id column is ignored.
    stream_raw = request.form.get('revenue_stream_id', '').strip()
    revenue_stream_id = int(stream_raw) if stream_raw.isdigit() else None
    if revenue_stream_id:
        chk = conn.execute(
            'SELECT id FROM revenue_streams WHERE id = ? AND is_active = 1',
            (revenue_stream_id,)
        ).fetchone()
        if not chk:
            revenue_stream_id = None

    conn.execute('''UPDATE products_services
                    SET name = ?, description = ?, type = ?, status = ?,
                        product_cost = ?, sale_price = ?,
                        cost_currency = ?, sale_currency = ?,
                        project_id = ?, revenue_stream_id = ?
                    WHERE id = ?''',
                (name, description, ps_type, status, product_cost, sale_price,
                 cost_currency, sale_currency, project_id_val, revenue_stream_id, ps_id))
    conn.commit()
    conn.close()
    flash('Updated successfully', 'success')
    src = request.form.get('source', '')
    if src == 'products':
        return redirect(url_for('products_list'))
    return redirect(url_for('project_detail', project_id=project_id_val))


@app.route('/products/<int:ps_id>/delete', methods=['POST'])
@admin_required
def delete_product(ps_id):
    """Admin-only hard delete of a product/service."""
    conn = get_db()
    ps = conn.execute('SELECT * FROM products_services WHERE id = ?', (ps_id,)).fetchone()
    if not ps:
        conn.close()
        flash('Product not found', 'error')
        return redirect(url_for('products_list'))
    try:
        conn.execute('DELETE FROM products_services WHERE id = ?', (ps_id,))
        conn.commit()
        flash('Product deleted', 'success')
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        flash(f'Could not delete: {e}', 'error')
    finally:
        conn.close()
    src = request.form.get('source', '')
    if src == 'project':
        return redirect(url_for('project_detail', project_id=ps['project_id']))
    return redirect(url_for('products_list'))


# ─── Sales News Routes ───

@app.route('/sales/news')
@login_required
@sales_access_required
def sales_news():
    user = get_user()
    conn = get_db()
    news = conn.execute('''
        SELECT n.*, e.name as posted_by_name, e.photo_url
        FROM sales_news n
        JOIN employees e ON n.posted_by = e.id
        WHERE n.is_active = 1
        ORDER BY n.created_at DESC
    ''').fetchall()
    conn.close()
    return render_template('sales_news.html', user=user, news=news,
                         can_post=user['is_admin'] or user['emp_code'] in MANAGEMENT_CODES)


@app.route('/sales/news/add', methods=['GET', 'POST'])
@login_required
@sales_access_required
def add_sales_news():
    user = get_user()
    if not user['is_admin'] and user['emp_code'] not in MANAGEMENT_CODES:
        flash('Only admin/management can post news', 'error')
        return redirect(url_for('sales_news'))

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        content = request.form.get('content', '').strip()
        category = request.form.get('category', 'general')

        if not title or not content:
            flash('Title and content are required', 'error')
            return redirect(url_for('add_sales_news'))

        conn = get_db()
        conn.execute('INSERT INTO sales_news (title, content, category, posted_by) VALUES (?, ?, ?, ?)',
                    (title, content, category, user['id']))
        conn.commit()
        conn.close()
        flash('News posted', 'success')
        return redirect(url_for('sales_news'))

    return render_template('add_sales_news.html', user=user)


@app.route('/sales/news/delete/<int:news_id>', methods=['POST'])
@login_required
@sales_access_required
def delete_sales_news(news_id):
    user = get_user()
    if not user['is_admin'] and user['emp_code'] not in MANAGEMENT_CODES:
        flash('Access denied', 'error')
        return redirect(url_for('sales_news'))

    conn = get_db()
    conn.execute('UPDATE sales_news SET is_active = 0 WHERE id = ?', (news_id,))
    conn.commit()
    conn.close()
    flash('News removed', 'success')
    return redirect(url_for('sales_news'))


# ─── Sales Dashboard ───

@app.route('/sales')
@login_required
@sales_access_required
def sales_dashboard():
    user = get_user()
    conn = get_db()

    now = datetime.now()
    month_start = now.strftime('%Y-%m-01')
    month_end = now.strftime('%Y-%m-') + str(calendar.monthrange(now.year, now.month)[1])
    month_name = calendar.month_name[now.month]

    # Determine if user sees company-wide or personal view
    is_company_view = user['is_admin'] == 1 or user['emp_code'] in MANAGEMENT_CODES
    emp_filter = '' if is_company_view else ' AND t.employee_id = ?'
    emp_params = [] if is_company_view else [user['id']]

    # Recent news (always company-wide)
    recent_news = conn.execute('''
        SELECT n.*, e.name as posted_by_name, e.photo_url
        FROM sales_news n JOIN employees e ON n.posted_by = e.id
        WHERE n.is_active = 1
        ORDER BY n.created_at DESC LIMIT 5
    ''').fetchall()

    # Recent B2B trips (filtered for employees)
    recent_trips = conn.execute('''
        SELECT t.*, e.name as employee_name, e.photo_url as emp_photo, e.emp_code, p.name as project_name,
               (SELECT COUNT(*) FROM b2b_meetings m WHERE m.trip_id = t.id) as meeting_count
        FROM b2b_trips t
        JOIN employees e ON t.employee_id = e.id
        LEFT JOIN projects p ON t.project_id = p.id
        WHERE 1=1''' + emp_filter + '''
        ORDER BY t.created_at DESC LIMIT 5
    ''', emp_params).fetchall()

    # Active projects count (company-wide, all can see)
    project_count = conn.execute("SELECT COUNT(*) as cnt FROM projects WHERE status = 'active'").fetchone()['cnt']

    # Total products/services count
    product_count = conn.execute("SELECT COUNT(*) as cnt FROM products_services").fetchone()['cnt']

    # Total trips count (filtered)
    if is_company_view:
        total_trips = conn.execute("SELECT COUNT(*) as cnt FROM b2b_trips").fetchone()['cnt']
        outstation_count = conn.execute("SELECT COUNT(*) as cnt FROM b2b_trips WHERE trip_type = 'outstation'").fetchone()['cnt']
        city_count = conn.execute("SELECT COUNT(*) as cnt FROM b2b_trips WHERE trip_type = 'city'").fetchone()['cnt']
    else:
        total_trips = conn.execute("SELECT COUNT(*) as cnt FROM b2b_trips WHERE employee_id = ?", (user['id'],)).fetchone()['cnt']
        outstation_count = conn.execute("SELECT COUNT(*) as cnt FROM b2b_trips WHERE trip_type = 'outstation' AND employee_id = ?", (user['id'],)).fetchone()['cnt']
        city_count = conn.execute("SELECT COUNT(*) as cnt FROM b2b_trips WHERE trip_type = 'city' AND employee_id = ?", (user['id'],)).fetchone()['cnt']

    # Meeting stats this month (filtered)
    if is_company_view:
        meeting_count = conn.execute('''
            SELECT COUNT(*) as cnt FROM b2b_meetings
            WHERE meeting_date >= ? AND meeting_date <= ?
        ''', (month_start, month_end)).fetchone()['cnt']
    else:
        meeting_count = conn.execute('''
            SELECT COUNT(*) as cnt FROM b2b_meetings m
            JOIN b2b_trips t ON m.trip_id = t.id
            WHERE m.meeting_date >= ? AND m.meeting_date <= ? AND t.employee_id = ?
        ''', (month_start, month_end, user['id'])).fetchone()['cnt']

    # Meeting type breakdown this month (filtered)
    if is_company_view:
        meeting_types = conn.execute('''
            SELECT mt.name, COUNT(m.id) as cnt
            FROM b2b_meetings m
            JOIN meeting_types mt ON m.meeting_type_id = mt.id
            WHERE m.meeting_date >= ? AND m.meeting_date <= ?
            GROUP BY mt.name ORDER BY cnt DESC
        ''', (month_start, month_end)).fetchall()
    else:
        meeting_types = conn.execute('''
            SELECT mt.name, COUNT(m.id) as cnt
            FROM b2b_meetings m
            JOIN meeting_types mt ON m.meeting_type_id = mt.id
            JOIN b2b_trips t ON m.trip_id = t.id
            WHERE m.meeting_date >= ? AND m.meeting_date <= ? AND t.employee_id = ?
            GROUP BY mt.name ORDER BY cnt DESC
        ''', (month_start, month_end, user['id'])).fetchall()

    # Top performers (always company-wide — everyone can see the leaderboard)
    top_performers = conn.execute('''
        SELECT e.name, e.photo_url, e.emp_code, COUNT(m.id) as meeting_count
        FROM b2b_meetings m
        JOIN b2b_trips t ON m.trip_id = t.id
        JOIN employees e ON t.employee_id = e.id
        WHERE m.meeting_date >= ? AND m.meeting_date <= ?
        GROUP BY e.id ORDER BY meeting_count DESC LIMIT 5
    ''', (month_start, month_end)).fetchall()

    # Active projects list (company-wide)
    active_projects = conn.execute('''
        SELECT p.*, (SELECT COUNT(*) FROM products_services ps WHERE ps.project_id = p.id) as product_count
        FROM projects p WHERE p.status = 'active' ORDER BY p.name LIMIT 6
    ''').fetchall()

    # Upcoming meetings (next 7 days, filtered)
    today_str = now.strftime('%Y-%m-%d')
    week_end = (now + timedelta(days=7)).strftime('%Y-%m-%d')
    if is_company_view:
        upcoming_meetings = conn.execute('''
            SELECT m.*, mt.name as type_name, e.name as emp_name, e.photo_url as emp_photo, t.trip_type
            FROM b2b_meetings m
            JOIN meeting_types mt ON m.meeting_type_id = mt.id
            JOIN b2b_trips t ON m.trip_id = t.id
            JOIN employees e ON t.employee_id = e.id
            WHERE m.meeting_date >= ? AND m.meeting_date <= ?
            ORDER BY m.meeting_date ASC LIMIT 8
        ''', (today_str, week_end)).fetchall()
    else:
        upcoming_meetings = conn.execute('''
            SELECT m.*, mt.name as type_name, e.name as emp_name, e.photo_url as emp_photo, t.trip_type
            FROM b2b_meetings m
            JOIN meeting_types mt ON m.meeting_type_id = mt.id
            JOIN b2b_trips t ON m.trip_id = t.id
            JOIN employees e ON t.employee_id = e.id
            WHERE m.meeting_date >= ? AND m.meeting_date <= ? AND t.employee_id = ?
            ORDER BY m.meeting_date ASC LIMIT 8
        ''', (today_str, week_end, user['id'])).fetchall()

    conn.close()
    return render_template('sales_dashboard.html', user=user,
                         recent_news=recent_news, recent_trips=recent_trips,
                         project_count=project_count, product_count=product_count,
                         total_trips=total_trips, meeting_count=meeting_count,
                         outstation_count=outstation_count, city_count=city_count,
                         meeting_types=meeting_types, top_performers=top_performers,
                         active_projects=active_projects, upcoming_meetings=upcoming_meetings,
                         month_name=month_name, current_year=now.year,
                         is_company_view=is_company_view)


# ─── B2B Meetings Routes ───

@app.route('/b2b')
@login_required
@sales_access_required
def b2b_trips_list():
    user = get_user()
    conn = get_db()

    trip_type = request.args.get('type', '')
    f_from = request.args.get('from_date', '')
    f_to = request.args.get('to_date', '')

    is_company_view = user['is_admin'] == 1 or user['emp_code'] in MANAGEMENT_CODES

    query = '''
        SELECT t.*, e.name as employee_name, e.emp_code, e.photo_url as emp_photo, p.name as project_name,
               t.meeting_category, (SELECT COUNT(*) FROM b2b_meetings m WHERE m.trip_id = t.id) as meeting_count
        FROM b2b_trips t
        JOIN employees e ON t.employee_id = e.id
        LEFT JOIN projects p ON t.project_id = p.id
        WHERE 1=1
    '''
    params = []

    # Non-admin/non-management employees only see their own trips
    if not is_company_view:
        query += ' AND t.employee_id = ?'
        params.append(user['id'])

    if trip_type:
        query += ' AND t.trip_type = ?'
        params.append(trip_type)
    if f_from:
        query += ' AND t.from_date >= ?'
        params.append(f_from)
    if f_to:
        query += ' AND t.to_date <= ?'
        params.append(f_to)

    query += ' ORDER BY t.from_date DESC'
    trips = conn.execute(query, params).fetchall()
    conn.close()

    return render_template('meetings_list.html', user=user, trips=trips,
                         trip_type=trip_type, f_from=f_from, f_to=f_to,
                         is_company_view=is_company_view)


@app.route('/b2b/add', methods=['GET', 'POST'])
@login_required
@sales_access_required
def add_b2b_trip():
    user = get_user()
    conn = get_db()

    if request.method == 'POST':
        trip_type = request.form.get('trip_type')
        meeting_category = request.form.get('meeting_category', 'face_to_face')
        from_date = request.form.get('from_date')
        to_date = request.form.get('to_date')
        travel_date = request.form.get('travel_date', '')
        project_id = request.form.get('project_id') or None
        notes = request.form.get('notes', '').strip()

        if not from_date or not to_date:
            flash('From date and to date are required', 'error')
            projects = conn.execute("SELECT id, name FROM projects WHERE status = 'active' ORDER BY name").fetchall()
            meeting_clients = conn.execute("SELECT * FROM meeting_types WHERE is_active = 1 ORDER BY name").fetchall()
            conn.close()
            return render_template('add_meeting.html', user=user, projects=projects, meeting_clients=meeting_clients)

        conn.execute('''
            INSERT INTO b2b_trips (employee_id, trip_type, meeting_category, from_date, to_date, travel_date, project_id, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (user['id'], trip_type, meeting_category, from_date, to_date, travel_date or None, project_id, notes))
        conn.commit()

        # Get the new trip ID
        trip = conn.execute('SELECT id FROM b2b_trips WHERE employee_id = ? ORDER BY id DESC LIMIT 1', (user['id'],)).fetchone()
        trip_id = trip['id']

        # Process multiple meetings
        meeting_dates = request.form.getlist('meeting_date[]')
        meeting_withs = request.form.getlist('meeting_with[]')
        meeting_type_ids = request.form.getlist('meeting_type_id[]')
        meeting_locations = request.form.getlist('meeting_location[]')
        meeting_contacts = request.form.getlist('contact_person[]')
        meeting_phones = request.form.getlist('contact_phone[]')
        meeting_project_ids = request.form.getlist('meeting_project_id[]')

        for i in range(len(meeting_dates)):
            if meeting_dates[i] and meeting_withs[i]:
                m_project_id = meeting_project_ids[i] if i < len(meeting_project_ids) and meeting_project_ids[i] else project_id
                m_type_id = meeting_type_ids[i] if i < len(meeting_type_ids) and meeting_type_ids[i] else None
                conn.execute('''
                    INSERT INTO b2b_meetings (trip_id, meeting_type_id, meeting_with, meeting_date, project_id, location, contact_person, contact_phone)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (trip_id, m_type_id, meeting_withs[i], meeting_dates[i],
                      m_project_id, meeting_locations[i] if i < len(meeting_locations) else '',
                      meeting_contacts[i] if i < len(meeting_contacts) else '',
                      meeting_phones[i] if i < len(meeting_phones) else ''))

        conn.commit()
        conn.close()
        flash('Meeting created successfully', 'success')
        return redirect(url_for('b2b_trip_detail', trip_id=trip_id))

    projects = conn.execute("SELECT id, name FROM projects WHERE status = 'active' ORDER BY name").fetchall()
    meeting_clients = conn.execute("SELECT * FROM meeting_types WHERE is_active = 1 ORDER BY name").fetchall()
    conn.close()
    return render_template('add_meeting.html', user=user, projects=projects, meeting_clients=meeting_clients)


@app.route('/b2b/<int:trip_id>')
@login_required
@sales_access_required
def b2b_trip_detail(trip_id):
    user = get_user()
    conn = get_db()
    trip = conn.execute('''
        SELECT t.*, e.name as employee_name, e.emp_code, p.name as project_name
        FROM b2b_trips t
        JOIN employees e ON t.employee_id = e.id
        LEFT JOIN projects p ON t.project_id = p.id
        WHERE t.id = ?
    ''', (trip_id,)).fetchone()

    if not trip:
        flash('Trip not found', 'error')
        conn.close()
        return redirect(url_for('b2b_trips_list'))

    meetings = conn.execute('''
        SELECT m.*, mt.name as meeting_type_name, p.name as project_name
        FROM b2b_meetings m
        LEFT JOIN meeting_types mt ON m.meeting_type_id = mt.id
        LEFT JOIN projects p ON m.project_id = p.id
        WHERE m.trip_id = ?
        ORDER BY m.meeting_date, m.id
    ''', (trip_id,)).fetchall()

    meeting_clients = conn.execute("SELECT * FROM meeting_types WHERE is_active = 1 ORDER BY name").fetchall()
    projects = conn.execute("SELECT id, name FROM projects WHERE status = 'active' ORDER BY name").fetchall()

    conn.close()
    return render_template('meeting_detail.html', user=user, trip=trip, meetings=meetings, meeting_clients=meeting_clients, projects=projects)


@app.route('/b2b/<int:trip_id>/add-meeting', methods=['POST'])
@login_required
@sales_access_required
def add_meeting_to_trip(trip_id):
    user = get_user()
    conn = get_db()

    trip = conn.execute('SELECT * FROM b2b_trips WHERE id = ?', (trip_id,)).fetchone()
    if not trip:
        flash('Trip not found', 'error')
        conn.close()
        return redirect(url_for('b2b_trips_list'))

    meeting_date = request.form.get('meeting_date')
    meeting_with = request.form.get('meeting_with', '').strip()
    meeting_type_id = request.form.get('meeting_type_id') or None
    project_id = request.form.get('project_id') or trip['project_id']
    location = request.form.get('location', '').strip()
    contact_person = request.form.get('contact_person', '').strip()
    contact_phone = request.form.get('contact_phone', '').strip()

    if not meeting_date or not meeting_with:
        flash('Meeting date and name are required', 'error')
        conn.close()
        return redirect(url_for('b2b_trip_detail', trip_id=trip_id))

    conn.execute('''
        INSERT INTO b2b_meetings (trip_id, meeting_type_id, meeting_with, meeting_date, project_id, location, contact_person, contact_phone)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (trip_id, meeting_type_id, meeting_with, meeting_date, project_id, location, contact_person, contact_phone))
    conn.commit()
    conn.close()
    flash('Meeting added', 'success')
    return redirect(url_for('b2b_trip_detail', trip_id=trip_id))


@app.route('/b2b/meeting/<int:meeting_id>/outcome', methods=['POST'])
@login_required
@sales_access_required
def update_meeting_outcome(meeting_id):
    user = get_user()
    conn = get_db()

    meeting = conn.execute('SELECT * FROM b2b_meetings WHERE id = ?', (meeting_id,)).fetchone()
    if not meeting:
        flash('Meeting not found', 'error')
        conn.close()
        return redirect(url_for('b2b_trips_list'))

    outcome = request.form.get('outcome', '').strip()
    notes = request.form.get('notes', '').strip()

    conn.execute('UPDATE b2b_meetings SET outcome = ?, notes = ? WHERE id = ?',
                (outcome, notes, meeting_id))
    conn.commit()
    conn.close()
    flash('Meeting outcome updated', 'success')
    return redirect(url_for('b2b_trip_detail', trip_id=meeting['trip_id']))


# ─── Meeting Types Management (Admin) ───

@app.route('/admin/meeting-types', methods=['GET', 'POST'])
@admin_required
def manage_meeting_types():
    user = get_user()
    conn = get_db()

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            name = request.form.get('name', '').strip()
            if name:
                conn.execute('INSERT INTO meeting_types (name) VALUES (?)', (name,))
                conn.commit()
                flash('Meeting type added', 'success')
        elif action == 'toggle':
            mt_id = request.form.get('id')
            conn.execute('UPDATE meeting_types SET is_active = CASE WHEN is_active = 1 THEN 0 ELSE 1 END WHERE id = ?', (mt_id,))
            conn.commit()
            flash('Meeting type updated', 'success')
        elif action == 'delete':
            mt_id = request.form.get('id')
            conn.execute('DELETE FROM meeting_types WHERE id = ?', (mt_id,))
            conn.commit()
            flash('Meeting client type deleted', 'success')
        conn.close()
        return redirect(url_for('manage_meeting_types'))

    types = conn.execute('SELECT * FROM meeting_types ORDER BY name').fetchall()
    conn.close()
    return render_template('manage_meeting_types.html', user=user, types=types, page_title='Meeting Client Types')


# ─── Module Access Management (Admin) ───

@app.route('/admin/module-access', methods=['GET', 'POST'])
@admin_required
def manage_module_access():
    user = get_user()
    conn = get_db()

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'grant':
            employee_ids = request.form.getlist('employee_ids')
            modules = request.form.getlist('modules')
            if employee_ids and modules:
                for emp_id in employee_ids:
                    for module in modules:
                        conn.execute('''
                            INSERT INTO module_access (employee_id, module, granted_by, is_active)
                            VALUES (?, ?, ?, 1)
                            ON CONFLICT (employee_id, module) DO UPDATE SET is_active = 1, granted_by = ?
                        ''', (emp_id, module, user['id'], user['id']))
                conn.commit()
                flash(f'Access granted to {len(employee_ids)} employee(s) for {len(modules)} module(s)', 'success')
            else:
                flash('Please select at least one employee and one module', 'error')

        elif action == 'toggle':
            access_id = request.form.get('access_id')
            new_status = request.form.get('new_status', '1')
            if access_id:
                conn.execute('UPDATE module_access SET is_active = ? WHERE id = ?', (int(new_status), access_id))
                conn.commit()
                flash('Access status updated', 'success')

        elif action == 'revoke':
            access_id = request.form.get('access_id')
            if access_id:
                conn.execute('DELETE FROM module_access WHERE id = ?', (access_id,))
                conn.commit()
                flash('Access revoked', 'success')

        conn.close()
        return redirect(url_for('manage_module_access'))

    employees = conn.execute("SELECT id, name, emp_code, department, photo_url FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name").fetchall()
    access_list = conn.execute('''
        SELECT ma.id as access_id, ma.employee_id, ma.module, ma.is_active, ma.created_at,
               e.name, e.emp_code, e.department, e.photo_url,
               g.name as granted_by_name
        FROM module_access ma
        JOIN employees e ON ma.employee_id = e.id
        LEFT JOIN employees g ON ma.granted_by = g.id
        ORDER BY e.name, ma.module
    ''').fetchall()

    modules = ['sales', 'projects', 'b2b_meetings']
    conn.close()
    return render_template('manage_module_access.html', user=user, employees=employees,
                         access_list=access_list, modules=modules)


def ensure_management_admins():
    """Ensure MANAGEMENT_CODES employees always have is_admin = 1."""
    try:
        conn = get_db()
        for code in MANAGEMENT_CODES:
            conn.execute('UPDATE employees SET is_admin = 1 WHERE emp_code = ? AND is_admin = 0', (code,))
        conn.commit()
        conn.close()
    except Exception as e:
        logging.error(f"ensure_management_admins: {e}")

# ─── Leave Applications (All) ─── Admin / Management view ───
@app.route('/admin/leave-applications')
@login_required
@admin_required
def admin_leave_applications():
    user = get_user()
    conn = get_db()

    # Read filter params
    f_employee = request.args.get('employee', '').strip()
    f_type = request.args.get('leave_type', '').strip()
    f_status = request.args.get('status', '').strip()
    f_from = request.args.get('from_date', '').strip()
    f_to = request.args.get('to_date', '').strip()

    # Build query
    query = '''
        SELECT lr.id, lr.employee_id, lr.leave_type, lr.leave_date, lr.days,
               lr.day_portion, lr.reason, lr.status, lr.is_late,
               lr.original_id, lr.modification_reason, lr.original_reason,
               lr.approved_by, lr.approved_at, lr.created_at,
               e.name        AS employee_name,
               e.emp_code    AS emp_code,
               e.department  AS department,
               e.designation AS designation,
               e.email       AS email,
               e.phone       AS mobile,
               e.photo_url   AS photo_url,
               a.name        AS approver_name
        FROM leave_records lr
        JOIN employees e ON lr.employee_id = e.id
        LEFT JOIN employees a ON lr.approved_by = a.id
        WHERE 1=1
    '''
    params = []

    if f_employee:
        query += " AND (LOWER(e.name) LIKE LOWER(?) OR LOWER(e.emp_code) LIKE LOWER(?))"
        params.extend([f'%{f_employee}%', f'%{f_employee}%'])
    if f_type:
        query += " AND lr.leave_type = ?"
        params.append(f_type)
    if f_status:
        query += " AND lr.status = ?"
        params.append(f_status)
    if f_from:
        query += " AND lr.leave_date >= ?"
        params.append(f_from)
    if f_to:
        query += " AND lr.leave_date <= ?"
        params.append(f_to)

    query += " ORDER BY lr.leave_date DESC, lr.created_at DESC"

    leaves_raw = conn.execute(query, params).fetchall()

    # Enrich rows
    leaves = []
    for r in leaves_raw:
        d = dict(r)
        p = d.get('photo_url') or ''
        d['photo_src'] = (p if p.startswith('http') else f'/static/photos/{p}') if p else ''
        raw = d.get('reason') or ''
        if '[Late Application Reason]:' in raw:
            parts = raw.split('\n\n[Late Application Reason]:', 1)
            d['reason_main'] = parts[0].strip()
            d['late_reason'] = parts[1].strip() if len(parts) > 1 else ''
        else:
            d['reason_main'] = raw
            d['late_reason'] = ''
        leaves.append(d)

    # Get employee list for filter dropdown
    employees = conn.execute(
        "SELECT id, name, emp_code FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name"
    ).fetchall()

    # Summary counts
    total = len(leaves)
    pending_count = sum(1 for l in leaves if l['status'] == 'pending')
    approved_count = sum(1 for l in leaves if l['status'] == 'approved')
    rejected_count = sum(1 for l in leaves if l['status'] == 'rejected')

    conn.close()
    today_str = datetime.now().strftime('%Y-%m-%d')
    return render_template('admin_leave_applications.html',
        user=user, leaves=leaves, employees=employees,
        total=total, pending_count=pending_count,
        approved_count=approved_count, rejected_count=rejected_count,
        f_employee=f_employee, f_type=f_type, f_status=f_status,
        f_from=f_from, f_to=f_to, today=today_str)


# ─── My Leave Applications ─── Employee's own leaves ───
@app.route('/my-applications')
@login_required
def my_leave_applications():
    user = get_user()

    # Admin control account has no personal applications
    if user['emp_code'] == 'admin':
        return redirect(url_for('admin_leave_applications'))

    conn = get_db()

    f_type = request.args.get('leave_type', '').strip()
    f_status = request.args.get('status', '').strip()
    f_from = request.args.get('from_date', '').strip()
    f_to = request.args.get('to_date', '').strip()

    query = '''
        SELECT lr.*, e.name, e.emp_code, e.department
        FROM leave_records lr
        JOIN employees e ON lr.employee_id = e.id
        WHERE lr.employee_id = ?
    '''
    params = [user['id']]

    if f_type:
        query += " AND lr.leave_type = ?"
        params.append(f_type)
    if f_status:
        query += " AND lr.status = ?"
        params.append(f_status)
    if f_from:
        query += " AND lr.leave_date >= ?"
        params.append(f_from)
    if f_to:
        query += " AND lr.leave_date <= ?"
        params.append(f_to)

    query += " ORDER BY lr.leave_date DESC, lr.created_at DESC"

    leaves = conn.execute(query, params).fetchall()

    total = len(leaves)
    pending_count = sum(1 for l in leaves if l['status'] == 'pending')
    approved_count = sum(1 for l in leaves if l['status'] == 'approved')
    rejected_count = sum(1 for l in leaves if l['status'] == 'rejected')

    conn.close()
    today_str = datetime.now().strftime('%Y-%m-%d')
    return render_template('my_leave_applications.html',
        user=user, leaves=leaves,
        total=total, pending_count=pending_count,
        approved_count=approved_count, rejected_count=rejected_count,
        f_type=f_type, f_status=f_status,
        f_from=f_from, f_to=f_to, today=today_str)


# ─── Team Leave Applications ─── Manager view of team leaves ───
@app.route('/team-applications')
@login_required
def team_leave_applications():
    user = get_user()
    conn = get_db()

    # Check if user is a manager or management
    direct_reports = conn.execute(
        "SELECT id FROM employees WHERE reporting_to = ? AND is_active = 1",
        (user['id'],)
    ).fetchall()

    is_mgmt = user['emp_code'] in MANAGEMENT_CODES

    if not direct_reports and not is_mgmt:
        flash('You do not have team members to view', 'error')
        conn.close()
        return redirect(url_for('dashboard'))

    # Build list of viewable employee IDs
    viewable_ids = [r['id'] for r in direct_reports]

    # Management can also see other management members' leaves
    if is_mgmt:
        mgmt_employees = conn.execute(
            "SELECT id FROM employees WHERE emp_code IN ({})".format(
                ','.join('?' * len(MANAGEMENT_CODES))
            ), MANAGEMENT_CODES
        ).fetchall()
        for m in mgmt_employees:
            if m['id'] not in viewable_ids and m['id'] != user['id']:
                viewable_ids.append(m['id'])

    f_employee = request.args.get('employee', '').strip()
    f_type = request.args.get('leave_type', '').strip()
    f_status = request.args.get('status', '').strip()
    f_from = request.args.get('from_date', '').strip()
    f_to = request.args.get('to_date', '').strip()

    placeholders = ','.join('?' * len(viewable_ids))
    query = f'''
        SELECT lr.id, lr.employee_id, lr.leave_type, lr.leave_date, lr.days,
               lr.day_portion, lr.reason, lr.status, lr.is_late,
               lr.original_id, lr.modification_reason, lr.original_reason,
               lr.approved_by, lr.approved_at, lr.created_at,
               e.name        AS employee_name,
               e.emp_code    AS emp_code,
               e.department  AS department,
               e.designation AS designation,
               e.email       AS email,
               e.phone       AS mobile,
               e.photo_url   AS photo_url,
               a.name        AS approver_name
        FROM leave_records lr
        JOIN employees e ON lr.employee_id = e.id
        LEFT JOIN employees a ON lr.approved_by = a.id
        WHERE lr.employee_id IN ({placeholders})
    '''
    params = list(viewable_ids)

    if f_employee:
        query += " AND (LOWER(e.name) LIKE LOWER(?) OR LOWER(e.emp_code) LIKE LOWER(?))"
        params.extend([f'%{f_employee}%', f'%{f_employee}%'])
    if f_type:
        query += " AND lr.leave_type = ?"
        params.append(f_type)
    if f_status:
        query += " AND lr.status = ?"
        params.append(f_status)
    if f_from:
        query += " AND lr.leave_date >= ?"
        params.append(f_from)
    if f_to:
        query += " AND lr.leave_date <= ?"
        params.append(f_to)

    query += " ORDER BY lr.leave_date DESC, lr.created_at DESC"

    leaves_raw = conn.execute(query, params).fetchall()

    # Enrich rows
    leaves = []
    for r in leaves_raw:
        d = dict(r)
        p = d.get('photo_url') or ''
        d['photo_src'] = (p if p.startswith('http') else f'/static/photos/{p}') if p else ''
        raw = d.get('reason') or ''
        if '[Late Application Reason]:' in raw:
            parts = raw.split('\n\n[Late Application Reason]:', 1)
            d['reason_main'] = parts[0].strip()
            d['late_reason'] = parts[1].strip() if len(parts) > 1 else ''
        else:
            d['reason_main'] = raw
            d['late_reason'] = ''
        leaves.append(d)

    # Get team member names for filter
    team_members = conn.execute(
        f"SELECT id, name, emp_code FROM employees WHERE id IN ({placeholders}) ORDER BY name",
        viewable_ids
    ).fetchall()

    total = len(leaves)
    pending_count = sum(1 for l in leaves if l['status'] == 'pending')
    approved_count = sum(1 for l in leaves if l['status'] == 'approved')
    rejected_count = sum(1 for l in leaves if l['status'] == 'rejected')

    conn.close()
    return render_template('team_leave_applications.html',
        user=user, leaves=leaves, team_members=team_members,
        total=total, pending_count=pending_count,
        approved_count=approved_count, rejected_count=rejected_count,
        f_employee=f_employee, f_type=f_type, f_status=f_status,
        f_from=f_from, f_to=f_to)


# ─── Notification Routes ───

@app.route('/notifications')
@login_required
def notifications_page():
    """Show all notifications for the logged-in employee."""
    user = get_user()
    conn = get_db()
    notifs = conn.execute(
        'SELECT * FROM notifications WHERE employee_id = ? ORDER BY created_at DESC LIMIT 50',
        (user['id'],)
    ).fetchall()
    unread = sum(1 for n in notifs if not n['is_read'])
    conn.close()
    return render_template('notifications.html', notifications=notifs, unread_count=unread)


@app.route('/notifications/mark-read/<int:notif_id>', methods=['POST'])
@login_required
def mark_notification_read(notif_id):
    """Mark a single notification as read."""
    user = get_user()
    conn = get_db()
    conn.execute('UPDATE notifications SET is_read = 1 WHERE id = ? AND employee_id = ?',
                 (notif_id, user['id']))
    conn.commit()
    conn.close()
    return redirect(request.referrer or url_for('notifications_page'))


@app.route('/notifications/mark-all-read', methods=['POST'])
@login_required
def mark_all_notifications_read():
    """Mark all notifications as read."""
    user = get_user()
    conn = get_db()
    conn.execute('UPDATE notifications SET is_read = 1 WHERE employee_id = ? AND is_read = 0',
                 (user['id'],))
    conn.commit()
    conn.close()
    return redirect(url_for('notifications_page'))


@app.route('/api/notifications/count')
@login_required
def api_notification_count():
    """Return unread notification count as JSON."""
    user = get_user()
    conn = get_db()
    row = conn.execute(
        'SELECT COUNT(*) as cnt FROM notifications WHERE employee_id = ? AND is_read = 0',
        (user['id'],)
    ).fetchone()
    conn.close()
    return jsonify({'unread': row['cnt'] if row else 0})


# ─── KRA (Key Result Area) Routes ───

@app.route('/kra/admin/templates')
@login_required
@admin_required
def kra_admin_templates():
    """Admin: list all KRA templates."""
    user = get_user()
    conn = get_db()
    templates = conn.execute('''
        SELECT t.*, e.name as creator_name,
               (SELECT COUNT(*) FROM kra_template_items WHERE template_id = t.id) as item_count,
               (SELECT COUNT(*) FROM kra_assignments WHERE template_id = t.id AND is_active = 1) as assign_count
        FROM kra_templates t
        LEFT JOIN employees e ON t.created_by = e.id
        ORDER BY t.fy_year DESC, t.name
    ''').fetchall()
    conn.close()
    return render_template('kra_admin_templates.html', user=user, templates=templates)


@app.route('/kra/admin/templates/create', methods=['GET', 'POST'])
@login_required
@admin_required
def kra_admin_template_create():
    """Admin: create a new KRA template."""
    user = get_user()
    conn = get_db()

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        department = request.form.get('department', '').strip()
        role_title = request.form.get('role_title', '').strip()
        now = datetime.now()
        fy_year = int(request.form.get('fy_year', now.year if now.month >= 4 else now.year - 1))

        if not name:
            flash('Template name is required', 'error')
            conn.close()
            return redirect(url_for('kra_admin_template_create'))

        conn.execute(
            'INSERT INTO kra_templates (name, department, role_title, fy_year, created_by) VALUES (?, ?, ?, ?, ?)',
            (name, department, role_title, fy_year, user['id'])
        )
        conn.commit()
        tid = conn.execute('SELECT MAX(id) as mid FROM kra_templates').fetchone()['mid']

        # Auto-add common KPIs (categories 2-6)
        common_cats = conn.execute('SELECT * FROM kra_categories WHERE is_common = 1 ORDER BY sort_order').fetchall()
        common_kpis = {
            'Knowledge': [
                ('Product Knowledge & Training', 0.5, 0.05),
                ('Process Knowledge', 0.5, 0.05),
                ('Industry news & awareness', 0.2, 0.025),
            ],
            'Customer Handling': [
                ('Call Quality', 0.6, 0.05),
                ('Handling Customer Queries', 0.6, 0.02),
                ('Issue Management', 0.5, 0.02),
                ('Customer Follow-up', 0.5, 0.02),
                ('Rapport with Customer', 0.6, 0.02),
            ],
            'HR': [
                ('Training', 0.8, 0.02),
                ('Time Management', 0.8, 0.02),
                ('HR & Leave Policy', 0.6, 0.02),
            ],
            'Extra Mile': [
                ('Idea Generation', 0.1, 0.02),
                ('Discussion & feedback', 0.2, 0.01),
                ('Inter functional relationship', 0.5, 0.025),
            ],
            'Interpersonal Skill': [
                ('Communication with Co-workers', 0.75, 0.015),
                ('Office Equipment Care', 0.75, 0.02),
                ('Table & Office hygiene', 0.7, 0.02),
            ],
        }
        for cat in common_cats:
            kpis = common_kpis.get(cat['name'], [])
            cat_num = cat['sort_order']
            for idx, (desc, pct, pct_share) in enumerate(kpis, 1):
                kpi_code = f"{cat_num}.1.{idx}"
                conn.execute(
                    '''INSERT INTO kra_template_items
                    (template_id, category_id, kpi_code, measure_description, percentage, percentage_sharing, is_target_based, sort_order)
                    VALUES (?, ?, ?, ?, ?, ?, 0, ?)''',
                    (tid, cat['id'], kpi_code, desc, pct, pct_share, idx)
                )
        conn.commit()
        conn.close()
        flash(f'Template "{name}" created with common KPIs. Add Target KPIs now.', 'success')
        return redirect(url_for('kra_admin_template_edit', template_id=tid))

    categories = conn.execute('SELECT * FROM kra_categories ORDER BY sort_order').fetchall()
    conn.close()
    now = datetime.now()
    fy_year = now.year if now.month >= 4 else now.year - 1
    return render_template('kra_admin_template_create.html', user=user, categories=categories, fy_year=fy_year)


@app.route('/kra/admin/templates/<int:template_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def kra_admin_template_edit(template_id):
    """Admin: edit a KRA template — add/remove/reorder KPIs."""
    user = get_user()
    conn = get_db()
    template = conn.execute('SELECT * FROM kra_templates WHERE id = ?', (template_id,)).fetchone()
    if not template:
        conn.close()
        flash('Template not found', 'error')
        return redirect(url_for('kra_admin_templates'))

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add_item':
            cat_id = int(request.form.get('category_id'))
            kpi_code = request.form.get('kpi_code', '').strip()
            desc = request.form.get('measure_description', '').strip()
            target_val = request.form.get('target_value', '').strip()
            pct = float(request.form.get('percentage', 0) or 0)
            pct_share = float(request.form.get('percentage_sharing', 0) or 0)
            is_target = 1 if request.form.get('is_target_based') else 0

            if not kpi_code or not desc:
                flash('KPI code and description are required', 'error')
            else:
                conn.execute(
                    '''INSERT INTO kra_template_items
                    (template_id, category_id, kpi_code, measure_description, target_value, percentage, percentage_sharing, is_target_based, sort_order)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (template_id, cat_id, kpi_code, desc,
                     float(target_val) if target_val else None, pct, pct_share, is_target,
                     conn.execute('SELECT COALESCE(MAX(sort_order),0)+1 as n FROM kra_template_items WHERE template_id = ?', (template_id,)).fetchone()['n'])
                )
                conn.commit()
                flash('KPI added', 'success')

        elif action == 'delete_item':
            item_id = int(request.form.get('item_id'))
            conn.execute('DELETE FROM kra_template_items WHERE id = ? AND template_id = ?', (item_id, template_id))
            conn.commit()
            flash('KPI removed', 'success')

        elif action == 'update_item':
            item_id = int(request.form.get('item_id'))
            desc = request.form.get('measure_description', '').strip()
            target_val = request.form.get('target_value', '').strip()
            pct = float(request.form.get('percentage', 0) or 0)
            pct_share = float(request.form.get('percentage_sharing', 0) or 0)
            conn.execute(
                '''UPDATE kra_template_items SET measure_description = ?, target_value = ?, percentage = ?, percentage_sharing = ?
                WHERE id = ? AND template_id = ?''',
                (desc, float(target_val) if target_val else None, pct, pct_share, item_id, template_id)
            )
            conn.commit()
            flash('KPI updated', 'success')

        elif action == 'update_template':
            name = request.form.get('name', '').strip()
            department = request.form.get('department', '').strip()
            role_title = request.form.get('role_title', '').strip()
            fy_year = int(request.form.get('fy_year', template['fy_year']))
            conn.execute(
                'UPDATE kra_templates SET name = ?, department = ?, role_title = ?, fy_year = ? WHERE id = ?',
                (name, department, role_title, fy_year, template_id)
            )
            conn.commit()
            flash('Template updated', 'success')

        return redirect(url_for('kra_admin_template_edit', template_id=template_id))

    # GET
    template = conn.execute('SELECT * FROM kra_templates WHERE id = ?', (template_id,)).fetchone()
    categories = conn.execute('SELECT * FROM kra_categories ORDER BY sort_order').fetchall()
    items = conn.execute('''
        SELECT ti.*, c.name as category_name, c.sort_order as cat_sort
        FROM kra_template_items ti
        JOIN kra_categories c ON ti.category_id = c.id
        WHERE ti.template_id = ?
        ORDER BY c.sort_order, ti.sort_order
    ''', (template_id,)).fetchall()

    total_weight = sum(item['percentage_sharing'] for item in items)
    conn.close()
    return render_template('kra_admin_template_edit.html', user=user, template=template,
                         categories=categories, items=items, total_weight=total_weight)


@app.route('/kra/admin/templates/<int:template_id>/duplicate')
@login_required
@admin_required
def kra_admin_template_duplicate(template_id):
    """Admin: duplicate a KRA template."""
    user = get_user()
    conn = get_db()
    orig = conn.execute('SELECT * FROM kra_templates WHERE id = ?', (template_id,)).fetchone()
    if not orig:
        conn.close()
        flash('Template not found', 'error')
        return redirect(url_for('kra_admin_templates'))

    now = datetime.now()
    fy_year = now.year if now.month >= 4 else now.year - 1
    conn.execute(
        'INSERT INTO kra_templates (name, department, role_title, fy_year, created_by) VALUES (?, ?, ?, ?, ?)',
        (orig['name'] + ' (Copy)', orig['department'], orig['role_title'], fy_year, user['id'])
    )
    conn.commit()
    new_id = conn.execute('SELECT MAX(id) as mid FROM kra_templates').fetchone()['mid']

    items = conn.execute('SELECT * FROM kra_template_items WHERE template_id = ?', (template_id,)).fetchall()
    for item in items:
        conn.execute(
            '''INSERT INTO kra_template_items
            (template_id, category_id, kpi_code, measure_description, target_value, percentage, percentage_sharing, is_target_based, sort_order)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (new_id, item['category_id'], item['kpi_code'], item['measure_description'],
             item['target_value'], item['percentage'], item['percentage_sharing'],
             item['is_target_based'], item['sort_order'])
        )
    conn.commit()
    conn.close()
    flash('Template duplicated', 'success')
    return redirect(url_for('kra_admin_template_edit', template_id=new_id))


@app.route('/kra/admin/assignments', methods=['GET', 'POST'])
@login_required
@admin_required
def kra_admin_assignments():
    """Admin: assign KRA templates to employees."""
    user = get_user()
    conn = get_db()

    if request.method == 'POST':
        template_id = int(request.form.get('template_id'))
        employee_ids = request.form.getlist('employee_ids')
        now = datetime.now()
        fy_year = int(request.form.get('fy_year', now.year if now.month >= 4 else now.year - 1))

        assigned_count = 0
        for eid in employee_ids:
            eid = int(eid)
            existing = conn.execute(
                'SELECT id FROM kra_assignments WHERE employee_id = ? AND fy_year = ? AND is_active = 1',
                (eid, fy_year)
            ).fetchone()
            if existing:
                conn.execute(
                    'UPDATE kra_assignments SET template_id = ?, assigned_by = ?, assigned_at = CURRENT_TIMESTAMP WHERE id = ?',
                    (template_id, user['id'], existing['id'])
                )
            else:
                conn.execute(
                    'INSERT INTO kra_assignments (template_id, employee_id, fy_year, assigned_by) VALUES (?, ?, ?, ?)',
                    (template_id, eid, fy_year, user['id'])
                )
            assigned_count += 1

        conn.commit()
        flash(f'KRA template assigned to {assigned_count} employee(s)', 'success')
        return redirect(url_for('kra_admin_assignments'))

    templates = conn.execute('SELECT * FROM kra_templates WHERE is_active = 1 ORDER BY fy_year DESC, name').fetchall()
    employees = conn.execute("SELECT * FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name").fetchall()
    now = datetime.now()
    fy_year = now.year if now.month >= 4 else now.year - 1

    assignments = conn.execute('''
        SELECT a.*, e.name as emp_name, e.emp_code, e.department, t.name as template_name, t.fy_year as t_fy
        FROM kra_assignments a
        JOIN employees e ON a.employee_id = e.id
        JOIN kra_templates t ON a.template_id = t.id
        WHERE a.is_active = 1
        ORDER BY t.fy_year DESC, e.name
    ''').fetchall()

    conn.close()
    return render_template('kra_admin_assignments.html', user=user, templates=templates,
                         employees=employees, assignments=assignments, fy_year=fy_year)


@app.route('/kra/admin/assignments/<int:assignment_id>/remove')
@login_required
@admin_required
def kra_admin_assignment_remove(assignment_id):
    """Admin: deactivate a KRA assignment."""
    conn = get_db()
    conn.execute('UPDATE kra_assignments SET is_active = 0 WHERE id = ?', (assignment_id,))
    conn.commit()
    conn.close()
    flash('Assignment removed', 'success')
    return redirect(url_for('kra_admin_assignments'))


@app.route('/kra/admin/report')
@login_required
@admin_required
def kra_admin_report():
    """Admin: KRA overview report across all employees."""
    user = get_user()
    conn = get_db()
    now = datetime.now()
    fy_year = int(request.args.get('fy', now.year if now.month >= 4 else now.year - 1))
    view_month = int(request.args.get('month', now.month))
    view_year = int(request.args.get('year', now.year))

    assignments = conn.execute('''
        SELECT a.id as assignment_id, a.employee_id, a.template_id,
               e.name as emp_name, e.emp_code, e.department, e.photo_url,
               t.name as template_name
        FROM kra_assignments a
        JOIN employees e ON a.employee_id = e.id
        JOIN kra_templates t ON a.template_id = t.id
        WHERE a.fy_year = ? AND a.is_active = 1
        ORDER BY e.name
    ''', (fy_year,)).fetchall()

    report_data = []
    for asn in assignments:
        ratings = conn.execute('''
            SELECT r.employee_result, r.manager_result, r.employee_submitted, r.manager_submitted
            FROM kra_monthly_ratings r
            WHERE r.assignment_id = ? AND r.month = ? AND r.year = ?
        ''', (asn['assignment_id'], view_month, view_year)).fetchall()

        emp_total = sum(r['employee_result'] or 0 for r in ratings)
        mgr_total = sum(r['manager_result'] or 0 for r in ratings)
        emp_submitted = all(r['employee_submitted'] for r in ratings) if ratings else False
        mgr_submitted = all(r['manager_submitted'] for r in ratings) if ratings else False

        report_data.append({
            'emp_name': asn['emp_name'],
            'emp_code': asn['emp_code'],
            'department': asn['department'],
            'photo_url': asn['photo_url'],
            'template_name': asn['template_name'],
            'employee_id': asn['employee_id'],
            'emp_score': round(emp_total, 4),
            'mgr_score': round(mgr_total, 4),
            'emp_submitted': emp_submitted,
            'mgr_submitted': mgr_submitted,
        })

    conn.close()
    return render_template('kra_admin_report.html', user=user, report_data=report_data,
                         fy_year=fy_year, view_month=view_month, view_year=view_year)


# ─── Employee KRA Routes ───

@app.route('/kra/dashboard')
@login_required
def kra_dashboard():
    """Employee KRA dashboard with stats and charts."""
    user = get_user()
    conn = get_db()
    now = datetime.now()
    fy_year = int(request.args.get('fy', now.year if now.month >= 4 else now.year - 1))

    assignment = conn.execute('''
        SELECT a.*, t.name as template_name, t.department, t.role_title
        FROM kra_assignments a
        JOIN kra_templates t ON a.template_id = t.id
        WHERE a.employee_id = ? AND a.fy_year = ? AND a.is_active = 1
    ''', (user['id'], fy_year)).fetchone()

    if not assignment:
        conn.close()
        return render_template('kra_dashboard.html', user=user, assignment=None, fy_year=fy_year,
                             monthly_scores=[], items=[], categories=[])

    items = conn.execute('''
        SELECT ti.*, c.name as category_name, c.sort_order as cat_sort
        FROM kra_template_items ti
        JOIN kra_categories c ON ti.category_id = c.id
        WHERE ti.template_id = ?
        ORDER BY c.sort_order, ti.sort_order
    ''', (assignment['template_id'],)).fetchall()

    # Monthly scores for chart
    monthly_scores = []
    fy_months = [(m, fy_year if m >= 4 else fy_year + 1) for m in [4,5,6,7,8,9,10,11,12,1,2,3]]
    for m, yr in fy_months:
        ratings = conn.execute('''
            SELECT SUM(employee_result) as emp_total, SUM(manager_result) as mgr_total,
                   MAX(employee_submitted) as emp_sub, MAX(manager_submitted) as mgr_sub
            FROM kra_monthly_ratings
            WHERE assignment_id = ? AND month = ? AND year = ?
        ''', (assignment['id'], m, yr)).fetchone()
        monthly_scores.append({
            'month': calendar.month_abbr[m],
            'month_num': m,
            'year': yr,
            'emp_score': round(ratings['emp_total'] or 0, 4),
            'mgr_score': round(ratings['mgr_total'] or 0, 4),
            'emp_submitted': bool(ratings['emp_sub']),
            'mgr_submitted': bool(ratings['mgr_sub']),
        })

    # Category-wise scores for current month
    current_month = now.month
    current_year = now.year
    cat_scores = []
    categories = conn.execute('SELECT * FROM kra_categories ORDER BY sort_order').fetchall()
    for cat in categories:
        cat_items = [i for i in items if i['category_name'] == cat['name']]
        if not cat_items:
            continue
        cat_ratings = conn.execute('''
            SELECT SUM(r.employee_result) as emp, SUM(r.manager_result) as mgr
            FROM kra_monthly_ratings r
            JOIN kra_template_items ti ON r.template_item_id = ti.id
            WHERE r.assignment_id = ? AND r.month = ? AND r.year = ? AND ti.category_id = ?
        ''', (assignment['id'], current_month, current_year, cat['id'])).fetchone()
        cat_weight = sum(i['percentage_sharing'] for i in cat_items)
        cat_scores.append({
            'name': cat['name'],
            'emp_score': round(cat_ratings['emp'] or 0, 4),
            'mgr_score': round(cat_ratings['mgr'] or 0, 4),
            'max_score': round(cat_weight * 5, 4),
            'weight': round(cat_weight * 100, 1),
        })

    conn.close()
    return render_template('kra_dashboard.html', user=user, assignment=assignment,
                         fy_year=fy_year, monthly_scores=monthly_scores,
                         items=items, cat_scores=cat_scores, categories=categories)


@app.route('/kra/rate/<int:month>/<int:year>', methods=['GET', 'POST'])
@login_required
def kra_employee_rate(month, year):
    """Employee: enter self-ratings for a month."""
    user = get_user()
    conn = get_db()
    fy_year = year if month >= 4 else year - 1

    assignment = conn.execute('''
        SELECT a.*, t.name as template_name, t.department, t.role_title
        FROM kra_assignments a
        JOIN kra_templates t ON a.template_id = t.id
        WHERE a.employee_id = ? AND a.fy_year = ? AND a.is_active = 1
    ''', (user['id'], fy_year)).fetchone()

    if not assignment:
        conn.close()
        flash('No KRA assigned for this period', 'error')
        return redirect(url_for('kra_dashboard'))

    items = conn.execute('''
        SELECT ti.*, c.name as category_name, c.sort_order as cat_sort, c.id as cat_id
        FROM kra_template_items ti
        JOIN kra_categories c ON ti.category_id = c.id
        WHERE ti.template_id = ?
        ORDER BY c.sort_order, ti.sort_order
    ''', (assignment['template_id'],)).fetchall()

    if request.method == 'POST':
        for item in items:
            rating_id_key = f"rating_id_{item['id']}"
            existing_id = request.form.get(rating_id_key)

            achieved = request.form.get(f"achieved_{item['id']}", '').strip()
            emp_rating = request.form.get(f"emp_rating_{item['id']}", '').strip()

            achieved_val = float(achieved) if achieved else None
            if item['is_target_based'] and item['target_value'] and achieved_val is not None:
                pct = min(achieved_val / item['target_value'], 1.0) if item['target_value'] > 0 else 0
                emp_rating_val = round(pct * 5, 6)
            else:
                emp_rating_val = float(emp_rating) if emp_rating else None

            emp_result = round(emp_rating_val * item['percentage_sharing'], 6) if emp_rating_val is not None else None

            if existing_id:
                conn.execute('''
                    UPDATE kra_monthly_ratings
                    SET achieved_value = ?, employee_rating = ?, employee_result = ?,
                        employee_submitted = 1, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (achieved_val, emp_rating_val, emp_result, int(existing_id)))
            else:
                conn.execute('''
                    INSERT INTO kra_monthly_ratings
                    (assignment_id, template_item_id, month, year, achieved_value, employee_rating, employee_result, employee_submitted)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 1)
                ''', (assignment['id'], item['id'], month, year, achieved_val, emp_rating_val, emp_result))

        conn.commit()
        conn.close()
        flash(f'Self-ratings submitted for {calendar.month_name[month]} {year}', 'success')
        return redirect(url_for('kra_dashboard'))

    # GET — load existing ratings
    existing_ratings = {}
    ratings = conn.execute('''
        SELECT * FROM kra_monthly_ratings
        WHERE assignment_id = ? AND month = ? AND year = ?
    ''', (assignment['id'], month, year)).fetchall()
    for r in ratings:
        existing_ratings[r['template_item_id']] = r

    # Notes
    notes = conn.execute('''
        SELECT * FROM kra_monthly_notes
        WHERE assignment_id = ? AND month = ? AND year = ?
    ''', (assignment['id'], month, year)).fetchall()

    conn.close()
    month_name = calendar.month_name[month]
    return render_template('kra_employee_rate.html', user=user, assignment=assignment,
                         items=items, existing_ratings=existing_ratings,
                         month=month, year=year, month_name=month_name, notes=notes)


# ─── Manager KRA Rating Routes ───

@app.route('/kra/manager/team')
@login_required
def kra_manager_team():
    """Manager: view team KRA assignments for rating."""
    user = get_user()
    conn = get_db()
    now = datetime.now()
    fy_year = int(request.args.get('fy', now.year if now.month >= 4 else now.year - 1))
    view_month = int(request.args.get('month', now.month))
    view_year = int(request.args.get('year', now.year))

    # Get direct reports
    direct_reports = conn.execute(
        'SELECT id, name, emp_code, department, photo_url FROM employees WHERE reporting_to = ? AND is_active = 1 ORDER BY name',
        (user['id'],)
    ).fetchall()

    if not direct_reports:
        conn.close()
        flash('You have no direct reports with KRA assignments', 'info')
        return redirect(url_for('kra_dashboard'))

    team_data = []
    for emp in direct_reports:
        asn = conn.execute('''
            SELECT a.id as assignment_id, t.name as template_name
            FROM kra_assignments a
            JOIN kra_templates t ON a.template_id = t.id
            WHERE a.employee_id = ? AND a.fy_year = ? AND a.is_active = 1
        ''', (emp['id'], fy_year)).fetchone()
        if not asn:
            continue

        ratings = conn.execute('''
            SELECT SUM(employee_result) as emp_total, SUM(manager_result) as mgr_total,
                   MAX(employee_submitted) as emp_sub, MAX(manager_submitted) as mgr_sub
            FROM kra_monthly_ratings
            WHERE assignment_id = ? AND month = ? AND year = ?
        ''', (asn['assignment_id'], view_month, view_year)).fetchone()

        team_data.append({
            'id': emp['id'],
            'name': emp['name'],
            'emp_code': emp['emp_code'],
            'department': emp['department'],
            'photo_url': emp['photo_url'],
            'template_name': asn['template_name'],
            'assignment_id': asn['assignment_id'],
            'emp_score': round(ratings['emp_total'] or 0, 4),
            'mgr_score': round(ratings['mgr_total'] or 0, 4),
            'emp_submitted': bool(ratings['emp_sub']),
            'mgr_submitted': bool(ratings['mgr_sub']),
        })

    conn.close()
    return render_template('kra_manager_team.html', user=user, team_data=team_data,
                         fy_year=fy_year, view_month=view_month, view_year=view_year)


@app.route('/kra/manager/rate/<int:employee_id>/<int:month>/<int:year>', methods=['GET', 'POST'])
@login_required
def kra_manager_rate(employee_id, month, year):
    """Manager: rate a direct report's KRA for a given month."""
    user = get_user()
    conn = get_db()
    fy_year = year if month >= 4 else year - 1

    # Verify this is a direct report (or admin)
    emp = conn.execute('SELECT * FROM employees WHERE id = ?', (employee_id,)).fetchone()
    if not emp:
        conn.close()
        flash('Employee not found', 'error')
        return redirect(url_for('kra_manager_team'))

    is_admin = user['is_admin']
    is_reporting_manager = (emp['reporting_to'] == user['id'])
    if not is_admin and not is_reporting_manager:
        conn.close()
        flash('You can only rate your direct reports', 'error')
        return redirect(url_for('kra_manager_team'))

    assignment = conn.execute('''
        SELECT a.*, t.name as template_name, t.department, t.role_title
        FROM kra_assignments a
        JOIN kra_templates t ON a.template_id = t.id
        WHERE a.employee_id = ? AND a.fy_year = ? AND a.is_active = 1
    ''', (employee_id, fy_year)).fetchone()

    if not assignment:
        conn.close()
        flash('No KRA assigned to this employee', 'error')
        return redirect(url_for('kra_manager_team'))

    items = conn.execute('''
        SELECT ti.*, c.name as category_name, c.sort_order as cat_sort
        FROM kra_template_items ti
        JOIN kra_categories c ON ti.category_id = c.id
        WHERE ti.template_id = ?
        ORDER BY c.sort_order, ti.sort_order
    ''', (assignment['template_id'],)).fetchall()

    if request.method == 'POST':
        for item in items:
            mgr_rating = request.form.get(f"mgr_rating_{item['id']}", '').strip()
            existing_id = request.form.get(f"rating_id_{item['id']}")

            if item['is_target_based'] and item['target_value']:
                # For target items, manager rating = same as employee (auto-calculated from achievement)
                existing_row = conn.execute(
                    'SELECT achieved_value FROM kra_monthly_ratings WHERE id = ?', (int(existing_id),)
                ).fetchone() if existing_id else None
                if existing_row and existing_row['achieved_value'] is not None:
                    pct = min(existing_row['achieved_value'] / item['target_value'], 1.0) if item['target_value'] > 0 else 0
                    mgr_rating_val = round(pct * 5, 6)
                else:
                    mgr_rating_val = float(mgr_rating) if mgr_rating else None
            else:
                mgr_rating_val = float(mgr_rating) if mgr_rating else None

            mgr_result = round(mgr_rating_val * item['percentage_sharing'], 6) if mgr_rating_val is not None else None

            if existing_id:
                conn.execute('''
                    UPDATE kra_monthly_ratings
                    SET manager_rating = ?, manager_result = ?, manager_submitted = 1, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (mgr_rating_val, mgr_result, int(existing_id)))
            else:
                conn.execute('''
                    INSERT INTO kra_monthly_ratings
                    (assignment_id, template_item_id, month, year, manager_rating, manager_result, manager_submitted)
                    VALUES (?, ?, ?, ?, ?, ?, 1)
                ''', (assignment['id'], item['id'], month, year, mgr_rating_val, mgr_result))

        # Save manager notes
        mgr_notes = request.form.get('manager_notes', '').strip()
        if mgr_notes:
            existing_note = conn.execute(
                'SELECT id FROM kra_monthly_notes WHERE assignment_id = ? AND month = ? AND year = ? AND created_by = ?',
                (assignment['id'], month, year, user['id'])
            ).fetchone()
            if existing_note:
                conn.execute('UPDATE kra_monthly_notes SET notes = ? WHERE id = ?', (mgr_notes, existing_note['id']))
            else:
                conn.execute(
                    'INSERT INTO kra_monthly_notes (assignment_id, month, year, notes, created_by) VALUES (?, ?, ?, ?, ?)',
                    (assignment['id'], month, year, mgr_notes, user['id'])
                )

        conn.commit()
        conn.close()
        flash(f'Manager ratings submitted for {emp["name"]} — {calendar.month_name[month]} {year}', 'success')
        return redirect(url_for('kra_manager_team', month=month, year=year))

    # GET
    existing_ratings = {}
    ratings = conn.execute('''
        SELECT * FROM kra_monthly_ratings
        WHERE assignment_id = ? AND month = ? AND year = ?
    ''', (assignment['id'], month, year)).fetchall()
    for r in ratings:
        existing_ratings[r['template_item_id']] = r

    notes = conn.execute('''
        SELECT * FROM kra_monthly_notes
        WHERE assignment_id = ? AND month = ? AND year = ?
    ''', (assignment['id'], month, year)).fetchall()

    conn.close()
    month_name = calendar.month_name[month]
    return render_template('kra_manager_rate.html', user=user, emp=emp, assignment=assignment,
                         items=items, existing_ratings=existing_ratings,
                         month=month, year=year, month_name=month_name, notes=notes)


@app.route('/kra/report')
@login_required
def kra_report():
    """Employee: annual KRA report with monthly scores."""
    user = get_user()
    conn = get_db()
    now = datetime.now()
    fy_year = int(request.args.get('fy', now.year if now.month >= 4 else now.year - 1))

    assignment = conn.execute('''
        SELECT a.*, t.name as template_name, t.department, t.role_title
        FROM kra_assignments a
        JOIN kra_templates t ON a.template_id = t.id
        WHERE a.employee_id = ? AND a.fy_year = ? AND a.is_active = 1
    ''', (user['id'], fy_year)).fetchone()

    if not assignment:
        conn.close()
        return render_template('kra_report.html', user=user, assignment=None, fy_year=fy_year,
                             monthly_data=[], avg_emp=0, avg_mgr=0)

    items = conn.execute('''
        SELECT ti.*, c.name as category_name, c.sort_order as cat_sort
        FROM kra_template_items ti
        JOIN kra_categories c ON ti.category_id = c.id
        WHERE ti.template_id = ?
        ORDER BY c.sort_order, ti.sort_order
    ''', (assignment['template_id'],)).fetchall()

    fy_months = [(m, fy_year if m >= 4 else fy_year + 1) for m in [4,5,6,7,8,9,10,11,12,1,2,3]]
    monthly_data = []
    total_emp = total_mgr = rated_months = 0

    for m, yr in fy_months:
        ratings = conn.execute('''
            SELECT r.*, ti.kpi_code, ti.measure_description, ti.percentage_sharing, ti.is_target_based,
                   ti.target_value, c.name as category_name
            FROM kra_monthly_ratings r
            JOIN kra_template_items ti ON r.template_item_id = ti.id
            JOIN kra_categories c ON ti.category_id = c.id
            WHERE r.assignment_id = ? AND r.month = ? AND r.year = ?
            ORDER BY c.sort_order, ti.sort_order
        ''', (assignment['id'], m, yr)).fetchall()

        emp_total = sum(r['employee_result'] or 0 for r in ratings)
        mgr_total = sum(r['manager_result'] or 0 for r in ratings)
        emp_sub = any(r['employee_submitted'] for r in ratings)
        mgr_sub = any(r['manager_submitted'] for r in ratings)

        if emp_sub or mgr_sub:
            total_emp += emp_total
            total_mgr += mgr_total
            rated_months += 1

        notes = conn.execute('''
            SELECT n.*, e.name as author_name FROM kra_monthly_notes n
            LEFT JOIN employees e ON n.created_by = e.id
            WHERE n.assignment_id = ? AND n.month = ? AND n.year = ?
        ''', (assignment['id'], m, yr)).fetchall()

        monthly_data.append({
            'month': calendar.month_name[m],
            'month_num': m,
            'year': yr,
            'ratings': ratings,
            'emp_score': round(emp_total, 4),
            'mgr_score': round(mgr_total, 4),
            'emp_submitted': emp_sub,
            'mgr_submitted': mgr_sub,
            'notes': notes,
        })

    avg_emp = round(total_emp / rated_months, 4) if rated_months > 0 else 0
    avg_mgr = round(total_mgr / rated_months, 4) if rated_months > 0 else 0

    conn.close()
    return render_template('kra_report.html', user=user, assignment=assignment,
                         fy_year=fy_year, monthly_data=monthly_data, items=items,
                         avg_emp=avg_emp, avg_mgr=avg_mgr, rated_months=rated_months)


# ─── Budget & Finance Module ───

def ensure_budget_tables():
    """Create budget/finance tables if they don't exist."""
    try:
        conn = get_db()
        tables_sql = [
            '''CREATE TABLE IF NOT EXISTS salary_items (
                id SERIAL PRIMARY KEY,
                employee_id INTEGER REFERENCES employees(id),
                name TEXT NOT NULL,
                department TEXT,
                project_id INTEGER REFERENCES projects(id),
                monthly_cost NUMERIC(14,2) DEFAULT 0,
                currency TEXT DEFAULT 'INR',
                is_active INTEGER DEFAULT 1,
                notes TEXT,
                created_by INTEGER REFERENCES employees(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS subscription_items (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                vendor TEXT,
                cost NUMERIC(14,2) DEFAULT 0,
                currency TEXT DEFAULT 'INR',
                frequency TEXT DEFAULT 'monthly',
                primary_department TEXT,
                shared_departments TEXT,
                project_id INTEGER REFERENCES projects(id),
                is_active INTEGER DEFAULT 1,
                notes TEXT,
                created_by INTEGER REFERENCES employees(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS budget_categories (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                parent_id INTEGER REFERENCES budget_categories(id),
                cat_type TEXT NOT NULL DEFAULT 'expense',
                department TEXT,
                sort_order INTEGER DEFAULT 0,
                is_active INTEGER DEFAULT 1,
                is_recurring INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS budget_entries (
                id SERIAL PRIMARY KEY,
                category_id INTEGER NOT NULL REFERENCES budget_categories(id),
                fy_year TEXT NOT NULL DEFAULT '2026-2027',
                month INTEGER NOT NULL,
                year INTEGER NOT NULL,
                budget_amount REAL DEFAULT 0,
                actual_amount REAL DEFAULT 0,
                notes TEXT,
                project_id INTEGER,
                is_locked INTEGER DEFAULT 0,
                created_by INTEGER REFERENCES employees(id),
                updated_by INTEGER REFERENCES employees(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS budget_settings (
                id SERIAL PRIMARY KEY,
                fy_year TEXT NOT NULL DEFAULT '2026-2027',
                setting_key TEXT NOT NULL,
                setting_value TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )'''
        ]
        for sql in tables_sql:
            try:
                conn.execute(sql)
                conn.commit()
            except Exception as _tbl_err:
                try:
                    conn.rollback()
                except Exception:
                    pass
                logging.error(f"ensure_budget_tables create failed: {_tbl_err}")
        # Migration: add is_recurring column if missing
        try:
            conn.execute("ALTER TABLE budget_categories ADD COLUMN is_recurring INTEGER DEFAULT 0")
            conn.commit()
            logging.info("Added is_recurring column to budget_categories")
        except Exception:
            try:
                conn.rollback()
            except:
                pass
        # Migration: add project_id column to budget_entries if missing
        try:
            conn.execute("ALTER TABLE budget_entries ADD COLUMN project_id INTEGER")
            conn.commit()
            logging.info("Added project_id column to budget_entries")
        except Exception:
            try:
                conn.rollback()
            except:
                pass
        # Migration: add stream_id column to budget_categories so revenue
        # categories can be linked to sales revenue_streams
        try:
            conn.execute("ALTER TABLE budget_categories ADD COLUMN stream_id INTEGER")
            conn.commit()
            logging.info("Added stream_id column to budget_categories")
        except Exception:
            try:
                conn.rollback()
            except:
                pass
        # Migration: add product_id column to budget_entries so revenue can
        # be tracked at product/service level (with stream + project derived)
        try:
            conn.execute("ALTER TABLE budget_entries ADD COLUMN product_id INTEGER")
            conn.commit()
            logging.info("Added product_id column to budget_entries")
        except Exception:
            try:
                conn.rollback()
            except:
                pass
        # Migration: relax NOT NULL on category_id so product-based revenue
        # entries can exist without a category
        try:
            conn.execute("ALTER TABLE budget_entries ALTER COLUMN category_id DROP NOT NULL")
            conn.commit()
            logging.info("Dropped NOT NULL on budget_entries.category_id")
        except Exception:
            try:
                conn.rollback()
            except:
                pass
        # Migration: add salary_id / subscription_id to budget_entries so
        # line-item expenses (salaries, subscriptions) can be tracked
        for col in ('salary_id', 'subscription_id'):
            try:
                conn.execute(f"ALTER TABLE budget_entries ADD COLUMN {col} INTEGER")
                conn.commit()
                logging.info(f"Added {col} column to budget_entries")
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass
        # Migration: add per-row unit counts so revenue rows can store units
        # sold and compute amount = units × sale_price at save time.
        for col in ('budget_units', 'actual_units'):
            try:
                conn.execute(f"ALTER TABLE budget_entries ADD COLUMN {col} NUMERIC(14,2) DEFAULT 0")
                conn.commit()
                logging.info(f"Added {col} column to budget_entries")
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass
        # Backfill: heal salary_items rows whose name/department got wiped by
        # the pre-fix edit handler (readonly fields without `name` attribute
        # sent empty strings). Pull canonical name/dept from the employee row.
        try:
            conn.execute('''
                UPDATE salary_items s
                SET name = e.name,
                    department = COALESCE(e.department, '')
                FROM employees e
                WHERE s.employee_id = e.id
                  AND (s.name IS NULL OR s.name = '' OR s.department IS NULL OR s.department = '')
            ''')
            conn.commit()
            logging.info("Backfilled salary_items name/department from employees table")
        except Exception as _bf_err:
            try:
                conn.rollback()
            except Exception:
                pass
            logging.error(f"salary_items backfill failed: {_bf_err}")
        conn.commit()
        conn.close()
        logging.info("Budget tables ensured.")
    except Exception as e:
        logging.error(f"ensure_budget_tables: {e}")


def seed_budget_categories():
    """Seed default budget categories for expense and revenue."""
    try:
        conn = get_db()
        # One-time rename: old "Software & Tools" label → new "Software Tool Subscription".
        # Safe to run every boot — it's a no-op after the first successful rename
        # because the WHERE clause won't match anymore.
        try:
            conn.execute(
                "UPDATE budget_categories SET name = ? "
                "WHERE LOWER(TRIM(name)) IN ('software & tools','software and tools','subscriptions') "
                "AND type = 'expense'",
                ('Software Tool Subscription',)
            )
            conn.commit()
        except Exception as _rn_err:
            try:
                conn.rollback()
            except Exception:
                pass
            logging.warning(f"budget category rename skipped: {_rn_err}")

        count = conn.execute('SELECT COUNT(*) as cnt FROM budget_categories').fetchone()
        if count['cnt'] > 0:
            conn.close()
            return

        # Expense categories (top-level)
        expense_cats = [
            ('Salaries & Wages', None, 'expense', None, 1),
            ('Rent & Utilities', None, 'expense', None, 2),
            ('Marketing & Advertising', None, 'expense', None, 3),
            ('Travel & Conveyance', None, 'expense', None, 4),
            ('Software Tool Subscription', None, 'expense', None, 5),
            ('Office Supplies', None, 'expense', None, 6),
            ('Professional Services', None, 'expense', None, 7),
            ('Communication & Internet', None, 'expense', None, 8),
            ('Insurance', None, 'expense', None, 9),
            ('Miscellaneous Expenses', None, 'expense', None, 10),
        ]

        # Department budget categories
        departments = ['Human Resources', 'Marketing', 'Sales', 'Operations', 'Technology']

        # Revenue categories (top-level)
        revenue_cats = [
            ('Student Consulting / Counseling Fees', None, 'revenue', None, 1),
            ('University Placement Commissions', None, 'revenue', None, 2),
            ('Medical PG International Pathways', None, 'revenue', None, 3),
            ('Online & Offline Trainings', None, 'revenue', None, 4),
            ('B2B Services', None, 'revenue', None, 5),
            ('Portfolio Services', None, 'revenue', None, 6),
            ('Other Products / Services', None, 'revenue', None, 7),
        ]

        # Insert expense categories
        for name, parent_id, cat_type, dept, sort_order in expense_cats:
            conn.execute(
                'INSERT INTO budget_categories (name, parent_id, cat_type, department, sort_order) VALUES (?, ?, ?, ?, ?)',
                (name, parent_id, cat_type, dept, sort_order)
            )

        # Insert department expense categories
        dept_sort = 20
        for dept in departments:
            conn.execute(
                'INSERT INTO budget_categories (name, parent_id, cat_type, department, sort_order) VALUES (?, ?, ?, ?, ?)',
                (f'{dept} - Department Budget', None, 'department', dept, dept_sort)
            )
            dept_sort += 1

        # Insert revenue categories
        for name, parent_id, cat_type, dept, sort_order in revenue_cats:
            conn.execute(
                'INSERT INTO budget_categories (name, parent_id, cat_type, department, sort_order) VALUES (?, ?, ?, ?, ?)',
                (name, parent_id, cat_type, dept, sort_order)
            )

        conn.commit()
        conn.close()
        logging.info("Budget categories seeded.")
    except Exception as e:
        logging.error(f"seed_budget_categories: {e}")


# Budget helper: get FY months in order (Apr 2026 - Mar 2027)
def get_fy_months(fy_year='2026-2027'):
    parts = fy_year.split('-')
    start_year = int(parts[0])
    end_year = int(parts[1]) if len(parts) > 1 else start_year + 1
    months = []
    for m in range(4, 13):  # Apr to Dec
        months.append({'month': m, 'year': start_year, 'label': f"{calendar.month_abbr[m]} {start_year}"})
    for m in range(1, 4):  # Jan to Mar
        months.append({'month': m, 'year': end_year, 'label': f"{calendar.month_abbr[m]} {end_year}"})
    return months


def get_quarter_label(month):
    if month in [4, 5, 6]:
        return 'Q1'
    elif month in [7, 8, 9]:
        return 'Q2'
    elif month in [10, 11, 12]:
        return 'Q3'
    else:
        return 'Q4'


# ─── Finance / Budget Routes ───

@app.route('/finance/budget')
@admin_required
def finance_dashboard():
    user = get_user()
    fy_year = request.args.get('fy', '2026-2027')
    fy_months = get_fy_months(fy_year)
    conn = get_db()

    # Monthly totals for expense categories (department cats are derived now)
    try:
        monthly_agg = conn.execute('''
            SELECT be.month, be.year, bc.cat_type,
                   COALESCE(SUM(be.budget_amount), 0) as total_budget,
                   COALESCE(SUM(be.actual_amount), 0) as total_actual
            FROM budget_entries be
            JOIN budget_categories bc ON be.category_id = bc.id
            WHERE be.fy_year = ? AND bc.is_active = 1
              AND bc.cat_type = 'expense'
            GROUP BY be.month, be.year, bc.cat_type
        ''', (fy_year,)).fetchall()
    except Exception as e:
        logging.error(f"finance_dashboard agg query: {e}")
        try:
            conn.rollback()
        except:
            pass
        monthly_agg = []

    # Monthly revenue totals from product-based entries
    try:
        revenue_agg = conn.execute('''
            SELECT be.month, be.year,
                   COALESCE(SUM(be.budget_amount), 0) as total_budget,
                   COALESCE(SUM(be.actual_amount), 0) as total_actual
            FROM budget_entries be
            WHERE be.fy_year = ? AND be.product_id IS NOT NULL
            GROUP BY be.month, be.year
        ''', (fy_year,)).fetchall()
    except Exception as e:
        logging.error(f"finance_dashboard revenue agg query: {e}")
        try:
            conn.rollback()
        except:
            pass
        revenue_agg = []

    # Monthly expense totals from salary + subscription line-items
    try:
        line_expense_agg = conn.execute('''
            SELECT be.month, be.year,
                   COALESCE(SUM(be.budget_amount), 0) as total_budget,
                   COALESCE(SUM(be.actual_amount), 0) as total_actual
            FROM budget_entries be
            WHERE be.fy_year = ?
              AND (be.salary_id IS NOT NULL OR be.subscription_id IS NOT NULL)
            GROUP BY be.month, be.year
        ''', (fy_year,)).fetchall()
    except Exception as e:
        logging.error(f"finance_dashboard line expense agg query: {e}")
        try:
            conn.rollback()
        except:
            pass
        line_expense_agg = []

    # Department totals — derived from linked expenses (salaries, subs, dept-tagged cats)
    dept_bucket = {}
    def _dept_key(name):
        k = (name or '').strip() or '— Unassigned —'
        if k not in dept_bucket:
            dept_bucket[k] = {'name': k, 'total_budget': 0.0, 'total_actual': 0.0}
        return dept_bucket[k]
    try:
        sal_rows = conn.execute('''
            SELECT si.department,
                   COALESCE(SUM(be.budget_amount), 0) as tb,
                   COALESCE(SUM(be.actual_amount), 0) as ta
            FROM salary_items si
            LEFT JOIN budget_entries be
              ON be.salary_id = si.id AND be.fy_year = ?
            WHERE si.is_active = 1
            GROUP BY si.department
        ''', (fy_year,)).fetchall()
        for r in sal_rows:
            b = _dept_key(r['department'])
            b['total_budget'] += float(r['tb'] or 0)
            b['total_actual'] += float(r['ta'] or 0)
    except Exception as e:
        logging.error(f"finance_dashboard salary dept query: {e}")
        try: conn.rollback()
        except: pass
    try:
        sub_rows = conn.execute('''
            SELECT si.primary_department,
                   COALESCE(SUM(be.budget_amount), 0) as tb,
                   COALESCE(SUM(be.actual_amount), 0) as ta
            FROM subscription_items si
            LEFT JOIN budget_entries be
              ON be.subscription_id = si.id AND be.fy_year = ?
            WHERE si.is_active = 1
            GROUP BY si.primary_department
        ''', (fy_year,)).fetchall()
        for r in sub_rows:
            b = _dept_key(r['primary_department'])
            b['total_budget'] += float(r['tb'] or 0)
            b['total_actual'] += float(r['ta'] or 0)
    except Exception as e:
        logging.error(f"finance_dashboard subscription dept query: {e}")
        try: conn.rollback()
        except: pass
    try:
        cat_rows = conn.execute('''
            SELECT bc.department,
                   COALESCE(SUM(be.budget_amount), 0) as tb,
                   COALESCE(SUM(be.actual_amount), 0) as ta
            FROM budget_categories bc
            LEFT JOIN budget_entries be
              ON be.category_id = bc.id AND be.fy_year = ?
            WHERE bc.cat_type = 'expense' AND bc.is_active = 1
              AND bc.department IS NOT NULL AND bc.department != ''
            GROUP BY bc.department
        ''', (fy_year,)).fetchall()
        for r in cat_rows:
            b = _dept_key(r['department'])
            b['total_budget'] += float(r['tb'] or 0)
            b['total_actual'] += float(r['ta'] or 0)
    except Exception as e:
        logging.error(f"finance_dashboard cat-dept query: {e}")
        try: conn.rollback()
        except: pass
    dept_agg = sorted(dept_bucket.values(), key=lambda x: (-x['total_budget'], x['name']))

    # Locked months
    try:
        locked_rows = conn.execute(
            "SELECT DISTINCT month, year FROM budget_entries WHERE fy_year = ? AND is_locked = 1",
            (fy_year,)
        ).fetchall()
    except Exception as e:
        logging.error(f"finance_dashboard locked query: {e}")
        try:
            conn.rollback()
        except:
            pass
        locked_rows = []

    conn.close()

    # Build monthly lookup
    monthly_data = {}
    for r in monthly_agg:
        key = (r['month'], r['year'])
        if key not in monthly_data:
            monthly_data[key] = {'expense_budget': 0, 'expense_actual': 0, 'dept_budget': 0, 'dept_actual': 0, 'revenue_budget': 0, 'revenue_actual': 0}
        ct = r['cat_type']
        if ct == 'expense':
            monthly_data[key]['expense_budget'] += float(r['total_budget'] or 0)
            monthly_data[key]['expense_actual'] += float(r['total_actual'] or 0)
        elif ct == 'department':
            monthly_data[key]['dept_budget'] += float(r['total_budget'] or 0)
            monthly_data[key]['dept_actual'] += float(r['total_actual'] or 0)
        elif ct == 'revenue':
            monthly_data[key]['revenue_budget'] += float(r['total_budget'] or 0)
            monthly_data[key]['revenue_actual'] += float(r['total_actual'] or 0)

    # Merge product-based revenue totals
    for r in revenue_agg:
        key = (r['month'], r['year'])
        if key not in monthly_data:
            monthly_data[key] = {'expense_budget': 0, 'expense_actual': 0, 'dept_budget': 0, 'dept_actual': 0, 'revenue_budget': 0, 'revenue_actual': 0}
        monthly_data[key]['revenue_budget'] += float(r['total_budget'] or 0)
        monthly_data[key]['revenue_actual'] += float(r['total_actual'] or 0)

    # Merge salary + subscription line-item expense totals
    for r in line_expense_agg:
        key = (r['month'], r['year'])
        if key not in monthly_data:
            monthly_data[key] = {'expense_budget': 0, 'expense_actual': 0, 'dept_budget': 0, 'dept_actual': 0, 'revenue_budget': 0, 'revenue_actual': 0}
        monthly_data[key]['expense_budget'] += float(r['total_budget'] or 0)
        monthly_data[key]['expense_actual'] += float(r['total_actual'] or 0)

    # Build arrays
    total_expense_budget = 0; total_expense_actual = 0
    total_revenue_budget = 0; total_revenue_actual = 0
    monthly_expense_budget = []; monthly_expense_actual = []
    monthly_revenue_budget = []; monthly_revenue_actual = []
    month_labels = []

    for fm in fy_months:
        key = (fm['month'], fm['year'])
        d = monthly_data.get(key, {})
        month_labels.append(fm['label'])
        eb = d.get('expense_budget', 0) + d.get('dept_budget', 0)
        ea = d.get('expense_actual', 0) + d.get('dept_actual', 0)
        rb = d.get('revenue_budget', 0)
        ra = d.get('revenue_actual', 0)
        total_expense_budget += eb; total_expense_actual += ea
        total_revenue_budget += rb; total_revenue_actual += ra
        monthly_expense_budget.append(round(eb, 2))
        monthly_expense_actual.append(round(ea, 2))
        monthly_revenue_budget.append(round(rb, 2))
        monthly_revenue_actual.append(round(ra, 2))

    # Quarterly aggregates
    quarters = {q: {'expense_budget': 0, 'expense_actual': 0, 'revenue_budget': 0, 'revenue_actual': 0} for q in ['Q1','Q2','Q3','Q4']}
    for i, fm in enumerate(fy_months):
        q = get_quarter_label(fm['month'])
        quarters[q]['expense_budget'] += monthly_expense_budget[i]
        quarters[q]['expense_actual'] += monthly_expense_actual[i]
        quarters[q]['revenue_budget'] += monthly_revenue_budget[i]
        quarters[q]['revenue_actual'] += monthly_revenue_actual[i]

    # Department breakdown (derived from linked expenses)
    dept_data = []
    for dc in dept_agg:
        b = float(dc['total_budget'] or 0)
        a = float(dc['total_actual'] or 0)
        dept_data.append({'name': dc['name'], 'budget': round(b, 2), 'actual': round(a, 2),
                          'variance': round(b - a, 2), 'utilization': round((a / b * 100) if b > 0 else 0, 1)})

    locked_months = set((r['month'], r['year']) for r in locked_rows)

    return render_template('finance_dashboard.html', user=user, fy_year=fy_year, fy_months=fy_months,
                           total_expense_budget=round(total_expense_budget, 2),
                           total_expense_actual=round(total_expense_actual, 2),
                           total_revenue_budget=round(total_revenue_budget, 2),
                           total_revenue_actual=round(total_revenue_actual, 2),
                           monthly_expense_budget=monthly_expense_budget,
                           monthly_expense_actual=monthly_expense_actual,
                           monthly_revenue_budget=monthly_revenue_budget,
                           monthly_revenue_actual=monthly_revenue_actual,
                           month_labels=month_labels,
                           quarters=quarters, dept_data=dept_data, locked_months=locked_months)


@app.route('/finance/budget/expenses')
@admin_required
def finance_expenses():
    user = get_user()
    fy_year = request.args.get('fy', '2026-2027')
    fy_months = get_fy_months(fy_year)
    conn = get_db()

    expense_cats = conn.execute(
        "SELECT * FROM budget_categories WHERE cat_type = 'expense' AND is_active = 1 ORDER BY sort_order"
    ).fetchall()
    dept_cats = conn.execute(
        "SELECT * FROM budget_categories WHERE cat_type = 'department' AND is_active = 1 ORDER BY sort_order"
    ).fetchall()
    entries = conn.execute("SELECT * FROM budget_entries WHERE fy_year = ?", (fy_year,)).fetchall()

    # Line-item rollup: salary_items and subscription_items feed their
    # respective category rows on the expense sheet. The master list's
    # monthly_cost is the default budget for every month; any per-month
    # budget_entries with salary_id / subscription_id override it, and
    # their actual_amount rolls up as the category actual.
    try:
        salary_items = conn.execute(
            "SELECT id, monthly_cost FROM salary_items WHERE is_active = 1"
        ).fetchall()
    except Exception:
        salary_items = []
    try:
        sub_items = conn.execute(
            "SELECT id, cost, frequency FROM subscription_items WHERE is_active = 1"
        ).fetchall()
    except Exception:
        sub_items = []
    conn.close()

    # Category entries keyed by (category_id, month, year)
    entry_map = {}
    # Line-item entries keyed by (salary_id/subscription_id, month, year)
    sal_entry_map = {}
    sub_entry_map = {}
    for e in entries:
        keys = e.keys() if hasattr(e, 'keys') else []
        sid = e['salary_id'] if 'salary_id' in keys else None
        subid = e['subscription_id'] if 'subscription_id' in keys else None
        cid = e['category_id'] if 'category_id' in keys else None
        if sid:
            sal_entry_map[(sid, e['month'], e['year'])] = e
        elif subid:
            sub_entry_map[(subid, e['month'], e['year'])] = e
        elif cid:
            entry_map[(cid, e['month'], e['year'])] = e

    # Identify the category rows that absorb line-item rollups
    salary_cat_id = None
    sub_cat_id = None
    for c in expense_cats:
        nm = (c['name'] or '').lower().strip()
        if nm in ('salaries & wages', 'salaries and wages', 'salary', 'salaries'):
            salary_cat_id = c['id']
        elif nm in ('software tool subscription', 'software tools subscription',
                    'software & tools', 'software and tools', 'software', 'subscriptions'):
            sub_cat_id = c['id']

    def _sub_monthly_equiv(cost, frequency):
        try:
            raw = float(cost or 0)
        except (TypeError, ValueError):
            raw = 0.0
        freq = (frequency or 'monthly').lower()
        if freq == 'monthly':
            return raw
        if freq == 'quarterly':
            return raw / 3.0
        if freq == 'annual':
            return raw / 12.0
        return 0.0

    # Merge line-item totals into the category rows per month
    for fm in fy_months:
        m, y = fm['month'], fm['year']
        if salary_cat_id and salary_items:
            s_budget = 0.0
            s_actual = 0.0
            for s in salary_items:
                e = sal_entry_map.get((s['id'], m, y))
                if e:
                    s_budget += float(e['budget_amount'] or 0)
                    s_actual += float(e['actual_amount'] or 0)
                else:
                    s_budget += float(s['monthly_cost'] or 0)
            existing = entry_map.get((salary_cat_id, m, y))
            merged = {
                'budget_amount': s_budget + (float(existing['budget_amount'] or 0) if existing else 0),
                'actual_amount': s_actual + (float(existing['actual_amount'] or 0) if existing else 0),
            }
            entry_map[(salary_cat_id, m, y)] = merged

        if sub_cat_id and sub_items:
            b_sum = 0.0
            a_sum = 0.0
            for s in sub_items:
                e = sub_entry_map.get((s['id'], m, y))
                if e:
                    b_sum += float(e['budget_amount'] or 0)
                    a_sum += float(e['actual_amount'] or 0)
                else:
                    b_sum += _sub_monthly_equiv(s['cost'], s['frequency'])
            existing = entry_map.get((sub_cat_id, m, y))
            merged = {
                'budget_amount': b_sum + (float(existing['budget_amount'] or 0) if existing else 0),
                'actual_amount': a_sum + (float(existing['actual_amount'] or 0) if existing else 0),
            }
            entry_map[(sub_cat_id, m, y)] = merged

    # Per-month totals across all expense categories (budget / actual / variance)
    monthly_totals = []
    annual_budget_total = 0.0
    annual_actual_total = 0.0
    for fm in fy_months:
        m_budget = 0.0
        m_actual = 0.0
        for cat in expense_cats:
            e = entry_map.get((cat['id'], fm['month'], fm['year']))
            if e:
                m_budget += float(e.get('budget_amount') or 0)
                m_actual += float(e.get('actual_amount') or 0)
        monthly_totals.append({
            'budget': m_budget,
            'actual': m_actual,
            'variance': m_budget - m_actual,
        })
        annual_budget_total += m_budget
        annual_actual_total += m_actual
    annual_variance_total = annual_budget_total - annual_actual_total

    return render_template('finance_expenses.html', user=user, fy_year=fy_year, fy_months=fy_months,
                           expense_cats=expense_cats, dept_cats=dept_cats, entry_map=entry_map,
                           monthly_totals=monthly_totals,
                           annual_budget_total=annual_budget_total,
                           annual_actual_total=annual_actual_total,
                           annual_variance_total=annual_variance_total,
                           get_quarter_label=get_quarter_label)


@app.route('/finance/budget/revenue')
@admin_required
def finance_revenue():
    user = get_user()
    fy_year = request.args.get('fy', '2026-2027')
    fy_months = get_fy_months(fy_year)
    conn = get_db()

    try:
        products_raw = conn.execute('''
            SELECT ps.id, ps.name, ps.type, ps.sale_price, ps.sale_currency,
                   ps.project_id, ps.revenue_stream_id,
                   p.name  AS project_name,
                   rs.name AS stream_name
            FROM products_services ps
            LEFT JOIN projects p         ON ps.project_id = p.id
            LEFT JOIN revenue_streams rs ON ps.revenue_stream_id = rs.id
            WHERE ps.status = 'active'
            ORDER BY rs.name NULLS LAST, p.name NULLS LAST, ps.name
        ''').fetchall()
    except Exception as e:
        logging.error(f"finance_revenue products query: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        products_raw = []

    try:
        entries = conn.execute(
            "SELECT * FROM budget_entries WHERE fy_year = ? AND product_id IS NOT NULL",
            (fy_year,)
        ).fetchall()
    except Exception:
        entries = []
    conn.close()

    entry_map = {}
    for e in entries:
        entry_map[(e['product_id'], e['month'], e['year'])] = e

    # Build per-product FY rows & group by stream
    revenue_groups = {}
    for p in products_raw:
        stream_key = p['stream_name'] or '— Unassigned stream —'
        if stream_key not in revenue_groups:
            revenue_groups[stream_key] = {
                'stream_name': stream_key,
                'products': [],
                'monthly_budget': [0.0] * len(fy_months),
                'monthly_actual': [0.0] * len(fy_months),
                'total_budget': 0.0,
                'total_actual': 0.0,
            }
        grp = revenue_groups[stream_key]
        product_monthly_budget = []
        product_monthly_actual = []
        p_total_b = 0.0
        p_total_a = 0.0
        for idx, fm in enumerate(fy_months):
            e = entry_map.get((p['id'], fm['month'], fm['year']))
            b = float(e['budget_amount'] or 0) if e else 0.0
            a = float(e['actual_amount'] or 0) if e else 0.0
            product_monthly_budget.append(b)
            product_monthly_actual.append(a)
            grp['monthly_budget'][idx] += b
            grp['monthly_actual'][idx] += a
            p_total_b += b
            p_total_a += a
        grp['total_budget'] += p_total_b
        grp['total_actual'] += p_total_a
        grp['products'].append({
            'name': p['name'],
            'project_name': p['project_name'] or '— No project —',
            'type': p['type'],
            'monthly_budget': product_monthly_budget,
            'monthly_actual': product_monthly_actual,
            'total_budget': p_total_b,
            'total_actual': p_total_a,
        })

    revenue_groups_list = sorted(revenue_groups.values(), key=lambda g: g['stream_name'])

    # FY totals
    fy_total_budget = sum(g['total_budget'] for g in revenue_groups_list)
    fy_total_actual = sum(g['total_actual'] for g in revenue_groups_list)
    fy_total_variance = fy_total_budget - fy_total_actual

    # Per-month totals across all streams
    monthly_totals = []
    for idx, fm in enumerate(fy_months):
        m_b = sum(g['monthly_budget'][idx] for g in revenue_groups_list)
        m_a = sum(g['monthly_actual'][idx] for g in revenue_groups_list)
        monthly_totals.append({'budget': m_b, 'actual': m_a, 'variance': m_b - m_a})

    return render_template('finance_revenue.html', user=user, fy_year=fy_year, fy_months=fy_months,
                           revenue_groups=revenue_groups_list,
                           fy_total_budget=fy_total_budget,
                           fy_total_actual=fy_total_actual,
                           fy_total_variance=fy_total_variance,
                           monthly_totals=monthly_totals,
                           get_quarter_label=get_quarter_label)


@app.route('/finance/budget/edit/<int:month>/<int:year>', methods=['GET', 'POST'])
@admin_required
def finance_edit_month(month, year):
    user = get_user()
    fy_year = request.args.get('fy', '2026-2027')
    conn = get_db()

    # Check if month is locked
    lock_check = conn.execute(
        "SELECT is_locked FROM budget_entries WHERE fy_year = ? AND month = ? AND year = ? AND is_locked = 1 LIMIT 1",
        (fy_year, month, year)
    ).fetchone()
    is_locked = lock_check is not None

    if request.method == 'POST' and not is_locked:
        # Only expense & department categories are entered by category now;
        # revenue is entered at product level (see product loop below).
        # Department cats are now display-only (computed from linked expenses),
        # so we only save expense categories here.
        all_cats = conn.execute(
            "SELECT * FROM budget_categories WHERE is_active = 1 AND cat_type = 'expense' ORDER BY sort_order"
        ).fetchall()

        recurring_fills = []  # Track recurring categories that need auto-fill

        for cat in all_cats:
            budget_val = request.form.get(f'budget_{cat["id"]}', '0')
            actual_val = request.form.get(f'actual_{cat["id"]}', '0')
            notes_val = request.form.get(f'notes_{cat["id"]}', '').strip()
            project_val = request.form.get(f'project_{cat["id"]}', '')
            project_id = int(project_val) if project_val and project_val.isdigit() else None

            try:
                budget_val = float(budget_val) if budget_val else 0
            except ValueError:
                budget_val = 0
            try:
                actual_val = float(actual_val) if actual_val else 0
            except ValueError:
                actual_val = 0

            existing = conn.execute(
                "SELECT id FROM budget_entries WHERE category_id = ? AND fy_year = ? AND month = ? AND year = ?",
                (cat['id'], fy_year, month, year)
            ).fetchone()

            if existing:
                conn.execute(
                    "UPDATE budget_entries SET budget_amount = ?, actual_amount = ?, notes = ?, project_id = ?, updated_by = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (budget_val, actual_val, notes_val, project_id, user['id'], existing['id'])
                )
            else:
                if budget_val > 0 or actual_val > 0 or notes_val:
                    conn.execute(
                        "INSERT INTO budget_entries (category_id, fy_year, month, year, budget_amount, actual_amount, notes, project_id, created_by, updated_by) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (cat['id'], fy_year, month, year, budget_val, actual_val, notes_val, project_id, user['id'], user['id'])
                    )

            # Track recurring categories with budget > 0 (must be inside the cat loop)
            try:
                _is_recurring = cat['is_recurring'] if 'is_recurring' in cat.keys() else 0
            except Exception:
                _is_recurring = 0
            if _is_recurring and budget_val > 0:
                recurring_fills.append({'cat_id': cat['id'], 'budget': budget_val, 'notes': notes_val})

        # ── Salary line-item expenses ──
        try:
            active_salaries = conn.execute(
                "SELECT id, department, project_id, monthly_cost FROM salary_items WHERE is_active = 1"
            ).fetchall()
        except Exception:
            active_salaries = []
        for sal in active_salaries:
            sid = sal['id']
            sbudget = request.form.get(f'sal_budget_{sid}', '0')
            sactual = request.form.get(f'sal_actual_{sid}', '0')
            snotes = request.form.get(f'sal_notes_{sid}', '').strip()
            try:
                sbudget = float(sbudget) if sbudget else 0
            except ValueError:
                sbudget = 0
            try:
                sactual = float(sactual) if sactual else 0
            except ValueError:
                sactual = 0
            existing_s = conn.execute(
                "SELECT id FROM budget_entries WHERE salary_id = ? AND fy_year = ? AND month = ? AND year = ?",
                (sid, fy_year, month, year)
            ).fetchone()
            if existing_s:
                conn.execute(
                    "UPDATE budget_entries SET budget_amount = ?, actual_amount = ?, notes = ?, project_id = ?, updated_by = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (sbudget, sactual, snotes, sal['project_id'], user['id'], existing_s['id'])
                )
            else:
                if sbudget > 0 or sactual > 0 or snotes:
                    conn.execute(
                        "INSERT INTO budget_entries (salary_id, fy_year, month, year, budget_amount, actual_amount, notes, project_id, created_by, updated_by) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (sid, fy_year, month, year, sbudget, sactual, snotes, sal['project_id'], user['id'], user['id'])
                    )

        # ── Subscription line-item expenses ──
        try:
            active_subs = conn.execute(
                "SELECT id, primary_department, project_id FROM subscription_items WHERE is_active = 1"
            ).fetchall()
        except Exception:
            active_subs = []
        for sub in active_subs:
            sid = sub['id']
            sbudget = request.form.get(f'sub_budget_{sid}', '0')
            sactual = request.form.get(f'sub_actual_{sid}', '0')
            snotes = request.form.get(f'sub_notes_{sid}', '').strip()
            try:
                sbudget = float(sbudget) if sbudget else 0
            except ValueError:
                sbudget = 0
            try:
                sactual = float(sactual) if sactual else 0
            except ValueError:
                sactual = 0
            existing_sub = conn.execute(
                "SELECT id FROM budget_entries WHERE subscription_id = ? AND fy_year = ? AND month = ? AND year = ?",
                (sid, fy_year, month, year)
            ).fetchone()
            if existing_sub:
                conn.execute(
                    "UPDATE budget_entries SET budget_amount = ?, actual_amount = ?, notes = ?, project_id = ?, updated_by = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (sbudget, sactual, snotes, sub['project_id'], user['id'], existing_sub['id'])
                )
            else:
                if sbudget > 0 or sactual > 0 or snotes:
                    conn.execute(
                        "INSERT INTO budget_entries (subscription_id, fy_year, month, year, budget_amount, actual_amount, notes, project_id, created_by, updated_by) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (sid, fy_year, month, year, sbudget, sactual, snotes, sub['project_id'], user['id'], user['id'])
                    )

        # ── Product-level revenue entries ──
        # When a product has a sale_price, the monthly form collects UNITS
        # (quantity) and the amount is auto-calculated as units × sale_price
        # (converted to INR). If sale_price is 0, the form falls back to
        # rupee amounts for backward compatibility.
        try:
            active_products = conn.execute(
                "SELECT id, project_id, sale_price, sale_currency FROM products_services WHERE status = 'active'"
            ).fetchall()
        except Exception:
            active_products = []
        for prod in active_products:
            pid = prod['id']
            try:
                unit_price_inr = float(to_inr(prod['sale_price'], prod['sale_currency'] or 'INR') or 0)
            except Exception:
                unit_price_inr = 0.0

            pnotes = request.form.get(f'prod_notes_{pid}', '').strip()

            if unit_price_inr > 0:
                # Units-based entry
                try:
                    pbudget_units = float(request.form.get(f'prod_budget_units_{pid}', '0') or 0)
                except ValueError:
                    pbudget_units = 0
                try:
                    pactual_units = float(request.form.get(f'prod_actual_units_{pid}', '0') or 0)
                except ValueError:
                    pactual_units = 0
                pbudget = pbudget_units * unit_price_inr
                pactual = pactual_units * unit_price_inr
            else:
                # Legacy rupee-amount entry (no sale_price set on product)
                try:
                    pbudget = float(request.form.get(f'prod_budget_{pid}', '0') or 0)
                except ValueError:
                    pbudget = 0
                try:
                    pactual = float(request.form.get(f'prod_actual_{pid}', '0') or 0)
                except ValueError:
                    pactual = 0
                pbudget_units = 0
                pactual_units = 0

            existing_p = conn.execute(
                "SELECT id FROM budget_entries WHERE product_id = ? AND fy_year = ? AND month = ? AND year = ?",
                (pid, fy_year, month, year)
            ).fetchone()
            if existing_p:
                conn.execute(
                    "UPDATE budget_entries SET budget_amount = ?, actual_amount = ?, budget_units = ?, actual_units = ?, notes = ?, project_id = ?, updated_by = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (pbudget, pactual, pbudget_units, pactual_units, pnotes, prod['project_id'], user['id'], existing_p['id'])
                )
            else:
                if pbudget > 0 or pactual > 0 or pnotes:
                    conn.execute(
                        "INSERT INTO budget_entries (product_id, fy_year, month, year, budget_amount, actual_amount, budget_units, actual_units, notes, project_id, created_by, updated_by) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (pid, fy_year, month, year, pbudget, pactual, pbudget_units, pactual_units, pnotes, prod['project_id'], user['id'], user['id'])
                    )

        # Auto-fill recurring budgets to other unlocked months
        if recurring_fills:
            all_fy_months = get_fy_months(fy_year)
            for fm in all_fy_months:
                if fm['month'] == month and fm['year'] == year:
                    continue  # Skip the current month
                # Check if month is locked
                lock_chk = conn.execute(
                    "SELECT is_locked FROM budget_entries WHERE fy_year = ? AND month = ? AND year = ? AND is_locked = 1 LIMIT 1",
                    (fy_year, fm['month'], fm['year'])
                ).fetchone()
                if lock_chk:
                    continue
                for rc in recurring_fills:
                    existing_rc = conn.execute(
                        "SELECT id, budget_amount FROM budget_entries WHERE category_id = ? AND fy_year = ? AND month = ? AND year = ?",
                        (rc['cat_id'], fy_year, fm['month'], fm['year'])
                    ).fetchone()
                    if existing_rc:
                        # Only fill if budget is still 0 (don't overwrite manual edits)
                        if float(existing_rc['budget_amount'] or 0) == 0:
                            conn.execute(
                                "UPDATE budget_entries SET budget_amount = ?, notes = ?, updated_by = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                                (rc['budget'], rc['notes'], user['id'], existing_rc['id'])
                            )
                    else:
                        conn.execute(
                            "INSERT INTO budget_entries (category_id, fy_year, month, year, budget_amount, actual_amount, notes, created_by, updated_by) VALUES (?, ?, ?, ?, ?, 0, ?, ?, ?)",
                            (rc['cat_id'], fy_year, fm['month'], fm['year'], rc['budget'], rc['notes'], user['id'], user['id'])
                        )

        conn.commit()
        recurring_msg = f' Recurring budgets auto-filled to other months.' if recurring_fills else ''
        flash(f'Budget data for {calendar.month_name[month]} {year} saved successfully!{recurring_msg}', 'success')
        conn.close()
        return redirect(url_for('finance_edit_month', month=month, year=year, fy=fy_year))

    # GET: load existing data
    # Hide legacy Salaries & Wages category — salaries are now line-item managed
    expense_cats = conn.execute(
        "SELECT * FROM budget_categories WHERE cat_type = 'expense' AND is_active = 1 AND LOWER(name) NOT IN ('salaries & wages', 'salaries and wages', 'salary') ORDER BY sort_order"
    ).fetchall()
    dept_cats = conn.execute(
        "SELECT * FROM budget_categories WHERE cat_type = 'department' AND is_active = 1 ORDER BY sort_order"
    ).fetchall()

    entries = conn.execute(
        "SELECT * FROM budget_entries WHERE fy_year = ? AND month = ? AND year = ?",
        (fy_year, month, year)
    ).fetchall()

    # Load active projects for the dropdown
    try:
        projects = conn.execute(
            "SELECT id, name FROM projects WHERE status = 'active' ORDER BY name"
        ).fetchall()
    except Exception:
        projects = []

    # Load active salary items
    try:
        salary_items_raw = conn.execute('''
            SELECT s.id, s.name, s.department, s.project_id, s.monthly_cost, s.currency,
                   p.name AS project_name
            FROM salary_items s
            LEFT JOIN projects p ON s.project_id = p.id
            WHERE s.is_active = 1
            ORDER BY s.department NULLS LAST, s.name
        ''').fetchall()
    except Exception as e:
        logging.error(f"finance_edit_month salary query: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        salary_items_raw = []

    # Load active subscription items
    try:
        subscription_items_raw = conn.execute('''
            SELECT s.id, s.name, s.vendor, s.cost, s.currency, s.frequency,
                   s.primary_department, s.shared_departments, s.project_id,
                   p.name AS project_name
            FROM subscription_items s
            LEFT JOIN projects p ON s.project_id = p.id
            WHERE s.is_active = 1
            ORDER BY s.primary_department NULLS LAST, s.name
        ''').fetchall()
    except Exception as e:
        logging.error(f"finance_edit_month subscription query: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        subscription_items_raw = []

    # Load active products (revenue is entered at product level)
    try:
        products_raw = conn.execute('''
            SELECT ps.id, ps.name, ps.type, ps.sale_price, ps.sale_currency,
                   ps.product_cost, ps.cost_currency,
                   ps.project_id, ps.revenue_stream_id,
                   p.name  AS project_name,
                   rs.name AS stream_name
            FROM products_services ps
            LEFT JOIN projects p         ON ps.project_id = p.id
            LEFT JOIN revenue_streams rs ON ps.revenue_stream_id = rs.id
            WHERE ps.status = 'active'
            ORDER BY rs.name NULLS LAST, p.name NULLS LAST, ps.name
        ''').fetchall()
    except Exception as e:
        logging.error(f"finance_edit_month products query: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        products_raw = []

    # Seed department rollup with active employee departments
    try:
        emp_depts_raw = conn.execute(
            "SELECT DISTINCT department FROM employees WHERE is_active = 1 AND department IS NOT NULL AND department != ''"
        ).fetchall()
    except Exception:
        emp_depts_raw = []

    conn.close()

    # Map category-based entries (expense/department)
    entry_map = {}
    # Map product-based entries (revenue)
    product_entry_map = {}
    # Map salary/subscription-based entries (line-item expenses)
    salary_entry_map = {}
    subscription_entry_map = {}
    for e in entries:
        keys = e.keys() if hasattr(e, 'keys') else []
        cid = e['category_id'] if 'category_id' in keys else None
        pid = e['product_id'] if 'product_id' in keys else None
        sid = e['salary_id'] if 'salary_id' in keys else None
        subid = e['subscription_id'] if 'subscription_id' in keys else None
        if sid:
            salary_entry_map[sid] = e
        elif subid:
            subscription_entry_map[subid] = e
        elif pid:
            product_entry_map[pid] = e
        elif cid:
            entry_map[cid] = e

    # Group salary items by department
    salary_groups = []
    sal_group_index = {}
    for s in salary_items_raw:
        dept_key = s['department'] or '— No department —'
        if dept_key not in sal_group_index:
            sal_group_index[dept_key] = {
                'department': dept_key,
                'line_items': [],
                'total_budget': 0.0,
                'total_actual': 0.0,
            }
            salary_groups.append(sal_group_index[dept_key])
        e = salary_entry_map.get(s['id'])
        sb = float(e['budget_amount'] or 0) if e else float(s['monthly_cost'] or 0)
        sa = float(e['actual_amount'] or 0) if e else 0.0
        sal_group_index[dept_key]['total_budget'] += sb
        sal_group_index[dept_key]['total_actual'] += sa
        sal_group_index[dept_key]['line_items'].append({
            'id': s['id'],
            'name': s['name'],
            'project_name': s['project_name'] or '— No project —',
            'monthly_cost': float(s['monthly_cost'] or 0),
            'currency': s['currency'] or 'INR',
            'budget': sb,
            'actual': sa,
            'notes': (e['notes'] if e else '') or '',
            'has_entry': e is not None,
        })

    # Group subscription items by primary department
    subscription_groups = []
    sub_group_index = {}
    for s in subscription_items_raw:
        dept_key = s['primary_department'] or '— No department —'
        if dept_key not in sub_group_index:
            sub_group_index[dept_key] = {
                'department': dept_key,
                'line_items': [],
                'total_budget': 0.0,
                'total_actual': 0.0,
            }
            subscription_groups.append(sub_group_index[dept_key])
        e = subscription_entry_map.get(s['id'])
        # Default monthly-equivalent if no entry yet
        freq = (s['frequency'] or 'monthly').lower()
        raw_cost = float(s['cost'] or 0)
        if freq == 'monthly':    default_cost = raw_cost
        elif freq == 'quarterly': default_cost = raw_cost / 3.0
        elif freq == 'annual':    default_cost = raw_cost / 12.0
        else:                     default_cost = 0.0
        sb = float(e['budget_amount'] or 0) if e else default_cost
        sa = float(e['actual_amount'] or 0) if e else 0.0
        sub_group_index[dept_key]['total_budget'] += sb
        sub_group_index[dept_key]['total_actual'] += sa
        sub_group_index[dept_key]['line_items'].append({
            'id': s['id'],
            'name': s['name'],
            'vendor': s['vendor'] or '',
            'frequency': freq,
            'raw_cost': raw_cost,
            'currency': s['currency'] or 'INR',
            'shared_departments': s['shared_departments'] or '',
            'project_name': s['project_name'] or '— No project —',
            'budget': sb,
            'actual': sa,
            'notes': (e['notes'] if e else '') or '',
            'has_entry': e is not None,
        })

    # Project-wise expense rollup from salaries + subscriptions for this month
    expense_project_rollup = {}
    for g in salary_groups:
        for it in g['line_items']:
            key = it['project_name']
            if key not in expense_project_rollup:
                expense_project_rollup[key] = {'project_name': key, 'salary': 0.0, 'subscription': 0.0}
            expense_project_rollup[key]['salary'] += it['budget']
    for g in subscription_groups:
        for it in g['line_items']:
            key = it['project_name']
            if key not in expense_project_rollup:
                expense_project_rollup[key] = {'project_name': key, 'salary': 0.0, 'subscription': 0.0}
            expense_project_rollup[key]['subscription'] += it['budget']
    expense_project_rollup_list = sorted(expense_project_rollup.values(), key=lambda x: x['project_name'])

    # Department rollup — derived from salaries + subscriptions + dept-tagged expense categories
    # (no more editable dept budgets). Seed from employee departments so every dept shows up.
    department_rollup = {}
    def _dept_bucket(name):
        k = name or '— Unassigned —'
        if k not in department_rollup:
            department_rollup[k] = {
                'department': k,
                'salary_budget': 0.0, 'salary_actual': 0.0,
                'subscription_budget': 0.0, 'subscription_actual': 0.0,
                'category_budget': 0.0, 'category_actual': 0.0,
                'salary_count': 0, 'subscription_count': 0, 'category_count': 0,
            }
        return department_rollup[k]

    # Seed with active employee departments so zero-expense depts still render
    for d in emp_depts_raw:
        _dept_bucket(d['department'])

    for g in salary_groups:
        b = _dept_bucket(g['department'])
        b['salary_budget'] += g['total_budget']
        b['salary_actual'] += g['total_actual']
        b['salary_count'] += len(g['line_items'])
    for g in subscription_groups:
        b = _dept_bucket(g['department'])
        b['subscription_budget'] += g['total_budget']
        b['subscription_actual'] += g['total_actual']
        b['subscription_count'] += len(g['line_items'])
    # Dept-tagged expense categories for this month
    for cat in expense_cats:
        dept_name = cat['department'] if 'department' in cat.keys() else None
        if not dept_name:
            continue
        e = entry_map.get(cat['id'], {})
        cb = float(e.get('budget_amount') or 0)
        ca = float(e.get('actual_amount') or 0)
        b = _dept_bucket(dept_name)
        b['category_budget'] += cb
        b['category_actual'] += ca
        b['category_count'] += 1

    for v in department_rollup.values():
        v['total_budget'] = v['salary_budget'] + v['subscription_budget'] + v['category_budget']
        v['total_actual'] = v['salary_actual'] + v['subscription_actual'] + v['category_actual']

    department_rollup_list = sorted(
        department_rollup.values(),
        key=lambda x: (-x['total_budget'], x['department'])
    )

    # Group products by revenue stream for display
    revenue_groups = []
    group_index = {}
    for p in products_raw:
        stream_key = p['stream_name'] or '— Unassigned stream —'
        if stream_key not in group_index:
            group_index[stream_key] = {
                'stream_name': stream_key,
                'stream_id': p['revenue_stream_id'],
                'products': [],
                'total_budget': 0.0,
                'total_actual': 0.0,
            }
            revenue_groups.append(group_index[stream_key])
        e = product_entry_map.get(p['id'])
        pb = float(e['budget_amount'] or 0) if e else 0.0
        pa = float(e['actual_amount'] or 0) if e else 0.0
        try:
            _ekeys = e.keys() if e and hasattr(e, 'keys') else []
        except Exception:
            _ekeys = []
        pb_units = float((e['budget_units'] or 0) if e and 'budget_units' in _ekeys else 0)
        pa_units = float((e['actual_units'] or 0) if e and 'actual_units' in _ekeys else 0)
        _sale_cur = p['sale_currency'] if 'sale_currency' in p.keys() else 'INR'
        _sale_price_inr = float(to_inr(p['sale_price'], _sale_cur) or 0)
        group_index[stream_key]['total_budget'] += pb
        group_index[stream_key]['total_actual'] += pa
        group_index[stream_key]['products'].append({
            'id': p['id'],
            'name': p['name'],
            'type': p['type'],
            'project_name': p['project_name'] or '— No project —',
            'project_id': p['project_id'],
            'sale_price': float(p['sale_price'] or 0),
            'sale_currency': _sale_cur,
            'sale_price_inr': _sale_price_inr,
            'budget': pb,
            'actual': pa,
            'budget_units': pb_units,
            'actual_units': pa_units,
            'notes': (e['notes'] if e else '') or '',
        })

    # Project-wise rollup for the month
    project_rollup = {}
    for g in revenue_groups:
        for pr in g['products']:
            key = pr['project_name']
            if key not in project_rollup:
                project_rollup[key] = {'project_name': key, 'budget': 0.0, 'actual': 0.0}
            project_rollup[key]['budget'] += pr['budget']
            project_rollup[key]['actual'] += pr['actual']
    project_rollup_list = sorted(project_rollup.values(), key=lambda x: x['project_name'])

    month_name = calendar.month_name[month]
    return render_template('finance_edit_month.html', user=user, fy_year=fy_year,
                           month=month, year=year, month_name=month_name,
                           expense_cats=expense_cats, dept_cats=dept_cats,
                           revenue_groups=revenue_groups,
                           project_rollup=project_rollup_list,
                           salary_groups=salary_groups,
                           subscription_groups=subscription_groups,
                           expense_project_rollup=expense_project_rollup_list,
                           department_rollup=department_rollup_list,
                           entry_map=entry_map,
                           is_locked=is_locked, get_fy_months=get_fy_months,
                           projects=projects)


@app.route('/finance/budget/lock/<int:month>/<int:year>', methods=['POST'])
@admin_required
def finance_lock_month(month, year):
    fy_year = request.form.get('fy', '2026-2027')
    conn = get_db()
    conn.execute(
        "UPDATE budget_entries SET is_locked = 1 WHERE fy_year = ? AND month = ? AND year = ?",
        (fy_year, month, year)
    )
    conn.commit()
    conn.close()
    flash(f'{calendar.month_name[month]} {year} has been locked.', 'success')
    return redirect(url_for('finance_edit_month', month=month, year=year, fy=fy_year))


@app.route('/finance/budget/unlock/<int:month>/<int:year>', methods=['POST'])
@admin_required
def finance_unlock_month(month, year):
    fy_year = request.form.get('fy', '2026-2027')
    conn = get_db()
    conn.execute(
        "UPDATE budget_entries SET is_locked = 0 WHERE fy_year = ? AND month = ? AND year = ?",
        (fy_year, month, year)
    )
    conn.commit()
    conn.close()
    flash(f'{calendar.month_name[month]} {year} has been unlocked.', 'success')
    return redirect(url_for('finance_edit_month', month=month, year=year, fy=fy_year))


@app.route('/finance/budget/report')
@admin_required
def finance_report():
    user = get_user()
    fy_year = request.args.get('fy', '2026-2027')
    fy_months = get_fy_months(fy_year)
    conn = get_db()

    all_cats = conn.execute(
        "SELECT * FROM budget_categories WHERE is_active = 1 ORDER BY cat_type, sort_order"
    ).fetchall()
    entries = conn.execute("SELECT * FROM budget_entries WHERE fy_year = ?", (fy_year,)).fetchall()
    conn.close()

    entry_map = {}
    for e in entries:
        entry_map[(e['category_id'], e['month'], e['year'])] = e

    expense_cats = [c for c in all_cats if c['cat_type'] == 'expense']
    dept_cats = [c for c in all_cats if c['cat_type'] == 'department']
    revenue_cats = [c for c in all_cats if c['cat_type'] == 'revenue']

    return render_template('finance_report.html', user=user, fy_year=fy_year, fy_months=fy_months,
                           expense_cats=expense_cats, dept_cats=dept_cats, revenue_cats=revenue_cats,
                           entry_map=entry_map, get_quarter_label=get_quarter_label)


@app.route('/finance/budget/category/add', methods=['POST'])
@admin_required
def finance_add_category():
    name = request.form.get('name', '').strip()
    cat_type = request.form.get('cat_type', 'expense')
    department = request.form.get('department', '').strip() or None
    fy = request.form.get('fy', '2026-2027')
    source = request.form.get('source', '')  # 'settings' if from settings page

    if not name:
        flash('Category name is required', 'error')
        if source == 'settings':
            return redirect(url_for('finance_settings', fy=fy))
        return redirect(url_for('finance_dashboard', fy=fy))

    conn = get_db()
    max_sort = conn.execute(
        "SELECT MAX(sort_order) as mx FROM budget_categories WHERE cat_type = ?", (cat_type,)
    ).fetchone()
    next_sort = (max_sort['mx'] or 0) + 1

    conn.execute(
        "INSERT INTO budget_categories (name, cat_type, department, sort_order) VALUES (?, ?, ?, ?)",
        (name, cat_type, department, next_sort)
    )
    conn.commit()
    conn.close()
    flash(f'Category "{name}" added successfully!', 'success')

    if source == 'settings':
        return redirect(url_for('finance_settings', fy=fy))
    if cat_type == 'revenue':
        return redirect(url_for('finance_revenue', fy=fy))
    return redirect(url_for('finance_expenses', fy=fy))


def sync_streams_to_budget_categories(conn):
    """Mirror every sales revenue_stream into budget_categories so the
    finance module's revenue list is driven by sales. Matches by stream_id.
    - New stream -> inserts a revenue budget_category with stream_id set
    - Renamed / (de)activated stream -> updates the linked category
    - Deleted stream -> the linked category (and its budget_entries) is
      deleted so Finance Settings mirrors Sales exactly
    """
    try:
        streams = conn.execute('SELECT id, name, is_active FROM revenue_streams').fetchall()
        linked = conn.execute(
            "SELECT id, stream_id, name, is_active FROM budget_categories "
            "WHERE cat_type = 'revenue' AND stream_id IS NOT NULL"
        ).fetchall()
        by_stream = {row['stream_id']: row for row in linked}
        active_stream_ids = set()
        for s in streams:
            active_stream_ids.add(s['id'])
            existing = by_stream.get(s['id'])
            if existing:
                if existing['name'] != s['name'] or int(existing['is_active'] or 0) != int(s['is_active'] or 0):
                    conn.execute(
                        "UPDATE budget_categories SET name = ?, is_active = ? WHERE id = ?",
                        (s['name'], int(s['is_active'] or 0), existing['id'])
                    )
            else:
                conn.execute(
                    "INSERT INTO budget_categories (name, cat_type, sort_order, is_active, stream_id) "
                    "VALUES (?, 'revenue', 0, ?, ?)",
                    (s['name'], int(s['is_active'] or 0), s['id'])
                )
        # Hard-delete categories whose stream has been deleted so Finance
        # Settings reflects Sales exactly. Also purge their budget_entries
        # (NOT NULL FK) so the delete doesn't fail.
        for sid, row in by_stream.items():
            if sid not in active_stream_ids:
                try:
                    conn.execute("DELETE FROM budget_entries WHERE category_id = ?", (row['id'],))
                except Exception:
                    pass
                conn.execute("DELETE FROM budget_categories WHERE id = ?", (row['id'],))
        conn.commit()
    except Exception as e:
        logging.error(f"sync_streams_to_budget_categories: {e}")
        try:
            conn.rollback()
        except Exception:
            pass


@app.route('/finance/budget/settings')
@admin_required
def finance_settings():
    user = get_user()
    fy_year = request.args.get('fy', '2026-2027')
    conn = get_db()
    # Keep finance revenue list in sync with sales streams on every load
    sync_streams_to_budget_categories(conn)
    expense_cats = conn.execute(
        "SELECT * FROM budget_categories WHERE cat_type = 'expense' ORDER BY sort_order"
    ).fetchall()
    dept_cats = conn.execute(
        "SELECT * FROM budget_categories WHERE cat_type = 'department' ORDER BY sort_order"
    ).fetchall()
    # Only show revenue categories that came from sales streams
    revenue_cats = conn.execute(
        "SELECT * FROM budget_categories WHERE cat_type = 'revenue' AND stream_id IS NOT NULL "
        "ORDER BY is_active DESC, name"
    ).fetchall()
    # Get unique departments for the dropdown
    departments = sorted(set(d['department'] for d in dept_cats if d['department']))
    conn.close()
    return render_template('finance_settings.html', user=user, fy_year=fy_year,
                           expense_cats=expense_cats, dept_cats=dept_cats,
                           revenue_cats=revenue_cats, departments=departments)


@app.route('/finance/budget/category/edit/<int:cat_id>', methods=['POST'])
@admin_required
def finance_edit_category(cat_id):
    fy = request.form.get('fy', '2026-2027')
    name = request.form.get('name', '').strip()
    department = request.form.get('department', '').strip() or None
    is_recurring = 1 if request.form.get('is_recurring') else 0
    is_active = 1 if request.form.get('is_active') else 0

    if not name:
        flash('Category name is required', 'error')
        return redirect(url_for('finance_settings', fy=fy))

    conn = get_db()
    conn.execute(
        "UPDATE budget_categories SET name = ?, department = ?, is_recurring = ?, is_active = ? WHERE id = ?",
        (name, department, is_recurring, is_active, cat_id)
    )
    conn.commit()
    conn.close()
    flash(f'Category "{name}" updated!', 'success')
    return redirect(url_for('finance_settings', fy=fy))


@app.route('/finance/budget/category/toggle/<int:cat_id>', methods=['POST'])
@admin_required
def finance_toggle_category(cat_id):
    fy = request.form.get('fy', '2026-2027')
    conn = get_db()
    cat = conn.execute("SELECT * FROM budget_categories WHERE id = ?", (cat_id,)).fetchone()
    if cat:
        new_status = 0 if cat['is_active'] else 1
        conn.execute("UPDATE budget_categories SET is_active = ? WHERE id = ?", (new_status, cat_id))
        conn.commit()
        action = 'activated' if new_status else 'deactivated'
        flash(f'Category "{cat["name"]}" {action}!', 'success')
    conn.close()
    return redirect(url_for('finance_settings', fy=fy))


@app.route('/finance/budget/department/add', methods=['POST'])
@admin_required
def finance_add_department():
    fy = request.form.get('fy', '2026-2027')
    name = request.form.get('name', '').strip()
    if not name:
        flash('Department name is required', 'error')
        return redirect(url_for('finance_settings', fy=fy))

    conn = get_db()
    # Check if dept already exists
    existing = conn.execute(
        "SELECT id FROM budget_categories WHERE cat_type = 'department' AND department = ?", (name,)
    ).fetchone()
    if existing:
        flash(f'Department "{name}" already exists', 'error')
        conn.close()
        return redirect(url_for('finance_settings', fy=fy))

    max_sort = conn.execute(
        "SELECT MAX(sort_order) as mx FROM budget_categories WHERE cat_type = 'department'"
    ).fetchone()
    next_sort = (max_sort['mx'] or 0) + 1
    conn.execute(
        "INSERT INTO budget_categories (name, cat_type, department, sort_order) VALUES (?, 'department', ?, ?)",
        (f'{name} - Department Budget', name, next_sort)
    )
    conn.commit()
    conn.close()
    flash(f'Department "{name}" added!', 'success')
    return redirect(url_for('finance_settings', fy=fy))


@app.route('/finance/budget/department/edit/<int:cat_id>', methods=['POST'])
@admin_required
def finance_edit_department(cat_id):
    fy = request.form.get('fy', '2026-2027')
    name = request.form.get('name', '').strip()
    is_active = 1 if request.form.get('is_active') else 0
    if not name:
        flash('Department name is required', 'error')
        return redirect(url_for('finance_settings', fy=fy))

    conn = get_db()
    conn.execute(
        "UPDATE budget_categories SET name = ?, department = ?, is_active = ? WHERE id = ?",
        (f'{name} - Department Budget', name, is_active, cat_id)
    )
    conn.commit()
    conn.close()
    flash(f'Department "{name}" updated!', 'success')
    return redirect(url_for('finance_settings', fy=fy))


@app.route('/finance/budget/projects')
@admin_required
def finance_project_pl():
    """Project-wise Profit & Loss view."""
    user = get_user()
    fy_year = request.args.get('fy', '2026-2027')
    fy_months = get_fy_months(fy_year)
    conn = get_db()

    # Get all projects that have budget entries
    try:
        projects = conn.execute(
            "SELECT id, name FROM projects WHERE status = 'active' ORDER BY name"
        ).fetchall()
    except Exception:
        projects = []

    # Get all budget entries with project_id set, joined with category info
    try:
        entries = conn.execute('''
            SELECT be.project_id, be.month, be.year, bc.cat_type,
                   COALESCE(SUM(be.budget_amount), 0) as total_budget,
                   COALESCE(SUM(be.actual_amount), 0) as total_actual
            FROM budget_entries be
            JOIN budget_categories bc ON be.category_id = bc.id
            WHERE be.fy_year = ? AND be.project_id IS NOT NULL AND bc.is_active = 1
            GROUP BY be.project_id, be.month, be.year, bc.cat_type
        ''', (fy_year,)).fetchall()
    except Exception as e:
        logging.error(f"finance_project_pl entries: {e}")
        entries = []

    # Also get untagged totals
    try:
        untagged = conn.execute('''
            SELECT be.month, be.year, bc.cat_type,
                   COALESCE(SUM(be.budget_amount), 0) as total_budget,
                   COALESCE(SUM(be.actual_amount), 0) as total_actual
            FROM budget_entries be
            JOIN budget_categories bc ON be.category_id = bc.id
            WHERE be.fy_year = ? AND be.project_id IS NULL AND bc.is_active = 1
            GROUP BY be.month, be.year, bc.cat_type
        ''', (fy_year,)).fetchall()
    except Exception:
        untagged = []

    conn.close()

    # Build project lookup: {project_id: {name, expense_budget, expense_actual, revenue_budget, revenue_actual, monthly: [{...}]}}
    project_map = {p['id']: {'name': p['name'], 'id': p['id'],
                              'expense_budget': 0, 'expense_actual': 0,
                              'revenue_budget': 0, 'revenue_actual': 0,
                              'monthly_expense': [0]*12, 'monthly_revenue': [0]*12}
                   for p in projects}

    month_index = {}
    for i, fm in enumerate(fy_months):
        month_index[(fm['month'], fm['year'])] = i

    for e in entries:
        pid = e['project_id']
        if pid not in project_map:
            continue
        idx = month_index.get((e['month'], e['year']))
        if idx is None:
            continue
        ct = e['cat_type']
        b = float(e['total_budget'] or 0)
        a = float(e['total_actual'] or 0)
        if ct in ('expense', 'department'):
            project_map[pid]['expense_budget'] += b
            project_map[pid]['expense_actual'] += a
            project_map[pid]['monthly_expense'][idx] += round(a, 2)
        elif ct == 'revenue':
            project_map[pid]['revenue_budget'] += b
            project_map[pid]['revenue_actual'] += a
            project_map[pid]['monthly_revenue'][idx] += round(a, 2)

    # Untagged totals
    untagged_data = {'name': 'Untagged (No Project)', 'id': None,
                     'expense_budget': 0, 'expense_actual': 0,
                     'revenue_budget': 0, 'revenue_actual': 0,
                     'monthly_expense': [0]*12, 'monthly_revenue': [0]*12}
    for e in untagged:
        idx = month_index.get((e['month'], e['year']))
        if idx is None:
            continue
        ct = e['cat_type']
        b = float(e['total_budget'] or 0)
        a = float(e['total_actual'] or 0)
        if ct in ('expense', 'department'):
            untagged_data['expense_budget'] += b
            untagged_data['expense_actual'] += a
            untagged_data['monthly_expense'][idx] += round(a, 2)
        elif ct == 'revenue':
            untagged_data['revenue_budget'] += b
            untagged_data['revenue_actual'] += a
            untagged_data['monthly_revenue'][idx] += round(a, 2)

    # Build project list with computed P&L
    project_list = []
    for pid, pd in project_map.items():
        pd['net_budget'] = round(pd['revenue_budget'] - pd['expense_budget'], 2)
        pd['net_actual'] = round(pd['revenue_actual'] - pd['expense_actual'], 2)
        pd['expense_budget'] = round(pd['expense_budget'], 2)
        pd['expense_actual'] = round(pd['expense_actual'], 2)
        pd['revenue_budget'] = round(pd['revenue_budget'], 2)
        pd['revenue_actual'] = round(pd['revenue_actual'], 2)
        project_list.append(pd)

    # Add untagged
    untagged_data['net_budget'] = round(untagged_data['revenue_budget'] - untagged_data['expense_budget'], 2)
    untagged_data['net_actual'] = round(untagged_data['revenue_actual'] - untagged_data['expense_actual'], 2)

    # Sort by actual revenue descending
    project_list.sort(key=lambda x: x['revenue_actual'], reverse=True)

    # ── Sales-side rollup: Revenue Streams → Products for each project (live FX → INR) ──
    # Streams are global; group products by (project_id, revenue_stream_id)
    sales_rollup = []
    try:
        conn2 = get_db()
        all_streams_raw = conn2.execute('SELECT * FROM revenue_streams ORDER BY name').fetchall()
        all_products = conn2.execute(
            'SELECT id, project_id, revenue_stream_id, name, type, product_cost, sale_price, cost_currency, sale_currency, status FROM products_services'
        ).fetchall()
        conn2.close()
        stream_lookup = {s['id']: dict(s) for s in all_streams_raw}
        # Bucket products by (project_id, stream_id)
        buckets = {}
        for p in all_products:
            d = dict(p)
            d['cost_inr'] = to_inr(d.get('product_cost'), d.get('cost_currency') or 'INR')
            d['revenue_inr'] = to_inr(d.get('sale_price'), d.get('sale_currency') or 'INR')
            d['margin_inr'] = d['revenue_inr'] - d['cost_inr']
            key = (d.get('project_id'), d.get('revenue_stream_id'))
            buckets.setdefault(key, []).append(d)
        # Build rollup per project
        project_name_lookup = {p['id']: p['name'] for p in projects}
        for pid, pname in project_name_lookup.items():
            streams_for_p = []
            unassigned_for_p = []
            for (bpid, sid), items in buckets.items():
                if bpid != pid:
                    continue
                if sid and sid in stream_lookup:
                    sd = dict(stream_lookup[sid])
                    sd['products'] = items
                    sd['total_cost'] = sum(i['cost_inr'] for i in items)
                    sd['total_revenue'] = sum(i['revenue_inr'] for i in items)
                    sd['total_margin'] = sd['total_revenue'] - sd['total_cost']
                    streams_for_p.append(sd)
                else:
                    unassigned_for_p.extend(items)
            if not streams_for_p and not unassigned_for_p:
                continue
            streams_for_p.sort(key=lambda s: s['name'])
            total_cost = sum(s['total_cost'] for s in streams_for_p) + sum(i['cost_inr'] for i in unassigned_for_p)
            total_rev = sum(s['total_revenue'] for s in streams_for_p) + sum(i['revenue_inr'] for i in unassigned_for_p)
            sales_rollup.append({
                'project_id': pid,
                'project_name': pname,
                'streams': streams_for_p,
                'unassigned': unassigned_for_p,
                'total_cost': total_cost,
                'total_revenue': total_rev,
                'total_margin': total_rev - total_cost,
            })
        sales_rollup.sort(key=lambda x: x['total_revenue'], reverse=True)
    except Exception as e:
        logging.error(f"finance_project_pl sales_rollup: {e}")
        sales_rollup = []

    return render_template('finance_project_pl.html', user=user, fy_year=fy_year,
                           fy_months=fy_months, project_list=project_list,
                           untagged_data=untagged_data,
                           month_labels=[fm['label'] for fm in fy_months],
                           sales_rollup=sales_rollup,
                           fx_rates=get_fx_rates_inr())


@app.route('/finance/products')
@admin_required
def finance_products():
    """Finance view of sales products — pick from the Sales product list,
    see project + revenue stream context, and rolled-up INR totals."""
    user = get_user()
    fy_year = request.args.get('fy', '2026-2027')
    conn = get_db()
    try:
        products_raw = conn.execute('''
            SELECT ps.id, ps.name, ps.type, ps.description, ps.status,
                   ps.product_cost, ps.sale_price, ps.cost_currency, ps.sale_currency,
                   ps.project_id, ps.revenue_stream_id,
                   p.name as project_name,
                   rs.name as stream_name
            FROM products_services ps
            LEFT JOIN projects p ON ps.project_id = p.id
            LEFT JOIN revenue_streams rs ON ps.revenue_stream_id = rs.id
            ORDER BY p.name, rs.name, ps.name
        ''').fetchall()
    except Exception as e:
        logging.error(f"finance_products: {e}")
        products_raw = []

    # Pull booked actual revenue per project from budget_entries (revenue cat_type)
    project_actuals = {}
    try:
        rows = conn.execute('''
            SELECT be.project_id,
                   COALESCE(SUM(be.actual_amount), 0) as total_actual,
                   COALESCE(SUM(be.budget_amount), 0) as total_budget
            FROM budget_entries be
            JOIN budget_categories bc ON be.category_id = bc.id
            WHERE be.fy_year = ? AND bc.cat_type = 'revenue' AND be.project_id IS NOT NULL
            GROUP BY be.project_id
        ''', (fy_year,)).fetchall()
        for r in rows:
            project_actuals[r['project_id']] = {
                'actual': float(r['total_actual'] or 0),
                'budget': float(r['total_budget'] or 0),
            }
    except Exception as e:
        logging.error(f"finance_products actuals: {e}")
    conn.close()

    # Convert products to INR and attach project actuals context
    products = []
    total_expected_cost = 0.0
    total_expected_rev = 0.0
    for p in products_raw:
        d = dict(p)
        d['cost_inr'] = to_inr(d.get('product_cost'), d.get('cost_currency') or 'INR')
        d['revenue_inr'] = to_inr(d.get('sale_price'), d.get('sale_currency') or 'INR')
        d['margin_inr'] = d['revenue_inr'] - d['cost_inr']
        total_expected_cost += d['cost_inr']
        total_expected_rev += d['revenue_inr']
        pact = project_actuals.get(d.get('project_id'), {})
        d['project_actual_inr'] = pact.get('actual', 0)
        d['project_budget_inr'] = pact.get('budget', 0)
        products.append(d)

    summary = {
        'total_products': len(products),
        'total_expected_cost': total_expected_cost,
        'total_expected_revenue': total_expected_rev,
        'total_expected_margin': total_expected_rev - total_expected_cost,
    }

    return render_template('finance_products.html', user=user, products=products,
                           summary=summary, fy_year=fy_year,
                           fx_rates=get_fx_rates_inr())


# ─── Salary master list ───

@app.route('/finance/salaries')
@admin_required
def finance_salaries():
    user = get_user()
    conn = get_db()
    try:
        salaries = conn.execute('''
            SELECT s.id, s.employee_id, s.name, s.department, s.project_id,
                   s.monthly_cost, s.currency, s.is_active, s.notes,
                   p.name AS project_name,
                   e.name AS emp_name
            FROM salary_items s
            LEFT JOIN projects p  ON s.project_id = p.id
            LEFT JOIN employees e ON s.employee_id = e.id
            ORDER BY s.is_active DESC, s.department NULLS LAST, s.name
        ''').fetchall()
    except Exception as e:
        logging.error(f"finance_salaries: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        salaries = []
    try:
        projects = conn.execute("SELECT id, name FROM projects WHERE status = 'active' ORDER BY name").fetchall()
    except Exception:
        projects = []
    try:
        unlinked_employees = conn.execute('''
            SELECT e.id, e.name, e.department
            FROM employees e
            WHERE (e.is_active IS NULL OR e.is_active = 1)
              AND NOT EXISTS (SELECT 1 FROM salary_items s WHERE s.employee_id = e.id)
            ORDER BY e.name
        ''').fetchall()
    except Exception:
        unlinked_employees = []
    conn.close()

    total_monthly = sum(float(s['monthly_cost'] or 0) for s in salaries if int(s['is_active'] or 0) == 1)
    return render_template('finance_salaries.html', user=user, salaries=salaries,
                           projects=projects, unlinked_employees=unlinked_employees,
                           total_monthly=total_monthly)


@app.route('/finance/salaries/sync-employees', methods=['POST'])
@admin_required
def finance_salaries_sync():
    """Create a salary_items row for every active employee that doesn't have one yet."""
    user = get_user()
    conn = get_db()
    added = 0
    try:
        to_add = conn.execute('''
            SELECT e.id, e.name, e.department
            FROM employees e
            WHERE (e.is_active IS NULL OR e.is_active = 1)
              AND NOT EXISTS (SELECT 1 FROM salary_items s WHERE s.employee_id = e.id)
        ''').fetchall()
        for emp in to_add:
            conn.execute(
                "INSERT INTO salary_items (employee_id, name, department, monthly_cost, currency, is_active, created_by) "
                "VALUES (?, ?, ?, 0, 'INR', 1, ?)",
                (emp['id'], emp['name'], emp['department'] or '', user['id'])
            )
            added += 1
        conn.commit()
        flash(f'Synced {added} employee{"s" if added != 1 else ""} into the salary list.', 'success')
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        flash(f'Could not sync employees: {e}', 'error')
    finally:
        conn.close()
    return redirect(url_for('finance_salaries'))


@app.route('/finance/salaries/add', methods=['GET', 'POST'])
@admin_required
def finance_salaries_add():
    user = get_user()
    conn = get_db()
    if request.method == 'POST':
        # Employee-linked salary: get name + department from the selected employee
        emp_val = request.form.get('employee_id', '')
        employee_id = int(emp_val) if emp_val and emp_val.isdigit() else None
        if not employee_id:
            flash('Please select an employee.', 'error')
            conn.close()
            return redirect(url_for('finance_salaries_add'))
        emp = conn.execute("SELECT name, department FROM employees WHERE id = ?", (employee_id,)).fetchone()
        if not emp:
            flash('Employee not found.', 'error')
            conn.close()
            return redirect(url_for('finance_salaries_add'))
        name = emp['name']
        department = emp['department'] or ''
        project_val = request.form.get('project_id', '')
        project_id = int(project_val) if project_val and project_val.isdigit() else None
        try:
            monthly_cost = float(request.form.get('monthly_cost', 0) or 0)
        except ValueError:
            monthly_cost = 0
        currency = request.form.get('currency', 'INR') or 'INR'
        notes = request.form.get('notes', '').strip()
        # Check duplicate
        existing = conn.execute("SELECT id FROM salary_items WHERE employee_id = ?", (employee_id,)).fetchone()
        if existing:
            flash(f'{name} already has a salary entry.', 'error')
            conn.close()
            return redirect(url_for('finance_salaries_add'))
        try:
            conn.execute(
                "INSERT INTO salary_items (employee_id, name, department, project_id, monthly_cost, currency, is_active, notes, created_by) "
                "VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?)",
                (employee_id, name, department, project_id, monthly_cost, currency, notes, user['id'])
            )
            conn.commit()
            flash(f'Salary for {name} added.', 'success')
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            flash(f'Could not add: {e}', 'error')
        conn.close()
        return redirect(url_for('finance_salaries'))
    # GET: employees without an active salary item yet
    available_employees = []
    query_error = False
    try:
        all_emps = conn.execute(
            "SELECT id, name, department FROM employees WHERE is_active = 1 AND emp_code != 'admin' ORDER BY name"
        ).fetchall()
        logging.info(f"finance_salaries_add: {len(all_emps)} active employees found")
    except Exception as e:
        logging.error(f"finance_salaries_add employees query: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        all_emps = []
        query_error = True

    if all_emps and not query_error:
        existing_ids = set()
        try:
            rows = conn.execute("SELECT employee_id FROM salary_items WHERE employee_id IS NOT NULL AND is_active = 1").fetchall()
            existing_ids = set(r['employee_id'] for r in rows)
            logging.info(f"finance_salaries_add: {len(existing_ids)} employees already have active salary items")
        except Exception as e2:
            logging.error(f"salary_items employee_id query: {e2}")
            try:
                conn.rollback()
            except Exception:
                pass
            # If salary_items query fails, show all employees (don't block)
            existing_ids = set()
        available_employees = [e for e in all_emps if e['id'] not in existing_ids]
        logging.info(f"finance_salaries_add: {len(available_employees)} employees available to add")
    try:
        projects = conn.execute("SELECT id, name FROM projects WHERE status = 'active' ORDER BY name").fetchall()
    except Exception:
        projects = []
    conn.close()
    return render_template('finance_salary_form.html', user=user, item=None, projects=projects,
                           available_employees=available_employees, query_error=query_error, mode='add')


@app.route('/finance/salaries/<int:item_id>/edit', methods=['GET', 'POST'])
@admin_required
def finance_salaries_edit(item_id):
    user = get_user()
    conn = get_db()
    item = conn.execute('SELECT * FROM salary_items WHERE id = ?', (item_id,)).fetchone()
    if not item:
        conn.close()
        flash('Not found', 'error')
        return redirect(url_for('finance_salaries'))
    if request.method == 'POST':
        # Name and department are NOT editable here — they're sourced from the
        # linked employee record. Preserve existing values, and if an employee
        # is linked, refresh from the employees table so stale entries self-heal.
        name = item['name']
        department = item['department'] or ''
        if item['employee_id']:
            try:
                _emp = conn.execute("SELECT name, department FROM employees WHERE id = ?", (item['employee_id'],)).fetchone()
                if _emp:
                    name = _emp['name']
                    department = _emp['department'] or ''
            except Exception:
                pass
        project_val = request.form.get('project_id', '')
        project_id = int(project_val) if project_val and project_val.isdigit() else None
        try:
            monthly_cost = float(request.form.get('monthly_cost', 0) or 0)
        except ValueError:
            monthly_cost = 0
        currency = request.form.get('currency', 'INR') or 'INR'
        notes = request.form.get('notes', '').strip()
        is_active = 1 if request.form.get('is_active') else 0
        try:
            conn.execute(
                "UPDATE salary_items SET name=?, department=?, project_id=?, monthly_cost=?, currency=?, notes=?, is_active=? WHERE id=?",
                (name, department, project_id, monthly_cost, currency, notes, is_active, item_id)
            )
            conn.commit()
            flash('Updated', 'success')
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            flash(f'Could not update: {e}', 'error')
        conn.close()
        return redirect(url_for('finance_salaries'))
    try:
        projects = conn.execute("SELECT id, name FROM projects WHERE status = 'active' ORDER BY name").fetchall()
    except Exception:
        projects = []
    conn.close()
    return render_template('finance_salary_form.html', user=user, item=item, projects=projects, mode='edit')


@app.route('/finance/salaries/<int:item_id>/delete', methods=['POST'])
@admin_required
def finance_salaries_delete(item_id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM budget_entries WHERE salary_id = ?", (item_id,))
        conn.execute('DELETE FROM salary_items WHERE id = ?', (item_id,))
        conn.commit()
        flash('Salary item deleted', 'success')
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        flash(f'Could not delete: {e}', 'error')
    finally:
        conn.close()
    return redirect(url_for('finance_salaries'))


# ─── Subscription master list ───

@app.route('/finance/subscriptions')
@admin_required
def finance_subscriptions():
    user = get_user()
    conn = get_db()
    try:
        subs = conn.execute('''
            SELECT s.id, s.name, s.vendor, s.cost, s.currency, s.frequency,
                   s.primary_department, s.shared_departments, s.project_id,
                   s.is_active, s.notes, p.name AS project_name
            FROM subscription_items s
            LEFT JOIN projects p ON s.project_id = p.id
            ORDER BY s.is_active DESC, s.primary_department NULLS LAST, s.name
        ''').fetchall()
    except Exception as e:
        logging.error(f"finance_subscriptions: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        subs = []
    conn.close()

    # Convert to monthly-equivalent for summary
    def monthly_equiv(cost, freq):
        c = float(cost or 0)
        f = (freq or 'monthly').lower()
        if f == 'monthly':    return c
        if f == 'quarterly':  return c / 3.0
        if f == 'annual':     return c / 12.0
        if f == 'one_time':   return 0.0
        return c
    total_monthly = sum(monthly_equiv(s['cost'], s['frequency']) for s in subs if int(s['is_active'] or 0) == 1)
    return render_template('finance_subscriptions.html', user=user, subs=subs, total_monthly=total_monthly)


@app.route('/finance/subscriptions/add', methods=['GET', 'POST'])
@admin_required
def finance_subscriptions_add():
    user = get_user()
    conn = get_db()
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        vendor = request.form.get('vendor', '').strip()
        try:
            cost = float(request.form.get('cost', 0) or 0)
        except ValueError:
            cost = 0
        currency = request.form.get('currency', 'INR') or 'INR'
        frequency = request.form.get('frequency', 'monthly') or 'monthly'
        primary_department = request.form.get('primary_department', '').strip()
        shared_list = request.form.getlist('shared_departments')
        shared_departments = ', '.join(d.strip() for d in shared_list if d.strip())
        project_val = request.form.get('project_id', '')
        project_id = int(project_val) if project_val and project_val.isdigit() else None
        notes = request.form.get('notes', '').strip()
        if not name:
            flash('Name is required', 'error')
            conn.close()
            return redirect(url_for('finance_subscriptions_add'))
        try:
            conn.execute(
                "INSERT INTO subscription_items (name, vendor, cost, currency, frequency, primary_department, shared_departments, project_id, is_active, notes, created_by) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)",
                (name, vendor, cost, currency, frequency, primary_department, shared_departments, project_id, notes, user['id'])
            )
            conn.commit()
            flash('Subscription added', 'success')
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            flash(f'Could not add: {e}', 'error')
        conn.close()
        return redirect(url_for('finance_subscriptions'))
    try:
        projects = conn.execute("SELECT id, name FROM projects WHERE status = 'active' ORDER BY name").fetchall()
    except Exception:
        projects = []
    try:
        dept_cats = conn.execute("SELECT department FROM budget_categories WHERE cat_type = 'department' AND is_active = 1 ORDER BY sort_order, department").fetchall()
        departments = sorted(set(d['department'] for d in dept_cats if d['department']))
    except Exception:
        departments = []
    conn.close()
    return render_template('finance_subscription_form.html', user=user, item=None, projects=projects, departments=departments, mode='add')


@app.route('/finance/subscriptions/<int:item_id>/edit', methods=['GET', 'POST'])
@admin_required
def finance_subscriptions_edit(item_id):
    user = get_user()
    conn = get_db()
    item = conn.execute('SELECT * FROM subscription_items WHERE id = ?', (item_id,)).fetchone()
    if not item:
        conn.close()
        flash('Not found', 'error')
        return redirect(url_for('finance_subscriptions'))
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        vendor = request.form.get('vendor', '').strip()
        try:
            cost = float(request.form.get('cost', 0) or 0)
        except ValueError:
            cost = 0
        currency = request.form.get('currency', 'INR') or 'INR'
        frequency = request.form.get('frequency', 'monthly') or 'monthly'
        primary_department = request.form.get('primary_department', '').strip()
        # Multi-select shared departments: join selected values with comma
        shared_list = request.form.getlist('shared_departments')
        shared_departments = ', '.join(d.strip() for d in shared_list if d.strip())
        project_val = request.form.get('project_id', '')
        project_id = int(project_val) if project_val and project_val.isdigit() else None
        notes = request.form.get('notes', '').strip()
        is_active = 1 if request.form.get('is_active') else 0
        try:
            conn.execute(
                "UPDATE subscription_items SET name=?, vendor=?, cost=?, currency=?, frequency=?, primary_department=?, shared_departments=?, project_id=?, notes=?, is_active=? WHERE id=?",
                (name, vendor, cost, currency, frequency, primary_department, shared_departments, project_id, notes, is_active, item_id)
            )
            conn.commit()
            flash('Updated', 'success')
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            flash(f'Could not update: {e}', 'error')
        conn.close()
        return redirect(url_for('finance_subscriptions'))
    try:
        projects = conn.execute("SELECT id, name FROM projects WHERE status = 'active' ORDER BY name").fetchall()
    except Exception:
        projects = []
    try:
        dept_cats = conn.execute("SELECT department FROM budget_categories WHERE cat_type = 'department' AND is_active = 1 ORDER BY sort_order, department").fetchall()
        departments = sorted(set(d['department'] for d in dept_cats if d['department']))
    except Exception:
        departments = []
    conn.close()
    return render_template('finance_subscription_form.html', user=user, item=item, projects=projects, departments=departments, mode='edit')


@app.route('/finance/subscriptions/<int:item_id>/delete', methods=['POST'])
@admin_required
def finance_subscriptions_delete(item_id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM budget_entries WHERE subscription_id = ?", (item_id,))
        conn.execute('DELETE FROM subscription_items WHERE id = ?', (item_id,))
        conn.commit()
        flash('Subscription deleted', 'success')
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        flash(f'Could not delete: {e}', 'error')
    finally:
        conn.close()
    return redirect(url_for('finance_subscriptions'))


# ─────────────────────────────────────────────────────────
#  OPERATIONS – PLAB Pathway Client Management
# ─────────────────────────────────────────────────────────

def ensure_ops_tables():
    """Create Operations / PLAB client tables."""
    try:
        conn = get_db()
        conn.execute('''CREATE TABLE IF NOT EXISTS plab_clients (
            id SERIAL PRIMARY KEY,
            customer_id TEXT,
            registration_number TEXT UNIQUE,
            registration_date TEXT,
            prefix TEXT DEFAULT 'Dr.',
            first_name TEXT NOT NULL,
            last_name TEXT,
            mobile TEXT,
            whatsapp1 TEXT,
            whatsapp2 TEXT,
            email TEXT,
            dob TEXT,
            city TEXT,
            state TEXT,
            instagram TEXT,
            facebook TEXT,
            linkedin TEXT,
            photo_path TEXT,
            father_name TEXT,
            father_phone TEXT,
            mother_name TEXT,
            mother_phone TEXT,
            parents_email TEXT,
            joined_stage TEXT,
            plan_type TEXT,
            english_training TEXT,
            account_status TEXT DEFAULT 'In Process',
            current_stage TEXT,
            dropped_date TEXT,
            upgraded_to TEXT,
            counsellor TEXT,
            counsellor_email TEXT,
            counsellor_number TEXT,
            lead_source TEXT,
            referral_type TEXT,
            operations_referral TEXT,
            package_amount NUMERIC(14,2) DEFAULT 0,
            discount_allowed NUMERIC(14,2) DEFAULT 0,
            additional_package_notes TEXT,
            final_package NUMERIC(14,2) DEFAULT 0,
            inst1_amount NUMERIC(14,2) DEFAULT 0,
            inst1_date TEXT,
            inst1_note TEXT,
            inst2_amount NUMERIC(14,2) DEFAULT 0,
            inst2_date TEXT,
            inst2_note TEXT,
            inst3_amount NUMERIC(14,2) DEFAULT 0,
            inst3_date TEXT,
            inst3_note TEXT,
            inst4_amount NUMERIC(14,2) DEFAULT 0,
            inst4_date TEXT,
            inst4_note TEXT,
            total_paid NUMERIC(14,2) DEFAULT 0,
            welcome_mail TEXT,
            welcome_call_by TEXT,
            welcome_call_date TEXT,
            english_book TEXT,
            english_book_date TEXT,
            oxford_book TEXT,
            oxford_book_date TEXT,
            plab_brochure INTEGER DEFAULT 0,
            ceo_letter INTEGER DEFAULT 0,
            refund_policy INTEGER DEFAULT 0,
            service_agreement INTEGER DEFAULT 0,
            goodie_pen INTEGER DEFAULT 0,
            goodie_diary INTEGER DEFAULT 0,
            goodie_laptop_bag INTEGER DEFAULT 0,
            goodie_stickers INTEGER DEFAULT 0,
            contract_path TEXT,
            additional_notes TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        # ── Coaching / Training ──
        conn.execute('''CREATE TABLE IF NOT EXISTS ops_coaching (
            id SERIAL PRIMARY KEY,
            registration_number TEXT REFERENCES plab_clients(registration_number),
            course_type TEXT,
            coaching_method TEXT,
            coaching_status TEXT,
            batch_month TEXT,
            batch_year TEXT,
            start_date TEXT,
            end_date TEXT,
            english_training TEXT,
            blueprint_stage TEXT,
            attendance TEXT,
            ielts_vendor TEXT,
            oet_vendor TEXT,
            other_vendor TEXT,
            plab1_partner TEXT,
            plab2_vendor TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        # ── English Login Details ──
        conn.execute('''CREATE TABLE IF NOT EXISTS ops_english_logins (
            id SERIAL PRIMARY KEY,
            registration_number TEXT REFERENCES plab_clients(registration_number),
            ielts_login_id TEXT,
            ielts_password TEXT,
            ielts_hint_question TEXT,
            ielts_security_answer TEXT,
            oet_login_id TEXT,
            oet_password TEXT,
            oet_hint_question TEXT,
            oet_security_answer TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        # ── Test Bookings ──
        conn.execute('''CREATE TABLE IF NOT EXISTS ops_test_bookings (
            id SERIAL PRIMARY KEY,
            registration_number TEXT REFERENCES plab_clients(registration_number),
            exam TEXT,
            exam_type TEXT,
            booking_date TEXT,
            exam_date TEXT,
            exam_status TEXT,
            exam_result TEXT,
            exam_result_date TEXT,
            score TEXT,
            test_center TEXT,
            city_state TEXT,
            country TEXT,
            booked_by TEXT,
            revaluation TEXT,
            reval_applied_date TEXT,
            reval_result TEXT,
            reval_score TEXT,
            notes TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        # ── Call Notes ──
        conn.execute('''CREATE TABLE IF NOT EXISTS ops_call_notes (
            id SERIAL PRIMARY KEY,
            registration_number TEXT REFERENCES plab_clients(registration_number),
            call_date TEXT,
            call_note TEXT,
            added_by TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        # ── Payments ──
        conn.execute('''CREATE TABLE IF NOT EXISTS ops_payments (
            id SERIAL PRIMARY KEY,
            registration_number TEXT REFERENCES plab_clients(registration_number),
            payment_date TEXT,
            amount_paid NUMERIC(14,2) DEFAULT 0,
            gst_paid NUMERIC(14,2) DEFAULT 0,
            total_amount_paid NUMERIC(14,2) DEFAULT 0,
            instalment TEXT,
            payment_method TEXT,
            total_package NUMERIC(14,2) DEFAULT 0,
            notes TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        conn.commit()
        conn.close()
    except Exception as e:
        logging.error(f"ensure_ops_tables: {e}")


# ── Stages & statuses for dropdowns ──
PLAB_STAGES = ['English Stage', 'PLAB 1 Stage', 'PLAB 2 Stage', 'GMC Stage', 'Job Hunt Stage', 'Completed']
ACCOUNT_STATUSES = ['In Process', 'Switched Program', 'Dropped and Refunded', 'Dropped Out', 'On Hold', 'Completed']
PLAN_TYPES = [
    'Full Spon', 'Integrated Consulting',
    '2022 UK - Full Sponsorship', '2022 UK - IC',
    '2023 UK - PGCP', '2024 UK - PGCP', '2025 UK - PGCP',
    '2023 UK - PGCP - Dual', '2024 UK - PGCP - Dual', '2025 UK - PGCP - Dual',
]
JOINED_STAGES = ['English Stage', 'PLAB 1 Stage', 'PLAB 2 Stage', 'GMC Stage']
ENGLISH_TRAINING_OPTIONS = ['IELTS', 'OET']
LEAD_SOURCES = ['Social Media', 'Website', 'Referral', 'Walk-in', 'Event', 'Other']

# ── Coaching / Training dropdowns ──
COACHING_COURSE_TYPES = ['Full Course', 'Crash Course']
COACHING_STATUSES = ['On Going', 'Completed']
COACHING_METHODS = ['Online', 'Offline', 'Hybrid']

# ── Test Booking dropdowns ──
EXAM_NAMES = ['OET', 'IELTS', 'PLAB 1', 'PLAB 2', 'MRCP', 'MRCS', 'MRCEM']
EXAM_STATUSES = ['Booked', 'Attended', 'Cancelled by Client', 'Cancelled by Authority', 'Rescheduled', 'Missed']
EXAM_RESULTS = ['Passed', 'Failed']

# ── Payment dropdowns ──
PAYMENT_METHODS = ['Bank Transfer', 'Cash Deposit', 'Discount', 'Shifted from Portfolio', 'Online Payment', 'Cheque']
INSTALMENT_OPTIONS = ['1st Instalment', '2nd Instalment', '3rd Instalment', '4th Instalment', '5th Instalment']


def _next_registration_number(conn):
    """Generate next GCUKIP/YY-YY/NNN registration number."""
    now = datetime.now()
    if now.month >= 4:
        y1, y2 = now.year % 100, (now.year + 1) % 100
    else:
        y1, y2 = (now.year - 1) % 100, now.year % 100
    prefix = f"GCUKIP/{y1:02d}-{y2:02d}/"
    row = conn.execute(
        "SELECT registration_number FROM plab_clients WHERE registration_number LIKE ? ORDER BY registration_number DESC LIMIT 1",
        (prefix + '%',)
    ).fetchone()
    if row:
        try:
            last_num = int(row['registration_number'].split('/')[-1])
        except (ValueError, IndexError):
            last_num = 0
    else:
        last_num = 0
    return f"{prefix}{last_num + 1:03d}"


@app.route('/operations/plab')
@admin_required
def ops_plab_list():
    """PLAB clients list with search and filters."""
    conn = get_db()
    try:
        search = (request.args.get('search', '') or request.args.get('q', '')).strip()
        status_filter = request.args.get('status_filter', '') or request.args.get('status', '')
        stage_filter = request.args.get('stage_filter', '') or request.args.get('stage', '')

        sql = "SELECT * FROM plab_clients WHERE 1=1"
        params = []
        if search:
            sql += """ AND (
                LOWER(first_name || ' ' || COALESCE(last_name,'')) LIKE LOWER(?)
                OR LOWER(COALESCE(prefix,'') || ' ' || first_name || ' ' || COALESCE(last_name,'')) LIKE LOWER(?)
                OR LOWER(registration_number) LIKE LOWER(?)
                OR LOWER(COALESCE(email,'')) LIKE LOWER(?)
                OR LOWER(COALESCE(mobile,'')) LIKE LOWER(?)
            )"""
            like = f"%{search}%"
            params += [like, like, like, like, like]
        if status_filter:
            sql += " AND account_status = ?"
            params.append(status_filter)
        if stage_filter:
            sql += " AND current_stage = ?"
            params.append(stage_filter)
        sql += " ORDER BY id DESC"
        clients_raw = conn.execute(sql, tuple(params)).fetchall()

        # Get actual payment totals per client from ops_payments
        payment_totals = {}
        pay_rows = conn.execute("""SELECT registration_number,
            COALESCE(SUM(amount_paid), 0) as paid,
            COALESCE(SUM(gst_paid), 0) as gst
            FROM ops_payments GROUP BY registration_number""").fetchall()
        for pr in pay_rows:
            payment_totals[pr['registration_number']] = {
                'paid': float(pr['paid'] or 0),
                'gst': float(pr['gst'] or 0)
            }

        # Attach computed paid/balance to each client
        clients = []
        for c in clients_raw:
            cd = dict(c)
            reg = cd['registration_number']
            pt = payment_totals.get(reg, {'paid': 0, 'gst': 0})
            pkg = float(cd.get('package_amount') or 0)
            disc = float(cd.get('discount_allowed') or 0)
            cd['computed_package'] = pkg - disc
            cd['computed_paid'] = pt['paid']
            cd['computed_gst'] = pt['gst']
            cd['computed_balance'] = cd['computed_package'] - pt['paid']
            clients.append(cd)

        # Summary stats
        total = len(clients)
        active_count = sum(1 for c in clients if c['account_status'] == 'In Process')
        total_package = sum(float(c['computed_package'] or 0) for c in clients)
        total_collected = sum(float(c['computed_paid'] or 0) for c in clients)
    except Exception as e:
        logging.error(f"ops_plab_list error: {e}")
        conn.close()
        flash(f'Error loading clients: {e}', 'error')
        clients, search, status_filter, stage_filter = [], '', '', ''
        total = active_count = 0
        total_package = total_collected = 0.0
        return render_template('ops_plab_list.html',
                               clients=clients, search=search, status_filter=status_filter,
                               stage_filter=stage_filter, total=total, active_count=active_count,
                               total_package=total_package, total_collected=total_collected,
                               account_statuses=ACCOUNT_STATUSES, plab_stages=PLAB_STAGES)

    conn.close()
    return render_template('ops_plab_list.html',
                           clients=clients, search=search, status_filter=status_filter,
                           stage_filter=stage_filter, total=total, active_count=active_count,
                           total_package=total_package, total_collected=total_collected,
                           account_statuses=ACCOUNT_STATUSES, plab_stages=PLAB_STAGES)


@app.route('/operations/plab/<int:client_id>')
@admin_required
def ops_plab_dashboard(client_id):
    """Individual client dashboard with all details and linked sections."""
    conn = get_db()
    client = conn.execute("SELECT * FROM plab_clients WHERE id = ?", (client_id,)).fetchone()
    if not client:
        conn.close()
        flash('Client not found', 'error')
        return redirect(url_for('ops_plab_list'))
    reg = client['registration_number']
    # Compute payment info from actual payments table
    payments_total = conn.execute("""SELECT
        COALESCE(SUM(amount_paid), 0) as amount_paid,
        COALESCE(SUM(gst_paid), 0) as gst_paid,
        COALESCE(SUM(total_amount_paid), 0) as total_paid
        FROM ops_payments WHERE registration_number = ?""", (reg,)).fetchone()
    amount_paid = float(payments_total['amount_paid'] or 0)
    gst_paid = float(payments_total['gst_paid'] or 0)
    total_paid = float(payments_total['total_paid'] or 0)
    # Package and discount from registration page
    pkg = float(client['package_amount'] or 0)
    discount = float(client['discount_allowed'] or 0)
    final_pkg = pkg - discount
    balance = final_pkg - amount_paid
    payment_pct = (amount_paid / final_pkg * 100) if final_pkg > 0 else 0
    # Linked sections
    coaching = conn.execute("SELECT * FROM ops_coaching WHERE registration_number = ? ORDER BY created_at DESC", (reg,)).fetchall()
    english_logins = conn.execute("SELECT * FROM ops_english_logins WHERE registration_number = ? ORDER BY created_at DESC", (reg,)).fetchall()
    test_bookings = conn.execute("SELECT * FROM ops_test_bookings WHERE registration_number = ? ORDER BY exam_date DESC NULLS LAST", (reg,)).fetchall()
    call_notes = conn.execute("SELECT * FROM ops_call_notes WHERE registration_number = ? ORDER BY call_date DESC NULLS LAST LIMIT 20", (reg,)).fetchall()
    call_notes_count = conn.execute("SELECT COUNT(*) as cnt FROM ops_call_notes WHERE registration_number = ?", (reg,)).fetchone()['cnt']
    payments = conn.execute("SELECT * FROM ops_payments WHERE registration_number = ? ORDER BY payment_date DESC NULLS LAST", (reg,)).fetchall()
    conn.close()
    return render_template('ops_plab_dashboard.html', client=client,
                           amount_paid=amount_paid, gst_paid=gst_paid,
                           total_paid=total_paid, balance=balance, payment_pct=payment_pct,
                           final_pkg=final_pkg, discount=discount,
                           plab_stages=PLAB_STAGES, account_statuses=ACCOUNT_STATUSES,
                           coaching=coaching, english_logins=english_logins,
                           test_bookings=test_bookings, call_notes=call_notes,
                           call_notes_count=call_notes_count, payments=payments)


@app.route('/operations/plab/add', methods=['GET', 'POST'])
@admin_required
def ops_plab_add():
    """Add new PLAB client."""
    conn = get_db()
    if request.method == 'POST':
        try:
            reg_num = _next_registration_number(conn)
            f = request.form
            pkg = float(f.get('package_amount') or 0)
            disc = float(f.get('discount_allowed') or 0)
            final_pkg = pkg - disc
            inst_total = sum(float(f.get(f'inst{i}_amount') or 0) for i in range(1, 5))

            conn.execute('''INSERT INTO plab_clients (
                registration_number, registration_date, customer_id,
                prefix, first_name, last_name,
                mobile, whatsapp1, whatsapp2, email, dob, city, state,
                instagram, facebook, linkedin,
                father_name, father_phone, mother_name, mother_phone, parents_email,
                joined_stage, plan_type, english_training,
                account_status, current_stage,
                counsellor, counsellor_email, counsellor_number,
                lead_source, referral_type, operations_referral,
                package_amount, discount_allowed, additional_package_notes,
                final_package, total_paid,
                inst1_amount, inst1_date, inst1_note,
                inst2_amount, inst2_date, inst2_note,
                inst3_amount, inst3_date, inst3_note,
                inst4_amount, inst4_date, inst4_note,
                additional_notes, created_by
            ) VALUES (
                ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?
            )''', (
                reg_num, f.get('registration_date') or datetime.now().strftime('%Y-%m-%d'),
                f.get('customer_id') or '',
                f.get('prefix', 'Dr.'), f.get('first_name'), f.get('last_name', ''),
                f.get('mobile', ''), f.get('whatsapp1', ''), f.get('whatsapp2', ''),
                f.get('email', ''), f.get('dob', ''), f.get('city', ''), f.get('state', ''),
                f.get('instagram', ''), f.get('facebook', ''), f.get('linkedin', ''),
                f.get('father_name', ''), f.get('father_phone', ''),
                f.get('mother_name', ''), f.get('mother_phone', ''), f.get('parents_email', ''),
                f.get('joined_stage', ''), f.get('plan_type', ''), f.get('english_training', ''),
                f.get('account_status', 'In Process'), f.get('joined_stage', ''),
                f.get('counsellor', ''), f.get('counsellor_email', ''), f.get('counsellor_number', ''),
                f.get('lead_source', ''), f.get('referral_type', ''), f.get('operations_referral', ''),
                pkg, disc, f.get('additional_package_notes', ''),
                final_pkg, inst_total,
                float(f.get('inst1_amount') or 0), f.get('inst1_date', ''), f.get('inst1_note', ''),
                float(f.get('inst2_amount') or 0), f.get('inst2_date', ''), f.get('inst2_note', ''),
                float(f.get('inst3_amount') or 0), f.get('inst3_date', ''), f.get('inst3_note', ''),
                float(f.get('inst4_amount') or 0), f.get('inst4_date', ''), f.get('inst4_note', ''),
                f.get('additional_notes', ''), session['user_id']
            ))
            conn.commit()
            flash(f'Client {reg_num} added successfully', 'success')
            conn.close()
            return redirect(url_for('ops_plab_list'))
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            logging.error(f"ops_plab_add: {e}")
            flash(f'Error adding client: {e}', 'error')

    conn.close()
    return render_template('ops_plab_form.html', mode='add', item=None,
                           plan_types=PLAN_TYPES, joined_stages=JOINED_STAGES,
                           english_options=ENGLISH_TRAINING_OPTIONS,
                           account_statuses=ACCOUNT_STATUSES, plab_stages=PLAB_STAGES,
                           lead_sources=LEAD_SOURCES)


@app.route('/operations/plab/<int:client_id>/edit', methods=['GET', 'POST'])
@admin_required
def ops_plab_edit(client_id):
    """Edit PLAB client."""
    conn = get_db()
    if request.method == 'POST':
        try:
            f = request.form
            pkg = float(f.get('package_amount') or 0)
            disc = float(f.get('discount_allowed') or 0)
            final_pkg = pkg - disc
            inst_total = sum(float(f.get(f'inst{i}_amount') or 0) for i in range(1, 5))

            conn.execute('''UPDATE plab_clients SET
                registration_date=?, customer_id=?,
                prefix=?, first_name=?, last_name=?,
                mobile=?, whatsapp1=?, whatsapp2=?, email=?, dob=?, city=?, state=?,
                instagram=?, facebook=?, linkedin=?,
                father_name=?, father_phone=?, mother_name=?, mother_phone=?, parents_email=?,
                joined_stage=?, plan_type=?, english_training=?,
                account_status=?, current_stage=?,
                dropped_date=?, upgraded_to=?,
                counsellor=?, counsellor_email=?, counsellor_number=?,
                lead_source=?, referral_type=?, operations_referral=?,
                package_amount=?, discount_allowed=?, additional_package_notes=?,
                final_package=?, total_paid=?,
                inst1_amount=?, inst1_date=?, inst1_note=?,
                inst2_amount=?, inst2_date=?, inst2_note=?,
                inst3_amount=?, inst3_date=?, inst3_note=?,
                inst4_amount=?, inst4_date=?, inst4_note=?,
                welcome_mail=?, welcome_call_by=?, welcome_call_date=?,
                english_book=?, english_book_date=?,
                oxford_book=?, oxford_book_date=?,
                plab_brochure=?, ceo_letter=?, refund_policy=?, service_agreement=?,
                goodie_pen=?, goodie_diary=?, goodie_laptop_bag=?, goodie_stickers=?,
                additional_notes=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            ''', (
                f.get('registration_date', ''), f.get('customer_id', ''),
                f.get('prefix', 'Dr.'), f.get('first_name'), f.get('last_name', ''),
                f.get('mobile', ''), f.get('whatsapp1', ''), f.get('whatsapp2', ''),
                f.get('email', ''), f.get('dob', ''), f.get('city', ''), f.get('state', ''),
                f.get('instagram', ''), f.get('facebook', ''), f.get('linkedin', ''),
                f.get('father_name', ''), f.get('father_phone', ''),
                f.get('mother_name', ''), f.get('mother_phone', ''), f.get('parents_email', ''),
                f.get('joined_stage', ''), f.get('plan_type', ''), f.get('english_training', ''),
                f.get('account_status', 'In Process'), f.get('current_stage', ''),
                f.get('dropped_date', ''), f.get('upgraded_to', ''),
                f.get('counsellor', ''), f.get('counsellor_email', ''), f.get('counsellor_number', ''),
                f.get('lead_source', ''), f.get('referral_type', ''), f.get('operations_referral', ''),
                pkg, disc, f.get('additional_package_notes', ''),
                final_pkg, inst_total,
                float(f.get('inst1_amount') or 0), f.get('inst1_date', ''), f.get('inst1_note', ''),
                float(f.get('inst2_amount') or 0), f.get('inst2_date', ''), f.get('inst2_note', ''),
                float(f.get('inst3_amount') or 0), f.get('inst3_date', ''), f.get('inst3_note', ''),
                float(f.get('inst4_amount') or 0), f.get('inst4_date', ''), f.get('inst4_note', ''),
                f.get('welcome_mail', ''), f.get('welcome_call_by', ''), f.get('welcome_call_date', ''),
                f.get('english_book', ''), f.get('english_book_date', ''),
                f.get('oxford_book', ''), f.get('oxford_book_date', ''),
                1 if f.get('plab_brochure') else 0,
                1 if f.get('ceo_letter') else 0,
                1 if f.get('refund_policy') else 0,
                1 if f.get('service_agreement') else 0,
                1 if f.get('goodie_pen') else 0,
                1 if f.get('goodie_diary') else 0,
                1 if f.get('goodie_laptop_bag') else 0,
                1 if f.get('goodie_stickers') else 0,
                f.get('additional_notes', ''),
                client_id
            ))
            conn.commit()
            flash('Client updated', 'success')
            conn.close()
            return redirect(url_for('ops_plab_dashboard', client_id=client_id))
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            logging.error(f"ops_plab_edit: {e}")
            flash(f'Error: {e}', 'error')

    client = conn.execute("SELECT * FROM plab_clients WHERE id = ?", (client_id,)).fetchone()
    conn.close()
    if not client:
        flash('Client not found', 'error')
        return redirect(url_for('ops_plab_list'))
    return render_template('ops_plab_form.html', mode='edit', item=client,
                           plan_types=PLAN_TYPES, joined_stages=JOINED_STAGES,
                           english_options=ENGLISH_TRAINING_OPTIONS,
                           account_statuses=ACCOUNT_STATUSES, plab_stages=PLAB_STAGES,
                           lead_sources=LEAD_SOURCES)


@app.route('/operations/plab/<int:client_id>/delete', methods=['POST'])
@admin_required
def ops_plab_delete(client_id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM plab_clients WHERE id = ?", (client_id,))
        conn.commit()
        flash('Client deleted', 'success')
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        flash(f'Error: {e}', 'error')
    finally:
        conn.close()
    return redirect(url_for('ops_plab_list'))


@app.route('/operations/plab/import', methods=['GET', 'POST'])
@admin_required
def ops_plab_import():
    """Import PLAB clients from CSV (combined Zoho export)."""
    if request.method == 'POST':
        import csv, io
        file = request.files.get('csv_file')
        if not file or not file.filename.endswith('.csv'):
            flash('Please upload a CSV file', 'error')
            return redirect(request.url)
        try:
            stream = io.StringIO(file.stream.read().decode('utf-8-sig'))
            reader = csv.DictReader(stream)
            conn = get_db()
            imported = 0
            skipped = 0
            # Status normalization map (fix Zoho typos)
            status_fix = {'Oh Hold': 'On Hold'}
            for row in reader:
                reg_num = row.get('registration_number', '').strip()
                first_name = row.get('first_name', '').strip()
                if not first_name and not reg_num:
                    skipped += 1
                    continue
                # Check duplicate by registration number
                if reg_num:
                    exists = conn.execute("SELECT id FROM plab_clients WHERE registration_number = ?", (reg_num,)).fetchone()
                    if exists:
                        skipped += 1
                        continue

                pkg = _safe_float(row.get('package_amount', '0'))
                disc = _safe_float(row.get('discount_allowed', '0'))
                final = _safe_float(row.get('final_package', '0')) or (pkg - disc)
                i1 = _safe_float(row.get('inst1_amount', '0'))
                i2 = _safe_float(row.get('inst2_amount', '0'))
                i3 = _safe_float(row.get('inst3_amount', '0'))
                i4 = _safe_float(row.get('inst4_amount', '0'))
                total_paid = i1 + i2 + i3 + i4
                raw_status = row.get('account_status', 'In Process').strip()
                account_status = status_fix.get(raw_status, raw_status)

                conn.execute('''INSERT INTO plab_clients (
                    registration_number, registration_date, customer_id,
                    prefix, first_name, last_name,
                    mobile, whatsapp1, whatsapp2, email, dob, city, state,
                    instagram, facebook, linkedin,
                    father_name, father_phone, mother_name, mother_phone, parents_email,
                    joined_stage, plan_type, english_training,
                    account_status, current_stage,
                    counsellor, counsellor_email, counsellor_number,
                    lead_source, referral_type, operations_referral,
                    package_amount, discount_allowed, additional_package_notes,
                    final_package, total_paid,
                    inst1_amount, inst1_date, inst1_note,
                    inst2_amount, inst2_date, inst2_note,
                    inst3_amount, inst3_date, inst3_note,
                    inst4_amount, inst4_date, inst4_note,
                    dropped_date, upgraded_to,
                    welcome_mail, welcome_call_by, welcome_call_date,
                    english_book, english_book_date,
                    oxford_book, oxford_book_date,
                    plab_brochure, ceo_letter, refund_policy, service_agreement,
                    goodie_pen, goodie_diary, goodie_laptop_bag, goodie_stickers,
                    additional_notes, created_by
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                    reg_num or _next_registration_number(conn),
                    row.get('registration_date', ''),
                    row.get('customer_id', ''),
                    row.get('prefix', 'Dr.'), first_name, row.get('last_name', ''),
                    row.get('mobile', ''), row.get('whatsapp1', ''), row.get('whatsapp2', ''),
                    row.get('email', ''), row.get('dob', ''), row.get('city', ''), row.get('state', ''),
                    row.get('instagram', ''), row.get('facebook', ''), row.get('linkedin', ''),
                    row.get('father_name', ''), row.get('father_phone', ''),
                    row.get('mother_name', ''), row.get('mother_phone', ''), row.get('parents_email', ''),
                    row.get('joined_stage', ''), row.get('plan_type', ''), row.get('english_training', ''),
                    account_status, row.get('current_stage', ''),
                    row.get('counsellor', ''), row.get('counsellor_email', ''), row.get('counsellor_number', ''),
                    row.get('lead_source', ''), row.get('referral_type', ''), row.get('operations_referral', ''),
                    pkg, disc, row.get('additional_package_notes', ''),
                    final, total_paid,
                    i1, row.get('inst1_date', ''), row.get('inst1_note', ''),
                    i2, row.get('inst2_date', ''), row.get('inst2_note', ''),
                    i3, row.get('inst3_date', ''), row.get('inst3_note', ''),
                    i4, row.get('inst4_date', ''), row.get('inst4_note', ''),
                    row.get('dropped_date', ''), row.get('upgraded_to', ''),
                    row.get('welcome_mail', ''), row.get('welcome_call_by', ''), row.get('welcome_call_date', ''),
                    row.get('english_book', ''), row.get('english_book_date', ''),
                    row.get('oxford_book', ''), row.get('oxford_book_date', ''),
                    int(row.get('plab_brochure', 0) or 0),
                    int(row.get('ceo_letter', 0) or 0),
                    int(row.get('refund_policy', 0) or 0),
                    int(row.get('service_agreement', 0) or 0),
                    int(row.get('goodie_pen', 0) or 0),
                    int(row.get('goodie_diary', 0) or 0),
                    int(row.get('goodie_laptop_bag', 0) or 0),
                    int(row.get('goodie_stickers', 0) or 0),
                    row.get('additional_notes', ''),
                    session['user_id']
                ))
                imported += 1
            conn.commit()
            conn.close()
            flash(f'Imported {imported} clients, skipped {skipped} (duplicates/empty)', 'success')
            return redirect(url_for('ops_plab_list'))
        except Exception as e:
            logging.error(f"ops_plab_import: {e}")
            flash(f'Import error: {e}', 'error')
    return render_template('ops_plab_import.html')



def _safe_float(val):
    """Parse a float from a string, stripping currency symbols and commas."""
    if not val:
        return 0.0
    try:
        return float(str(val).replace(',', '').replace('₹', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return 0.0


# ─────────────────────────────────────────────────────────
#  OPERATIONS – Coaching / Training
# ─────────────────────────────────────────────────────────

@app.route('/operations/coaching')
@admin_required
def ops_coaching_list():
    """List all coaching/training records, optionally filtered by client."""
    conn = get_db()
    reg = request.args.get('client', '')
    search = request.args.get('q', '')
    status_filter = request.args.get('status', '')
    try:
        sql = '''SELECT c.*, p.first_name, p.last_name, p.prefix
                 FROM ops_coaching c
                 LEFT JOIN plab_clients p ON c.registration_number = p.registration_number
                 WHERE 1=1'''
        params = []
        if reg:
            sql += ' AND c.registration_number = ?'
            params.append(reg)
        if status_filter:
            sql += ' AND c.coaching_status = ?'
            params.append(status_filter)
        if search:
            sql += ' AND (p.first_name ILIKE ? OR p.last_name ILIKE ? OR c.course_type ILIKE ?)'
            params.extend([f'%{search}%'] * 3)
        sql += ' ORDER BY c.created_at DESC'
        records = conn.execute(sql, params).fetchall()
    except Exception as e:
        logging.error(f"ops_coaching_list: {e}")
        records = []
    conn.close()
    return render_template('ops_coaching_list.html', records=records, client_reg=reg,
                           search=search, status_filter=status_filter,
                           coaching_statuses=COACHING_STATUSES)


@app.route('/operations/coaching/add', methods=['GET', 'POST'])
@admin_required
def ops_coaching_add():
    """Add coaching/training record."""
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        conn.execute('''INSERT INTO ops_coaching (
            registration_number, course_type, coaching_method, coaching_status,
            batch_month, batch_year, start_date, end_date, english_training,
            blueprint_stage, attendance, ielts_vendor, oet_vendor, other_vendor,
            plab1_partner, plab2_vendor, created_by
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
            f.get('registration_number'), f.get('course_type'), f.get('coaching_method'),
            f.get('coaching_status'), f.get('batch_month'), f.get('batch_year'),
            f.get('start_date'), f.get('end_date'), f.get('english_training'),
            f.get('blueprint_stage'), f.get('attendance'), f.get('ielts_vendor'),
            f.get('oet_vendor'), f.get('other_vendor'), f.get('plab1_partner'),
            f.get('plab2_vendor'), session.get('user_id', 0)
        ))
        conn.commit()
        conn.close()
        flash('Coaching record added', 'success')
        return redirect(request.args.get('next') or url_for('ops_coaching_list'))
    clients = conn.execute("SELECT registration_number, prefix, first_name, last_name FROM plab_clients ORDER BY first_name").fetchall()
    conn.close()
    pre_reg = request.args.get('client', '')
    return render_template('ops_coaching_form.html', record=None, clients=clients,
                           course_types=COACHING_COURSE_TYPES, coaching_statuses=COACHING_STATUSES,
                           coaching_methods=COACHING_METHODS, pre_reg=pre_reg)


@app.route('/operations/coaching/<int:record_id>/edit', methods=['GET', 'POST'])
@admin_required
def ops_coaching_edit(record_id):
    """Edit coaching/training record."""
    conn = get_db()
    record = conn.execute("SELECT * FROM ops_coaching WHERE id = ?", (record_id,)).fetchone()
    if not record:
        conn.close()
        flash('Record not found', 'error')
        return redirect(url_for('ops_coaching_list'))
    if request.method == 'POST':
        f = request.form
        conn.execute('''UPDATE ops_coaching SET
            registration_number=?, course_type=?, coaching_method=?, coaching_status=?,
            batch_month=?, batch_year=?, start_date=?, end_date=?, english_training=?,
            blueprint_stage=?, attendance=?, ielts_vendor=?, oet_vendor=?, other_vendor=?,
            plab1_partner=?, plab2_vendor=? WHERE id=?''', (
            f.get('registration_number'), f.get('course_type'), f.get('coaching_method'),
            f.get('coaching_status'), f.get('batch_month'), f.get('batch_year'),
            f.get('start_date'), f.get('end_date'), f.get('english_training'),
            f.get('blueprint_stage'), f.get('attendance'), f.get('ielts_vendor'),
            f.get('oet_vendor'), f.get('other_vendor'), f.get('plab1_partner'),
            f.get('plab2_vendor'), record_id
        ))
        conn.commit()
        conn.close()
        flash('Coaching record updated', 'success')
        return redirect(request.args.get('next') or url_for('ops_coaching_list'))
    clients = conn.execute("SELECT registration_number, prefix, first_name, last_name FROM plab_clients ORDER BY first_name").fetchall()
    conn.close()
    return render_template('ops_coaching_form.html', record=record, clients=clients,
                           course_types=COACHING_COURSE_TYPES, coaching_statuses=COACHING_STATUSES,
                           coaching_methods=COACHING_METHODS, pre_reg='')


@app.route('/operations/coaching/<int:record_id>/delete', methods=['POST'])
@admin_required
def ops_coaching_delete(record_id):
    conn = get_db()
    conn.execute("DELETE FROM ops_coaching WHERE id = ?", (record_id,))
    conn.commit()
    conn.close()
    flash('Coaching record deleted', 'success')
    return redirect(request.args.get('next') or url_for('ops_coaching_list'))


# ─────────────────────────────────────────────────────────
#  OPERATIONS – English Login Details
# ─────────────────────────────────────────────────────────

@app.route('/operations/english-logins')
@admin_required
def ops_english_logins_list():
    """List all English login details."""
    conn = get_db()
    search = request.args.get('q', '')
    try:
        sql = '''SELECT e.*, p.first_name, p.last_name, p.prefix
                 FROM ops_english_logins e
                 LEFT JOIN plab_clients p ON e.registration_number = p.registration_number
                 WHERE 1=1'''
        params = []
        if search:
            sql += ' AND (p.first_name ILIKE ? OR p.last_name ILIKE ? OR e.registration_number ILIKE ?)'
            params.extend([f'%{search}%'] * 3)
        sql += ' ORDER BY e.created_at DESC'
        records = conn.execute(sql, params).fetchall()
    except Exception as e:
        logging.error(f"ops_english_logins_list: {e}")
        records = []
    conn.close()
    return render_template('ops_english_logins_list.html', records=records, search=search)


@app.route('/operations/english-logins/add', methods=['GET', 'POST'])
@admin_required
def ops_english_logins_add():
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        conn.execute('''INSERT INTO ops_english_logins (
            registration_number, ielts_login_id, ielts_password, ielts_hint_question,
            ielts_security_answer, oet_login_id, oet_password, oet_hint_question,
            oet_security_answer, created_by
        ) VALUES (?,?,?,?,?,?,?,?,?,?)''', (
            f.get('registration_number'), f.get('ielts_login_id'), f.get('ielts_password'),
            f.get('ielts_hint_question'), f.get('ielts_security_answer'),
            f.get('oet_login_id'), f.get('oet_password'),
            f.get('oet_hint_question'), f.get('oet_security_answer'),
            session.get('user_id', 0)
        ))
        conn.commit()
        conn.close()
        flash('English login details added', 'success')
        return redirect(request.args.get('next') or url_for('ops_english_logins_list'))
    clients = conn.execute("SELECT registration_number, prefix, first_name, last_name FROM plab_clients ORDER BY first_name").fetchall()
    conn.close()
    pre_reg = request.args.get('client', '')
    return render_template('ops_english_logins_form.html', record=None, clients=clients, pre_reg=pre_reg)


@app.route('/operations/english-logins/<int:record_id>/edit', methods=['GET', 'POST'])
@admin_required
def ops_english_logins_edit(record_id):
    conn = get_db()
    record = conn.execute("SELECT * FROM ops_english_logins WHERE id = ?", (record_id,)).fetchone()
    if not record:
        conn.close()
        flash('Record not found', 'error')
        return redirect(url_for('ops_english_logins_list'))
    if request.method == 'POST':
        f = request.form
        conn.execute('''UPDATE ops_english_logins SET
            registration_number=?, ielts_login_id=?, ielts_password=?, ielts_hint_question=?,
            ielts_security_answer=?, oet_login_id=?, oet_password=?, oet_hint_question=?,
            oet_security_answer=? WHERE id=?''', (
            f.get('registration_number'), f.get('ielts_login_id'), f.get('ielts_password'),
            f.get('ielts_hint_question'), f.get('ielts_security_answer'),
            f.get('oet_login_id'), f.get('oet_password'),
            f.get('oet_hint_question'), f.get('oet_security_answer'),
            record_id
        ))
        conn.commit()
        conn.close()
        flash('English login details updated', 'success')
        return redirect(request.args.get('next') or url_for('ops_english_logins_list'))
    clients = conn.execute("SELECT registration_number, prefix, first_name, last_name FROM plab_clients ORDER BY first_name").fetchall()
    conn.close()
    return render_template('ops_english_logins_form.html', record=record, clients=clients, pre_reg='')


@app.route('/operations/english-logins/<int:record_id>/delete', methods=['POST'])
@admin_required
def ops_english_logins_delete(record_id):
    conn = get_db()
    conn.execute("DELETE FROM ops_english_logins WHERE id = ?", (record_id,))
    conn.commit()
    conn.close()
    flash('English login record deleted', 'success')
    return redirect(request.args.get('next') or url_for('ops_english_logins_list'))


# ─────────────────────────────────────────────────────────
#  OPERATIONS – Test Bookings
# ─────────────────────────────────────────────────────────

@app.route('/operations/test-bookings')
@admin_required
def ops_test_bookings_list():
    """List all test bookings."""
    conn = get_db()
    reg = request.args.get('client', '')
    search = request.args.get('q', '')
    exam_filter = request.args.get('exam', '')
    status_filter = request.args.get('status', '')
    try:
        sql = '''SELECT t.*, p.first_name, p.last_name, p.prefix
                 FROM ops_test_bookings t
                 LEFT JOIN plab_clients p ON t.registration_number = p.registration_number
                 WHERE 1=1'''
        params = []
        if reg:
            sql += ' AND t.registration_number = ?'
            params.append(reg)
        if exam_filter:
            sql += ' AND t.exam = ?'
            params.append(exam_filter)
        if status_filter:
            sql += ' AND t.exam_status = ?'
            params.append(status_filter)
        if search:
            sql += ' AND (p.first_name ILIKE ? OR p.last_name ILIKE ? OR t.test_center ILIKE ?)'
            params.extend([f'%{search}%'] * 3)
        sql += ' ORDER BY t.exam_date DESC NULLS LAST, t.created_at DESC'
        records = conn.execute(sql, params).fetchall()
    except Exception as e:
        logging.error(f"ops_test_bookings_list: {e}")
        records = []
    conn.close()
    return render_template('ops_test_bookings_list.html', records=records, client_reg=reg,
                           search=search, exam_filter=exam_filter, status_filter=status_filter,
                           exam_names=EXAM_NAMES, exam_statuses=EXAM_STATUSES)


@app.route('/operations/test-bookings/add', methods=['GET', 'POST'])
@admin_required
def ops_test_bookings_add():
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        conn.execute('''INSERT INTO ops_test_bookings (
            registration_number, exam, exam_type, booking_date, exam_date,
            exam_status, exam_result, exam_result_date, score, test_center,
            city_state, country, booked_by, revaluation, reval_applied_date,
            reval_result, reval_score, notes, created_by
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
            f.get('registration_number'), f.get('exam'), f.get('exam_type'),
            f.get('booking_date'), f.get('exam_date'), f.get('exam_status'),
            f.get('exam_result'), f.get('exam_result_date'), f.get('score'),
            f.get('test_center'), f.get('city_state'), f.get('country'),
            f.get('booked_by'), f.get('revaluation'), f.get('reval_applied_date'),
            f.get('reval_result'), f.get('reval_score'), f.get('notes'),
            session.get('user_id', 0)
        ))
        conn.commit()
        conn.close()
        flash('Test booking added', 'success')
        return redirect(request.args.get('next') or url_for('ops_test_bookings_list'))
    clients = conn.execute("SELECT registration_number, prefix, first_name, last_name FROM plab_clients ORDER BY first_name").fetchall()
    conn.close()
    pre_reg = request.args.get('client', '')
    return render_template('ops_test_bookings_form.html', record=None, clients=clients,
                           exam_names=EXAM_NAMES, exam_statuses=EXAM_STATUSES,
                           exam_results=EXAM_RESULTS, pre_reg=pre_reg)


@app.route('/operations/test-bookings/<int:record_id>/edit', methods=['GET', 'POST'])
@admin_required
def ops_test_bookings_edit(record_id):
    conn = get_db()
    record = conn.execute("SELECT * FROM ops_test_bookings WHERE id = ?", (record_id,)).fetchone()
    if not record:
        conn.close()
        flash('Record not found', 'error')
        return redirect(url_for('ops_test_bookings_list'))
    if request.method == 'POST':
        f = request.form
        conn.execute('''UPDATE ops_test_bookings SET
            registration_number=?, exam=?, exam_type=?, booking_date=?, exam_date=?,
            exam_status=?, exam_result=?, exam_result_date=?, score=?, test_center=?,
            city_state=?, country=?, booked_by=?, revaluation=?, reval_applied_date=?,
            reval_result=?, reval_score=?, notes=? WHERE id=?''', (
            f.get('registration_number'), f.get('exam'), f.get('exam_type'),
            f.get('booking_date'), f.get('exam_date'), f.get('exam_status'),
            f.get('exam_result'), f.get('exam_result_date'), f.get('score'),
            f.get('test_center'), f.get('city_state'), f.get('country'),
            f.get('booked_by'), f.get('revaluation'), f.get('reval_applied_date'),
            f.get('reval_result'), f.get('reval_score'), f.get('notes'),
            record_id
        ))
        conn.commit()
        conn.close()
        flash('Test booking updated', 'success')
        return redirect(request.args.get('next') or url_for('ops_test_bookings_list'))
    clients = conn.execute("SELECT registration_number, prefix, first_name, last_name FROM plab_clients ORDER BY first_name").fetchall()
    conn.close()
    return render_template('ops_test_bookings_form.html', record=record, clients=clients,
                           exam_names=EXAM_NAMES, exam_statuses=EXAM_STATUSES,
                           exam_results=EXAM_RESULTS, pre_reg='')


@app.route('/operations/test-bookings/<int:record_id>/delete', methods=['POST'])
@admin_required
def ops_test_bookings_delete(record_id):
    conn = get_db()
    conn.execute("DELETE FROM ops_test_bookings WHERE id = ?", (record_id,))
    conn.commit()
    conn.close()
    flash('Test booking deleted', 'success')
    return redirect(request.args.get('next') or url_for('ops_test_bookings_list'))


# ─────────────────────────────────────────────────────────
#  OPERATIONS – Call Notes
# ─────────────────────────────────────────────────────────

@app.route('/operations/call-notes')
@admin_required
def ops_call_notes_list():
    """List call notes with detailed filters: registration number, client name, added_by, note text."""
    conn = get_db()
    # Filter parameters
    reg = request.args.get('reg', '').strip()
    client_name = request.args.get('client_name', '').strip()
    added_by_filter = request.args.get('added_by', '').strip()
    note_search = request.args.get('note_search', '').strip()
    # Legacy support: 'client' param maps to reg, 'q' maps to note_search
    if not reg and request.args.get('client', ''):
        reg = request.args.get('client', '').strip()
    if not note_search and request.args.get('q', ''):
        note_search = request.args.get('q', '').strip()

    page = request.args.get('page', 1, type=int)
    if page < 1:
        page = 1

    per_page = 50
    offset = (page - 1) * per_page

    try:
        # Get distinct added_by values for dropdown
        added_by_list = conn.execute("SELECT DISTINCT added_by FROM ops_call_notes WHERE added_by IS NOT NULL AND added_by != '' ORDER BY added_by").fetchall()
        added_by_options = [r['added_by'] for r in added_by_list]

        # Build base query
        sql_base = '''SELECT n.*, p.first_name, p.last_name, p.prefix
                      FROM ops_call_notes n
                      LEFT JOIN plab_clients p ON n.registration_number = p.registration_number
                      WHERE 1=1'''
        params = []
        if reg:
            sql_base += ' AND n.registration_number ILIKE ?'
            params.append(f'%{reg}%')
        if client_name:
            sql_base += ' AND (p.first_name ILIKE ? OR p.last_name ILIKE ? OR (p.first_name || \' \' || p.last_name) ILIKE ?)'
            params.extend([f'%{client_name}%'] * 3)
        if added_by_filter:
            sql_base += ' AND n.added_by = ?'
            params.append(added_by_filter)
        if note_search:
            sql_base += ' AND n.call_note ILIKE ?'
            params.append(f'%{note_search}%')

        # Get total count
        count_sql = f'SELECT COUNT(*) as total FROM (SELECT 1 {sql_base[sql_base.find("FROM"):]})'
        total_count = conn.execute(count_sql, params).fetchone()['total']
        total_pages = (total_count + per_page - 1) // per_page

        # Ensure page is within valid range
        if page > total_pages and total_pages > 0:
            page = total_pages
            offset = (page - 1) * per_page

        # Get paginated records
        sql = sql_base + ' ORDER BY n.call_date DESC NULLS LAST, n.created_at DESC LIMIT ? OFFSET ?'
        sql_params = params + [per_page, offset]
        records = conn.execute(sql, sql_params).fetchall()
    except Exception as e:
        logging.error(f"ops_call_notes_list: {e}")
        records = []
        total_count = 0
        total_pages = 0
        added_by_options = []
    conn.close()

    # Check if any filter is active
    has_filters = bool(reg or client_name or added_by_filter or note_search)

    return render_template('ops_call_notes_list.html',
        records=records, reg=reg, client_name=client_name,
        added_by_filter=added_by_filter, note_search=note_search,
        added_by_options=added_by_options, has_filters=has_filters,
        page=page, per_page=per_page, total_count=total_count, total_pages=total_pages)


@app.route('/operations/api/client-search')
@admin_required
def ops_client_search_api():
    """Return matching client names for autocomplete (JSON)."""
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    conn = get_db()
    try:
        rows = conn.execute('''
            SELECT id, registration_number, prefix, first_name, last_name
            FROM plab_clients
            WHERE first_name ILIKE ? OR last_name ILIKE ?
               OR (first_name || ' ' || last_name) ILIKE ?
               OR registration_number ILIKE ?
            ORDER BY first_name, last_name
            LIMIT 15
        ''', (f'%{q}%', f'%{q}%', f'%{q}%', f'%{q}%')).fetchall()
        results = []
        for r in rows:
            name = f"{r['prefix']} {r['first_name']} {r['last_name']}" if r['prefix'] else f"{r['first_name']} {r['last_name']}"
            results.append({'id': r['id'], 'name': name.strip(), 'reg': r['registration_number']})
    except Exception as e:
        logging.error(f"ops_client_search_api: {e}")
        results = []
    conn.close()
    return jsonify(results)


@app.route('/operations/call-notes/add', methods=['GET', 'POST'])
@admin_required
def ops_call_notes_add():
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        conn.execute('''INSERT INTO ops_call_notes (
            registration_number, call_date, call_note, added_by, created_by
        ) VALUES (?,?,?,?,?)''', (
            f.get('registration_number'), f.get('call_date'), f.get('call_note'),
            f.get('added_by', ''), session.get('user_id', 0)
        ))
        conn.commit()
        conn.close()
        flash('Call note added', 'success')
        return redirect(request.args.get('next') or url_for('ops_call_notes_list'))
    clients = conn.execute("SELECT registration_number, prefix, first_name, last_name FROM plab_clients ORDER BY first_name").fetchall()
    conn.close()
    pre_reg = request.args.get('client', '')
    return render_template('ops_call_notes_form.html', record=None, clients=clients, pre_reg=pre_reg)


@app.route('/operations/call-notes/<int:record_id>/edit', methods=['GET', 'POST'])
@admin_required
def ops_call_notes_edit(record_id):
    conn = get_db()
    record = conn.execute("SELECT * FROM ops_call_notes WHERE id = ?", (record_id,)).fetchone()
    if not record:
        conn.close()
        flash('Record not found', 'error')
        return redirect(url_for('ops_call_notes_list'))
    if request.method == 'POST':
        f = request.form
        conn.execute('''UPDATE ops_call_notes SET
            registration_number=?, call_date=?, call_note=?, added_by=? WHERE id=?''', (
            f.get('registration_number'), f.get('call_date'), f.get('call_note'),
            f.get('added_by', ''), record_id
        ))
        conn.commit()
        conn.close()
        flash('Call note updated', 'success')
        return redirect(request.args.get('next') or url_for('ops_call_notes_list'))
    clients = conn.execute("SELECT registration_number, prefix, first_name, last_name FROM plab_clients ORDER BY first_name").fetchall()
    conn.close()
    return render_template('ops_call_notes_form.html', record=record, clients=clients, pre_reg='')


@app.route('/operations/call-notes/<int:record_id>/delete', methods=['POST'])
@admin_required
def ops_call_notes_delete(record_id):
    conn = get_db()
    conn.execute("DELETE FROM ops_call_notes WHERE id = ?", (record_id,))
    conn.commit()
    conn.close()
    flash('Call note deleted', 'success')
    return redirect(request.args.get('next') or url_for('ops_call_notes_list'))


# ═════════════════════════════════════════════════════════════════════════════
# PAYMENTS
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/operations/payments')
@admin_required
def ops_payments_list():
    """List all payments with filters: registration number, client name, payment method, instalment."""
    conn = get_db()
    # Filter parameters
    reg = request.args.get('reg', '').strip()
    client_name = request.args.get('client_name', '').strip()
    payment_method = request.args.get('payment_method', '').strip()
    instalment = request.args.get('instalment', '').strip()

    page = request.args.get('page', 1, type=int)
    if page < 1:
        page = 1

    per_page = 50
    offset = (page - 1) * per_page

    try:
        # Build base query
        sql_base = '''SELECT p.*, c.first_name, c.last_name, c.prefix
                      FROM ops_payments p
                      LEFT JOIN plab_clients c ON p.registration_number = c.registration_number
                      WHERE 1=1'''
        params = []
        if reg:
            sql_base += ' AND p.registration_number ILIKE ?'
            params.append(f'%{reg}%')
        if client_name:
            sql_base += """ AND (c.first_name ILIKE ? OR c.last_name ILIKE ?
                OR (c.first_name || ' ' || COALESCE(c.last_name,'')) ILIKE ?
                OR (COALESCE(c.prefix,'') || ' ' || c.first_name || ' ' || COALESCE(c.last_name,'')) ILIKE ?)"""
            params.extend([f'%{client_name}%'] * 4)
        if payment_method:
            sql_base += ' AND p.payment_method = ?'
            params.append(payment_method)
        if instalment:
            sql_base += ' AND p.instalment = ?'
            params.append(instalment)

        # Get total count
        count_sql = f'SELECT COUNT(*) as total FROM (SELECT 1 {sql_base[sql_base.find("FROM"):]})'
        total_count = conn.execute(count_sql, params).fetchone()['total']
        total_pages = (total_count + per_page - 1) // per_page

        # Ensure page is within valid range
        if page > total_pages and total_pages > 0:
            page = total_pages
            offset = (page - 1) * per_page

        # Get paginated records
        sql = sql_base + ' ORDER BY p.payment_date DESC NULLS LAST, p.created_at DESC LIMIT ? OFFSET ?'
        sql_params = params + [per_page, offset]
        records = conn.execute(sql, sql_params).fetchall()

        # Stats
        stats_base = sql_base.replace('SELECT p.*, c.first_name, c.last_name, c.prefix', 'SELECT COUNT(*) as total, COALESCE(SUM(p.total_amount_paid), 0) as total_collected, COALESCE(SUM(p.gst_paid), 0) as total_gst')
        stats = conn.execute(stats_base, params).fetchone()

    except Exception as e:
        logging.error(f"ops_payments_list: {e}")
        records = []
        total_count = 0
        total_pages = 0
        stats = {'total': 0, 'total_collected': 0, 'total_gst': 0}

    conn.close()

    # Check if any filter is active
    has_filters = bool(reg or client_name or payment_method or instalment)

    return render_template('ops_payments_list.html',
        records=records, reg=reg, client_name=client_name,
        payment_method=payment_method, instalment=instalment,
        payment_methods=PAYMENT_METHODS, instalment_options=INSTALMENT_OPTIONS,
        has_filters=has_filters, stats=stats,
        page=page, per_page=per_page, total_count=total_count, total_pages=total_pages)


@app.route('/operations/payments/add', methods=['GET', 'POST'])
@admin_required
def ops_payments_add():
    conn = get_db()
    if request.method == 'POST':
        registration_number = request.form.get('registration_number')
        payment_date = request.form.get('payment_date')
        amount_paid = request.form.get('amount_paid', '0')
        gst_paid = request.form.get('gst_paid', '0')
        instalment = request.form.get('instalment')
        payment_method = request.form.get('payment_method')
        total_package = request.form.get('total_package', '0')
        notes = request.form.get('notes', '')

        try:
            amount_paid = float(amount_paid or 0)
            gst_paid = float(gst_paid or 0)
            total_amount_paid = amount_paid + gst_paid
            total_package = float(total_package or 0)

            conn.execute('''INSERT INTO ops_payments
                (registration_number, payment_date, amount_paid, gst_paid, total_amount_paid,
                 instalment, payment_method, total_package, notes, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (registration_number, payment_date, amount_paid, gst_paid, total_amount_paid,
                 instalment, payment_method, total_package, notes, session.get('user_id')))
            conn.commit()
            flash('Payment added successfully', 'success')
            next_url = request.args.get('next')
            if next_url:
                return redirect(next_url)
            return redirect(url_for('ops_payments_list'))
        except Exception as e:
            logging.error(f"ops_payments_add: {e}")
            flash('Error adding payment', 'error')

    # Get all clients for dropdown
    clients = conn.execute('SELECT id, registration_number, prefix, first_name, last_name FROM plab_clients ORDER BY first_name').fetchall()
    pre_reg = request.args.get('reg', '')
    conn.close()

    return render_template('ops_payments_form.html',
        clients=clients, record=None, pre_reg=pre_reg,
        payment_methods=PAYMENT_METHODS, instalment_options=INSTALMENT_OPTIONS)


@app.route('/operations/payments/<int:record_id>/edit', methods=['GET', 'POST'])
@admin_required
def ops_payments_edit(record_id):
    conn = get_db()
    record = conn.execute('SELECT * FROM ops_payments WHERE id = ?', (record_id,)).fetchone()
    if not record:
        conn.close()
        flash('Payment record not found', 'error')
        return redirect(url_for('ops_payments_list'))

    if request.method == 'POST':
        registration_number = request.form.get('registration_number')
        payment_date = request.form.get('payment_date')
        amount_paid = request.form.get('amount_paid', '0')
        gst_paid = request.form.get('gst_paid', '0')
        instalment = request.form.get('instalment')
        payment_method = request.form.get('payment_method')
        total_package = request.form.get('total_package', '0')
        notes = request.form.get('notes', '')

        try:
            amount_paid = float(amount_paid or 0)
            gst_paid = float(gst_paid or 0)
            total_amount_paid = amount_paid + gst_paid
            total_package = float(total_package or 0)

            conn.execute('''UPDATE ops_payments SET
                registration_number = ?, payment_date = ?, amount_paid = ?, gst_paid = ?,
                total_amount_paid = ?, instalment = ?, payment_method = ?, total_package = ?,
                notes = ?
                WHERE id = ?''',
                (registration_number, payment_date, amount_paid, gst_paid, total_amount_paid,
                 instalment, payment_method, total_package, notes, record_id))
            conn.commit()
            flash('Payment updated successfully', 'success')
            return redirect(url_for('ops_payments_list'))
        except Exception as e:
            logging.error(f"ops_payments_edit: {e}")
            flash('Error updating payment', 'error')

    # Get all clients for dropdown
    clients = conn.execute('SELECT id, registration_number, prefix, first_name, last_name FROM plab_clients ORDER BY first_name').fetchall()
    conn.close()

    return render_template('ops_payments_form.html',
        clients=clients, record=record, pre_reg=None,
        payment_methods=PAYMENT_METHODS, instalment_options=INSTALMENT_OPTIONS)


@app.route('/operations/payments/<int:record_id>/delete', methods=['POST'])
@admin_required
def ops_payments_delete(record_id):
    conn = get_db()
    conn.execute('DELETE FROM ops_payments WHERE id = ?', (record_id,))
    conn.commit()
    conn.close()
    flash('Payment deleted', 'success')
    return redirect(request.args.get('next') or url_for('ops_payments_list'))


# Run on startup
ensure_crm_tables()
ensure_kra_tables()
ensure_notification_tables()
ensure_budget_tables()
ensure_ops_tables()
seed_kra_categories()
seed_default_meeting_types()
seed_budget_categories()
# Backfill finance revenue categories from sales streams at boot
try:
    _sync_conn = get_db()
    sync_streams_to_budget_categories(_sync_conn)
    _sync_conn.close()
except Exception as _sync_err:
    logging.error(f"Startup stream->category sync failed: {_sync_err}")
ensure_management_admins()

if __name__ == '__main__':
    PORT = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=PORT, debug=False)
