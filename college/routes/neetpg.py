"""College module — neetpg routes. Extracted from app.py (modularization)."""
import os, re, io, csv, json, time, random, hashlib, logging, threading, traceback
from io import BytesIO
from datetime import datetime, timedelta
import requests
import requests as http_requests
import json as _json
import time as _time
from flask import (render_template, request, session, redirect, url_for, flash,
                   jsonify, make_response, send_file, abort, Response)
from db import get_db
from core.auth import admin_required, login_required
from core.users import get_user
try:
    import pytz
except Exception:
    pytz = None
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except Exception:
    Workbook = Font = PatternFill = Alignment = Border = Side = None
from college.utils import _make_slug, _to_inr, _get_exchange_rates, _currency_sym, _currency_cache

# Document type is a separate dimension from `category` (exam type: neetpg/dnb).
# Order here drives the tab order on admin + landing pages.
NEETPG_DOC_TYPES = [
    ('cutoff', 'Cut Off'),
    ('stipend_bond_penalty', 'Stipend, Bond & Penalty'),
    ('mcc_profile', 'AIQ MCC College Profiles'),
    ('mcc_bond_doc', 'AIQ MCC College Bond Doc'),
]
DOC_TYPE_LABELS = dict(NEETPG_DOC_TYPES)
VALID_DOC_TYPES = set(DOC_TYPE_LABELS)
# Doc types whose "specialty" column actually holds a College Name (not a course).
COLLEGE_NAME_DOC_TYPES = {'mcc_profile', 'mcc_bond_doc'}


def _neetpg_schedule_map(conn):
    """pdf_id -> 'May 13, 10:00 AM IST' ETA for queued (unpublished, auto-schedule)
    drafts. Slots are PER TAB (category, doc_type): each tab publishes at 10am / 1pm /
    5pm IST independently, so a draft's ETA = the slot matching its 0-based rank within
    its own tab's FIFO queue (tabs roll out in parallel, not one global queue)."""
    schedule_map = {}
    try:
        import pytz
        from datetime import timedelta
        ist = pytz.timezone('Asia/Kolkata')
        now_ist = datetime.now(ist)
        queued = conn.execute(
            "SELECT id, category, doc_type FROM neetpg_pdfs "
            "WHERE is_published = 0 AND is_active = 1 AND auto_schedule = 1 ORDER BY upload_date ASC"
        ).fetchall()
        if not queued:
            return schedule_map
        # Rank each draft within its own (category, doc_type) tab.
        tab_seen, ranks, max_rank = {}, {}, 0
        for d in queued:
            key = (d['category'], d['doc_type'])
            r = tab_seen.get(key, 0)
            ranks[d['id']] = r
            tab_seen[key] = r + 1
            max_rank = max(max_rank, r + 1)
        # Build the first `max_rank` future slots (today's remaining first, then next days).
        slots, day = [], 0
        while len(slots) < max_rank and day < 400:
            base = now_ist + timedelta(days=day)
            for h in (10, 13, 17):
                s = base.replace(hour=h, minute=0, second=0, microsecond=0)
                if s > now_ist:
                    slots.append(s)
            day += 1
        for d in queued:
            idx = ranks[d['id']]
            if idx < len(slots):
                schedule_map[d['id']] = slots[idx].strftime('%b %d, %I:%M %p') + ' IST'
    except Exception as e:
        logging.error(f"neetpg schedule_map build error: {e}")
    return schedule_map


def _build_neetpg_title(doc_type, category, state, specialty, quota=''):
    """Auto-compose a PDF title from the upload fields.
    Format: "<State> <Specialty/College> <NEET PG|DNB> <Quota?> <Doc Type> 2025".
    Quota only applies to Cut Off. For College Profile / Bond Doc, `specialty`
    holds the College Name. Mirrors the admin-form JS so the saved title matches
    the live preview."""
    cat_label = 'DNB' if category == 'dnb' else 'NEET PG'
    dt_label = DOC_TYPE_LABELS.get(doc_type, 'Cut Off')
    parts = [state, specialty, cat_label, quota, dt_label, '2025']
    return ' '.join(p.strip() for p in parts if p and p.strip())


def _parse_quota_from_filename(file_name):
    """Extract the quota segment from a Cut-Off filename (mirrors the admin JS).
    Pattern: 'State - Course - <NEET PG|DNB> - <Quota> - Cut Off 2025'. The quota
    is whatever sits between the category segment and the trailing 'Cut Off' part.
    Returns '' if the name isn't a parseable cut-off name."""
    if not file_name:
        return ''
    name = re.sub(r'\.pdf$', '', file_name, flags=re.I).strip()
    if not re.search(r'cut ?off', name, flags=re.I):
        return ''
    # Split only on a spaced hyphen so internal hyphens (Radio-diagnosis) survive.
    parts = [p.strip() for p in re.split(r'\s+-\s*|\s*-\s+', name) if p.strip()]
    if len(parts) < 3:
        return ''
    cat_idx = -1
    for i in range(1, len(parts)):
        pl = parts[i].lower()
        if pl == 'dnb' or pl in ('neet pg', 'neetpg'):
            cat_idx = i
            break
    if cat_idx == -1:
        return ''
    tail = parts[cat_idx + 1:]                       # [quota?, 'Cut Off 2025']
    return ' - '.join(tail[:-1]).strip() if len(tail) > 1 else ''


# ── Server-side filename parse + specialty snap (mirror of the admin-form JS) ──
_SPLIT_RE = re.compile(r'\s+-\s*|\s*-\s+')            # split only on a spaced hyphen

# Official lists + alias bridge live in the generated data module (single source
# of truth, shared with the admin-form JS which receives them as JSON).
from college.data.neetpg_lists import NEETPG_COURSES as _NEETPG_COURSES
from college.data.neetpg_lists import DNB_COURSES as _DNB_COURSES
from college.data.neetpg_lists import SPECIALTY_ALIAS as _SPECIALTY_ALIAS
from college.data.neetpg_lists import MCC_COLLEGES as _MCC_COLLEGES


def _flat(s):
    return re.sub(r'[^a-z0-9]', '', (s or '').lower())


def _course_core(e):
    e = re.sub(r'^\(NBEMS\)\s*(Diploma\s+)?', '', e, flags=re.I)
    e = re.sub(r'^(MD/MS|MD|MS|MCh|MPH)\s*-\s*', '', e, flags=re.I)
    e = re.sub(r'^Diploma in\s+', '', e, flags=re.I)
    return e


def _snap_specialty(raw, category):
    """Snap a general course name to the canonical filter value. Mirrors the JS."""
    is_dnb = (category == 'dnb')
    lst = _DNB_COURSES if is_dnb else _NEETPG_COURSES
    t = _flat(raw)
    if not t:
        return raw
    al = _SPECIALTY_ALIAS['dnb' if is_dnb else 'neetpg']
    if t in al:
        return al[t]

    def _is_dip(e):
        el = e.lower()
        return el.startswith('diploma') or el.startswith('(nbems) diploma')
    # A bare name (no "Diploma" prefix) means the degree, not the diploma — so
    # match non-diploma entries first (both passes: exact core, then exact full).
    for pref in (False, True):
        for e in lst:
            if (pref or not _is_dip(e)) and _flat(_course_core(e)) == t:
                return e
    for pref in (False, True):
        for e in lst:
            if (pref or not _is_dip(e)) and _flat(e) == t:
                return e
    if is_dnb:
        m = re.match(r'^diploma\s+(.+)$', (raw or '').strip(), flags=re.I)
        return '(Diploma) ' + m.group(1).strip() if m else '(NBEMS) ' + (raw or '').strip()
    return raw


def _parse_neetpg_filename(file_name):
    """Parse a doc filename into fields, or None if unrecognised. Mirrors the JS."""
    if not file_name:
        return None
    name = re.sub(r'\.pdf$', '', file_name, flags=re.I).strip()
    low = name.lower()
    if re.search(r'college bond|bond doc', low):
        doc_type = 'mcc_bond_doc'
    elif re.search(r'college profile', low):
        doc_type = 'mcc_profile'
    elif re.search(r'cut ?off', low):
        doc_type = 'cutoff'
    elif re.search(r'stipend|penalty', low):
        doc_type = 'stipend_bond_penalty'
    else:
        return None
    parts = [p.strip() for p in _SPLIT_RE.split(name) if p.strip()]
    if len(parts) < 3:
        return None
    cat_idx, category = -1, None
    for i in range(1, len(parts)):
        pl = parts[i].lower()
        if pl == 'dnb':
            cat_idx, category = i, 'dnb'
            break
        if pl in ('neet pg', 'neetpg'):
            cat_idx, category = i, 'neetpg'
            break
    if cat_idx == -1:
        return None
    state = parts[0]
    middle = ' - '.join(parts[1:cat_idx]).strip()
    tail = parts[cat_idx + 1:]
    quota = ' - '.join(tail[:-1]).strip() if len(tail) > 1 else ''
    if doc_type in ('cutoff', 'stipend_bond_penalty'):
        specialty = _snap_specialty(middle, category)
    else:
        specialty = middle
    return {
        'doc_type': doc_type, 'state': state, 'category': category,
        'specialty': specialty, 'quota': quota if doc_type == 'cutoff' else '',
    }


def neetpg_landing():
    conn = get_db()
    # Track page visit
    try:
        ip_raw = request.remote_addr or ''
        ip_hash = hashlib.md5(ip_raw.encode()).hexdigest()[:12]
        ua = (request.headers.get('User-Agent', '') or '')[:200]
        conn.execute("INSERT INTO neetpg_page_visits (ip_hash, user_agent) VALUES (?, ?)", (ip_hash, ua))
        conn.commit()
    except Exception as e:
        logging.error(f"neetpg visit track: {e}")
    # Only show published + active PDFs, newest published first
    all_pdfs = conn.execute(
        "SELECT id, title, category, doc_type, specialty, quota_category, state, file_name, file_size, upload_date, download_count, published_at FROM neetpg_pdfs WHERE is_published = 1 AND is_active = 1 ORDER BY published_at DESC"
    ).fetchall()
    neetpg_pdfs = [p for p in all_pdfs if p['category'] == 'neetpg']
    dnb_pdfs = [p for p in all_pdfs if p['category'] == 'dnb']

    # ── Fetch scheduled (coming soon) PDFs ──
    scheduled_pdfs = conn.execute(
        "SELECT id, title, category, doc_type, specialty, quota_category, state, file_name, file_size, upload_date, download_count, published_at FROM neetpg_pdfs WHERE is_published = 0 AND is_active = 1 AND auto_schedule = 1 ORDER BY upload_date ASC"
    ).fetchall()

    # ── Coming-soon ETAs, per tab (each category+doc_type publishes in parallel) ──
    schedule_map = _neetpg_schedule_map(conn)

    neetpg_scheduled = [p for p in scheduled_pdfs if p['category'] == 'neetpg']
    dnb_scheduled = [p for p in scheduled_pdfs if p['category'] == 'dnb']

    # ── Sidebar filter items, tagged by both category (neetpg/dnb) AND doc_type ──
    # The landing page shows/hides these in JS based on the active doc-type tab
    # and the active NEET/DNB sub-tab, so each item carries both keys.
    def _dt(p):
        return (p['doc_type'] or 'cutoff') if 'doc_type' in p.keys() else 'cutoff'

    state_counts = {}    # (cat, doc_type, state) -> count
    course_seen = set()  # (cat, doc_type, course)
    courses_items = []
    for p in all_pdfs:
        cat = p['category']
        dt = _dt(p)
        st = p['state'] or 'AIQ MCC'
        state_counts[(cat, dt, st)] = state_counts.get((cat, dt, st), 0) + 1
        sp = p['specialty']
        if sp and (cat, dt, sp) not in course_seen:
            course_seen.add((cat, dt, sp))
            courses_items.append({'course': sp, 'cat': cat, 'doc_type': dt})
    states_items = [
        {'state': st, 'cat': cat, 'doc_type': dt, 'count': cnt}
        for (cat, dt, st), cnt in state_counts.items()
    ]
    states_items.sort(key=lambda x: (x['cat'], x['doc_type'], x['state']))
    courses_items.sort(key=lambda x: (x['cat'], x['doc_type'], x['course']))

    # Quota filter items (Cut Off only) — comma lists like 'GM, ST, OBC' split
    # into individual tags so a file shows under each. Tagged by category.
    def _quota(p):
        return (p['quota_category'] or '') if 'quota_category' in p.keys() else ''
    quota_seen = set()
    quota_items = []
    for p in all_pdfs:
        if _dt(p) != 'cutoff':
            continue
        for tag in [t.strip() for t in _quota(p).split(',') if t.strip()]:
            key = (p['category'], tag)
            if key not in quota_seen:
                quota_seen.add(key)
                quota_items.append({'quota': tag, 'cat': p['category']})
    quota_items.sort(key=lambda x: (x['cat'], x['quota']))

    # Published count per doc_type (for the top-level doc-type tab badges)
    doctype_counts = {k: 0 for k in DOC_TYPE_LABELS}
    for p in all_pdfs:
        dt = _dt(p)
        doctype_counts[dt] = doctype_counts.get(dt, 0) + 1
    conn.close()
    # Check if lead already captured (cookie)
    lead_captured = request.cookies.get('neetpg_lead', '')
    lead_name = request.cookies.get('neetpg_lead_name', '')
    resp = make_response(render_template('college/neetpg_landing.html',
                           neetpg_pdfs=neetpg_pdfs,
                           dnb_pdfs=dnb_pdfs,
                           all_pdfs=all_pdfs,
                           neetpg_scheduled=neetpg_scheduled,
                           dnb_scheduled=dnb_scheduled,
                           schedule_map=schedule_map,
                           states_items=states_items,
                           courses_items=courses_items,
                           quota_items=quota_items,
                           doc_types=NEETPG_DOC_TYPES,
                           doctype_counts=doctype_counts,
                           lead_captured=lead_captured,
                           lead_name=lead_name))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return resp


def neetpg_capture_lead():
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    whatsapp = data.get('whatsapp', '').strip()
    email = data.get('email', '').strip()
    source = data.get('source', 'whatsapp').strip()

    # At least one contact method required
    if not name:
        return jsonify({'error': 'Name is required'}), 400
    if not whatsapp and not email:
        return jsonify({'error': 'WhatsApp number or Google sign-in is required'}), 400

    whatsapp_clean = None
    if whatsapp:
        whatsapp_clean = re.sub(r'[^0-9]', '', whatsapp)
        if len(whatsapp_clean) < 10:
            return jsonify({'error': 'Please enter a valid WhatsApp number'}), 400
        whatsapp_clean = whatsapp_clean[-10:]

    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO neetpg_leads (name, whatsapp, email, source) VALUES (?, ?, ?, ?)",
            (name, whatsapp_clean, email or None, source)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        logging.error(f"neetpg_capture_lead: {e}")
    resp = jsonify({'success': True})
    resp.set_cookie('neetpg_lead', 'captured', max_age=365*24*60*60, samesite='Lax')
    return resp


def neetpg_check_phone():
    data = request.get_json() or {}
    whatsapp = data.get('whatsapp', '').strip()

    if not whatsapp:
        return jsonify({'error': 'WhatsApp number is required'}), 400

    whatsapp_clean = re.sub(r'[^0-9]', '', whatsapp)
    if len(whatsapp_clean) < 10:
        return jsonify({'error': 'Please enter a valid 10-digit number'}), 400
    whatsapp_clean = whatsapp_clean[-10:]

    try:
        conn = get_db()
        existing = conn.execute(
            "SELECT name FROM neetpg_leads WHERE whatsapp = ? ORDER BY id DESC LIMIT 1",
            (whatsapp_clean,)
        ).fetchone()
        conn.close()
        if existing:
            return jsonify({'exists': True, 'name': existing['name']})
        else:
            return jsonify({'exists': False})
    except Exception as e:
        logging.error(f"neetpg_check_phone: {e}")
        return jsonify({'exists': False})


def neetpg_send_otp():
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    whatsapp = data.get('whatsapp', '').strip()
    is_returning = data.get('returning', False)

    if not whatsapp:
        return jsonify({'error': 'WhatsApp number is required'}), 400

    whatsapp_clean = re.sub(r'[^0-9]', '', whatsapp)
    if len(whatsapp_clean) < 10:
        return jsonify({'error': 'Please enter a valid WhatsApp number'}), 400
    whatsapp_clean = whatsapp_clean[-10:]

    # For new users, name is required
    if not is_returning and not name:
        return jsonify({'error': 'Name is required'}), 400

    # For returning users, fetch name from DB
    if is_returning:
        try:
            conn = get_db()
            existing = conn.execute(
                "SELECT name FROM neetpg_leads WHERE whatsapp = ? ORDER BY id DESC LIMIT 1",
                (whatsapp_clean,)
            ).fetchone()
            conn.close()
            if existing:
                name = existing['name']
        except Exception as e:
            logging.error(f"neetpg_send_otp fetch name: {e}")

    # Generate 6-digit OTP
    otp_code = str(random.randint(100000, 999999))

    # Store OTP in session
    session['neetpg_otp'] = otp_code
    session['neetpg_otp_phone'] = whatsapp_clean
    session['neetpg_otp_name'] = name
    session['neetpg_otp_returning'] = is_returning
    session['neetpg_otp_time'] = datetime.now().isoformat()

    # Send OTP via Infobip WhatsApp
    infobip_key = os.environ.get('INFOBIP_API_KEY', '')
    infobip_base = os.environ.get('INFOBIP_BASE_URL', '')

    if not infobip_key or not infobip_base:
        logging.error("Infobip credentials not configured")
        return jsonify({'error': 'OTP service not configured. Please try again later.'}), 500

    try:
        import requests as http_requests
        url = f"https://{infobip_base}/whatsapp/1/message/template"
        headers = {
            "Authorization": f"App {infobip_key}",
            "Content-Type": "application/json"
        }
        payload = {
            "messages": [{
                "from": "15558246314",
                "to": f"91{whatsapp_clean}",
                "content": {
                    "templateName": "goocampus_otp_verify",
                    "templateData": {
                        "body": {
                            "placeholders": [otp_code]
                        },
                        "buttons": [{
                            "type": "URL",
                            "parameter": otp_code
                        }]
                    },
                    "language": "en"
                }
            }]
        }
        resp = http_requests.post(url, json=payload, headers=headers, timeout=15)
        resp_data = resp.json()
        logging.info(f"Infobip OTP response for {whatsapp_clean}: status={resp.status_code} body={resp_data}")

        if resp.status_code >= 400:
            # Try fallback template
            payload["messages"][0]["content"]["templateName"] = "otp_verify"
            payload["messages"][0]["content"]["language"] = "en_GB"
            resp2 = http_requests.post(url, json=payload, headers=headers, timeout=15)
            resp2_data = resp2.json()
            logging.info(f"Infobip OTP fallback response: status={resp2.status_code} body={resp2_data}")
            if resp2.status_code >= 400:
                return jsonify({'error': 'Could not send OTP. Please check your WhatsApp number and try again.'}), 500

        return jsonify({'success': True, 'message': 'OTP sent to your WhatsApp'})
    except Exception as e:
        logging.error(f"Infobip OTP send error: {e}")
        return jsonify({'error': 'Failed to send OTP. Please try again.'}), 500


def neetpg_verify_otp():
    data = request.get_json() or {}
    entered_otp = data.get('otp', '').strip()

    if not entered_otp:
        return jsonify({'error': 'Please enter the OTP'}), 400

    stored_otp = session.get('neetpg_otp')
    stored_phone = session.get('neetpg_otp_phone')
    stored_name = session.get('neetpg_otp_name')
    stored_time = session.get('neetpg_otp_time')
    is_returning = session.get('neetpg_otp_returning', False)

    if not stored_otp or not stored_phone:
        return jsonify({'error': 'OTP expired. Please request a new one.'}), 400

    # Check OTP expiry (10 minutes)
    try:
        otp_time = datetime.fromisoformat(stored_time)
        if (datetime.now() - otp_time).total_seconds() > 600:
            session.pop('neetpg_otp', None)
            return jsonify({'error': 'OTP expired. Please request a new one.'}), 400
    except Exception:
        pass

    if entered_otp != stored_otp:
        return jsonify({'error': 'Invalid OTP. Please try again.'}), 400

    # Only insert lead for new users (not returning)
    if not is_returning:
        try:
            conn = get_db()
            # Check once more to avoid race condition duplicates
            existing = conn.execute(
                "SELECT id FROM neetpg_leads WHERE whatsapp = ? LIMIT 1",
                (stored_phone,)
            ).fetchone()
            if not existing:
                conn.execute(
                    "INSERT INTO neetpg_leads (name, whatsapp, source) VALUES (?, ?, ?)",
                    (stored_name, stored_phone, 'whatsapp_otp')
                )
                conn.commit()
            conn.close()
        except Exception as e:
            logging.error(f"neetpg OTP lead save: {e}")

    # Clear OTP from session
    session.pop('neetpg_otp', None)
    session.pop('neetpg_otp_phone', None)
    session.pop('neetpg_otp_name', None)
    session.pop('neetpg_otp_time', None)
    session.pop('neetpg_otp_returning', None)

    resp = jsonify({'success': True, 'name': stored_name})
    resp.set_cookie('neetpg_lead', 'captured', max_age=365*24*60*60, samesite='Lax')
    resp.set_cookie('neetpg_lead_name', stored_name or '', max_age=365*24*60*60, samesite='Lax')
    return resp


def neetpg_logout():
    resp = jsonify({'success': True})
    resp.set_cookie('neetpg_lead', '', max_age=0, samesite='Lax')
    resp.set_cookie('neetpg_lead_name', '', max_age=0, samesite='Lax')
    return resp


def neetpg_submit_request():
    phone = request.cookies.get('neetpg_lead', '')
    name = request.cookies.get('neetpg_lead_name', '')
    if not phone:
        return jsonify({'error': 'Please login first'}), 401
    data = request.get_json() or {}
    category = data.get('category', '').strip()
    state = data.get('state', '').strip()
    message = data.get('message', '').strip()
    if not category or not message:
        return jsonify({'error': 'Category and message are required'}), 400
    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO neetpg_requests (phone, name, category, state, message) VALUES (?, ?, ?, ?, ?)",
            (phone, name or 'Doctor', category, state or None, message)
        )
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Your request has been submitted. We will notify you when the document is available.'})
    except Exception as e:
        logging.error(f"neetpg request submit error: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'error': 'Failed to submit request. Please try again.'}), 500


def neetpg_check_notifications():
    phone = request.cookies.get('neetpg_lead', '')
    if not phone:
        return jsonify({'notifications': []})
    try:
        conn = get_db()
        notifs = conn.execute(
            "SELECT id, category, state, message, admin_note, completed_at, fulfilled_pdf_title FROM neetpg_requests "
            "WHERE phone = ? AND status = 'completed' AND notified = 0 ORDER BY completed_at DESC",
            (phone,)
        ).fetchall()
        results = []
        for n in notifs:
            results.append({
                'id': n['id'],
                'category': n['category'],
                'state': n['state'],
                'message': n['message'],
                'admin_note': n['admin_note'],
                'completed_at': str(n['completed_at']) if n['completed_at'] else '',
                'fulfilled_pdf_title': n['fulfilled_pdf_title'] or ''
            })
        # Mark as notified
        if results:
            conn.execute(
                "UPDATE neetpg_requests SET notified = 1 WHERE phone = ? AND status = 'completed' AND notified = 0",
                (phone,)
            )
            conn.commit()
        conn.close()
        return jsonify({'notifications': results})
    except Exception as e:
        logging.error(f"neetpg check notifications error: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'notifications': []})


@login_required
def neetpg_complete_request(req_id):
    user = get_user()
    if not user.get('is_admin'):
        return jsonify({'error': 'Admin required'}), 403
    data = request.get_json() or {}
    admin_note = data.get('admin_note', '').strip()
    pdf_id = data.get('pdf_id')
    pdf_title = data.get('pdf_title', '').strip()
    try:
        conn = get_db()
        req = conn.execute("SELECT * FROM neetpg_requests WHERE id = ?", (req_id,)).fetchone()
        if not req:
            conn.close()
            return jsonify({'error': 'Request not found'}), 404
        conn.execute(
            "UPDATE neetpg_requests SET status = 'completed', completed_at = CURRENT_TIMESTAMP, "
            "admin_note = ?, fulfilled_pdf_id = ?, fulfilled_pdf_title = ? WHERE id = ?",
            (admin_note, pdf_id, pdf_title, req_id)
        )
        conn.commit()

        # Send WhatsApp notification via Infobip
        # Set NEETPG_WA_NOTIFY=1 in env to enable (disabled until template is approved)
        phone = req['phone']
        wa_sent = False
        wa_error = ''
        wa_enabled = os.environ.get('NEETPG_WA_NOTIFY', '') == '1'
        if not wa_enabled:
            wa_error = 'WhatsApp notifications disabled (enable via NEETPG_WA_NOTIFY=1)'
            logging.info(f"WA notification skipped for request {req_id} — NEETPG_WA_NOTIFY not set")
        elif phone and len(phone) >= 10:
            try:
                import requests as http_requests
                infobip_key = os.environ.get('INFOBIP_API_KEY', '')
                infobip_base = os.environ.get('INFOBIP_BASE_URL', '')
                if infobip_key and infobip_base:
                    wa_headers = {
                        "Authorization": f"App {infobip_key}",
                        "Content-Type": "application/json"
                    }
                    to_number = f"91{phone[-10:]}"
                    pdf_line = pdf_title if pdf_title else f"{req['category'].upper()} - {req['state'] or 'General'}"
                    msg_text = (
                        f"Hi {req['name']},\n\n"
                        f"Great news! The document you requested is now available:\n\n"
                        f"*{pdf_line}*\n\n"
                        f"Access it here: https://goocampus.org/neet-pg-2025\n\n"
                        f"— GooCampus Team"
                    )

                    # Try free-form text first (works if 24h session window is open)
                    wa_url = f"https://{infobip_base}/whatsapp/1/message/text"
                    wa_payload = {
                        "from": "15558246314",
                        "to": to_number,
                        "content": {"text": msg_text}
                    }
                    resp = http_requests.post(wa_url, json=wa_payload, headers=wa_headers, timeout=15)
                    logging.info(f"WA text msg to {phone}: status={resp.status_code}, body={resp.text[:300]}")

                    if resp.status_code >= 400:
                        # Free-form failed (likely no 24h session window) — try template
                        logging.info(f"Free-form WA failed, trying template for {phone}")
                        tmpl_url = f"https://{infobip_base}/whatsapp/1/message/template"
                        tmpl_payload = {
                            "messages": [{
                                "from": "15558246314",
                                "to": to_number,
                                "content": {
                                    "templateName": "joincommunity",
                                    "templateData": {
                                        "body": {
                                            "placeholders": [req['name'], pdf_line, "https://goocampus.org/neet-pg-2025"]
                                        },
                                        "buttons": [
                                            {"type": "URL", "parameter": "https://goocampus.org/neet-pg-2025"}
                                        ]
                                    },
                                    "language": "en_IN"
                                }
                            }]
                        }
                        resp2 = http_requests.post(tmpl_url, json=tmpl_payload, headers=wa_headers, timeout=15)
                        logging.info(f"WA template msg to {phone}: status={resp2.status_code}, body={resp2.text[:300]}")

                        if resp2.status_code >= 400:
                            # Template also failed — try with generic OTP template as last resort
                            # Just send a simple notification text via SMS-like approach
                            wa_error = f"WhatsApp delivery failed (text: {resp.status_code}, template: {resp2.status_code})"
                            logging.warning(wa_error)
                        else:
                            wa_sent = True
                    else:
                        wa_sent = True
                else:
                    wa_error = 'Infobip credentials not configured'
            except Exception as wa_err:
                wa_error = str(wa_err)
                logging.error(f"WA notification for request {req_id} failed: {wa_err}")

        conn.close()
        return jsonify({'success': True, 'wa_sent': wa_sent, 'wa_error': wa_error})
    except Exception as e:
        logging.error(f"neetpg complete request error: {e}")
        return jsonify({'error': 'Failed to complete request'}), 500


def neetpg_download(pdf_id):
    conn = get_db()
    pdf = conn.execute("SELECT file_name, file_data FROM neetpg_pdfs WHERE id = ?", (pdf_id,)).fetchone()
    if not pdf:
        conn.close()
        flash('PDF not found', 'error')
        return redirect(url_for('neetpg_landing'))
    # Increment download count
    conn.execute("UPDATE neetpg_pdfs SET download_count = download_count + 1 WHERE id = ?", (pdf_id,))
    conn.commit()
    conn.close()
    file_data = pdf['file_data']
    if isinstance(file_data, memoryview):
        file_data = bytes(file_data)
    return send_file(
        BytesIO(file_data),
        download_name=pdf['file_name'],
        as_attachment=True,
        mimetype='application/pdf'
    )


def neetpg_view(pdf_id):
    conn = get_db()
    pdf = conn.execute("SELECT file_name, file_data FROM neetpg_pdfs WHERE id = ?", (pdf_id,)).fetchone()
    if not pdf:
        conn.close()
        flash('PDF not found', 'error')
        return redirect(url_for('neetpg_landing'))
    conn.close()
    file_data = pdf['file_data']
    if isinstance(file_data, memoryview):
        file_data = bytes(file_data)
    return send_file(
        BytesIO(file_data),
        download_name=pdf['file_name'],
        as_attachment=False,
        mimetype='application/pdf'
    )


@login_required
def neetpg_admin():
    user = get_user()
    if not user.get('is_admin'):
        flash('Admin access required', 'error')
        return redirect(url_for('dashboard'))
    conn = get_db()
    # ── Filters (sub-tabs): NEET PG/DNB category · document type · publish status ──
    active_doc_type = request.args.get('doc_type', 'all').strip()
    if active_doc_type not in VALID_DOC_TYPES:
        active_doc_type = 'all'
    active_category = request.args.get('category', 'all').strip()
    if active_category not in ('neetpg', 'dnb'):
        active_category = 'all'
    active_status = request.args.get('status', 'all').strip()
    if active_status not in ('published', 'scheduled', 'draft'):
        active_status = 'all'

    def _admin_count(cnds, prms):
        w = (" WHERE " + " AND ".join(cnds)) if cnds else ""
        return conn.execute(f"SELECT COUNT(*) as c FROM neetpg_pdfs{w}", prms).fetchone()['c']

    def _status_conds(status):
        if status == 'published':
            return ["is_published = 1"]
        if status == 'scheduled':                    # queued for auto-publish
            return ["is_published = 0", "auto_schedule = 1"]
        if status == 'draft':                        # unpublished, auto-publish off
            return ["is_published = 0", "auto_schedule = 0"]
        return []

    def _filter_conds(include_dt=True, include_cat=True, include_status=True):
        c, p = [], []
        if include_dt and active_doc_type != 'all':
            c.append("doc_type = ?"); p.append(active_doc_type)
        if include_cat and active_category != 'all':
            c.append("category = ?"); p.append(active_category)
        if include_status and active_status != 'all':
            c += _status_conds(active_status)
        return c, p

    # Combined filter for the listing
    conds, params = _filter_conds()
    flt_where = (" WHERE " + " AND ".join(conds)) if conds else ""

    # Each tab row's counts reflect the OTHER two active filters.
    dc, dp = _filter_conds(include_dt=False)
    doctype_counts = {'all': _admin_count(dc, dp)}
    for key in DOC_TYPE_LABELS:
        doctype_counts[key] = _admin_count(dc + ["doc_type = ?"], dp + [key])
    cc, cp = _filter_conds(include_cat=False)
    category_counts = {'all': _admin_count(cc, cp)}
    for c in ('neetpg', 'dnb'):
        category_counts[c] = _admin_count(cc + ["category = ?"], cp + [c])
    sc, sp = _filter_conds(include_status=False)
    status_counts = {'all': _admin_count(sc, sp)}
    for s in ('published', 'scheduled', 'draft'):
        status_counts[s] = _admin_count(sc + _status_conds(s), sp)

    # Pagination for PDFs (within the active filters)
    page = int(request.args.get('page', 1))
    per_page = 20
    total_pdfs_count = _admin_count(conds, params)
    total_pages = max(1, (total_pdfs_count + per_page - 1) // per_page)
    if page < 1:
        page = 1
    if page > total_pages:
        page = total_pages
    offset = (page - 1) * per_page
    pdfs = conn.execute(
        f"SELECT * FROM neetpg_pdfs{flt_where} ORDER BY upload_date DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()
    leads = conn.execute("SELECT * FROM neetpg_leads ORDER BY created_at DESC").fetchall()
    lead_count = conn.execute("SELECT COUNT(*) as c FROM neetpg_leads").fetchone()['c']
    # Requests
    try:
        doc_requests = conn.execute("SELECT * FROM neetpg_requests ORDER BY created_at DESC").fetchall()
        request_count = conn.execute("SELECT COUNT(*) as c FROM neetpg_requests").fetchone()['c']
        pending_requests = conn.execute("SELECT COUNT(*) as c FROM neetpg_requests WHERE status = 'pending'").fetchone()['c']
    except Exception as req_err:
        logging.error(f"neetpg requests query error: {req_err}")
        try:
            conn.rollback()
        except Exception:
            pass
        doc_requests = []
        request_count = 0
        pending_requests = 0

    # Recent 10 PDFs for request completion dropdown
    recent_pdfs = conn.execute(
        "SELECT id, title, state, category FROM neetpg_pdfs ORDER BY upload_date DESC LIMIT 10"
    ).fetchall()

    # College-name suggestions for the MCC Profile / Bond Doc field (from existing data)
    try:
        college_names = [r['specialty'] for r in conn.execute(
            "SELECT DISTINCT specialty FROM neetpg_pdfs WHERE doc_type IN ('mcc_profile','mcc_bond_doc') AND specialty IS NOT NULL AND specialty != '' ORDER BY specialty"
        ).fetchall()]
    except Exception:
        college_names = []

    # Analytics
    total_visits = conn.execute("SELECT COUNT(*) as c FROM neetpg_page_visits").fetchone()['c']
    today_visits = conn.execute(
        "SELECT COUNT(*) as c FROM neetpg_page_visits WHERE visited_at::date = CURRENT_DATE"
    ).fetchone()['c']
    # Avg daily visits
    first_visit = conn.execute("SELECT MIN(visited_at) as fv FROM neetpg_page_visits").fetchone()['fv']
    if first_visit and total_visits > 0:
        from datetime import datetime as dt2
        if isinstance(first_visit, str):
            first_visit = datetime.strptime(first_visit[:10], '%Y-%m-%d')
        days_since = max((datetime.now() - (first_visit if isinstance(first_visit, datetime) else datetime.combine(first_visit, datetime.min.time()))).days, 1)
        avg_daily = round(total_visits / days_since, 1)
    else:
        avg_daily = 0

    total_downloads = conn.execute("SELECT COALESCE(SUM(download_count),0) as c FROM neetpg_pdfs").fetchone()['c']
    active_pdfs = conn.execute("SELECT COUNT(*) as c FROM neetpg_pdfs WHERE is_active = 1").fetchone()['c']
    published_pdfs = conn.execute("SELECT COUNT(*) as c FROM neetpg_pdfs WHERE is_published = 1").fetchone()['c']
    draft_pdfs = conn.execute("SELECT COUNT(*) as c FROM neetpg_pdfs WHERE is_published = 0 AND is_active = 1 AND auto_schedule = 1").fetchone()['c']
    leads_today = conn.execute(
        "SELECT COUNT(*) as c FROM neetpg_leads WHERE created_at::date = CURRENT_DATE"
    ).fetchone()['c']

    analytics = {
        'total_visits': total_visits,
        'today_visits': today_visits,
        'avg_daily_visits': avg_daily,
        'total_downloads': total_downloads,
        'active_pdfs': active_pdfs,
        'published_pdfs': published_pdfs,
        'draft_pdfs': draft_pdfs,
        'leads_today': leads_today
    }

    # ── Compute scheduled time slots per TAB (each category+doc_type publishes
    #    independently at 10am/1pm/5pm IST, so a draft's slot = its rank within its tab) ──
    schedule_map = _neetpg_schedule_map(conn)

    conn.close()
    return render_template('college/neetpg_admin.html', pdfs=pdfs, leads=leads, lead_count=lead_count, user=user, analytics=analytics,
                           page=page, total_pages=total_pages, total_pdfs_count=total_pdfs_count, schedule_map=schedule_map,
                           doc_requests=doc_requests, request_count=request_count, pending_requests=pending_requests,
                           recent_pdfs=recent_pdfs,
                           doc_types=NEETPG_DOC_TYPES, doc_type_labels=DOC_TYPE_LABELS,
                           college_names=college_names,
                           neetpg_courses=_NEETPG_COURSES, dnb_courses=_DNB_COURSES,
                           specialty_alias=_SPECIALTY_ALIAS,
                           active_doc_type=active_doc_type, doctype_counts=doctype_counts,
                           active_category=active_category, category_counts=category_counts,
                           active_status=active_status, status_counts=status_counts,
                    active_section='colleges')


@login_required
def neetpg_upload():
    user = get_user()
    if not user.get('is_admin'):
        return jsonify({'error': 'Admin required'}), 403
    edit_id = request.form.get('edit_id', '').strip()   # set → update in place, keep stats
    title = request.form.get('title', '').strip()
    category = request.form.get('category', 'neetpg')
    doc_type = request.form.get('doc_type', 'cutoff').strip()
    if doc_type not in VALID_DOC_TYPES:
        doc_type = 'cutoff'
    # For MCC College Profile the "specialty" field holds the College Name.
    specialty = request.form.get('specialty', '').strip()
    state = request.form.get('state', 'AIQ MCC').strip()
    # Quota & Category applies to Cut Off only (e.g. 'GM', 'OBC', 'Paid Seats').
    quota_category = request.form.get('quota_category', '').strip() if doc_type == 'cutoff' else ''
    file = request.files.get('pdf_file')
    has_file = bool(file and file.filename)
    if not specialty:
        label = 'college name' if doc_type in COLLEGE_NAME_DOC_TYPES else 'specialty'
        flash(f'{label.capitalize()} is required', 'error')
        return redirect(url_for('neetpg_admin'))
    # A file is required when adding; optional when editing (keep the current one).
    if not edit_id and not has_file:
        flash('A PDF file is required', 'error')
        return redirect(url_for('neetpg_admin'))
    if has_file and not file.filename.lower().endswith('.pdf'):
        flash('Only PDF files are allowed', 'error')
        return redirect(url_for('neetpg_admin'))
    # Title is auto-generated from the fields; only used as-is if admin overrode it.
    if not title:
        title = _build_neetpg_title(doc_type, category, state, specialty, quota_category)
    file_data = file_size = file_name = None
    if has_file:
        file_data = file.read()
        file_size = len(file_data)
        file_name = file.filename
    try:
        conn = get_db()
        if edit_id:
            # ── Update in place — keeps id, download_count, is_published, published_at ──
            if has_file:
                conn.execute(
                    "UPDATE neetpg_pdfs SET title=?, category=?, doc_type=?, specialty=?, quota_category=?, state=?, file_name=?, file_size=?, file_data=? WHERE id=?",
                    (title, category, doc_type, specialty, quota_category, state, file_name, file_size, file_data, int(edit_id))
                )
            else:
                conn.execute(
                    "UPDATE neetpg_pdfs SET title=?, category=?, doc_type=?, specialty=?, quota_category=?, state=? WHERE id=?",
                    (title, category, doc_type, specialty, quota_category, state, int(edit_id))
                )
            conn.commit()
            flash('PDF updated. Stats preserved.', 'success')
        else:
            conn.execute(
                "INSERT INTO neetpg_pdfs (title, category, doc_type, specialty, quota_category, file_name, file_size, file_data, state, is_published, auto_schedule) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 1)",
                (title, category, doc_type, specialty, quota_category, file_name, file_size, file_data, state)
            )
            conn.commit()
            flash('PDF uploaded as draft. Use Publish to make it live.', 'success')
        conn.close()
    except Exception as e:
        logging.error(f"neetpg_upload: {e}")
        flash('Save failed', 'error')
    return redirect(url_for('neetpg_admin'))


@login_required
def neetpg_toggle(pdf_id):
    user = get_user()
    if not user.get('is_admin'):
        return jsonify({'error': 'Admin required'}), 403
    conn = get_db()
    pdf = conn.execute("SELECT is_active FROM neetpg_pdfs WHERE id = ?", (pdf_id,)).fetchone()
    if pdf:
        new_status = 0 if pdf['is_active'] else 1
        conn.execute("UPDATE neetpg_pdfs SET is_active = ? WHERE id = ?", (new_status, pdf_id))
        conn.commit()
    conn.close()
    return jsonify({'success': True})


@login_required
def neetpg_publish(pdf_id):
    user = get_user()
    if not user.get('is_admin'):
        return jsonify({'error': 'Admin required'}), 403
    conn = get_db()
    pdf = conn.execute("SELECT is_published, title, state FROM neetpg_pdfs WHERE id = ?", (pdf_id,)).fetchone()
    if pdf:
        if pdf['is_published']:
            # Unpublish
            conn.execute("UPDATE neetpg_pdfs SET is_published = 0, published_at = NULL WHERE id = ?", (pdf_id,))
            print(f"[PUBLISH] Unpublished PDF id={pdf_id}", flush=True)
        else:
            # Publish now
            conn.execute("UPDATE neetpg_pdfs SET is_published = 1, published_at = CURRENT_TIMESTAMP, is_active = 1 WHERE id = ?", (pdf_id,))
            print(f"[PUBLISH] Published PDF id={pdf_id} title='{pdf['title']}' — triggering WA notify", flush=True)
            # Send WhatsApp notification to all leads
            neetpg_wa_notify_leads(pdf['title'], pdf.get('state', ''))
        conn.commit()
    conn.close()
    return jsonify({'success': True})


@login_required
def neetpg_toggle_schedule(pdf_id):
    user = get_user()
    if not user.get('is_admin'):
        return jsonify({'error': 'Admin required'}), 403
    conn = get_db()
    pdf = conn.execute("SELECT auto_schedule FROM neetpg_pdfs WHERE id = ?", (pdf_id,)).fetchone()
    if pdf:
        new_val = 0 if pdf['auto_schedule'] else 1
        conn.execute("UPDATE neetpg_pdfs SET auto_schedule = ? WHERE id = ?", (new_val, pdf_id))
        conn.commit()
    conn.close()
    return jsonify({'success': True})


@login_required
def neetpg_wa_blast(pdf_id):
    """Manually trigger WhatsApp notification for a specific published PDF."""
    user = get_user()
    if not user.get('is_admin'):
        return jsonify({'error': 'Admin required'}), 403
    conn = get_db()
    pdf = conn.execute("SELECT title, state, is_published FROM neetpg_pdfs WHERE id = ?", (pdf_id,)).fetchone()
    conn.close()
    if not pdf:
        return jsonify({'error': 'PDF not found'}), 404
    if not pdf['is_published']:
        return jsonify({'error': 'PDF must be published before sending notifications'}), 400
    neetpg_wa_notify_leads(pdf['title'], pdf.get('state', ''))
    return jsonify({'success': True, 'message': 'WhatsApp notifications are being sent in background'})


@login_required
def neetpg_delete(pdf_id):
    user = get_user()
    if not user.get('is_admin'):
        return jsonify({'error': 'Admin required'}), 403
    conn = get_db()
    conn.execute("DELETE FROM neetpg_pdfs WHERE id = ?", (pdf_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@login_required
def neetpg_delete_lead(lead_id):
    user = get_user()
    if not user.get('is_admin'):
        return jsonify({'error': 'Admin required'}), 403
    conn = get_db()
    conn.execute("DELETE FROM neetpg_leads WHERE id = ?", (lead_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@login_required
def neetpg_delete_request(req_id):
    user = get_user()
    if not user.get('is_admin'):
        return jsonify({'error': 'Admin required'}), 403
    conn = get_db()
    conn.execute("DELETE FROM neetpg_requests WHERE id = ?", (req_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@login_required
def neetpg_quota_backfill():
    """Preview (GET) / apply (POST) filling quota_category on existing Cut-Off
    files by parsing their stored filenames. Only fills BLANK quotas — never
    overwrites a value already set — and keeps all stats (id/downloads/publish)."""
    user = get_user()
    if not user.get('is_admin'):
        return jsonify({'error': 'Admin required'}), 403
    conn = get_db()
    rows = conn.execute(
        "SELECT id, title, file_name, quota_category FROM neetpg_pdfs WHERE doc_type = 'cutoff' ORDER BY upload_date DESC"
    ).fetchall()
    items = []
    for r in rows:
        current = (r['quota_category'] or '').strip()
        detected = _parse_quota_from_filename(r['file_name'])
        will_update = bool(detected) and not current
        items.append({'id': r['id'], 'title': r['title'], 'file_name': r['file_name'] or '',
                      'current': current, 'detected': detected, 'will_update': will_update})

    if request.method == 'POST':
        updated = 0
        for it in items:
            if it['will_update']:
                conn.execute("UPDATE neetpg_pdfs SET quota_category = ? WHERE id = ?",
                             (it['detected'], it['id']))
                updated += 1
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'updated': updated})

    conn.close()
    return jsonify({
        'total': len(items),
        'to_update': sum(1 for it in items if it['will_update']),
        'items': items,
    })


@login_required
def neetpg_college_suggest():
    """Type-ahead for the MCC Profile / Bond Doc College Name field, from the
    official AIQ-MCC college lists (per category + doc_type, optional state)."""
    user = get_user()
    if not user.get('is_admin'):
        return jsonify([]), 403
    cat = 'dnb' if request.args.get('cat') == 'dnb' else 'neetpg'
    doc = request.args.get('doc', 'mcc_profile')
    if doc not in ('mcc_profile', 'mcc_bond_doc'):
        doc = 'mcc_profile'
    state = (request.args.get('state', '') or '').strip().lower()
    q = (request.args.get('q', '') or '').strip().lower()
    rows = _MCC_COLLEGES.get(f"{cat}|{doc}", [])
    use_state = state and state not in ('aiq mcc', 'all india/mcc')
    out, seen = [], set()
    for st, nm in rows:
        if use_state and (st or '').strip().lower() != state:
            continue
        if q and q not in nm.lower():
            continue
        k = nm.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(nm)
        if len(out) >= 60:
            break
    return jsonify(out)


@login_required
def neetpg_all_uploads():
    """Persistent, searchable listing of every uploaded PDF (all doc types)."""
    user = get_user()
    if not user.get('is_admin'):
        flash('Admin access required', 'error')
        return redirect(url_for('dashboard'))
    conn = get_db()
    rows = conn.execute(
        "SELECT id, title, category, doc_type, specialty, quota_category, state, is_published, download_count FROM neetpg_pdfs ORDER BY upload_date DESC"
    ).fetchall()
    conn.close()
    return render_template('college/neetpg_all_uploads.html', rows=rows,
                           doc_type_labels=DOC_TYPE_LABELS, total=len(rows),
                           user=user, active_section='colleges')


@login_required
def neetpg_bulk_upload():
    """Upload many PDFs at once; each is auto-filled from its filename
    (doc type, state, category, snapped specialty, quota, title). Skips
    unrecognised names. Optionally publishes on upload (no WA blast)."""
    user = get_user()
    if not user.get('is_admin'):
        return jsonify({'error': 'Admin required'}), 403
    files = request.files.getlist('pdf_files')
    publish = request.form.get('publish') == '1'
    files = [f for f in files if f and f.filename]
    if not files:
        flash('No files selected for bulk upload', 'error')
        return redirect(url_for('neetpg_admin'))
    conn = get_db()
    results = []
    counts = {'added': 0, 'exists': 0, 'skipped': 0, 'error': 0}
    for f in files:
        name = f.filename
        if not name.lower().endswith('.pdf'):
            results.append({'name': name, 'type': '', 'status': 'skipped', 'detail': 'not a PDF'})
            counts['skipped'] += 1
            continue
        parsed = _parse_neetpg_filename(name)
        if not parsed:
            results.append({'name': name, 'type': '', 'status': 'skipped', 'detail': 'filename not recognised'})
            counts['skipped'] += 1
            continue
        dt_label = DOC_TYPE_LABELS.get(parsed['doc_type'], parsed['doc_type'])
        # Duplicate guard — skip files already uploaded (same filename).
        try:
            dup = conn.execute("SELECT id FROM neetpg_pdfs WHERE LOWER(file_name) = LOWER(?)", (name,)).fetchone()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
            dup = None
        if dup:
            results.append({'name': name, 'type': dt_label, 'status': 'exists', 'detail': 'already uploaded'})
            counts['exists'] += 1
            continue
        try:
            data = f.read()
            title = _build_neetpg_title(parsed['doc_type'], parsed['category'],
                                        parsed['state'], parsed['specialty'], parsed['quota'])
            if publish:
                conn.execute(
                    "INSERT INTO neetpg_pdfs (title, category, doc_type, specialty, quota_category, file_name, file_size, file_data, state, is_published, published_at, auto_schedule) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, CURRENT_TIMESTAMP, 0)",
                    (title, parsed['category'], parsed['doc_type'], parsed['specialty'], parsed['quota'], name, len(data), data, parsed['state'])
                )
            else:
                conn.execute(
                    "INSERT INTO neetpg_pdfs (title, category, doc_type, specialty, quota_category, file_name, file_size, file_data, state, is_published, auto_schedule) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 1)",
                    (title, parsed['category'], parsed['doc_type'], parsed['specialty'], parsed['quota'], name, len(data), data, parsed['state'])
                )
            conn.commit()
            detail = parsed['specialty'] + (' · ' + parsed['quota'] if parsed['quota'] else '')
            results.append({'name': name, 'type': dt_label, 'status': ('published' if publish else 'added'), 'detail': detail})
            counts['added'] += 1
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            results.append({'name': name, 'type': dt_label, 'status': 'error', 'detail': 'save failed'})
            counts['error'] += 1
            logging.error(f"neetpg_bulk_upload {name}: {e}")
    conn.close()
    return render_template('college/neetpg_bulk_result.html',
                           results=results, counts=counts, published=publish,
                           user=user, active_section='colleges')


@login_required
def neetpg_bulk_delete():
    """Preview (GET) / apply (POST) a scoped bulk delete of PDF records, filtered
    by doc_type (and optional exact state). Admin power tool for clearing a batch
    before a fresh bulk upload."""
    user = get_user()
    if not user.get('is_admin'):
        return jsonify({'error': 'Admin required'}), 403
    src = request.form if request.method == 'POST' else request.args
    doc_type = (src.get('doc_type', '') or '').strip()
    state = (src.get('state', '') or '').strip()
    conds, params = [], []
    if doc_type in VALID_DOC_TYPES:
        conds.append("doc_type = ?")
        params.append(doc_type)
    if state:
        conds.append("state = ?")
        params.append(state)
    # Safety: never allow an unscoped delete-everything.
    if not conds:
        return jsonify({'error': 'Pick a document type (and optionally a state) first'}), 400
    where = " WHERE " + " AND ".join(conds)
    conn = get_db()
    if request.method == 'POST':
        n = conn.execute(f"SELECT COUNT(*) as c FROM neetpg_pdfs{where}", params).fetchone()['c']
        conn.execute(f"DELETE FROM neetpg_pdfs{where}", params)
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'deleted': n})
    rows = conn.execute(
        f"SELECT title, state, quota_category FROM neetpg_pdfs{where} ORDER BY upload_date DESC", params
    ).fetchall()
    conn.close()
    return jsonify({'total': len(rows), 'items': [dict(r) for r in rows[:150]]})


@login_required
def neetpg_find_missing():
    """Given a list of filenames (from the admin's folder), report which are NOT
    present in neetpg_pdfs (by file_name) — i.e. which files didn't get uploaded.
    Only filenames are sent, not the files."""
    user = get_user()
    if not user.get('is_admin'):
        return jsonify({'error': 'Admin required'}), 403
    data = request.get_json(silent=True) or {}
    names = data.get('names', [])
    if not isinstance(names, list):
        return jsonify({'error': 'names must be a list'}), 400
    conn = get_db()
    rows = conn.execute("SELECT file_name FROM neetpg_pdfs").fetchall()
    conn.close()
    uploaded = set((r['file_name'] or '').strip() for r in rows)
    uploaded_lc = set(n.lower() for n in uploaded)
    missing = []
    seen = set()
    for n in names:
        nm = (n or '').strip()
        if not nm or nm.lower() in seen:
            continue
        seen.add(nm.lower())
        if nm in uploaded or nm.lower() in uploaded_lc:
            continue
        missing.append(nm)
    return jsonify({'checked': len(seen), 'uploaded_total': len(uploaded), 'missing': missing})


@login_required
def neetpg_export_leads():
    user = get_user()
    if not user.get('is_admin'):
        flash('Admin access required', 'error')
        return redirect(url_for('dashboard'))
    conn = get_db()
    leads = conn.execute("SELECT name, whatsapp, email, source, created_at FROM neetpg_leads ORDER BY created_at DESC").fetchall()
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = "NEET PG Leads"
    headers = ['Name', 'WhatsApp', 'Email', 'Source', 'Captured At']
    header_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    for row, lead in enumerate(leads, 2):
        ws.cell(row=row, column=1, value=lead['name'])
        ws.cell(row=row, column=2, value=lead['whatsapp'] or '')
        ws.cell(row=row, column=3, value=lead['email'] or '')
        ws.cell(row=row, column=4, value=lead['source'] or 'whatsapp')
        ws.cell(row=row, column=5, value=str(lead['created_at']))
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 22
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='neetpg_leads.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def neetpg_wa_notify_leads(pdf_title, pdf_state):
    """Send WhatsApp template notification to all NEET PG leads when a PDF is published.
    Uses the 'joincommunity' template (ID: 757749590683369) with 3 placeholders:
    {{1}} = doctor name, {{2}} = PDF title/state, {{3}} = access link.
    Runs in background thread to avoid blocking the publish response."""
    import threading

    def _send():
        try:
            import requests as http_requests
            print(f"[WA_NOTIFY] Starting WhatsApp notification for PDF: {pdf_title}", flush=True)
            infobip_key = os.environ.get('INFOBIP_API_KEY', '')
            infobip_base = os.environ.get('INFOBIP_BASE_URL', '')
            if not infobip_key or not infobip_base:
                print(f"[WA_NOTIFY] ERROR: Infobip credentials not configured — key={bool(infobip_key)}, base={bool(infobip_base)}", flush=True)
                return

            wa_headers = {
                "Authorization": f"App {infobip_key}",
                "Content-Type": "application/json"
            }
            tmpl_url = f"https://{infobip_base}/whatsapp/1/message/template"

            conn = get_db()
            leads = conn.execute(
                "SELECT DISTINCT name, whatsapp FROM neetpg_leads WHERE whatsapp IS NOT NULL AND whatsapp != ''"
            ).fetchall()
            conn.close()

            if not leads:
                print("[WA_NOTIFY] No leads with WhatsApp numbers found, skipping", flush=True)
                return

            print(f"[WA_NOTIFY] Found {len(leads)} leads to notify", flush=True)
            pdf_label = f"{pdf_title}" if pdf_title else pdf_state
            access_link = "https://goocampus.org/neet-pg-2025"
            sent_count = 0
            fail_count = 0

            for lead in leads:
                phone = lead['whatsapp'].strip()
                if len(phone) < 10:
                    continue
                phone_clean = phone[-10:]
                to_number = f"91{phone_clean}"
                doctor_name = lead['name'] or 'Doctor'

                payload = {
                    "messages": [{
                        "from": "15558246314",
                        "to": to_number,
                        "content": {
                            "templateName": "joincommunity",
                            "templateData": {
                                "body": {
                                    "placeholders": [doctor_name, pdf_label, access_link]
                                },
                                "buttons": [
                                    {"type": "URL", "parameter": access_link}
                                ]
                            },
                            "language": "en_IN"
                        }
                    }]
                }

                try:
                    resp = http_requests.post(tmpl_url, json=payload, headers=wa_headers, timeout=15)
                    resp_body = resp.text[:500] if resp.text else ''
                    if resp.status_code < 400:
                        sent_count += 1
                        print(f"[WA_NOTIFY] OK {phone_clean}: {resp.status_code} — {resp_body[:200]}", flush=True)
                    else:
                        fail_count += 1
                        print(f"[WA_NOTIFY] FAIL {phone_clean}: HTTP {resp.status_code} — {resp.text[:300]}", flush=True)
                except Exception as e:
                    fail_count += 1
                    print(f"[WA_NOTIFY] ERR {phone_clean}: {e}", flush=True)

            print(f"[WA_NOTIFY] DONE PDF '{pdf_label}' — sent={sent_count}, failed={fail_count}, total_leads={len(leads)}", flush=True)
        except Exception as e:
            print(f"[WA_NOTIFY] CRITICAL ERROR: {e}", flush=True)
            import traceback
            traceback.print_exc()

    thread = threading.Thread(target=_send, daemon=True)
    thread.start()
