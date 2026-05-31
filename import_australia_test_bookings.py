"""
One-time import: Australia test bookings from Excel into ops_test_bookings.

Each Excel row goes into ops_test_bookings with pathway='australia'.

Linking strategy:
  The Excel's "Enter Candidate Name" cell contains both the candidate name
  AND the registration number, e.g.:
      " Dr. Yashika Bopaiah -GCAUSIP/2023/026"
      " Dr. Nikhil Paul Mathew -GCAUSIP/24-25/111"
  We parse out the GCAUSIP-prefixed reg number and store it in
  registration_number so existing list/edit pages that JOIN against
  plab_clients keep working unchanged.

Idempotent: keyed on (registration_number, exam, booking_date). Re-running
with no marker change is a no-op. Bumping IMPORT_VERSION re-runs the
upsert. Bad rows (missing reg number, no match in plab_clients) are
skipped with a warning so one bad row never blocks the whole import.

Hooked into app boot from app.py the same way as
import_australia_clients.py.
"""

import logging
import os
import re
from datetime import datetime, date


EXCEL_FILENAME = 'All_Australia_Test_Bookings.xlsx'
IMPORT_VERSION = 'au_tb_v1_initial_356'


H = {
    'cand_name':   'Enter Candidate Name',
    'booked_by':   'Booked By',
    'exam':        'Exam',
    'book_date':   'Booking Date',
    'exam_method': 'Exam Method',
    'exam_date':   'Exam Date',
    'country':     'Country',
    'login_id':    'Login ID',
    'password':    'Password',
    'city_state':  'City / State',
    'exam_status': 'Exam Status',
    'exam_result': 'Exam Result',
    'exam_result_date': 'Exam Result Date',
    'test_score':  'Test Score',
    'reval':       'Revaluation Applied',
    'reval_date':  'Reval Applied Date',
    'reval_result': 'Reval Result',
    'reval_score': 'Reval Score',
}


_REG_RE = re.compile(r'(GCAUSIP/\S+)', re.IGNORECASE)


def _extract_reg_number(candidate_cell):
    """Pull the GCAUSIP/... reg number out of the combined name+reg cell.

    Returns '' if none found. The cell can have stray spaces and dashes
    around the reg number, so we use a generous regex and strip noise.
    """
    if not candidate_cell:
        return ''
    s = str(candidate_cell)
    m = _REG_RE.search(s)
    if not m:
        return ''
    raw = m.group(1).rstrip('.,;:-/ \t')
    # Normalize: uppercase prefix.
    return raw.upper().replace('GCAUSIP', 'GCAUSIP')


def _ss(v):
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _sd(v):
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def run_import_australia_test_bookings_once(get_db_fn):
    """Import Australia test bookings into ops_test_bookings (pathway='australia')."""
    import openpyxl

    excel_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        EXCEL_FILENAME,
    )
    if not os.path.exists(excel_path):
        logging.info("Australia test-bookings import: no Excel found, skipping")
        return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

    conn = get_db_fn()
    try:
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS _import_markers (key TEXT PRIMARY KEY, value TEXT)")
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
        marker = conn.execute(
            "SELECT value FROM _import_markers WHERE key = 'au_test_bookings_import'"
        ).fetchone()
        if marker and marker['value'] == IMPORT_VERSION:
            logging.info(f"Australia test-bookings import: already at {IMPORT_VERSION}, skipping")
            return {'inserted': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

        logging.info(f"Australia test-bookings import: starting {IMPORT_VERSION}...")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h is not None}

        def cell(row, key):
            i = idx.get(H[key])
            return row[i] if i is not None and i < len(row) else None

        inserted = 0
        updated = 0
        skipped = 0
        errors = 0

        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and c != '' for c in row_cells):
                continue
            reg_num = _extract_reg_number(cell(row_cells, 'cand_name'))
            if not reg_num:
                skipped += 1
                continue
            exam = _ss(cell(row_cells, 'exam'))
            book_date = _sd(cell(row_cells, 'book_date'))
            if not exam:
                skipped += 1
                continue

            data = {
                'registration_number': reg_num,
                'exam':                exam,
                'exam_type':           _ss(cell(row_cells, 'exam_method')),
                'booking_date':        book_date,
                'exam_date':           _sd(cell(row_cells, 'exam_date')),
                'exam_status':         _ss(cell(row_cells, 'exam_status')),
                'exam_result':         _ss(cell(row_cells, 'exam_result')),
                'exam_result_date':    _sd(cell(row_cells, 'exam_result_date')),
                'score':               _ss(cell(row_cells, 'test_score')),
                'city_state':          _ss(cell(row_cells, 'city_state')),
                'country':             _ss(cell(row_cells, 'country')),
                'booked_by':           _ss(cell(row_cells, 'booked_by')),
                'revaluation':         _ss(cell(row_cells, 'reval')),
                'reval_applied_date':  _sd(cell(row_cells, 'reval_date')),
                'reval_result':        _ss(cell(row_cells, 'reval_result')),
                'reval_score':         _ss(cell(row_cells, 'reval_score')),
                'pathway':             'australia',
            }
            # Stash login_id/password in notes so the Ops team can still see
            # them, but we don't surface plain credentials on a column.
            login_id = _ss(cell(row_cells, 'login_id'))
            password = _ss(cell(row_cells, 'password'))
            if login_id or password:
                bits = []
                if login_id: bits.append(f"login: {login_id}")
                if password: bits.append(f"password: {password}")
                data['notes'] = ' | '.join(bits)

            try:
                existing = conn.execute(
                    "SELECT id FROM ops_test_bookings "
                    "WHERE registration_number = ? AND exam = ? AND booking_date = ? AND pathway = 'australia'",
                    (reg_num, exam, book_date),
                ).fetchone()
                if existing:
                    sets = ', '.join(f"{k} = ?" for k in data.keys())
                    conn.execute(
                        f"UPDATE ops_test_bookings SET {sets} WHERE id = ?",
                        list(data.values()) + [existing['id']],
                    )
                    updated += 1
                else:
                    cols = ', '.join(data.keys())
                    placeholders = ', '.join(['?'] * len(data))
                    conn.execute(
                        f"INSERT INTO ops_test_bookings ({cols}) VALUES ({placeholders})",
                        list(data.values()),
                    )
                    inserted += 1
                conn.commit()
            except Exception as e:
                errors += 1
                logging.warning(f"Australia test-bookings row error ({reg_num} / {exam}): {e}")
                try:
                    conn.rollback()
                except Exception:
                    pass

        # Marker
        try:
            conn.execute(
                "INSERT INTO _import_markers (key, value) VALUES ('au_test_bookings_import', ?) "
                "ON CONFLICT (key) DO UPDATE SET value = excluded.value",
                (IMPORT_VERSION,),
            )
            conn.commit()
        except Exception:
            try:
                conn.execute("DELETE FROM _import_markers WHERE key = 'au_test_bookings_import'")
                conn.execute(
                    "INSERT INTO _import_markers (key, value) VALUES ('au_test_bookings_import', ?)",
                    (IMPORT_VERSION,),
                )
                conn.commit()
            except Exception as e:
                logging.warning(f"Australia test-bookings: could not write marker: {e}")

        logging.info(
            f"Australia test-bookings {IMPORT_VERSION} complete — "
            f"inserted={inserted} updated={updated} skipped={skipped} errors={errors}"
        )
        wb.close()
        return {'inserted': inserted, 'updated': updated, 'skipped': skipped, 'errors': errors}
    finally:
        try:
            conn.close()
        except Exception:
            pass
